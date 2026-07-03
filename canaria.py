"""
Global Multi-Asset Risk Scoring Model (단일 파일 버전)
=========================================================================
위험 민감 자산(canary) 13종을 가중 합산해 일별 Risk On/Off 레짐을 산출.

기존 risk_scoring_model.py(본체) + risk_scoring_model_2.py(v2 어댑터)를
하나로 통합하고 다음을 제거한 재작성 버전:
  - Bloomberg(xbbg) 데이터 경로 전부 → yfinance + FRED 단일 소스
  - 합성 데이터/데모 모드 전부 → 데이터 미수신 시 즉시 중단(fail-fast)
  - 단종된 ISM PMI(NAPM) 매크로 → 제거 (CPI·실업률 2종만 사용)
  - FRED API 키 하드코딩 → 환경변수(FRED_API_KEY)만, 없으면 에러

데이터 소스
----------
  • Yahoo Finance (yfinance) — 가격/지수/변동성 11종 + S&P500
  • FRED          (fredapi)  — HY OAS + 매크로 2종 (CPI, 실업률)

시그널 13종 (조건 충족=+가중치, 미충족=0 / SKEW만 역방향 페널티)
  가격(MA10): Dollar Index, EM Bond, EM Currency, US Treasury Agg,
              TIPS, Dev.Market Eq, Gold, Commodity
  변동성(MA60): VIX, VVIX
  신용(MA60): US HY OAS
  매크로(임계값+ffill): CPI YoY
  꼬리위험(절대임계값): SKEW Penalty (>140 → -2점)

복합 스코어 = Σ(가중 시그널)  →  레짐(in-sample 백분위: ≥p70 On / ≤p30 Off)

실행:  py -3.12 risk_scoring_model.py
환경변수:  FRED_API_KEY 필수
=========================================================================
"""
from __future__ import annotations

import os
import sys
import time
import warnings
from datetime import datetime, timedelta

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.gridspec import GridSpec
from matplotlib.lines import Line2D
from matplotlib.patches import Patch

warnings.filterwarnings("ignore")

# Windows 콘솔 UTF-8
try:
    sys.stdout.reconfigure(encoding="utf-8")   # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")   # type: ignore[attr-defined]
except Exception:
    pass

# ── 필수 외부 라이브러리 (없으면 즉시 중단) ──────────────────────────────────
try:
    import yfinance as yf
except ImportError as exc:
    raise ImportError("yfinance가 필요합니다:  pip install yfinance") from exc

try:
    from fredapi import Fred
except ImportError as exc:
    raise ImportError("fredapi가 필요합니다:  pip install fredapi") from exc


# ── .env 로더 (python-dotenv 없이 직접 파싱) ─────────────────────────────────
def load_dotenv(path: str | None = None) -> None:
    """
    .env 파일을 읽어 os.environ에 주입한다 (이미 설정된 환경변수는 덮어쓰지 않음).

    path 미지정 시 이 스크립트 폴더에서 시작해 상위로 거슬러 올라가며 .env를 탐색.
    → igis 폴더(또는 상위)에 .env 하나만 두면 여러 프로젝트가 공유한다.

    형식 예시 (.env):
        FRED_API_KEY=7469ba1eab0c6971b8a3802634163bc9
        # 주석 가능
    """
    candidates = []
    if path:
        candidates.append(path)
    else:
        here = os.path.dirname(os.path.abspath(__file__))
        # 현재 폴더부터 최대 4단계 상위까지 .env 탐색
        d = here
        for _ in range(5):
            candidates.append(os.path.join(d, ".env"))
            parent = os.path.dirname(d)
            if parent == d:
                break
            d = parent

    for env_path in candidates:
        if not os.path.exists(env_path):
            continue
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, val = line.partition("=")
                    key, val = key.strip(), val.strip().strip('"').strip("'")
                    # 이미 설정된 실제 환경변수가 우선 (덮어쓰지 않음)
                    if key and key not in os.environ:
                        os.environ[key] = val
        except Exception:
            pass
        return   # 첫 번째로 찾은 .env만 사용


# 모듈 로드 시 자동으로 .env 적용
load_dotenv()


# ══════════════════════════════════════════════════════════════════════════════

class RiskScoringModel:
    """
    위험 민감 자산 13종을 가중 합산해 일별 Risk On/Off 레짐을 산출하는 모델.
    데이터: yfinance(가격) + FRED(스프레드·매크로). 데이터 미수신 시 중단.
    """

    # ── 자산 유니버스 ─────────────────────────────────────────────────────────
    # PRICE_ASSETS: yf_ticker → (ma_period, weight, direction, label)
    #   direction "above": 가격 > MA → Risk On / "below": 가격 < MA → Risk On
    PRICE_ASSETS: dict[str, tuple] = {
        "DX-Y.NYB": (10, 2.0, "below", "Dollar Index"),     # 달러 약세 → risk-on, 고감도
        "EMB":      (10, 1.0, "above", "EM Bond"),
        "CEW":      (10, 1.0, "above", "EM Currency"),
        "BND":      (10, 1.0, "below", "US Treasury Agg"),   # 금리 상승 맥락 → risk-on
        "TIP":      (10, 1.0, "below", "TIPS"),
        "VEA":      (10, 1.0, "above", "Dev. Market Eq"),
        "GLD":      (10, 1.0, "below", "Gold"),              # 안전자산 유출 → risk-on
        "DBC":      (10, 1.0, "above", "Commodity"),
        "^VIX":     (60, 2.0, "below", "VIX"),               # 고감도
        "^VVIX":    (60, 1.0, "below", "VVIX"),
    }

    # SKEW: MA 미사용, 절대 임계값 기반 페널티
    SKEW_TICKER:    str   = "^SKEW"
    SKEW_THRESHOLD: float = 140.0
    SKEW_PENALTY:   float = -2.0

    # SPREAD_ASSETS: label_key → (fred_id, ma_period, weight, direction, label)
    SPREAD_ASSETS: dict[str, tuple] = {
        "US HY OAS": ("BAMLH0A0HYM2", 60, 2.0, "below", "US HY OAS"),
    }

    # MACRO_ASSETS: label_key → (fred_id, condition, threshold, weight, label, is_yoy)
    #   ISM PMI(NAPM)는 FRED 단종으로 제거. CPI(YoY 계산) + 실업률만 사용.
    MACRO_ASSETS: dict[str, tuple] = {
        "CPI YoY":      ("CPIAUCNS", "below", 3.0, 1.0, "CPI YoY", True),
        "Unemployment": ("UNRATE",   "below", 4.0, 1.0, "Unemployment", False),
    }

    SPX_TICKER: str = "^GSPC"

    # FRED 월간 시리즈 발표 지연(참조월말 + N일) — look-ahead bias 방지
    FRED_RELEASE_LAG_DAYS: dict[str, int] = {
        "CPIAUCNS": 15,   # CPI: 다음 달 10–15일
        "UNRATE":    7,   # 실업률: 다음 달 첫 금요일
    }

    # ── 레짐 분류 / 워밍업 / FRED 재시도 ──────────────────────────────────────
    RISK_ON_PCT:  int = 70      # score ≥ p70 → Risk On
    RISK_OFF_PCT: int = 30      # score ≤ p30 → Risk Off
    MA_WARMUP_DAYS: int = 120   # 60일 MA + 버퍼

    FRED_MAX_RETRIES: int = 4
    FRED_RETRY_BACKOFF_SEC: float = 1.5
    MACRO_STALENESS_WARN_DAYS: int = 60

    PALETTE: dict[str, str] = {
        "Risk On": "#27ae60", "Transition": "#e67e22", "Risk Off": "#c0392b",
        "score": "#2c3e50", "score_ma": "#8e44ad", "spx": "#2980b9",
    }

    # ═════════════════════════════════════════════════════════════════════════
    def __init__(self, fred_api_key: str | None = None) -> None:
        key = fred_api_key or os.environ.get("FRED_API_KEY")
        if not key:
            raise RuntimeError(
                "FRED API 키가 없습니다. 다음 중 하나로 설정하세요:\n"
                "  1) .env 파일(권장): igis 폴더에 .env 생성 후  FRED_API_KEY=your_key\n"
                "  2) 환경변수(PowerShell):  $env:FRED_API_KEY=\"your_key\"\n"
                "  3) 환경변수(mac/linux):   export FRED_API_KEY=your_key"
            )
        self._fred = Fred(api_key=key)

        self._full_price:  pd.DataFrame | None = None
        self._full_spread: pd.DataFrame | None = None
        self._full_macro:  pd.DataFrame | None = None
        self._start_date:  str | None = None
        self._end_date:    str | None = None

        self.signals:  pd.DataFrame | None = None
        self.scores:   pd.Series     | None = None
        self.regimes:  pd.Series     | None = None
        self.spx_data: pd.DataFrame  | None = None

    # ═════════════════════════════════════════════════════════════════════════
    # FRED 재시도 래퍼
    # ═════════════════════════════════════════════════════════════════════════
    def _fred_get_series(self, series_id: str, **kwargs) -> pd.Series | None:
        last_exc: Exception | None = None
        for attempt in range(self.FRED_MAX_RETRIES):
            try:
                return self._fred.get_series(series_id, **kwargs)
            except Exception as exc:
                msg = str(exc)
                last_exc = exc
                if "does not exist" in msg or "Bad Request" in msg:
                    return None
                if attempt < self.FRED_MAX_RETRIES - 1:
                    time.sleep(self.FRED_RETRY_BACKOFF_SEC * (2 ** attempt))
        print(f"       [WARN] FRED '{series_id}' 모든 재시도 실패: {last_exc}")
        return None

    # ═════════════════════════════════════════════════════════════════════════
    # 1. fetch_data
    # ═════════════════════════════════════════════════════════════════════════
    def fetch_data(self, start_date: str, end_date: str) -> None:
        """yfinance + FRED에서 데이터 수집. MA 워밍업 위해 start_date 앞으로 확장."""
        self._start_date = start_date
        self._end_date = end_date
        ext_start = (
            datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=self.MA_WARMUP_DAYS)
        ).strftime("%Y-%m-%d")

        sep = "─" * 62
        print(f"\n{sep}\n  YFINANCE + FRED DATA FETCH\n{sep}")
        print(f"  Requested : {start_date} → {end_date}")
        print(f"  Fetch     : {ext_start} → {end_date}  (+{self.MA_WARMUP_DAYS}d warm-up)")

        self._full_price = self._fetch_prices(ext_start, end_date)
        self._full_spread = self._fetch_spreads(ext_start, end_date)
        self._full_macro = self._fetch_macro(ext_start, end_date)
        self.spx_data = self._fetch_spx(start_date, end_date)

        # fail-fast: 가격 데이터가 비면 분석 불가
        if self._full_price is None or self._full_price.empty:
            raise RuntimeError(
                "가격 데이터를 받지 못했습니다(yfinance). 네트워크/티커를 확인하세요."
            )
        print(f"{sep}\n")

    @staticmethod
    def _extract_close(raw: pd.DataFrame, yf_ticker: str) -> pd.Series | None:
        """yfinance.download 결과(구조 가변)에서 (ticker, 'Close') 시리즈 추출."""
        if raw is None or raw.empty:
            return None
        if isinstance(raw.columns, pd.MultiIndex):
            for key in [(yf_ticker, "Close"), ("Close", yf_ticker)]:
                if key in raw.columns:
                    s = raw[key]
                    return s if isinstance(s, pd.Series) else s.iloc[:, 0]
            if "Close" in raw.columns.get_level_values(0):
                sub = raw["Close"]
                if yf_ticker in sub.columns:
                    return sub[yf_ticker]
            return None
        if "Close" in raw.columns:
            return raw["Close"]
        return None

    def _fetch_prices(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """가격 자산 + SKEW의 일별 종가를 yfinance에서 수집."""
        tickers = list(self.PRICE_ASSETS.keys()) + [self.SKEW_TICKER]
        print(f"\n  [1/4] Price assets  ({len(tickers)} tickers via yfinance) …")
        yf_end = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        try:
            raw = yf.download(tickers=tickers, start=ext_start, end=yf_end,
                              interval="1d", auto_adjust=False, progress=False,
                              group_by="ticker", threads=True)
        except Exception as exc:
            print(f"       [WARN] yfinance.download failed: {exc}")
            return pd.DataFrame()

        out = pd.DataFrame()
        for t in tickers:
            s = self._extract_close(raw, t)
            if s is None or s.dropna().empty:
                print(f"       [WARN] No data for {t}")
                continue
            out[t] = s
        if out.empty:
            return out
        out.index = pd.to_datetime(out.index)
        out = out.sort_index().ffill()
        print(f"       → {out.shape[0]} rows × {out.shape[1]} cols (last={out.index[-1].date()})")
        return out

    def _fetch_spreads(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """HY OAS를 FRED에서 수집(일별, 시프트 불필요)."""
        print(f"\n  [2/4] Spread assets ({len(self.SPREAD_ASSETS)} via FRED) …")
        frames = []
        for _, (fred_id, *_rest) in self.SPREAD_ASSETS.items():
            label = _rest[-1]
            s = self._fred_get_series(fred_id, observation_start=ext_start, observation_end=end_date)
            if s is None or s.empty:
                print(f"       [WARN] {label} ({fred_id}) no data")
                continue
            s.index = pd.to_datetime(s.index)
            s = s.sort_index().ffill().dropna()
            frames.append(pd.DataFrame({label: s}))
            print(f"       {label} → {fred_id} ({len(s)} rows, last={s.index[-1].date()})")
        return pd.concat(frames, axis=1) if frames else pd.DataFrame()

    def _fetch_macro(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """월간 매크로(CPI YoY, 실업률)를 FRED에서 수집 후 발표일로 시프트."""
        print(f"\n  [3/4] Macro assets  ({len(self.MACRO_ASSETS)} via FRED) …")
        # CPI YoY 계산용 13개월 추가 lookback
        ext_macro_start = (
            datetime.strptime(ext_start, "%Y-%m-%d") - timedelta(days=400)
        ).strftime("%Y-%m-%d")
        ext_start_ts = pd.Timestamp(ext_start)
        today_ts = pd.Timestamp.utcnow().tz_localize(None).normalize()
        out: dict[str, pd.Series] = {}

        for _, (fred_id, _cond, _thr, _w, label, is_yoy) in self.MACRO_ASSETS.items():
            s = self._fred_get_series(fred_id, observation_start=ext_macro_start, observation_end=end_date)
            if s is None or s.empty:
                print(f"       [WARN] {label} ({fred_id}) no data")
                continue
            s.index = pd.to_datetime(s.index)
            s = s.sort_index().dropna()

            if is_yoy:                                   # 지수값 → YoY%
                s = (s / s.shift(12) - 1.0) * 100.0
                s = s.dropna()
            if s.empty:
                continue

            # 발표일 시프트 (참조월말 + lag), 미래 release 제거
            lag = self.FRED_RELEASE_LAG_DAYS.get(fred_id, 30)
            release_idx = s.index + pd.offsets.MonthEnd(0) + pd.Timedelta(days=lag)
            mask = release_idx <= today_ts
            s_shift = pd.Series(s.values[mask], index=release_idx[mask], name=label)
            s_shift = s_shift[s_shift.index >= ext_start_ts]
            if s_shift.empty:
                continue

            stale = (today_ts - s_shift.index[-1]).days
            if stale > self.MACRO_STALENESS_WARN_DAYS:
                print(f"       [WARN] {label} last release {s_shift.index[-1].date()} → {stale}일 stale")
            out[label] = s_shift
            print(f"       {label} → {fred_id} ({len(s_shift)} releases, +{lag}d, last={s_shift.index[-1].date()})")

        if not out:
            return pd.DataFrame()
        df = pd.concat(out, axis=1).sort_index().dropna(how="all")
        print(f"       → {df.shape[0]} observations (release-dated)")
        return df

    def _fetch_spx(self, start_date: str, end_date: str) -> pd.DataFrame:
        """S&P 500(^GSPC) — 오버레이 차트용."""
        print(f"\n  [4/4] Benchmark     ({self.SPX_TICKER}) …")
        yf_end = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        try:
            raw = yf.download(tickers=self.SPX_TICKER, start=start_date, end=yf_end,
                              interval="1d", auto_adjust=False, progress=False)
        except Exception as exc:
            print(f"       [WARN] SPX fetch failed: {exc}")
            return pd.DataFrame()
        s = self._extract_close(raw, self.SPX_TICKER)
        if s is None or s.dropna().empty:
            return pd.DataFrame()
        df = pd.DataFrame({self.SPX_TICKER: s})
        df.index = pd.to_datetime(df.index)
        df = df.sort_index().ffill()
        print(f"       → {df.shape[0]} rows (last={df.index[-1].date()})")
        return df

    # ═════════════════════════════════════════════════════════════════════════
    # 2. calculate_signals
    # ═════════════════════════════════════════════════════════════════════════
    def calculate_signals(self) -> pd.DataFrame:
        """가격/스프레드 MA 크로스 + SKEW 페널티 + 매크로 임계값 → 가중 시그널."""
        if self._full_price is None:
            raise RuntimeError("데이터가 없습니다 — fetch_data()를 먼저 호출하세요.")

        frames: list[pd.Series] = []

        # 2a. 가격 MA 시그널
        for ticker, (ma_p, w, direction, label) in self.PRICE_ASSETS.items():
            s = self._full_price.get(ticker)
            if s is None or s.dropna().empty:
                print(f"  [WARN] {label} unavailable — skipped.")
                continue
            ma = s.rolling(window=ma_p, min_periods=ma_p).mean()
            cond = (s > ma) if direction == "above" else (s < ma)
            frames.append((cond.astype(float) * w).rename(label))

        # 2b. SKEW 페널티
        skew = self._full_price.get(self.SKEW_TICKER)
        if skew is not None and not skew.dropna().empty:
            pen = np.where(skew > self.SKEW_THRESHOLD, self.SKEW_PENALTY, 0.0)
            frames.append(pd.Series(pen, index=skew.index, name="SKEW Penalty"))
        else:
            print(f"  [WARN] SKEW unavailable — penalty skipped.")

        # 2c. 스프레드 MA 시그널
        if self._full_spread is not None and not self._full_spread.empty:
            for _, (_fid, ma_p, w, direction, label) in self.SPREAD_ASSETS.items():
                s = self._full_spread.get(label)
                if s is None or s.dropna().empty:
                    print(f"  [WARN] {label} spread unavailable — skipped.")
                    continue
                ma = s.rolling(window=ma_p, min_periods=ma_p).mean()
                cond = (s > ma) if direction == "above" else (s < ma)
                frames.append((cond.astype(float) * w).rename(label))

        # 2d. 매크로 시그널 (월별 → 일별 ffill)
        if self._full_macro is not None and not self._full_macro.empty:
            daily_idx = self._full_price.index
            for _, (_fid, cond_dir, thr, w, label, _yoy) in self.MACRO_ASSETS.items():
                s = self._full_macro.get(label)
                if s is None or s.dropna().empty:
                    print(f"  [WARN] {label} macro missing — skipped.")
                    continue
                daily_s = (s.dropna()
                           .reindex(daily_idx.union(s.dropna().index))
                           .ffill()
                           .reindex(daily_idx))
                cond = (daily_s > thr) if cond_dir == "above" else (daily_s < thr)
                frames.append((cond.astype(float) * w).rename(label))

        if not frames:
            raise RuntimeError("시그널이 하나도 계산되지 않았습니다 — 데이터를 확인하세요.")

        all_sig = pd.concat(frames, axis=1).sort_index()
        self.signals = all_sig.loc[self._start_date:self._end_date].copy()
        self.signals = self.signals.ffill()   # bfill 금지(look-ahead 방지)
        return self.signals

    # ═════════════════════════════════════════════════════════════════════════
    # 3. get_regime
    # ═════════════════════════════════════════════════════════════════════════
    def get_regime(self) -> pd.DataFrame:
        """가중 시그널 합산 → 복합 스코어 → in-sample 백분위로 레짐 분류."""
        if self.signals is None:
            self.calculate_signals()

        self.scores = self.signals.sum(axis=1, min_count=1)
        self.scores.name = "Total Risk Score"

        p_on = float(self.scores.quantile(self.RISK_ON_PCT / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        def _classify(s: float) -> str:
            if pd.isna(s):
                return "Unknown"
            if s >= p_on:
                return "Risk On"
            if s <= p_off:
                return "Risk Off"
            return "Transition"

        self.regimes = self.scores.apply(_classify)
        self.regimes.name = "Regime"
        result = pd.concat([self.scores, self.regimes], axis=1)
        result.columns = ["score", "regime"]
        self._print_regime_summary(p_on, p_off)
        return result

    def _print_regime_summary(self, p_on: float, p_off: float) -> None:
        sep = "═" * 62
        print(f"\n{sep}\n  RISK SCORING MODEL — REGIME SUMMARY\n{sep}")
        print(f"  Period      : {self._start_date} → {self._end_date}")
        print(f"  Score range : {self.scores.min():.2f} → {self.scores.max():.2f}")
        print(f"  Mean / Std  : {self.scores.mean():.2f} / {self.scores.std():.2f}")
        print(f"  Risk On (≥p{self.RISK_ON_PCT})  : {p_on:.2f}")
        print(f"  Risk Off(≤p{self.RISK_OFF_PCT}) : {p_off:.2f}\n")
        counts = self.regimes.value_counts()
        total = counts.sum()
        for regime in ["Risk On", "Transition", "Risk Off", "Unknown"]:
            if regime not in counts:
                continue
            cnt = counts[regime]
            pct = cnt / total * 100
            print(f"  {regime:<12} {cnt:5d} days  {pct:5.1f}%  {'█' * int(pct / 2)}")
        print(f"\n  Latest ({self.scores.index[-1].date()}) : "
              f"Score = {self.scores.iloc[-1]:.2f}  →  [ {self.regimes.iloc[-1]} ]")
        print(f"{sep}\n")

    # ═════════════════════════════════════════════════════════════════════════
    # 4. plot_results
    # ═════════════════════════════════════════════════════════════════════════
    def plot_results(self, figsize=(18, 11), save_path="risk_score_chart.png", show=False) -> None:
        """3패널 차트: SPX+레짐배경 / 스코어+임계선 / 시그널 기여 스택."""
        if self.scores is None or self.regimes is None:
            raise RuntimeError("plot_results() 전에 get_regime()을 실행하세요.")
        idx, C = self.scores.index, self.PALETTE
        p_on = float(self.scores.quantile(self.RISK_ON_PCT / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        fig = plt.figure(figsize=figsize, facecolor="#f7f8fa")
        fig.suptitle("Global Multi-Asset Risk Scoring Model", fontsize=15, fontweight="bold", y=0.995)
        gs = GridSpec(3, 1, figure=fig, height_ratios=[2.4, 2.0, 1.2], hspace=0.05)
        ax1, ax2, ax3 = (fig.add_subplot(gs[0]),
                         fig.add_subplot(gs[1]), fig.add_subplot(gs[2]))
        ax2.sharex(ax1); ax3.sharex(ax1)
        for ax in (ax1, ax2, ax3):
            self._shade_regimes(ax, idx, self.regimes, C)

        # Panel 1: SPX
        if self.spx_data is not None and not self.spx_data.empty:
            spx = self.spx_data.iloc[:, 0].reindex(idx).ffill()
            ax1.plot(idx, spx, color=C["spx"], lw=1.8, label="S&P 500")
        ax1.set_ylabel("SPX Level", fontsize=10)
        ax1.grid(True, alpha=0.25, linestyle="--")
        ax1.set_title("S&P 500 · Risk Regime Background", fontsize=10, pad=4)
        patches = [Patch(facecolor=C[r], alpha=0.30, label=r) for r in ["Risk On", "Transition", "Risk Off"]]
        ax1.legend(handles=[Line2D([0], [0], color=C["spx"], lw=1.8, label="S&P 500")] + patches,
                   loc="upper left", fontsize=8, framealpha=0.7)
        plt.setp(ax1.get_xticklabels(), visible=False)

        # Panel 2: Score
        ax2.plot(idx, self.scores, color=C["score"], lw=1.6, label="Risk Score", zorder=4)
        ma20 = self.scores.rolling(20, min_periods=5).mean()
        ax2.plot(idx, ma20, color=C["score_ma"], lw=1.2, ls="--", alpha=0.85, label="20-day MA", zorder=4)
        ax2.axhline(p_on, color=C["Risk On"], ls=":", lw=1.5, alpha=0.9, label=f"Risk On (p{self.RISK_ON_PCT}={p_on:.1f})")
        ax2.axhline(p_off, color=C["Risk Off"], ls=":", lw=1.5, alpha=0.9, label=f"Risk Off (p{self.RISK_OFF_PCT}={p_off:.1f})")
        ax2.axhline(0, color="black", lw=0.5, alpha=0.25)
        ax2.fill_between(idx, self.scores, p_on, where=(self.scores >= p_on), alpha=0.08, color=C["Risk On"])
        ax2.fill_between(idx, self.scores, p_off, where=(self.scores <= p_off), alpha=0.08, color=C["Risk Off"])
        ax2.set_ylabel("Composite Score", fontsize=10)
        ax2.grid(True, alpha=0.25, linestyle="--")
        ax2.legend(loc="upper left", fontsize=8, ncol=2, framealpha=0.7)
        plt.setp(ax2.get_xticklabels(), visible=False)

        # Panel 3: 시그널 기여 스택
        pos_cols = [c for c in self.signals.columns if "Penalty" not in c]
        neg_cols = [c for c in self.signals.columns if "Penalty" in c]
        cmap = plt.cm.tab20
        pos_arrays = [self.signals[c].reindex(idx).fillna(0).values for c in pos_cols]
        if pos_arrays:
            ax3.stackplot(idx, pos_arrays, labels=pos_cols,
                          colors=[cmap(i / max(len(pos_cols), 1)) for i in range(len(pos_cols))], alpha=0.72)
        for col in neg_cols:
            vals = self.signals[col].reindex(idx).fillna(0)
            ax3.fill_between(idx, vals, 0, color=C["Risk Off"], alpha=0.65, label=col)
        ax3.axhline(0, color="black", lw=0.7, alpha=0.5)
        ax3.set_ylabel("Signal Contrib.", fontsize=9)
        ax3.grid(True, alpha=0.20, linestyle="--")
        interval = max(1, round(len(idx) / 252 / 4))
        ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=interval * 3))
        ax3.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        plt.setp(ax3.get_xticklabels(), rotation=35, ha="right", fontsize=8)

        plt.tight_layout()
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            print(f"[INFO] Chart saved → {save_path}")
        plt.show() if show else plt.close(fig)

    @staticmethod
    def _shade_regimes(ax, idx, regimes, colors) -> None:
        if len(idx) < 2:
            return
        segments, cur, seg_start = [], regimes.iloc[0], idx[0]
        for date, regime in zip(idx[1:], regimes.iloc[1:]):
            if regime != cur:
                segments.append((seg_start, date, cur))
                cur, seg_start = regime, date
        segments.append((seg_start, idx[-1], cur))
        for s, e, r in segments:
            ax.axvspan(s, e, alpha=0.12, color=colors.get(r, "#cccccc"), linewidth=0)

    # ═════════════════════════════════════════════════════════════════════════
    # 5. run (end-to-end)
    # ═════════════════════════════════════════════════════════════════════════
    def run(self, start_date: str | None = None, end_date: str | None = None,
            plot: bool = False) -> pd.DataFrame:
        """fetch_data → calculate_signals → get_regime [→ plot_results]."""
        today = pd.Timestamp.now().normalize()
        end_date = end_date or today.strftime("%Y-%m-%d")
        start_date = start_date or (today - pd.DateOffset(years=2)).strftime("%Y-%m-%d")
        self.fetch_data(start_date, end_date)
        self.calculate_signals()
        results = self.get_regime()
        if plot:
            self.plot_results()
        return results

    # ═════════════════════════════════════════════════════════════════════════
    # 6. daily_snapshot (핵심 출력)
    # ═════════════════════════════════════════════════════════════════════════
    def daily_snapshot(self, date: str | None = None) -> dict:
        """오늘(또는 지정일)의 레짐·스코어·역사적 맥락·시그널 현황을 콘솔 출력."""
        if self.scores is None or self.signals is None:
            raise RuntimeError("get_regime()을 먼저 실행하세요.")

        if date is None:
            dt = self.scores.index[-1]
        else:
            dt = pd.Timestamp(date)
            if dt not in self.scores.index:
                dt = self.scores.index[self.scores.index.get_indexer([dt], method="ffill")[0]]

        score = float(self.scores.loc[dt])
        regime = str(self.regimes.loc[dt])
        pct_rank = float((self.scores <= score).mean() * 100)
        pos = int(self.scores.index.get_loc(dt))
        avg_30 = float(self.scores.iloc[max(0, pos - 30):pos].mean()) if pos > 0 else float("nan")
        avg_90 = float(self.scores.iloc[max(0, pos - 90):pos].mean()) if pos > 0 else float("nan")
        vs_30 = score - avg_30 if not np.isnan(avg_30) else float("nan")
        vs_90 = score - avg_90 if not np.isnan(avg_90) else float("nan")

        sig = self.signals.loc[dt]
        on_sigs = [(c, float(sig[c])) for c in sig.index if float(sig[c]) > 0 and "Penalty" not in c]
        off_sigs = [(c, float(sig[c])) for c in sig.index if float(sig[c]) == 0 and "Penalty" not in c]
        pen_sigs = [(c, float(sig[c])) for c in sig.index if "Penalty" in c]

        # 동적 최대값(가중치 합) — 매크로 누락에도 정확
        max_score = sum(w for (_, w, _, _) in self.PRICE_ASSETS.values()) \
            + sum(t[2] for t in self.SPREAD_ASSETS.values()) \
            + sum(t[3] for t in self.MACRO_ASSETS.values())
        bar = max(int(score / max_score * 20), 0) if max_score else 0
        score_bar = "█" * bar + "░" * (20 - bar)
        icon = {"Risk On": "🟢", "Transition": "🟡", "Risk Off": "🔴"}.get(regime, "⚪")

        sep, sep_s = "═" * 64, "─" * 64
        print(f"\n{sep}\n  오늘의 리스크 현황   {dt.strftime('%Y-%m-%d (%a)')}\n{sep}")
        print(f"\n  {icon}  레짐 : [ {regime} ]")
        print(f"     스코어 : {score:.1f} / {max_score:.0f}   [{score_bar}]")
        print(f"     역사적 백분위 : {pct_rank:.1f}%  "
              f"({'하위권' if pct_rank < 33 else '중간' if pct_rank < 67 else '상위권'})")

        def _d(v):
            if np.isnan(v):
                return "N/A"
            return f"{'▲' if v > 0 else '▼' if v < 0 else '━'} {abs(v):.2f}점"
        print(f"     30일 평균 대비 : {_d(vs_30)}   (30d avg = {avg_30:.1f})")
        print(f"     90일 평균 대비 : {_d(vs_90)}   (90d avg = {avg_90:.1f})")

        print(f"\n{sep_s}\n  ✔  RISK ON  시그널\n{sep_s}")
        for name, val in on_sigs:
            print(f"     ✔  {name:<22}  +{val:.1f}")
        if not on_sigs:
            print("     (없음)")
        print(f"\n{sep_s}\n  ✘  RISK OFF 시그널\n{sep_s}")
        for name, val in off_sigs:
            print(f"     ✘  {name:<22}   {val:.1f}")
        if not off_sigs:
            print("     (없음 — 전부 Risk On)")
        if pen_sigs:
            print(f"\n{sep_s}\n  ⚠  PENALTY\n{sep_s}")
            for name, val in pen_sigs:
                print(f"     {'⚠  발동 중!' if val < 0 else '   미발동'}  {name:<22}  {val:.1f}")
        print(f"\n{sep}\n  복합 스코어 합계 : {score:.2f}   →   [ {regime} ]\n{sep}\n")

        return {"date": dt.strftime("%Y-%m-%d"), "score": score, "regime": regime,
                "pct_rank": pct_rank, "vs_30d": vs_30, "vs_90d": vs_90,
                "on_signals": on_sigs, "off_signals": off_sigs, "penalties": pen_sigs}

    # ═════════════════════════════════════════════════════════════════════════
    # 7. signal_summary (진단)
    # ═════════════════════════════════════════════════════════════════════════
    def signal_summary(self, date: str | None = None) -> pd.DataFrame:
        """지정일의 시그널별 값·상태 테이블."""
        if self.signals is None:
            raise RuntimeError("calculate_signals()를 먼저 실행하세요.")
        if date is None:
            row, dt = self.signals.iloc[-1], self.signals.index[-1].date()
        else:
            row, dt = self.signals.loc[date], date

        wmap: dict[str, float] = {lbl: w for (_, w, _, lbl) in self.PRICE_ASSETS.values()}
        wmap["SKEW Penalty"] = self.SKEW_PENALTY
        for t in self.SPREAD_ASSETS.values():
            wmap[t[-1]] = t[2]
        for t in self.MACRO_ASSETS.values():
            wmap[t[4]] = t[3]

        records = []
        for col in self.signals.columns:
            val = row[col]
            if "Penalty" in col:
                status = "PENALTY ACTIVE" if val < 0 else "No Penalty"
            elif pd.isna(val):
                status = "N/A"
            else:
                status = "Risk On" if val > 0 else "Risk Off"
            records.append({"Signal": col, "Weight": wmap.get(col, float("nan")),
                            "Value": round(val, 2), "Status": status})
        df = pd.DataFrame(records).set_index("Signal")
        print(f"\nSignal snapshot — {dt}\n{'─' * 50}")
        print(df.to_string())
        print(f"{'─' * 50}\nComposite Score : {row.sum():.2f}")
        return df

    # ═════════════════════════════════════════════════════════════════════════
    # 8. to_excel
    # ═════════════════════════════════════════════════════════════════════════
    def to_excel(self, path: str = "risk_score_report.xlsx") -> None:
        """3시트 보고서(오늘의 현황 / 이력 비교 / 모델 정의) 생성."""
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise ImportError("openpyxl이 필요합니다:  pip install openpyxl") from exc
        if self.scores is None or self.regimes is None:
            raise RuntimeError("get_regime()을 먼저 실행하세요.")
        import openpyxl
        wb = openpyxl.Workbook()
        ws_today = wb.active
        ws_today.title = "오늘의 리스크 현황"
        ws_hist = wb.create_sheet("이력 비교")
        ws_def = wb.create_sheet("모델 정의")
        self._excel_today_sheet(ws_today)
        self._excel_history_sheet(ws_hist)
        self._excel_definition_sheet(ws_def)
        wb.save(path)
        print(f"[INFO] Excel report saved → {path}")

    # ── Sheet 1: 오늘의 리스크 현황 ──────────────────────────────────────────
    def _excel_today_sheet(self, ws) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        NAVY, DBLUE, MBLUE = "1A3650", "2E5077", "3D7EAA"
        LBLUE, LGREY, WHITE = "EBF4FB", "F4F6F8", "FFFFFF"
        REG_BG = {"Risk On": "D5F5E3", "Transition": "FDEBD0", "Risk Off": "FADBD8", "Unknown": "F5F5F5"}
        REG_FG = {"Risk On": "1D6A39", "Transition": "784212", "Risk Off": "922B21", "Unknown": "555555"}
        ICON = {"Risk On": "▲ Risk On", "Transition": "━ Transition", "Risk Off": "▼ Risk Off"}
        FN = "맑은 고딕"

        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)

        def bg(h):
            return PatternFill("solid", fgColor=h)

        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        _t = Side(style="thin", color="D0D0D0")
        _m = Side(style="medium", color="999999")

        def bdr():
            return Border(left=_t, right=_t, top=_t, bottom=_t)

        def bdr_m():
            return Border(left=_m, right=_m, top=_m, bottom=_m)

        def fill_row(r, n, h):
            for ci in range(1, n + 1):
                ws.cell(r, ci).fill = bg(h)

        def sec_hdr(row, title, n=6):
            fill_row(row, n, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64 + n)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = fnt(True, 10, WHITE); c.fill = bg(DBLUE); c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def col_hdr(row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = fnt(True, 9, WHITE); c.fill = bg(MBLUE)
                c.alignment = al("center", "center"); c.border = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        for col, w in zip("ABCDEF", [24, 14, 16, 22, 14, 14]):
            ws.column_dimensions[col].width = w

        dt = self.scores.index[-1]
        score = float(self.scores.iloc[-1])
        regime = str(self.regimes.iloc[-1])
        pct_rank = float((self.scores <= score).mean() * 100)
        pos = len(self.scores) - 1
        avg_30 = float(self.scores.iloc[max(0, pos - 30):pos].mean()) if pos > 0 else float("nan")
        avg_90 = float(self.scores.iloc[max(0, pos - 90):pos].mean()) if pos > 0 else float("nan")
        vs_30, vs_90 = score - avg_30, score - avg_90
        max_score = sum(w for (_, w, _, _) in self.PRICE_ASSETS.values()) \
            + sum(t[2] for t in self.SPREAD_ASSETS.values()) \
            + sum(t[3] for t in self.MACRO_ASSETS.values())
        bar = max(int(score / max_score * 20), 0) if max_score else 0
        score_bar = "█" * bar + "░" * (20 - bar)
        sig_row = self.signals.iloc[-1]
        wmap = {lbl: w for (_, w, _, lbl) in self.PRICE_ASSETS.values()}
        wmap["SKEW Penalty"] = self.SKEW_PENALTY
        for t in self.SPREAD_ASSETS.values():
            wmap[t[-1]] = t[2]
        for t in self.MACRO_ASSETS.values():
            wmap[t[4]] = t[3]

        row = 1
        for r in range(1, 4):
            fill_row(r, 6, NAVY)
        ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 20; ws.row_dimensions[3].height = 14
        ws.merge_cells("A1:F2")
        t = ws.cell(1, 1, "오늘의 리스크 현황")
        t.font = fnt(True, 18, WHITE); t.fill = bg(NAVY); t.alignment = al("center", "center")
        ws.merge_cells("A3:F3")
        s = ws.cell(3, 1, f"기준일: {dt.strftime('%Y-%m-%d (%A)')}    |    분석 기간: {self._start_date} ~ {self._end_date}    |    생성: {pd.Timestamp.now():%Y-%m-%d %H:%M}")
        s.font = fnt(sz=9, color="BBBBBB", italic=True); s.fill = bg(NAVY); s.alignment = al("center", "center")
        row = 5

        row = sec_hdr(row, "  ① 오늘의 레짐 & 스코어")
        rbg, rfg = REG_BG.get(regime, WHITE), REG_FG.get(regime, "000000")
        ws.merge_cells(f"A{row}:B{row+1}")
        lc = ws.cell(row=row, column=1, value="레짐")
        lc.font = fnt(True, 11, rfg); lc.fill = bg(rbg); lc.alignment = al("center", "center"); lc.border = bdr_m()
        ws.merge_cells(f"C{row}:F{row+1}")
        rc = ws.cell(row=row, column=3, value=ICON.get(regime, regime))
        rc.font = fnt(True, 22, rfg); rc.fill = bg(rbg); rc.alignment = al("center", "center"); rc.border = bdr_m()
        ws.row_dimensions[row].height = 28; ws.row_dimensions[row + 1].height = 28
        row += 2
        for ci, (lbl, val, bold_col) in enumerate([
            ("스코어", f"{score:.1f} / {max_score:.0f}", True),
            ("스코어 바", score_bar, False),
            ("백분위", f"{pct_rank:.1f}%  ({'하위권' if pct_rank < 33 else '중간' if pct_rank < 67 else '상위권'})", False),
        ], 0):
            ws.merge_cells(f"A{row}:B{row}")
            c1 = ws.cell(row=row, column=1, value=lbl)
            c1.font = fnt(True, 9); c1.fill = bg(LBLUE); c1.alignment = al("left", "center"); c1.border = bdr()
            ws.merge_cells(f"C{row}:F{row}")
            c2 = ws.cell(row=row, column=3, value=val)
            c2.font = fnt(bold_col, 10, rfg if bold_col else "222222")
            c2.fill = bg(rbg if bold_col else WHITE)
            c2.alignment = al("left" if ci > 0 else "center", "center"); c2.border = bdr()
            ws.row_dimensions[row].height = 18; row += 1
        row += 1

        row = sec_hdr(row, "  ② 역사적 맥락 비교  (오늘 vs 과거 평균)")
        row = col_hdr(row, ["구분", "스코어", "오늘 대비", "해석", "", ""])
        ctx = [
            ("오늘", score, float("nan"), "기준"),
            ("30일 평균", avg_30, vs_30, "최근 1개월 흐름"),
            ("90일 평균", avg_90, vs_90, "최근 3개월 흐름"),
            ("역사적 최고", float(self.scores.max()), score - float(self.scores.max()), "전체 구간 최고"),
            ("역사적 최저", float(self.scores.min()), score - float(self.scores.min()), "전체 구간 최저"),
        ]
        for i, (lbl, val, diff, interp) in enumerate(ctx):
            rbg_i = rbg if i == 0 else (WHITE if i % 2 == 0 else LGREY)

            def _ds(d):
                if np.isnan(d):
                    return "─"
                return f"▲ +{d:.1f}" if d > 0 else f"▼ {d:.1f}" if d < 0 else "━  0.0"
            dc = "1D6A39" if not np.isnan(diff) and diff > 0 else "922B21" if not np.isnan(diff) and diff < 0 else "555555"
            for ci, v in enumerate([lbl, f"{val:.1f}" if not np.isnan(val) else "─", _ds(diff), interp, "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(rbg_i); c.border = bdr(); c.alignment = al("center" if ci == 2 else "left", "center")
                c.font = fnt(True, 9, rfg if i == 0 else dc) if ci in (1, 3) else fnt(sz=9)
            ws.row_dimensions[row].height = 16; row += 1
        row += 1

        row = sec_hdr(row, f"  ③ 시그널별 현황  ({dt:%Y-%m-%d})")
        row = col_hdr(row, ["시그널", "가중치", "당일 기여값", "상태", "조건 해석", ""])
        for i, col in enumerate(self.signals.columns):
            val = float(sig_row[col]); w = wmap.get(col, float("nan")); is_pen = "Penalty" in col
            if is_pen:
                status = "⚠ PENALTY 발동" if val < 0 else "정상"
                sfg = REG_FG["Risk Off"] if val < 0 else REG_FG["Risk On"]
                sbg = REG_BG["Risk Off"] if val < 0 else WHITE
                interp = f"SKEW > {self.SKEW_THRESHOLD} → {self.SKEW_PENALTY}점 차감"
            elif val > 0:
                status, sfg = "✔ Risk On", REG_FG["Risk On"]
                sbg = REG_BG["Risk On"] if i % 2 == 0 else "E8F8F0"; interp = "MA/임계값 조건 충족"
            else:
                status, sfg = "✘ Risk Off", REG_FG["Risk Off"]
                sbg = REG_BG["Risk Off"] if i % 2 == 0 else "FBEAEA"; interp = "조건 미충족"
            for ci, v in enumerate([col, w, round(val, 2), status, interp, ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(sbg); c.border = bdr(); c.alignment = al("center" if ci in (2, 3) else "left", "center")
                c.font = fnt(True, 9, sfg) if ci == 4 else fnt(sz=9)
            ws.row_dimensions[row].height = 15; row += 1
        for ci, v in enumerate(["합계 (Composite Score)", "", round(score, 2), ICON.get(regime, regime), "", ""], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = fnt(True, 11, rfg); c.fill = bg(rbg); c.border = bdr_m()
            c.alignment = al("center" if ci > 1 else "left", "center")
        ws.row_dimensions[row].height = 22
        ws.freeze_panes = "A6"
        ws.sheet_properties.tabColor = REG_BG.get(regime, "2E5077")

    # ── Sheet 2: 이력 비교 ───────────────────────────────────────────────────
    def _excel_history_sheet(self, ws) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        NAVY, DBLUE, MBLUE, LBLUE, WHITE = "1A3650", "2E5077", "3D7EAA", "EBF4FB", "FFFFFF"
        REG_BG = {"Risk On": "D5F5E3", "Transition": "FDEBD0", "Risk Off": "FADBD8", "Unknown": "F5F5F5"}
        REG_FG = {"Risk On": "1D6A39", "Transition": "784212", "Risk Off": "922B21", "Unknown": "555555"}
        FN = "맑은 고딕"

        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)

        def bg(h):
            return PatternFill("solid", fgColor=h)

        def al(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)
        _t = Side(style="thin", color="D0D0D0")

        def bdr():
            return Border(left=_t, right=_t, top=_t, bottom=_t)

        def lighten(h, pct=0.45):
            return "".join(f"{int(int(h[i:i+2],16)+(255-int(h[i:i+2],16))*pct):02X}" for i in (0, 2, 4))

        def fill_row(r, n, h):
            for ci in range(1, n + 1):
                ws.cell(r, ci).fill = bg(h)

        def sec_hdr(row, title, n=5):
            fill_row(row, n, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64 + n)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = fnt(True, 10, WHITE); c.fill = bg(DBLUE); c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def col_hdr(row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = fnt(True, 9, WHITE); c.fill = bg(MBLUE)
                c.alignment = al("center", "center"); c.border = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        for col, w in zip("ABCDE", [18, 11, 16, 34, 14]):
            ws.column_dimensions[col].width = w
        p_on = float(self.scores.quantile(self.RISK_ON_PCT / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        row = 1
        for r in range(1, 4):
            fill_row(r, 5, NAVY)
        ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 14
        ws.merge_cells("A1:E2")
        t = ws.cell(1, 1, "이력 비교 — 일별 리스크 스코어 전체 이력")
        t.font = fnt(True, 14, WHITE); t.fill = bg(NAVY); t.alignment = al("center", "center")
        ws.merge_cells("A3:E3")
        s = ws.cell(3, 1, f"분석 기간: {self._start_date} ~ {self._end_date}    |    Risk On(≥p{self.RISK_ON_PCT}): {p_on:.1f}    |    Risk Off(≤p{self.RISK_OFF_PCT}): {p_off:.1f}")
        s.font = fnt(sz=9, color="BBBBBB", italic=True); s.fill = bg(NAVY); s.alignment = al("center", "center")
        row = 5

        row = sec_hdr(row, "  ① 레짐 분포 요약")
        row = col_hdr(row, ["레짐", "일수", "비율", "막대", "최신 위치"])
        counts = self.regimes.value_counts(); total = int(counts.sum())
        latest = float(self.scores.iloc[-1])
        for regime in ["Risk On", "Transition", "Risk Off"]:
            cnt = int(counts.get(regime, 0)); pct = cnt / total * 100
            rfg, rbg = REG_FG.get(regime, "000000"), REG_BG.get(regime, WHITE)
            here = "◀ 현재" if str(self.regimes.iloc[-1]) == regime else ""
            for ci, v in enumerate([regime, cnt, f"{pct:.1f}%", "▇" * int(pct / 2.5), here], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = fnt(sz=9, bold=(ci in (1, 5)), color=rfg); c.fill = bg(rbg); c.border = bdr()
                c.alignment = al("center" if ci > 1 else "left", "center")
            ws.row_dimensions[row].height = 16; row += 1
        row += 1
        for label, val in [("평균", f"{self.scores.mean():.2f}"), ("표준편차", f"{self.scores.std():.2f}"),
                           ("최고", f"{self.scores.max():.2f}"), ("최저", f"{self.scores.min():.2f}"),
                           ("오늘", f"{latest:.2f}"), ("오늘 백분위", f"{float((self.scores <= latest).mean()*100):.1f}%")]:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = fnt(True, 9); lc.fill = bg(LBLUE); lc.alignment = al("left", "center"); lc.border = bdr()
            ws.merge_cells(f"B{row}:E{row}")
            vc = ws.cell(row=row, column=2, value=val); vc.font = fnt(sz=9); vc.alignment = al("left", "center"); vc.border = bdr()
            ws.row_dimensions[row].height = 15; row += 1
        row += 1

        row = sec_hdr(row, "  ② 전체 일별 스코어 이력 (필터 ▼)")
        tbl_start = row
        row = col_hdr(row, ["날짜", "스코어", "레짐", "스코어 바", "오늘 대비"])
        max_sc = max(float(self.scores.max()), 1.0)
        for i, (dt, sc, rg) in enumerate(zip(self.scores.index, self.scores.values, self.regimes.values)):
            sc_f = float(sc); filled = max(int(sc_f / max_sc * 28), 0)
            diff = sc_f - latest
            diff_s = f"▲ +{diff:.1f}" if diff > 0 else f"▼ {diff:.1f}" if diff < 0 else "━  0.0"
            rbg = REG_BG.get(rg, WHITE); row_bg = rbg if i % 2 == 0 else lighten(rbg, 0.5)
            is_today = (i == len(self.scores) - 1)
            for ci, v in enumerate([dt.strftime("%Y-%m-%d"), round(sc_f, 2), rg, "█" * filled + "░" * (28 - filled), diff_s], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(row_bg); c.border = bdr(); c.alignment = al("center" if ci != 1 else "left", "center")
                c.font = fnt(True, 9, REG_FG.get(rg, "000000")) if ci == 3 else fnt(True, 9) if is_today else fnt(sz=9)
            ws.row_dimensions[row].height = 14; row += 1
        ws.freeze_panes = f"A{tbl_start + 1}"
        ws.auto_filter.ref = f"A{tbl_start}:E{row - 1}"
        ws.sheet_properties.tabColor = "2E5077"

    # ── Sheet 3: 모델 정의 ───────────────────────────────────────────────────
    def _excel_definition_sheet(self, ws) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        NAVY, DBLUE, MBLUE, LGREY, WHITE = "1A3650", "2E5077", "3D7EAA", "F4F6F8", "FFFFFF"
        GREEN, ORANGE, RED = "D5F5E3", "FDEBD0", "FADBD8"
        FG_G, FG_O, FG_R = "1D6A39", "784212", "922B21"
        FN = "맑은 고딕"

        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)

        def bg(h):
            return PatternFill("solid", fgColor=h)

        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        _t = Side(style="thin", color="D0D0D0")

        def bdr():
            return Border(left=_t, right=_t, top=_t, bottom=_t)

        for col, w in zip("ABCDEFGH", [14, 18, 18, 18, 10, 10, 26, 30]):
            ws.column_dimensions[col].width = w

        def fill_row(r, n, h):
            for ci in range(1, n + 1):
                ws.cell(r, ci).fill = bg(h)

        def sec(row, title, n=8):
            fill_row(row, n, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64 + n)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = fnt(True, 10, WHITE); c.fill = bg(DBLUE); c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def chdr(row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = fnt(True, 9, WHITE); c.fill = bg(MBLUE); c.alignment = al("center", "center"); c.border = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        def txt(row, text, bold=False, span=8, row_bg=WHITE, color="000000"):
            ws.merge_cells(f"A{row}:{chr(64 + span)}{row}")
            c = ws.cell(row=row, column=1, value=text)
            c.font = fnt(bold, 9, color); c.fill = bg(row_bg); c.alignment = al("left", "center", True)
            ws.row_dimensions[row].height = 16
            return row + 1

        row = 1
        for r in range(1, 4):
            fill_row(r, 8, NAVY)
        ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 14
        ws.merge_cells("A1:H2")
        t = ws.cell(1, 1, "모델 정의 및 산출 방식")
        t.font = fnt(True, 16, WHITE); t.fill = bg(NAVY); t.alignment = al("center", "center")
        ws.merge_cells("A3:H3")
        s = ws.cell(3, 1, "Global Multi-Asset Risk Scoring Model — Logic Reference")
        s.font = fnt(sz=9, color="BBBBBB", italic=True); s.fill = bg(NAVY); s.alignment = al("center", "center")
        row = 5

        row = sec(row, "  ① 모델 개요")
        max_score = sum(w for (_, w, _, _) in self.PRICE_ASSETS.values()) \
            + sum(t[2] for t in self.SPREAD_ASSETS.values()) \
            + sum(t[3] for t in self.MACRO_ASSETS.values())
        for text, bold, rbg in [
            ("위험 민감 자산(canary) 13종을 가중 합산해 매일 Risk On/Off 레짐을 수치화하는 모델.", False, WHITE),
            ("FX·EM·금리·주식·원자재·변동성·신용·매크로 자산을 yfinance + FRED에서 수집.", False, WHITE),
            ("각 자산의 MA 크로스 또는 임계값으로 Risk On 시그널(0/+가중치)을 만들고 SKEW는 −2점 페널티.", False, WHITE),
            ("복합 스코어 = Σ(개별 가중 시그널) + SKEW 패널티", True, DBLUE),
            (f"이론적 최대: +{max_score:.0f}점  |  최솟값: {self.SKEW_PENALTY:.0f}점 (SKEW 발동 시)", False, LGREY),
        ]:
            row = txt(row, text, bold, color=WHITE if rbg == DBLUE else "222222", row_bg=rbg)
        row += 1

        row = sec(row, "  ② 가격 기반 시그널")
        row = chdr(row, ["자산명", "yfinance 티커", "MA 기간", "가중치", "Risk On 조건", "", "", ""])
        for ticker, (ma_p, w, direction, label) in self.PRICE_ASSETS.items():
            cond = f"가격 {'<' if direction == 'below' else '>'} {ma_p}일 MA"
            for ci, v in enumerate([label, ticker, f"{ma_p}일", w, cond, "", "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(LGREY if row % 2 else WHITE); c.border = bdr()
                c.alignment = al("center" if ci in (3, 4) else "left", "center")
                c.font = fnt(sz=9, bold=(ci == 4))
            ws.row_dimensions[row].height = 15; row += 1
        # SKEW
        for ci, v in enumerate([self.SKEW_TICKER, self.SKEW_TICKER, "—", f"{self.SKEW_PENALTY}(페널티)", f"SKEW > {self.SKEW_THRESHOLD} → 차감", "", "", ""], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = bg(RED); c.border = bdr(); c.alignment = al("center" if ci in (3, 4) else "left", "center")
            c.font = fnt(sz=9, bold=(ci == 4), color=FG_R)
        ws.row_dimensions[row].height = 15; row += 1
        # HY OAS
        for _, (fid, ma_p, w, direction, label) in self.SPREAD_ASSETS.items():
            cond = f"스프레드 {'<' if direction == 'below' else '>'} {ma_p}일 MA"
            for ci, v in enumerate([label, fid, f"{ma_p}일", w, cond, "", "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(WHITE); c.border = bdr(); c.alignment = al("center" if ci in (3, 4) else "left", "center")
                c.font = fnt(sz=9, bold=(ci == 4))
            ws.row_dimensions[row].height = 15; row += 1
        row += 1

        row = sec(row, "  ③ 매크로 필터 (월별 → ffill)")
        row = chdr(row, ["지표명", "FRED ID", "Risk On 조건", "가중치", "처리", "", "", ""])
        for _, (fid, cond_dir, thr, w, label, is_yoy) in self.MACRO_ASSETS.items():
            cond = f"값 {'>' if cond_dir == 'above' else '<'} {thr}{'%' if is_yoy else ''}"
            note = "지수→YoY% + 발표일 시프트" if is_yoy else "발표일 시프트"
            for ci, v in enumerate([label, fid, cond, w, note, "", "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(LGREY if row % 2 else WHITE); c.border = bdr()
                c.alignment = al("center" if ci == 4 else "left", "center"); c.font = fnt(sz=9, bold=(ci == 4))
            ws.row_dimensions[row].height = 15; row += 1
        row = txt(row, "※ FRED 월간 데이터를 발표일로 시프트해 look-ahead bias를 방지. (ISM PMI는 FRED 단종으로 제외)", row_bg=ORANGE, color=FG_O)
        row += 1

        row = sec(row, "  ④ 레짐 분류 기준")
        row = chdr(row, ["레짐", "조건", "해석", "차트색", "", "", "", ""])
        for regime, cond, interp, cname, rbg, rfg in [
            ("Risk On", f"스코어 ≥ p{self.RISK_ON_PCT}", "위험자산 선호", "초록", GREEN, FG_G),
            ("Transition", f"p{self.RISK_OFF_PCT} < 스코어 < p{self.RISK_ON_PCT}", "중립·관망", "주황", ORANGE, FG_O),
            ("Risk Off", f"스코어 ≤ p{self.RISK_OFF_PCT}", "안전자산 선호", "빨강", RED, FG_R),
        ]:
            for ci, v in enumerate([regime, cond, interp, cname, "", "", "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(rbg); c.border = bdr(); c.alignment = al("left", "center", ci == 3)
                c.font = fnt(sz=9, bold=(ci == 1), color=rfg)
            ws.row_dimensions[row].height = 18; row += 1
        txt(row, f"※ 임계값은 분석 기간 전체의 백분위로 자동 계산됩니다 (p{self.RISK_ON_PCT}/p{self.RISK_OFF_PCT}).", row_bg="EBF4FB", color="2E5077")
        ws.sheet_properties.tabColor = "1A3650"


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    today_str = datetime.now().strftime("%Y-%m-%d")
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(base_dir, "outputs", "canaria", today_str)
    os.makedirs(out_dir, exist_ok=True)

    model = RiskScoringModel()                 # 환경변수 FRED_API_KEY 필요
    results = model.run(plot=False)            # 기간 미지정 → 최근 2년

    model.daily_snapshot()                     # 핵심: 오늘의 리스크 현황
    print("최근 10 영업일:")
    print(results.tail(10).to_string())

    model.plot_results(save_path=os.path.join(out_dir, "risk_score_chart.png"))
    last = model.scores.index[-1].strftime("%Y%m%d")
    model.to_excel(os.path.join(out_dir, f"risk_score_report_{last}.xlsx"))