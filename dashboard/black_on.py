# -*- coding: utf-8 -*-
"""black_on_1.xlsx -> black_on_dashboard.html generator.

mp_dashboard 의 레이아웃/CSS 틀을 그대로 계승하되, black_on 펀드(단일 펀드,
국내개별주식 / 미국개별주식 / 해외선물 3개 자산군)의 보유 스냅샷을 표시한다.

데이터 레이어 설계
- 시트는 코드명(9907/4409/4509)이 아니라 헤더 내용으로 자동 분류한다.
- NAV 는 순종목비 합산 역산: NAV = Σ|당일평가액| / Σ순종목비.
- 총계 행과 자체 합산을 대사(reconciliation)하여 파싱 누락을 감지한다.
- 국내/미국 현재비중 = 평가금액 / NAV, 현금·기타 = 100% - 주식비중합.
- 선물은 노셔널이라 평가금액/비중을 산정하지 않고 포지션만 표기한다.
"""
from __future__ import annotations

import base64
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


def clean_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def clean_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def excel_serial_to_date(value):
    """Excel 날짜 시리얼(1899-12-30 기준) 또는 datetime -> 'YYYY-MM-DD'."""
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    num = clean_number(value)
    if num is None:
        return clean_text(value)
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(num))).strftime("%Y-%m-%d")
    except (OverflowError, ValueError):
        return clean_text(value)


# 통합 단일시트(Sheet1) 1-indexed 컬럼 맵
COL = {
    "no": 1, "group": 2, "kind": 3, "sub": 4,
    "fundCode": 5, "fundName": 6, "code": 7, "name": 8,
    "prevQty": 10, "dQty": 11, "price": 12, "prevValue": 13,
    "pnl": 15, "maturity": 16, "weightPrev": 17,
    "ticker": 28, "value": 30, "weight": 31,
}
SUBTOTAL_KINDS = {"종류별소계", "총계", "소계", "합계"}
CATEGORIES = ["domesticStock", "domesticFutures", "overseasStock", "overseasFutures", "options", "cashOther"]
DERIV_CATS = ("domesticFutures", "overseasFutures", "options")
_CASH_KIND_TOKENS = ("예치금", "예금", "증거금")


def classify_row(group, kind):
    """자산그룹/종류로 자산군 분류.

    5탭(국내주식/국내선물/해외주식/해외선물/옵션) + 현금기타.
    반환: (category, known). known=False 면 화이트리스트 밖의 새 종류
    → cashOther 로 보내되 미분류 경고 대상.
    """
    k = kind or ""
    if k.endswith("선물"):
        return ("domesticFutures" if group == "선물그룹" else "overseasFutures"), True
    if k.endswith("옵션"):
        return "options", True
    if k == "주식":
        return "domesticStock", True
    if k == "해외주식":
        return "overseasStock", True
    if any(t in k for t in _CASH_KIND_TOKENS):
        return "cashOther", True
    return "cashOther", False


_OPT_NAME = re.compile(r"(\d{4})-(\d{2})\s*(콜|풋)\s*(\d+(?:\.\d+)?)")
_STOCKFUT_NAME = re.compile(r"^(.*?)\s+F\s+\d{6}")


_OVS_OPT_CODE = re.compile(r"^[A-Z][A-Z0-9 ]*?[FGHJKMNQUVXZ]\d(C|P)\d+(?:\.\d+)?$")
_OVS_OPT_NAME = re.compile(r"^(.*?)\s+(C|P)(\d+(?:\.\d+)?)\s*$", re.I)


def is_overseas_option(code, name):
    """해외 파생 중 옵션 판별. 종목코드가 (루트월물)(C|P)(행사가) 패턴이면 옵션.

    예) 'CLQ6C90' → 옵션(콜, 90) / 'CLQ6' → 선물.
    """
    if _OVS_OPT_CODE.match(clean_text(code).upper()):
        return True
    # 코드가 비어있을 때 종목명 보조 판정 ('Crude Oil C90.0')
    return bool(_OVS_OPT_NAME.match(clean_text(name)))


def parse_overseas_option(code, name):
    """해외옵션 종목명/코드에서 (기초자산, 'C'/'P', 행사가) 추출.

    'Crude Oil C90.0' → ('Crude Oil', 'C', 90.0). 실패 시 (name, None, None).
    """
    m = _OVS_OPT_NAME.match(clean_text(name))
    if m:
        under, cp, strike = m.group(1).strip(), m.group(2).upper(), float(m.group(3))
        return under, cp, strike
    # 코드에서 CP만이라도
    cm = re.search(r"(C|P)(\d+(?:\.\d+)?)$", clean_text(code).upper())
    if cm:
        return clean_text(name), cm.group(1), float(cm.group(2))
    return clean_text(name), None, None


def parse_stock_future_underlying(name):
    """개별주식선물 종목명에서 기초자산명 추출.

    '삼성전자   F 202607 (  10)' → '삼성전자'. 파싱 실패 시 None (미매칭 표시).
    """
    m = _STOCKFUT_NAME.match(name or "")
    return m.group(1).strip() if m else None


def parse_option_name(name):
    """옵션 종목명 파싱: '2026-07 콜1500 ff' → (만기 '2026-07', 'C'/'P', 행사가 1500.0).

    포맷이 다르면 (None, None, None) — 원문 표시 폴백.
    """
    m = _OPT_NAME.search(name or "")
    if not m:
        return None, None, None
    yy, mm, cp, strike = m.groups()
    return f"{yy}-{mm}", ("C" if cp == "콜" else "P"), float(strike)


def _cell(ws, r, key):
    return ws.cell(r, COL[key]).value


def read_sheet1(ws, cash_etf_codes=None):
    """통합 단일시트를 자산군별로 분류해 읽는다. 소계/총계/빈 행은 제외.

    - 순종목비(%NAV[당일]) 를 비중으로 직접 사용 (역산 불필요).
    - 보유수량 = 전일보유수량 + 당일증감수량 (선물은 부호 = 방향).
    - 선물 평가액은 부호가 있을 수 있어 절대값(valueAbs)도 함께 보관.
    - 취득원가 = 전일평가액 - 전일평가손익 (두 값 모두 '전일' 컬럼이라 날짜 정합).
    - 수익률 = (당일평가액 - 취득원가) / 취득원가.  (전일손익을 당일평가액에서 빼던 버그 수정)
    """
    cash_etf_codes = cash_etf_codes or {}
    cats = {c: [] for c in CATEGORIES}
    unclassified = []
    reclassified_cash = []   # 초단기채권 ETF → 현금 재분류 로그
    for r in range(2, ws.max_row + 1):
        kind = clean_text(_cell(ws, r, "kind"))
        group = clean_text(_cell(ws, r, "group"))
        name = clean_text(_cell(ws, r, "name"))
        if not name or kind in SUBTOTAL_KINDS or group in SUBTOTAL_KINDS or group == "총계":
            continue
        cat, known = classify_row(group, kind)
        # 해외 파생: 종류에 선물/옵션 구분이 없어 종목코드로 옵션 여부 판별
        _code_raw = clean_text(_cell(ws, r, "code"))
        if cat == "overseasFutures" and is_overseas_option(_code_raw, name):
            cat = "options"
        # 초단기채권 ETF 는 주식으로 분류돼도 현금성으로 재분류
        code_norm = clean_text(_cell(ws, r, "code")).upper().lstrip("A")
        if cat in ("domesticStock", "overseasStock") and code_norm in cash_etf_codes:
            reclassified_cash.append({"row": r, "code": code_norm, "name": name})
            cat, known = "cashOther", True
        if not known:
            unclassified.append({"row": r, "group": group, "kind": kind, "name": name})
        is_deriv = cat in DERIV_CATS
        is_futures = is_deriv  # 방향/절대값 처리는 선물·옵션 공통
        qty = (clean_number(_cell(ws, r, "prevQty")) or 0) + (clean_number(_cell(ws, r, "dQty")) or 0)
        value = clean_number(_cell(ws, r, "value"))
        weight = clean_number(_cell(ws, r, "weight"))
        weight = weight / 100.0 if weight is not None else None
        weight_prev = clean_number(_cell(ws, r, "weightPrev"))
        weight_prev = weight_prev / 100.0 if weight_prev is not None else None
        pnl = clean_number(_cell(ws, r, "pnl"))
        sub = clean_text(_cell(ws, r, "sub"))

        direction = ""
        dir_conflict = False
        if is_deriv:
            if sub in ("매수", "매도"):
                direction = sub
                # 교차검증: 구분 vs 수량 부호 모순 감지 (백오피스 포맷 변경 대비)
                if qty:
                    qty_dir = "매도" if qty < 0 else "매수"
                    dir_conflict = (qty_dir != direction)
            elif qty is not None:
                direction = "매도" if qty < 0 else "매수"

        prev_value = clean_number(_cell(ws, r, "prevValue"))  # 전일평가액(원)
        return_pct = None
        if not is_deriv and prev_value is not None and pnl is not None:
            cost = prev_value - pnl  # 취득원가 = 전일평가액 − 전일평가손익
            if cost and value is not None:
                return_pct = (value - cost) / cost * 100.0  # 당일평가액 기준 수익률

        opt_expiry, opt_cp, opt_strike = (None, None, None)
        opt_under = None
        if cat == "options":
            if kind.endswith("옵션"):          # 국내 주가지수옵션: '2026-07 콜1500 ff'
                opt_expiry, opt_cp, opt_strike = parse_option_name(name)
            else:                                # 해외옵션: 'Crude Oil C90.0'
                opt_under, opt_cp, opt_strike = parse_overseas_option(_code_raw, name)
        fut_under = parse_stock_future_underlying(name) if kind == "개별주식선물" else None

        cats[cat].append({
            "no": clean_number(_cell(ws, r, "no")),
            "category": cat,
            "group": group,
            "kind": kind,
            "sub": sub,
            "fundCode": clean_text(_cell(ws, r, "fundCode")),
            "fundName": clean_text(_cell(ws, r, "fundName")),
            "code": clean_text(_cell(ws, r, "code")),
            "ticker": clean_text(_cell(ws, r, "ticker")),
            "stockName": name,
            "qty": qty,
            "price": clean_number(_cell(ws, r, "price")),
            "value": value,
            "valueAbs": abs(value) if value is not None else None,
            "prevValue": prev_value,
            "weight": weight,
            "weightPrev": weight_prev,
            "pnl": pnl,
            "returnPct": return_pct,
            "direction": direction,
            "dirConflict": dir_conflict,
            "maturity": excel_serial_to_date(_cell(ws, r, "maturity")),
            "optExpiry": opt_expiry,
            "optCP": opt_cp,
            "optStrike": opt_strike,
            "optUnder": opt_under,
            "futUnder": fut_under,
            "isCashEtf": (cat == "cashOther" and clean_text(_cell(ws, r, "code")).upper().lstrip("A") in cash_etf_codes),
            "sourceRow": r,
        })
    _merge_duplicate_positions(cats)
    return cats, unclassified, reclassified_cash


def _merge_duplicate_positions(cats):
    """백오피스가 같은 종목을 여러 행(전일보유분/당일증감분)으로 내보내는 경우 병합.

    병합 키: (종목코드 또는 종목명) + 방향(파생). 수량·평가액·손익·비중을 합산해
    최종 포지션 한 행으로 만든다. 방향이 다른 파생(매수/매도 병존)은 별도 유지.
    """
    for cat, recs in cats.items():
        merged = {}
        order = []
        for rec in recs:
            key = (rec.get("code") or rec.get("stockName"), rec.get("direction") or "")
            if key not in merged:
                merged[key] = rec
                rec["mergedRows"] = 1
                order.append(key)
                continue
            m = merged[key]
            m["mergedRows"] += 1
            for f in ("qty", "value", "prevValue", "pnl", "weight", "weightPrev"):
                a, b = m.get(f), rec.get(f)
                if a is None and b is None:
                    continue
                m[f] = (a or 0) + (b or 0)
            m["valueAbs"] = abs(m["value"]) if m.get("value") is not None else None
            # 수익률 재계산 (주식류): 취득원가 = Σ전일평가액 − Σ전일손익
            if cat in ("domesticStock", "overseasStock"):
                pv, pnl, v = m.get("prevValue"), m.get("pnl"), m.get("value")
                m["returnPct"] = None
                if pv is not None and pnl is not None:
                    cost = pv - pnl
                    if cost and v is not None:
                        m["returnPct"] = (v - cost) / cost * 100.0
        cats[cat] = [merged[k] for k in order]


def derive_nav(cats):
    """NAV 역산 (합산 방식).

    순종목비(%NAV) = |당일평가액| / NAV 관계에서
      NAV = Σ|당일평가액| / Σ순종목비
    개별 비중은 소수 2자리 반올림이라 저비중 종목에서 오차가 크게 증폭되지만,
    합산하면 분자(원 단위)는 반올림이 없고 분모의 반올림 오차는 서로 상쇄된다.
    (구 방식: 종목별 |value|/weight 중앙값 → 저비중 종목이 많아지면 오염)
    """
    abs_value_sum, weight_sum = 0.0, 0.0
    for recs in cats.values():
        for rec in recs:
            v, w = rec.get("value"), rec.get("weight")
            if v and w and abs(w) > 0:
                abs_value_sum += abs(v)
                weight_sum += w
    return (abs_value_sum / weight_sum) if weight_sum else None


def read_total_row(ws):
    """총계 행(자산그룹=='총계')을 대사(reconciliation) 기준값으로 읽는다."""
    for r in range(2, ws.max_row + 1):
        if clean_text(_cell(ws, r, "group")) == "총계":
            w = clean_number(_cell(ws, r, "weight"))
            return {
                "value": clean_number(_cell(ws, r, "value")),      # 당일평가액 부호합
                "weight": (w / 100.0) if w is not None else None,  # 순종목비 합(그로스)
                "pnl": clean_number(_cell(ws, r, "pnl")),
                "hr": clean_number(ws.cell(r, 21).value),          # (전일)HR — 백오피스 헤지비율
                "sourceRow": r,
            }
    return None


def load_cash_etf_codes(base_dir):
    """같은 경로의 cash_etf.xlsx > 'cash' 시트에서 초단기채권 ETF 코드를 읽는다.

    시트 구조: code / name / class. class 가 '초단기채권'인 종목코드를
    현금 처리 대상으로 반환한다 (코드 6자리, 대문자 정규화).
    파일이 없거나 시트가 없으면 빈 집합을 반환하고 경고만 출력한다.
    """
    path = base_dir / "cash_etf.xlsx"
    if not path.exists():
        print("[안내] cash_etf.xlsx 없음 — 초단기채권 ETF 현금 재분류 생략")
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[경고] cash_etf.xlsx 열기 실패({str(e)[:50]}) — 현금 재분류 생략")
        return {}
    if "cash" not in wb.sheetnames:
        print("[경고] cash_etf.xlsx 에 'cash' 시트 없음 — 현금 재분류 생략")
        return {}
    ws = wb["cash"]
    header = [clean_text(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    def col_idx(*names):
        for i, h in enumerate(header):
            if h.lower() in names:
                return i
        return None
    ci, ni, xi = col_idx("code", "코드"), col_idx("name", "코드명"), col_idx("class", "유형분류", "유형분류(축소)")
    if ci is None or xi is None:
        print("[경고] cash 시트에 code/class 컬럼을 찾지 못함 — 현금 재분류 생략")
        return {}
    codes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if xi >= len(row) or ci >= len(row):
            continue
        cls = clean_text(row[xi])
        if cls != "초단기채권":
            continue
        code = clean_text(row[ci]).upper().lstrip("A")   # 혹시 접두 A가 있으면 제거
        if code:
            codes[code] = clean_text(row[ni]) if ni is not None and ni < len(row) else ""
    print(f"[안내] 초단기채권 ETF {len(codes)}종목 현금 재분류 대상 로드")
    return codes


def reconcile(cats, totals, tol_weight=0.005, tol_value_ratio=0.001):
    """자체 합산 vs 총계 행 대사. 불일치 리스트 반환 (빈 리스트 = 정합).

    자체 합이 총계와 어긋나면 행 파싱 누락/중복 신호다.
    주의: 총계 행도 단순 열합이므로 '손익 정의의 오염'(외화예치금 KRW값이
    pnl 컬럼에 들어오는 문제)은 공유한다. 이 대사는 파싱 무결성 검증용이다.
    """
    if not totals:
        return [{"check": "총계행", "msg": "총계 행 없음 — 대사 불가"}]
    my_w = sum((r["weight"] or 0) for recs in cats.values() for r in recs)
    my_v = sum((r["value"] or 0) for recs in cats.values() for r in recs)
    my_p = sum((r["pnl"] or 0) for recs in cats.values() for r in recs)
    issues = []
    for label, mine, ref, tol in (
        ("순종목비 합", my_w, totals["weight"], tol_weight),
        ("당일평가액 부호합", my_v, totals["value"], abs(totals["value"] or 1) * tol_value_ratio),
        ("전일평가손익 합", my_p, totals["pnl"], abs(totals["pnl"] or 1) * tol_value_ratio),
    ):
        if ref is not None and abs(mine - ref) > tol:
            issues.append({"check": label, "mine": mine, "totalRow": ref, "diff": mine - ref})
    return issues


def _os_copy(src: Path, dst: Path):
    """Excel 의 공유잠금에서도 복사되도록 OS 복사 명령을 사용한다.

    절대경로 + PowerShell 로 한글 경로/잠금을 모두 견딘다.
    """
    src, dst = str(src.resolve()), str(dst)
    if os.name == "nt":
        ps = f"Copy-Item -LiteralPath '{src}' -Destination '{dst}' -Force"
        subprocess.run(["powershell", "-NoProfile", "-Command", ps], check=True, capture_output=True)
    else:
        subprocess.run(["cp", src, dst], check=True, capture_output=True)


def load_safe(path: Path):
    """Excel 이 열려 있어 잠겨도 임시 복사본으로 읽는다. (wb, raw_bytes) 반환."""
    try:
        raw = path.read_bytes()
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb, raw
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / f"_blackon_{path.name}"
        try:
            shutil.copy(path, tmp)
        except PermissionError:
            _os_copy(path, tmp)  # Excel 배타 잠금 우회
        raw = tmp.read_bytes()
        wb = openpyxl.load_workbook(tmp, data_only=True)
        return wb, raw


def _signed_exposure(recs):
    """선물 순노출 = Σ 방향부호(매수+/매도−) × |평가액|."""
    total = 0.0
    for r in recs:
        sign = -1 if r["direction"] == "매도" else 1
        total += sign * (r["valueAbs"] or 0)
    return total


def build_data(workbook_path: Path):
    wb, raw = load_safe(workbook_path)
    sheet_name = wb.sheetnames[0]  # 통합 단일시트
    ws = wb[sheet_name]
    cash_etf_codes = load_cash_etf_codes(workbook_path.parent)
    cats, unclassified, reclassified_cash = read_sheet1(ws, cash_etf_codes)
    nav = derive_nav(cats)
    totals = read_total_row(ws)
    recon_issues = reconcile(cats, totals)

    def vsum(cat):
        return sum((r["value"] or 0) for r in cats[cat])

    def absum(cat):
        return sum((r["valueAbs"] or 0) for r in cats[cat])

    domestic_stock_value = vsum("domesticStock")
    overseas_stock_value = vsum("overseasStock")
    stock_value = domestic_stock_value + overseas_stock_value
    cash_value = vsum("cashOther")

    # 현금 = 외화예치금 + 예금잔고 (증거금 제외).
    # 외화예금/US MARGIN 은 보유수량이 USD 이고 당일평가액(col30)=0 이라, KRW 환산액이
    # col15(전일평가손익) 자리에 음수로 들어있다 → 절대값을 KRW 가치로 사용한다.
    def _cash_krw(r):
        return abs(r["value"]) if r["value"] else (abs(r["pnl"]) if r["pnl"] else 0)

    def _is_deposit(r):
        if r.get("isCashEtf"):
            return True   # 초단기채권 ETF = 현금성
        k = r["kind"] or ""
        return ("예치금" in k or "예금" in k) and "증거금" not in k

    for r in cats["cashOther"]:
        r["valueKrw"] = _cash_krw(r)
        r["isDeposit"] = _is_deposit(r)
    cash_deposit_value = sum(r["valueKrw"] for r in cats["cashOther"] if r["isDeposit"])
    margin_value = sum(r["valueKrw"] for r in cats["cashOther"] if "증거금" in (r["kind"] or ""))

    futures_net = _signed_exposure(cats["domesticFutures"]) + _signed_exposure(cats["overseasFutures"])
    futures_gross = absum("domesticFutures") + absum("overseasFutures")
    options_premium = absum("options")   # 옵션 프리미엄 평가액 합 (노출 아님)
    # 총 평가손익 = 주식(국내+해외) + 선물(국내+해외) + 옵션.
    # 현금성(cashOther) 제외 — 외화예치금/증거금의 pnl 컬럼에는 KRW 환산액이 들어와 손익이 아니다.
    PNL_CATS = ("domesticStock", "overseasStock", "domesticFutures", "overseasFutures", "options")
    # 손익 = 주식+선물+옵션 + (현금성 중) 초단기채권 ETF 수익. 예금/증거금은 제외.
    total_pnl = sum((r["pnl"] or 0) for c in PNL_CATS for r in cats[c])
    total_pnl += sum((r["pnl"] or 0) for r in cats["cashOther"] if r.get("isCashEtf"))
    dir_conflicts = [{"row": r["sourceRow"], "name": r["stockName"], "sub": r["sub"], "qty": r["qty"]}
                     for c in DERIV_CATS for r in cats[c] if r.get("dirConflict")]

    stat = workbook_path.stat()
    return {
        "workbook": workbook_path.name,
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbookModifiedAt": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "workbookSize": stat.st_size,
        "sheetName": sheet_name,
        "nav": nav,
        "navSource": "순종목비 합산 역산",
        "totalsRow": totals,
        "reconIssues": recon_issues,
        "reconOk": not recon_issues,
        "domesticStockValue": domestic_stock_value,
        "overseasStockValue": overseas_stock_value,
        "stockValue": stock_value,
        "cashValue": cash_value,
        "cashOtherWeight": (1.0 - stock_value / nav) if nav else None,
        "cashDepositValue": cash_deposit_value,
        "cashDepositWeight": (cash_deposit_value / nav) if nav else None,
        "marginValue": margin_value,
        "futuresNet": futures_net,
        "futuresGross": futures_gross,
        "optionsPremium": options_premium,
        "totalPnl": total_pnl,
        "unclassified": unclassified,
        "reclassifiedCash": reclassified_cash,
        "dirConflicts": dir_conflicts,
        "domesticStock": cats["domesticStock"],
        "domesticFutures": cats["domesticFutures"],
        "overseasStock": cats["overseasStock"],
        "overseasFutures": cats["overseasFutures"],
        "options": cats["options"],
        "cashOther": cats["cashOther"],
        "workbookBase64": base64.b64encode(raw).decode("ascii"),
    }


def _debug_dump(data):
    keys = ["workbook", "sheetName", "nav", "navSource", "reconOk", "domesticStockValue",
            "overseasStockValue", "stockValue", "cashValue", "cashOtherWeight",
            "cashDepositValue", "cashDepositWeight", "futuresNet", "futuresGross",
            "optionsPremium", "totalPnl"]
    out = {k: data[k] for k in keys}
    out["totalsRow"] = data["totalsRow"]
    out["reconIssues"] = data["reconIssues"]
    out["unclassified"] = data["unclassified"]
    out["options"] = [(r["stockName"], r["direction"], r["qty"], r["optExpiry"], r["optCP"], r["optStrike"], r["pnl"]) for r in data["options"]]
    out["counts"] = {c: len(data[c]) for c in CATEGORIES}
    out["base64_len"] = len(data["workbookBase64"])
    out["weightSums"] = {
        c: sum((r["weight"] or 0) for r in data[c])
        for c in CATEGORIES
    }
    out["sample_domStock"] = data["domesticStock"][0] if data["domesticStock"] else None
    out["sample_domFut"] = data["domesticFutures"][0] if data["domesticFutures"] else None
    out["sample_ovsFut"] = data["overseasFutures"][0] if data["overseasFutures"] else None
    out["cashRows"] = [(r["stockName"], r["value"], r["weight"]) for r in data["cashOther"]]
    out["domFut_dirs"] = [(r["stockName"], r["direction"], r["valueAbs"]) for r in data["domesticFutures"]]
    out["ovsFut_dirs"] = [(r["stockName"], r["direction"], r["qty"], r["valueAbs"]) for r in data["overseasFutures"]]
    print(json.dumps(out, ensure_ascii=True, indent=2, default=str))


HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>블랙ON 포트폴리오 대시보드</title>
  <style>
    :root {
      color-scheme: dark;
      --bg: #0a0a0a; --panel: #141414; --panel-2: #1b1b1b; --line: #2a2a2a; --line-strong: #3d3d3d;
      --text: #e8e6e1; --muted: #8a877f; --head: #f5a623; --head-2: #ff9500;
      --amber: #ffa028; --amber-soft: #3a2a10;
      --ok: #26a269; --warn: #f5a623; --bad: #e5484d;
      --soft-ok: #0f2a1f; --soft-warn: #332610; --soft-bad: #33151a; --edit: #2a2410;
      --shadow: 0 1px 2px rgba(0, 0, 0, 0.5);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0; background: var(--bg); color: var(--text);
      font-family: "Pretendard Variable", "Pretendard", -apple-system, BlinkMacSystemFont,
        system-ui, "Apple SD Gothic Neo", "Noto Sans KR", "Malgun Gothic", "Segoe UI", Arial, sans-serif;
      font-size: 14px; line-height: 1.45; -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility; font-feature-settings: "tnum" 1;
    }
    header {
      position: sticky; top: 0; z-index: 20; background: var(--panel);
      border-bottom: 1px solid var(--line); box-shadow: var(--shadow);
    }
    .topbar { max-width: 1680px; margin: 0 auto; padding: 14px 18px 12px; display: grid; gap: 10px; }
    .title-row { display: flex; align-items: baseline; justify-content: space-between; gap: 16px; flex-wrap: wrap; }
    h1 { margin: 0; font-size: 20px; font-weight: 700; color: var(--head); }
    .date-badge { display: inline-block; margin-left: 8px; padding: 3px 12px; font-size: 14px; font-weight: 700; color: #0a0a0a; background: var(--amber); border: 1px solid var(--amber); border-radius: 999px; vertical-align: middle; }
    .meta { color: var(--muted); font-size: 12px; display: flex; gap: 10px; flex-wrap: wrap; }
    .toolbar { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
    .main-menu { display: flex; gap: 6px; }
    .menu-btn { padding: 0 22px; height: 36px; font-weight: 700; font-size: 14px; border: 1px solid var(--line-strong); background: var(--panel-2); color: var(--muted); border-radius: 6px; cursor: pointer; }
    .menu-btn:hover { color: var(--amber); border-color: var(--amber); background: #1f1a10; }
    .menu-btn.active { background: var(--amber); border-color: var(--amber); color: #0a0a0a; }
    .view.view-hidden { display: none; }
    .toolbar-divider { width: 1px; height: 22px; background: var(--line); margin: 0 2px; }
    button, select, input {
      font: inherit; border: 1px solid var(--line-strong); background: #fff; color: var(--text);
      height: 32px; border-radius: 6px;
    }
    button { padding: 0 11px; cursor: pointer; font-weight: 600; }
    button.primary { background: var(--head-2); border-color: var(--head-2); color: #fff; }
    button.accent { background: var(--ok); border-color: var(--ok); color: #fff; }
    button.danger { color: var(--bad); border-color: #e0aaa5; background: #fff; }
    input, select { padding: 0 9px; }
    main { max-width: 1680px; margin: 0 auto; padding: 16px 18px 28px; display: grid; gap: 14px; }
    .notice-bar { display: flex; align-items: center; gap: 10px; background: #1c2530; color: #fff; border: 1px solid #2c3a49; border-radius: 8px; padding: 10px 14px; font-size: 13px; line-height: 1.4; box-shadow: var(--shadow); }
    .notice-bar b { color: #ffd45e; font-weight: 700; }
    .summary { display: grid; grid-template-columns: repeat(7, minmax(120px, 1fr)); gap: 10px; }
    .metric, .panel { background: var(--panel); border: 1px solid var(--line); box-shadow: var(--shadow); }
    .metric { padding: 11px 12px; min-height: 70px; }
    .metric .label { color: var(--muted); font-size: 12px; margin-bottom: 6px; }
    .metric .value { font-size: 19px; font-weight: 700; color: var(--text); }
    .metric .sub { font-size: 11px; color: var(--muted); margin-top: 2px; font-weight: 600; }
    .metric.bad .value { color: var(--bad); } .metric.ok .value { color: var(--ok); }
    .panel { overflow: hidden; }
    .panel-head {
      display: flex; justify-content: space-between; align-items: center; gap: 12px;
      padding: 10px 12px; background: var(--panel-2); border-bottom: 1px solid var(--line-strong);
      color: var(--amber); font-weight: 700; min-height: 50px;
    }
    .panel-body { padding: 10px 12px; }
    .workspace { display: grid; grid-template-columns: 300px minmax(0, 1fr); gap: 14px; align-items: stretch; }
    .workspace > section.panel { display: flex; flex-direction: column; max-height: calc(100vh - 130px); min-height: 0; }
    .fund-tabs { display: inline-flex; border: 1px solid var(--line-strong); border-radius: 6px; overflow: hidden; background: #fff; }
    .fund-tab { all: unset; cursor: pointer; padding: 4px 14px; font-weight: 600; font-size: 12.5px; color: var(--muted); border-right: 1px solid var(--line); }
    .fund-tab:last-child { border-right: 0; }
    .fund-tab:hover { background: #1f1a10; color: var(--amber); }
    .fund-tab.active { background: var(--amber); color: #0a0a0a; }
    .table-wrap { overflow: auto; flex: 1 1 auto; min-height: 0; background: #fff; }
    table { width: 100%; border-collapse: separate; border-spacing: 0; }
    th, td { border-right: 1px solid var(--line); border-bottom: 1px solid var(--line); padding: 8px 10px; vertical-align: middle; white-space: nowrap; }
    th { position: sticky; top: 0; z-index: 5; background: #000; color: var(--amber); font-weight: 700; text-align: center; padding: 10px; border-bottom: 1px solid var(--line-strong); }
    tbody td { background: var(--panel); } tbody tr:nth-child(even) td { background: var(--panel-2); }
    tbody tr:hover td { background: #241f14; }
    tbody tr.hidden { display: none; }
    td.number { text-align: right; font-variant-numeric: tabular-nums; }
    td.center { text-align: center; }
    td.gap-pos { color: var(--ok); font-weight: 700; } td.gap-neg { color: var(--bad); font-weight: 700; } td.gap-zero { color: var(--muted); }
    td.ret-pos { color: var(--ok); font-weight: 600; } td.ret-neg { color: var(--bad); font-weight: 600; }
    .col-hidden { display: none; }
    tbody tr.row-bad td.col-name { box-shadow: inset 3px 0 0 var(--bad); }
    td.col-name { max-width: 360px; overflow: hidden; text-overflow: ellipsis; }
    .editable { width: 100%; min-width: 70px; border: 1px solid transparent; background: transparent; color: var(--text); padding: 4px 6px; border-radius: 4px; }
    .editable:hover { background: var(--panel-2); border-color: var(--line); }
    .editable:focus { outline: none; background: var(--edit); border-color: #e1c96a; box-shadow: 0 0 0 2px rgba(225, 201, 106, 0.25); }
    .num-input { text-align: right; }
    .name-input { min-width: 160px; font-weight: 700; }
    .badge { display: inline-flex; align-items: center; justify-content: center; min-width: 40px; height: 22px; padding: 0 8px; border-radius: 999px; font-size: 12px; font-weight: 700; }
    .badge.b { color: var(--ok); background: var(--soft-ok); } .badge.s { color: var(--bad); background: var(--soft-bad); }
    .opt-pill { display: inline-flex; align-items: center; gap: 3px; padding: 2px 9px; border-radius: 999px; font-size: 11.5px; font-weight: 700; margin: 1px 2px; white-space: nowrap; }
    .opt-pill.up { color: var(--ok); background: var(--soft-ok); }
    .opt-pill.down { color: var(--bad); background: var(--soft-bad); }
    .opt-pill.flat { color: var(--muted); background: #eef1f5; }
    .deriv-line { line-height: 1.9; }
    .deriv-line + .deriv-line { margin-top: 2px; padding-top: 2px; border-top: 1px dashed var(--line); }
    tfoot td { position: sticky; bottom: 0; background: #000; font-weight: 700; color: var(--amber); border-top: 2px solid var(--amber); }
    .delete-row { width: 30px; padding: 0; }
    .grid-2 { display: grid; grid-template-columns: minmax(0, 1fr) minmax(320px, 0.7fr); gap: 14px; }
    .chart-layout { display: grid; grid-template-columns: minmax(460px, 1.6fr) minmax(300px, max-content); gap: 18px; align-items: center; }
    .alloc-detail { border: 1px solid var(--line); border-radius: 8px; background: var(--panel-2); min-height: 120px; }
    .alloc-detail-empty { padding: 18px 14px; color: var(--muted); font-size: 12.5px; text-align: center; }
    .alloc-detail-head { display: flex; justify-content: space-between; align-items: baseline; padding: 10px 12px; border-bottom: 1px solid var(--line-strong); background: #000; border-radius: 8px 8px 0 0; font-weight: 700; color: var(--amber); }
    .alloc-detail-body { padding: 8px 12px; display: grid; gap: 5px; font-size: 12.5px; }
    .alloc-detail-row { display: flex; justify-content: space-between; gap: 10px; }
    .alloc-detail-row .k { color: var(--muted); }
    .alloc-detail-row .v { font-variant-numeric: tabular-nums; font-weight: 600; }
    .alloc-detail-list { border-top: 1px dashed var(--line); margin-top: 4px; padding-top: 6px; display: grid; gap: 4px; }
    .alloc-detail-scroll { max-height: 420px; overflow-y: auto; margin-top: 6px; border: 1px solid var(--line); border-radius: 6px; }
    .mini-table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: 12.5px; }
    .mini-table th { position: sticky; top: 0; background: #000; color: var(--amber); font-weight: 700; padding: 7px 10px; text-align: right; font-size: 12px; z-index: 2; }
    .mini-table th:first-child { text-align: left; }
    .mini-table td { padding: 6px 10px; border-bottom: 1px solid var(--line); background: var(--panel); white-space: nowrap; }
    .mini-table tbody tr:nth-child(even) td { background: var(--panel-2); }
    .mini-table tbody tr:hover td { background: #241f14; }
    .mini-table td.nm { font-weight: 600; max-width: 300px; overflow: hidden; text-overflow: ellipsis; }
    .mini-table td.number { text-align: right; font-variant-numeric: tabular-nums; }
    .mini-table td.center { text-align: center; }
    .alloc-close { all: unset; cursor: pointer; color: var(--muted); font-size: 13px; font-weight: 700; padding: 0 4px; border-radius: 4px; }
    .alloc-close:hover { color: var(--bad); background: var(--soft-bad); }
    #allocPie path.seg { cursor: pointer; transition: opacity .12s, filter .12s; }
    #allocPie path.seg:hover { opacity: .9; filter: brightness(1.12); }
    #allocPie path.seg.selected { stroke: #ffd45e; stroke-width: 3.5; filter: brightness(1.15) drop-shadow(0 0 8px rgba(255,160,40,0.6)); }
    .pie-wrap { display: grid; justify-items: center; gap: 9px; }
    .pie-label-in { font-weight: 700; }
    .pie-label-out { font-size: 16px; font-weight: 600; fill: var(--text); }
    .pie-leader { stroke: var(--line-strong); stroke-width: 1; fill: none; }
    .legend { display: grid; gap: 4px; align-content: center; }
    .legend-item { display: grid; grid-template-columns: 12px 1fr 64px 84px; align-items: center; gap: 8px; font-size: 13px; padding: 5px 8px; border-radius: 6px; }
    .legend-item:hover { background: var(--panel-2); }
    .legend-swatch { width: 12px; height: 12px; border-radius: 3px; }
    .legend-label { white-space: nowrap; font-weight: 600; }
    .legend-pct { text-align: right; font-variant-numeric: tabular-nums; font-weight: 700; }
    .legend-eok { text-align: right; font-variant-numeric: tabular-nums; color: var(--muted); font-size: 12px; }
    .fut-summary { display: grid; gap: 6px; }
    .fut-summary .row { display: flex; justify-content: space-between; padding: 6px 8px; border: 1px solid var(--line); border-radius: 6px; background: #fafbfd; }
    @media (max-width: 1100px) {
      .summary { grid-template-columns: repeat(3, minmax(120px, 1fr)); }
      .workspace { grid-template-columns: 1fr; } .grid-2 { grid-template-columns: 1fr; }
      .chart-layout { grid-template-columns: 1fr; } .table-wrap { max-height: none; }
      }
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <div class="title-row">
        <h1>블랙ON #1 포트폴리오 대시보드 <span class="date-badge" id="dateBadge"></span></h1>
        <div class="meta">
          <span id="workbookMeta"></span>
          <span id="navMeta"></span>
          <span id="saveMeta"></span>
        </div>
      </div>
      <div class="main-menu" id="mainMenu">
        <button class="menu-btn active" type="button" data-view="overview">개요</button>
        <button class="menu-btn" type="button" data-view="holdings">보유내역</button>
        <button class="menu-btn" type="button" data-view="pnl">손익</button>
        <button class="menu-btn" type="button" data-view="exposure">선물 · 옵션 익스포저</button>
      </div>
    </div>
  </header>

  <main>
    <div class="notice-bar" role="note">
      <svg width="17" height="17" viewBox="0 0 24 24" fill="none" aria-hidden="true" style="flex:0 0 auto;">
        <path d="M12 9v4M12 17h.01" stroke="#ffd45e" stroke-width="2.2" stroke-linecap="round"/>
        <path d="M10.29 3.86 1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0Z" stroke="#ffd45e" stroke-width="2" stroke-linejoin="round"/>
      </svg>
      <span><b>장 마감 전</b> 대시보드를 이용하는 경우 실시간 시세를 반영하지 못해 평가액·수익률은 정확하지 않을 수 있습니다. 포지션 파악 용도로만 활용하세요.</span>
    </div>
    <section class="summary view" id="viewSummary" aria-label="요약">
      <div class="metric"><div class="label">펀드 NAV</div><div class="value" id="mNav">-</div><div class="sub" id="mNavSub"></div></div>
      <div class="metric" id="mPnlBox"><div class="label">총 평가손익</div><div class="value" id="mPnl">-</div><div class="sub" id="mPnlSub"></div></div>
      <div class="metric"><div class="label">종목 수 (주식)</div><div class="value" id="mCount">-</div><div class="sub" id="mCountSub"></div></div>
      <div class="metric" id="mNetBox"><div class="label">주식 순노출</div><div class="value" id="mNet">-</div><div class="sub" id="mNetSub"></div></div>
      <div class="metric"><div class="label">파생 노출 (NAV 대비)</div><div class="value" id="mDeriv">-</div><div class="sub" id="mDerivSub"></div></div>
      <div class="metric"><div class="label">그로스 익스포저</div><div class="value" id="mGross">-</div><div class="sub" id="mGrossSub"></div></div>
      <div class="metric"><div class="label">현금 · 증거금</div><div class="value" id="mCash">-</div><div class="sub" id="mCashSub"></div></div>
    </section>

    <section class="grid-2 view" id="viewOverview">
      <section class="panel">
        <div class="panel-head"><span>자산 배분</span><span class="muted">자본 100% + 파생 노셔널</span></div>
        <div class="panel-body">
          <div class="chart-layout">
            <div class="pie-wrap">
              <svg id="allocPie" viewBox="0 0 920 760" role="img" aria-label="자산배분 파이" style="width:100%;max-width:940px;display:block;margin:0 auto;"></svg>
            </div>
            <div class="legend" id="allocLegend"></div>
          </div>
        </div>
      </section>
      <section class="panel" id="allocDetailPanel">
        <div class="panel-head"><span>자산군 요약</span><span class="muted">다이어그램 항목 클릭</span></div>
        <div class="panel-body"><div class="alloc-detail" id="allocDetail"></div></div>
      </section>
    </section>

    <section class="view view-hidden" id="viewHoldings">
      <section class="panel">
        <div class="panel-head">
          <div style="display:inline-flex;align-items:center;gap:10px;flex-wrap:wrap;">
            <span id="tableTitle">국내주식</span>
            <div class="fund-tabs" id="assetTabs" role="tablist">
              <button class="fund-tab active" type="button" data-asset="domesticStock">국내주식</button>
              <button class="fund-tab" type="button" data-asset="domesticFutures">국내선물</button>
              <button class="fund-tab" type="button" data-asset="overseasStock">해외주식</button>
              <button class="fund-tab" type="button" data-asset="overseasFutures">해외선물</button>
              <button class="fund-tab" type="button" data-asset="options">옵션</button>
            </div>
          </div>
          <span class="muted" id="visibleCount"></span>
        </div>
        <div class="table-wrap">
          <table id="mainTable">
            <thead><tr id="theadRow"></tr></thead>
            <tbody id="tbody"></tbody>
            <tfoot><tr id="tfootRow"></tr></tfoot>
          </table>
        </div>
      </section>
    </section>

    <section class="view view-hidden" id="viewExposure">
      <section class="panel">
        <div class="panel-head"><span>기초자산별 포지션 구조 (현물 + 선물 + 옵션)</span><span class="muted" id="nettingHint"></span></div>
        <div class="table-wrap" style="max-height:calc(100vh - 220px);">
          <table>
            <thead><tr><th>기초자산</th><th>현물</th><th>선물 순</th><th>구성</th><th>순노출</th><th>%NAV</th><th>포지션 판정</th></tr></thead>
            <tbody id="nettingBody"></tbody>
          </table>
        </div>
      </section>
    </section>

    <section class="grid-2 view view-hidden" id="viewPnl">
      <section class="panel">
        <div class="panel-head"><span>누적수익률 상위 Top 10</span><span class="muted">종목 단독 성과 · 주식</span></div>
        <div class="table-wrap" style="max-height:560px;">
          <table class="mini-table"><thead><tr><th style="text-align:left;">종목명</th><th>누적수익률</th><th>평가액</th><th>손익</th></tr></thead><tbody id="pnlTop"></tbody></table>
        </div>
      </section>
      <section class="panel">
        <div class="panel-head"><span>누적수익률 하위 Top 10</span><span class="muted">종목 단독 성과 · 주식</span></div>
        <div class="table-wrap" style="max-height:560px;">
          <table class="mini-table"><thead><tr><th style="text-align:left;">종목명</th><th>누적수익률</th><th>평가액</th><th>손익</th></tr></thead><tbody id="pnlBottom"></tbody></table>
        </div>
      </section>
    </section>

  </main>

  <script>__SHEETJS__</script>
  <script>__BLACKON_XLSX__</script>
  <script>
    const DATA = __DATA_JSON__;
    const STORAGE_KEY = `blackon-dashboard:v2:${DATA.workbook}:${DATA.workbookModifiedAt}:${DATA.workbookSize}`;
    const CHART_COLORS = { domestic: "#ff9500", overseas: "#c77b0f", domesticFut: "#8a5a12", overseasFut: "#5c4a2a", opt: "#e5484d", cash: "#3a3a3a" };
    const CHART_TEXT = { domestic: "#0a0a0a", overseas: "#0a0a0a", domesticFut: "#ffffff", overseasFut: "#ffffff", opt: "#ffffff", cash: "#e8e6e1" };
    const TABS = ["domesticStock", "domesticFutures", "overseasStock", "overseasFutures", "options"];
    const TAB_LABEL = { domesticStock: "국내주식", domesticFutures: "국내선물", overseasStock: "해외주식", overseasFutures: "해외선물", options: "옵션" };
    const isFut = (t) => t === "domesticFutures" || t === "overseasFutures";
    const isOpt = (t) => t === "options";

    const state = { view: "overview", allocSel: null, assetTab: "domesticStock", search: "", group: "all", viewMode: "simple",
      assets: { domesticStock: [], domesticFutures: [], overseasStock: [], overseasFutures: [], options: [] } };

    document.addEventListener("DOMContentLoaded", init);

    function init() {
      // 기준일: 파일명 YYYYMMDD > 워크북 수정일. 배지로 강조
      const dm = (DATA.workbook || "").match(/(20\d{2})[._-]?(\d{2})[._-]?(\d{2})/);
      const baseDate = dm ? `${dm[1]}-${dm[2]}-${dm[3]}` : (DATA.workbookModifiedAt || "").slice(0, 10);
      document.getElementById("dateBadge").textContent = baseDate ? `기준일 ${baseDate}` : "";
      document.getElementById("workbookMeta").textContent = `${DATA.workbook} | 생성 ${DATA.generatedAt}`;
      const fund = (DATA.domesticStock[0] || {}).fundName || "";
      const recon = DATA.reconOk ? " · 대사 ✓" : ` · ⚠ 대사 불일치 ${DATA.reconIssues.length}건`;
      const navEl = document.getElementById("navMeta");
      navEl.textContent = `NAV ${formatEok(DATA.nav)}${fund ? " · " + fund : ""}${recon}`;
      if (!DATA.reconOk) {
        navEl.style.color = "var(--bad)";
        navEl.style.fontWeight = "700";
        navEl.title = DATA.reconIssues.map(i => `${i.check}: 자체 ${i.mine} vs 총계 ${i.totalRow}`).join("\n");
        console.warn("총계 행 대사 불일치:", DATA.reconIssues);
      }
      if ((DATA.unclassified || []).length) {
        const u = document.getElementById("workbookMeta");
        u.textContent += ` · ⚠ 미분류 종류 ${DATA.unclassified.length}건`;
        u.style.color = "var(--warn)"; u.style.fontWeight = "700";
        u.title = DATA.unclassified.map(x => `행${x.row} [${x.group}/${x.kind}] ${x.name}`).join("\n");
        console.warn("미분류 종류 (현금·기타로 임시 분류):", DATA.unclassified);
      }
      if ((DATA.dirConflicts || []).length) {
        const u = document.getElementById("workbookMeta");
        u.textContent += ` · ⚠ 방향모순 ${DATA.dirConflicts.length}건`;
        u.style.color = "var(--warn)"; u.style.fontWeight = "700";
        u.title = (u.title ? u.title + "\n" : "") + DATA.dirConflicts.map(x => `행${x.row} ${x.name}: 구분 ${x.sub} vs 수량 ${x.qty}`).join("\n");
        console.warn("파생 방향 모순 (구분 vs 수량부호):", DATA.dirConflicts);
      }
      loadAssets();
      bindEvents();
      render();
      applyView();
    }

    // ---- 편집 상태 빌드/저장 ----
    function mapStock(h) {
      return {
        sourceRow: h.sourceRow, deleted: false, isNew: false, category: h.category,
        group: h.group, kind: h.kind, code: h.code, ticker: h.ticker, stockName: h.stockName,
        qty: h.qty, qty0: h.qty, price: h.price, value0: h.value,
        cost0: (h.prevValue != null && h.pnl != null) ? h.prevValue - h.pnl : null,
        weightPrev: h.weightPrev,
      };
    }
    function mapFut(h) {
      return {
        sourceRow: h.sourceRow, deleted: false, isNew: false, category: h.category,
        group: h.group, kind: h.kind, sub: h.sub, code: h.code, ticker: h.ticker, stockName: h.stockName,
        qty: h.qty, qty0: h.qty, value0: h.value, valueAbs0: h.valueAbs,
        direction: h.direction, maturity: h.maturity, pnl0: h.pnl, weightPrev: h.weightPrev,
        futUnder: h.futUnder,
      };
    }
    function mapOpt(h) {
      return {
        sourceRow: h.sourceRow, deleted: false, isNew: false, category: h.category,
        group: h.group, kind: h.kind, sub: h.sub, code: h.code, stockName: h.stockName,
        qty: h.qty, qty0: h.qty, value0: h.value, valueAbs0: h.valueAbs,
        direction: h.direction, pnl0: h.pnl, weightPrev: h.weightPrev,
        optExpiry: h.optExpiry, optCP: h.optCP, optStrike: h.optStrike, optUnder: h.optUnder,
      };
    }
    function buildAssets() {
      return {
        domesticStock: DATA.domesticStock.map(mapStock),
        overseasStock: DATA.overseasStock.map(mapStock),
        domesticFutures: DATA.domesticFutures.map(mapFut),
        overseasFutures: DATA.overseasFutures.map(mapFut),
        options: DATA.options.map(mapOpt),
      };
    }

    function loadAssets() {
      const stored = localStorage.getItem(STORAGE_KEY);
      if (stored) {
        try {
          const parsed = JSON.parse(stored);
          if (TABS.every((t) => Array.isArray(parsed[t]))) {   // 스키마 검증
            state.assets = parsed;
            document.getElementById("saveMeta").textContent = "저장본 적용";
            return;
          }
          localStorage.removeItem(STORAGE_KEY);
        }
        catch (e) { localStorage.removeItem(STORAGE_KEY); }
      }
      state.assets = buildAssets();
    }

    function saveAssets() {
      localStorage.setItem(STORAGE_KEY, JSON.stringify(state.assets));
      const meta = document.getElementById("saveMeta");
      meta.textContent = `✓ 저장 ${new Date().toLocaleTimeString("ko-KR", { hour12: false })}`;
      meta.style.color = "var(--ok)"; meta.style.fontWeight = "700";
      setTimeout(() => { meta.style.color = ""; meta.style.fontWeight = ""; }, 1500);
    }

    function resetAssets() {
      if (!confirm("브라우저 저장본을 지우고 엑셀 기준으로 되돌릴까요?")) return;
      localStorage.removeItem(STORAGE_KEY);
      state.assets = buildAssets();
      document.getElementById("saveMeta").textContent = "";
      render();
    }

    // ---- 계산 ----
    function liveStock(row) {
      const qty = num(row.qty), price = num(row.price);
      let value;
      if (price != null && price > 0 && qty != null) value = qty * price;
      else if (qty != null && row.qty0) value = row.value0 * qty / row.qty0;
      else value = row.value0 || 0;
      const cost = row.cost0;
      const pnl = cost != null ? value - cost : null;
      const returnPct = cost ? pnl / cost * 100 : null;
      const weight = DATA.nav ? value / DATA.nav : 0;
      return { value, cost, pnl, returnPct, weight };
    }
    function liveFut(row) {
      const qty = num(row.qty);
      let value;
      if (qty != null && row.qty0) value = row.value0 * qty / row.qty0;
      else value = row.value0 || 0;
      const valueAbs = Math.abs(value || 0);
      let direction;
      if (qty != null && qty !== 0) direction = qty < 0 ? "매도" : "매수";
      else direction = row.sub || row.direction || "매수";
      const weight = DATA.nav ? valueAbs / DATA.nav : (row.valueAbs0 && DATA.nav ? row.valueAbs0 / DATA.nav : 0);
      return { value, valueAbs, direction, weight, pnl: row.pnl0 };
    }
    const dirSign = (d) => (d === "매도" ? -1 : 1);
    function liveOpt(row) {
      const qty = num(row.qty);
      let value;
      if (qty != null && row.qty0) value = row.value0 * qty / row.qty0;
      else value = row.value0 || 0;
      const valueAbs = Math.abs(value || 0);              // 프리미엄 평가액
      const direction = row.sub || row.direction || "매수";
      const weight = DATA.nav ? valueAbs / DATA.nav : 0;
      return { value, valueAbs, direction, weight, pnl: row.pnl0 };
    }

    // ---- 이벤트 ----
    const VIEW_IDS = { overview: "viewOverview", holdings: "viewHoldings", exposure: "viewExposure", pnl: "viewPnl" };
    function applyView() {
      document.querySelectorAll("#mainMenu .menu-btn").forEach((b) => b.classList.toggle("active", b.dataset.view === state.view));
      Object.entries(VIEW_IDS).forEach(([v, id]) => document.getElementById(id).classList.toggle("view-hidden", v !== state.view));
      document.getElementById("viewSummary").classList.toggle("view-hidden", state.view !== "overview");
    }

    function bindEvents() {
      document.querySelectorAll("#mainMenu .menu-btn").forEach((b) => {
        b.addEventListener("click", () => { state.view = b.dataset.view; applyView(); });
      });
      document.querySelectorAll("#assetTabs .fund-tab").forEach((tab) => {
        tab.addEventListener("click", () => { state.assetTab = tab.dataset.asset; state.group = "all"; render(); });
      });
    }

    function addRow() {
      const tab = state.assetTab;
      const base = { sourceRow: null, deleted: false, isNew: true, category: tab };
      if (tab === "domesticStock") Object.assign(base, { group: "주식그룹", kind: "주식", code: "", ticker: "", stockName: "", qty: null, qty0: null, price: null, value0: 0, cost0: null });
      else if (tab === "overseasStock") Object.assign(base, { group: "해외자산그룹", kind: "해외주식", code: "", ticker: "", stockName: "", qty: null, qty0: null, price: null, value0: 0, cost0: null });
      else if (tab === "domesticFutures") Object.assign(base, { group: "선물그룹", kind: "", sub: "매수", code: "", ticker: "", stockName: "", qty: null, qty0: null, value0: 0, valueAbs0: 0, direction: "매수", maturity: "", pnl0: null });
      else if (tab === "options") Object.assign(base, { group: "선물그룹", kind: "주가지수옵션", sub: "매수", code: "", stockName: "", qty: null, qty0: null, value0: 0, valueAbs0: 0, direction: "매수", pnl0: null, optExpiry: null, optCP: null, optStrike: null });
      else Object.assign(base, { group: "해외기타그룹", kind: "", sub: "", code: "", ticker: "", stockName: "", qty: null, qty0: null, value0: 0, valueAbs0: 0, direction: "매수", maturity: "", pnl0: null });
      state.assets[tab].push(base);
      render();
      const inputs = document.querySelectorAll("#tbody tr:last-child input");
      if (inputs.length) inputs[0].focus();
    }

    // ---- 렌더 ----
    function render() {
      document.querySelectorAll("#assetTabs .fund-tab").forEach((t) => t.classList.toggle("active", t.dataset.asset === state.assetTab));
      document.getElementById("tableTitle").textContent = TAB_LABEL[state.assetTab];
      renderHead();
      renderBody();
      renderFoot();
      renderMetrics();
      renderAlloc();
      renderNetting();
      renderPnl();
      applyViewMode();
      applyFilters();
    }

    const HEADERS = {
      domesticStock: [["종목명",""],["보유수량",""],["현재가",""],["평가액",""],["비중",""],["손익","detail"],["누적수익률",""],["전일비중","detail"],["",""]],
      overseasStock: [["Ticker",""],["종목명",""],["보유수량",""],["평가액",""],["비중",""],["손익","detail"],["누적수익률",""],["전일비중","detail"],["",""]],
      domesticFutures: [["종류",""],["방향",""],["종목명",""],["계약수",""],["평가액",""],["비중",""],["만기","detail"],["손익","detail"],["",""]],
      overseasFutures: [["종류",""],["방향",""],["종목명",""],["계약수",""],["평가액",""],["비중",""],["만기","detail"],["손익","detail"],["",""]],
      options: [["방향",""],["종목명",""],["만기/기초",""],["C/P",""],["행사가",""],["계약수",""],["프리미엄",""],["비중",""],["손익",""],["",""]],
    };

    function renderHead() {
      const tr = document.getElementById("theadRow");
      tr.innerHTML = "";
      HEADERS[state.assetTab].forEach(([label, cls]) => {
        const th = document.createElement("th");
        th.textContent = label;
        if (cls === "detail") th.classList.add("col-detail");
        tr.appendChild(th);
      });
    }

    function renderBody() {
      const body = document.getElementById("tbody");
      body.innerHTML = "";
      const frag = document.createDocumentFragment();
      const tab = state.assetTab;
      // 평가액 내림차순 정렬 (편집/삭제용 원본 인덱스는 보존)
      const valOf = isOpt(tab) ? (r) => liveOpt(r).valueAbs
        : isFut(tab) ? (r) => liveFut(r).valueAbs : (r) => liveStock(r).value;
      const entries = state.assets[tab].map((row, idx) => ({ row, idx })).filter((e) => !e.row.deleted);
      entries.sort((a, b) => (valOf(b.row) || 0) - (valOf(a.row) || 0));
      entries.forEach(({ row, idx }) => {
        frag.appendChild(isOpt(tab) ? optRow(row, idx) : isFut(tab) ? futRow(row, idx, tab) : stockRow(row, idx, tab));
      });
      body.appendChild(frag);
      body.querySelectorAll("input[data-field]").forEach((inp) => inp.addEventListener("change", onEdit));
    }

    function onEdit(e) {
      const inp = e.target;
      const idx = Number(inp.dataset.idx), field = inp.dataset.field, kind = inp.dataset.kind;
      const row = state.assets[state.assetTab][idx];
      if (kind === "num") row[field] = parseNumOrNull(inp.value);
      else if (kind === "pct") row[field] = parsePctOrNull(inp.value);
      else row[field] = inp.value;
      render();
    }

    function stockRow(row, idx, tab) {
      const c = liveStock(row);
      const tr = document.createElement("tr");
      tr.dataset.search = `${row.ticker || ""} ${row.stockName || ""}`.toLowerCase();
      if (tab === "overseasStock") tr.appendChild(textTd(row.ticker));
      tr.appendChild(cls(textTd(row.stockName), "col-name"));
      tr.appendChild(numTd(row.qty, "raw"));
      if (tab === "domesticStock") tr.appendChild(numTd(row.price, "raw"));
      tr.appendChild(numTd(c.value, "eok"));
      tr.appendChild(numTd(c.weight, "pct"));
      tr.appendChild(cls(retTd(c.pnl, "won0"), "col-detail"));
      tr.appendChild(retTd(c.returnPct, "ret"));
      tr.appendChild(cls(numTd(row.weightPrev, "pct"), "col-detail"));
      tr.appendChild(deleteTd(idx, row.stockName));
      return tr;
    }

    function futRow(row, idx, tab) {
      const c = liveFut(row);
      const tr = document.createElement("tr");
      tr.dataset.search = `${row.kind || ""} ${row.stockName || ""} ${row.ticker || ""}`.toLowerCase();
      tr.dataset.group = row.kind || "";
      tr.appendChild(textTd(row.kind));
      tr.appendChild(badgeTd(c.direction));
      tr.appendChild(cls(textTd(row.stockName), "col-name"));
      tr.appendChild(numTd(row.qty, "raw"));
      tr.appendChild(numTd(c.valueAbs, "eok"));
      tr.appendChild(numTd(c.weight, "pct"));
      tr.appendChild(cls(textTd(row.maturity), "col-detail"));
      tr.appendChild(cls(retTd(c.pnl, "won0"), "col-detail"));
      tr.appendChild(deleteTd(idx, row.stockName));
      return tr;
    }

    function optRow(row, idx) {
      const c = liveOpt(row);
      const tr = document.createElement("tr");
      tr.dataset.search = `${row.kind || ""} ${row.stockName || ""}`.toLowerCase();
      tr.dataset.group = row.kind || "";
      tr.appendChild(badgeTd(c.direction));
      tr.appendChild(cls(textTd(row.stockName), "col-name"));
      tr.appendChild(textTd(row.optExpiry || row.optUnder || ""));
      tr.appendChild(textTd(row.optCP || ""));
      tr.appendChild(numTd(row.optStrike, "raw"));
      tr.appendChild(numTd(row.qty, "raw"));
      tr.appendChild(numTd(c.valueAbs, "won0"));
      tr.appendChild(numTd(c.weight, "pct"));
      tr.appendChild(retTd(c.pnl, "won0"));
      tr.appendChild(deleteTd(idx, row.stockName));
      return tr;
    }

    // ---- 셀 헬퍼 ----
    function cls(td, c) { td.classList.add(c); return td; }
    function textTd(v) { const td = document.createElement("td"); td.textContent = v == null ? "" : v; if (v != null && String(v).length > 12) td.title = v; return td; }
    function numTd(v, type) { const td = document.createElement("td"); td.className = "number"; td.textContent = fmt(v, type); return td; }
    function retTd(v, type) { const td = numTd(v, type); if (v != null && !Number.isNaN(v)) td.classList.add(v >= 0 ? "ret-pos" : "ret-neg"); return td; }
    function badgeTd(dir) {
      const td = document.createElement("td"); td.className = "center";
      const span = document.createElement("span"); span.className = "badge " + (dir === "매수" ? "b" : "s"); span.textContent = dir;
      td.appendChild(span); return td;
    }
    function editTd(idx, field, value, kind, extra) {
      const td = document.createElement("td");
      if (kind === "num") td.className = "number";
      const inp = document.createElement("input");
      inp.className = "editable" + (kind === "num" ? " num-input" : "") + (extra ? " " + extra : "");
      inp.value = (kind === "num") ? (value == null ? "" : value) : (value || "");
      inp.dataset.idx = idx; inp.dataset.field = field; inp.dataset.kind = kind;
      td.appendChild(inp);
      return td;
    }
    function deleteTd(idx, name) {
      const td = document.createElement("td"); td.className = "center";
      const b = document.createElement("button"); b.type = "button"; b.className = "delete-row danger"; b.textContent = "✕"; b.title = "행 삭제";
      b.addEventListener("click", () => {
        if (name && !confirm(`'${name}' 행을 삭제할까요?`)) return;
        const row = state.assets[state.assetTab][idx];
        if (row.isNew || !row.sourceRow) state.assets[state.assetTab].splice(idx, 1);
        else row.deleted = true;
        render();
      });
      td.appendChild(b);
      return td;
    }

    function renderFoot() {
      const tr = document.getElementById("tfootRow");
      tr.innerHTML = "";
      const tab = state.assetTab;
      const heads = HEADERS[tab];
      const rows = activeRows();
      const cells = new Array(heads.length).fill("");
      cells[0] = `합계 (${rows.length})`;
      if (isOpt(tab)) {
        const lives = rows.map(liveOpt);
        cells[5] = fmt(sum(rows.map((r) => Math.abs(num(r.qty) || 0))), "raw");
        cells[6] = fmt(sum(lives.map((c) => c.valueAbs)), "won0");
        cells[7] = formatPct(sum(lives.map((c) => c.weight)));
        cells[8] = fmt(sum(lives.map((c) => c.pnl || 0)), "won0");
      } else if (isFut(tab)) {
        const lives = rows.map(liveFut);
        cells[3] = fmt(sum(rows.map((r) => Math.abs(num(r.qty) || 0))), "raw");
        cells[4] = fmt(sum(lives.map((c) => c.valueAbs)), "eok");
        cells[5] = formatPct(sum(lives.map((c) => c.weight)));
        cells[7] = fmt(sum(lives.map((c) => c.pnl || 0)), "won0");
      } else {
        const lives = rows.map(liveStock);
        cells[3] = fmt(sum(lives.map((c) => c.value)), "eok");
        cells[4] = formatPct(sum(lives.map((c) => c.weight)));
        cells[5] = fmt(sum(lives.map((c) => c.pnl || 0)), "won0");
      }
      cells.forEach((txt, i) => { const td = document.createElement("td"); td.textContent = txt; if (i > 0) td.className = "number"; if (heads[i] && heads[i][1] === "detail") td.classList.add("col-detail"); tr.appendChild(td); });
    }

    function activeRows() { return state.assets[state.assetTab].filter((r) => !r.deleted); }
    function activeList(tab) { return state.assets[tab].filter((r) => !r.deleted); }
    function activeCount(tab) { return activeList(tab).length; }
    function tabWeight(tab) {
      const calc = isOpt(tab) ? liveOpt : isFut(tab) ? liveFut : liveStock;
      return sum(activeList(tab).map((r) => calc(r).weight));
    }

    // ---- 좌측 패널 ----
    // ---- 메트릭/차트 ----
    function renderMetrics() {
      const ds = activeList("domesticStock").map(liveStock);
      const os = activeList("overseasStock").map(liveStock);
      const domVal = sum(ds.map((c) => c.value));
      const ovsVal = sum(os.map((c) => c.value));
      const stockVal = domVal + ovsVal;
      const df = activeList("domesticFutures").map(liveFut);
      const of = activeList("overseasFutures").map(liveFut);
      const op = activeList("options").map(liveOpt);
      const futNet = sum(df.map((c) => dirSign(c.direction) * c.valueAbs)) + sum(of.map((c) => dirSign(c.direction) * c.valueAbs));
      const futGross = sum(df.map((c) => c.valueAbs)) + sum(of.map((c) => c.valueAbs));
      const optPrem = sum(op.map((c) => c.valueAbs));
      const cashEtfPnl = sum((DATA.cashOther || []).filter((r) => r.isCashEtf).map((r) => r.pnl || 0));
      const pnl = sum(ds.map((c) => c.pnl || 0)) + sum(os.map((c) => c.pnl || 0)) + sum(df.map((c) => c.pnl || 0)) + sum(of.map((c) => c.pnl || 0)) + sum(op.map((c) => c.pnl || 0)) + cashEtfPnl;
      const cnt = TABS.reduce((a, t) => a + activeCount(t), 0);
      const nav = DATA.nav || 1;

      // 주식 순노출 = 주식 롱 + 선물 순 (선물까지 고려한 포지션)
      const equityNet = stockVal + futNet;
      // 그로스 = 주식 + 선물 노셔널 + 옵션 프리미엄 (총계 행 col31 정의와 동일)
      const gross = stockVal + futGross + optPrem;

      document.getElementById("mNav").textContent = formatEok(DATA.nav);
      document.getElementById("mNavSub").textContent = DATA.navSource || "";
      document.getElementById("mPnl").textContent = formatEok(pnl);
      document.getElementById("mPnlBox").className = "metric " + (pnl >= 0 ? "ok" : "bad");
      document.getElementById("mPnlSub").textContent = "주식+선물+옵션 · 현금성 제외";
      const stockCnt = activeCount("domesticStock") + activeCount("overseasStock");   // ETF도 1종목
      document.getElementById("mCount").textContent = stockCnt.toLocaleString("ko-KR") + "개";
      document.getElementById("mCountSub").textContent = `국내 ${activeCount("domesticStock")} · 해외 ${activeCount("overseasStock")} (ETF 포함, 선물·옵션 제외)`;
      document.getElementById("mNet").textContent = formatPct(equityNet / nav);
      document.getElementById("mNetSub").textContent = `${formatEok(equityNet)} · 주식 ${formatPct(stockVal / nav)} + 선물순 ${formatPct(futNet / nav)}`;
      document.getElementById("mDeriv").textContent = formatPct((futGross + optPrem) / nav);
      document.getElementById("mDerivSub").textContent = `선물 ${formatPct(futGross / nav)} (${formatEok(futGross)}) · 옵션 ${formatPct(optPrem / nav)} (${formatEok(optPrem)})`;
      document.getElementById("mNetBox").className = "metric " + (equityNet >= 0 ? "ok" : "bad");
      document.getElementById("mGross").textContent = formatPct(gross / nav);
      document.getElementById("mGrossSub").textContent = `${formatEok(gross)}${DATA.totalsRow && DATA.totalsRow.weight ? " · 총계행 " + formatPct(DATA.totalsRow.weight) : ""}`;
      document.getElementById("mCash").textContent = DATA.cashDepositValue == null ? "-" : formatEok(DATA.cashDepositValue);
      document.getElementById("mCashSub").textContent = `현금 ${formatPct(DATA.cashDepositWeight)} · 증거금 ${formatEok(DATA.marginValue || 0)}`;
    }

    // ---- SVG 파이 (가득 찬 원 + 내부 라벨, 좁은 조각은 리더라인) ----
    const SVGNS = "http://www.w3.org/2000/svg";
    function _el(tag, attrs) {
      const e = document.createElementNS(SVGNS, tag);
      for (const k in attrs) e.setAttribute(k, attrs[k]);
      return e;
    }
    const PIE_SHADE = { domestic: ["#ffb347", "#e07d00"], overseas: ["#e0900f", "#9c5f08"], domesticFut: ["#a56a16", "#6e470d"], overseasFut: ["#75603a", "#463719"], opt: ["#ff5f63", "#c1373b"], cash: ["#4a4a4a", "#2b2b2b"] };
    function _pieDefs(keys) {
      const ns = "http://www.w3.org/2000/svg";
      const defs = document.createElementNS(ns, "defs");
      // 조각별 방사형 그라디언트 (안쪽 밝게 → 바깥 어둡게)
      keys.forEach((k) => {
        const g = document.createElementNS(ns, "radialGradient");
        g.setAttribute("id", "pieG_" + k);
        g.setAttribute("cx", "50%"); g.setAttribute("cy", "50%"); g.setAttribute("r", "72%");
        const sh = PIE_SHADE[k] || ["#888", "#555"];
        const s1 = document.createElementNS(ns, "stop"); s1.setAttribute("offset", "35%"); s1.setAttribute("stop-color", sh[0]);
        const s2 = document.createElementNS(ns, "stop"); s2.setAttribute("offset", "100%"); s2.setAttribute("stop-color", sh[1]);
        g.appendChild(s1); g.appendChild(s2); defs.appendChild(g);
      });
      // 드롭 섀도 필터
      const f = document.createElementNS(ns, "filter");
      f.setAttribute("id", "pieShadow"); f.setAttribute("x", "-20%"); f.setAttribute("y", "-20%"); f.setAttribute("width", "140%"); f.setAttribute("height", "140%");
      f.innerHTML = '<feDropShadow dx="0" dy="6" stdDeviation="10" flood-color="#000000" flood-opacity="0.55"/>';
      defs.appendChild(f);
      // 홀 내부 은은한 글로우
      const hg = document.createElementNS(ns, "radialGradient");
      hg.setAttribute("id", "pieHoleGlow"); hg.setAttribute("cx", "50%"); hg.setAttribute("cy", "42%"); hg.setAttribute("r", "60%");
      hg.innerHTML = '<stop offset="0%" stop-color="#ffa028" stop-opacity="0.18"/><stop offset="100%" stop-color="#0a0a0a" stop-opacity="0"/>';
      defs.appendChild(hg);
      return defs;
    }
    function drawPieSVG(svg, parts, total, nav) {
      svg.innerHTML = "";
      const CX = 460, CY = 380, R = 310;
      svg.appendChild(_pieDefs(parts.map((p) => p.key)));
      // 바닥 그림자용 원반 (입체 받침)
      svg.appendChild(_el("circle", { cx: CX, cy: CY + 4, r: R, fill: "#000000", opacity: "0.35", filter: "url(#pieShadow)" }));
      let ang = -90;                         // 12시 방향 시작
      const outLabels = [];
      const MIN_FS = 17, MAX_FS = 26;
      const lr = R * 0.72;
      // 1-pass: 각 조각의 수용 가능 글자크기 계산 → 내부 라벨 공통 크기 결정
      const metas = [];
      let a = -90;
      parts.forEach((p) => {
        const sweep = p.val / total * 360;
        const name = p.label.replace(/ \(.*\)/, "");
        const chord = 2 * lr * Math.sin(Math.min(sweep, 180) / 2 * Math.PI / 180);
        const fitFs = chord * 0.92 / name.length;      // 한글 글자폭 ≈ 폰트크기
        metas.push({ p, sweep, name, inside: fitFs >= MIN_FS, fitFs });
        a += sweep;
      });
      const insideFits = metas.filter((m) => m.inside).map((m) => m.fitFs);
      const fsName = insideFits.length ? Math.min(MAX_FS, Math.min.apply(null, insideFits)) : MAX_FS;
      const fsPct = Math.max(15, fsName - 4);
      // 2-pass: 렌더 (내부 라벨은 전부 동일 크기)
      metas.forEach(({ p, sweep, name, inside }) => {
        const a0 = ang * Math.PI / 180, a1 = (ang + sweep) * Math.PI / 180;
        const mid = (ang + sweep / 2) * Math.PI / 180;
        const x0 = CX + R * Math.cos(a0), y0 = CY + R * Math.sin(a0);
        const x1 = CX + R * Math.cos(a1), y1 = CY + R * Math.sin(a1);
        const large = sweep > 180 ? 1 : 0;
        const path = sweep >= 359.99
          ? `M ${CX - R} ${CY} A ${R} ${R} 0 1 1 ${CX + R} ${CY} A ${R} ${R} 0 1 1 ${CX - R} ${CY} Z`
          : `M ${CX} ${CY} L ${x0} ${y0} A ${R} ${R} 0 ${large} 1 ${x1} ${y1} Z`;
        const seg = _el("path", { d: path, fill: `url(#pieG_${p.key})`, stroke: "#0a0a0a", "stroke-width": 2, class: "seg" + (state.allocSel === p.key ? " selected" : ""), "data-key": p.key });
        seg.addEventListener("click", () => { state.allocSel = (state.allocSel === p.key) ? null : p.key; renderAlloc(); });
        svg.appendChild(seg);
        const pctTxt = formatPct(p.val / nav);
        if (inside) {
          const lx = CX + lr * Math.cos(mid), ly = CY + lr * Math.sin(mid);
          const t = _el("text", { x: lx, y: ly - fsName * 0.32, "text-anchor": "middle", class: "pie-label-in", fill: CHART_TEXT[p.key], "pointer-events": "none", "font-size": fsName });
          t.textContent = name;
          const t2 = _el("text", { x: lx, y: ly + fsPct * 1.05, "text-anchor": "middle", class: "pie-label-in", fill: CHART_TEXT[p.key], "pointer-events": "none", "font-size": fsPct });
          t2.textContent = pctTxt;
          svg.appendChild(t); svg.appendChild(t2);
        } else {
          outLabels.push({ mid, key: p.key, label: `${name} ${pctTxt}` });
        }
        ang += sweep;
      });
      // 중앙 도넛 홀 (입체 + 정보 표시)
      const holeR = R * 0.42;
      svg.appendChild(_el("circle", { cx: CX, cy: CY, r: holeR, fill: "#0a0a0a", stroke: "#2a2a2a", "stroke-width": 1.5 }));
      svg.appendChild(_el("circle", { cx: CX, cy: CY, r: holeR, fill: "url(#pieHoleGlow)", opacity: "0.6" }));
      const navLabel = _el("text", { x: CX, y: CY - 10, "text-anchor": "middle", "font-size": 20, "font-weight": "700", fill: "#8a877f", "pointer-events": "none" });
      navLabel.textContent = "NAV";
      const navVal = _el("text", { x: CX, y: CY + 22, "text-anchor": "middle", "font-size": 30, "font-weight": "800", fill: "#ffa028", "pointer-events": "none" });
      navVal.textContent = formatEok(nav);
      svg.appendChild(navLabel); svg.appendChild(navVal);

      // 리더라인 라벨: 좌/우로 나눠 세로 겹침 방지
      const sides = { left: [], right: [] };
      outLabels.forEach((o) => { (Math.cos(o.mid) >= 0 ? sides.right : sides.left).push(o); });
      ["left", "right"].forEach((side) => {
        const list = sides[side].sort((a, b) => Math.sin(a.mid) - Math.sin(b.mid));
        let prevY = -1e9;
        list.forEach((o) => {
          const sx = CX + R * Math.cos(o.mid), sy = CY + R * Math.sin(o.mid);
          const ex = CX + (R + 26) * Math.cos(o.mid);
          let ey = CY + (R + 26) * Math.sin(o.mid);
          if (ey - prevY < 24) ey = prevY + 24;        // 겹침 방지
          prevY = ey;
          const hx = side === "right" ? ex + 16 : ex - 16;
          svg.appendChild(_el("polyline", { points: `${sx},${sy} ${ex},${ey} ${hx},${ey}`, class: "pie-leader" }));
          const t = _el("text", { x: side === "right" ? hx + 4 : hx - 4, y: ey + 4, "text-anchor": side === "right" ? "start" : "end", class: "pie-label-out", "data-key": o.key });
          t.textContent = o.label;
          t.style.cursor = "pointer";
          t.addEventListener("click", () => { state.allocSel = (state.allocSel === o.key) ? null : o.key; renderAlloc(); });
          svg.appendChild(t);
        });
      });
    }

    function renderAlloc() {
      const domVal = sum(activeList("domesticStock").map((r) => liveStock(r).value));
      const ovsVal = sum(activeList("overseasStock").map((r) => liveStock(r).value));
      const nav = DATA.nav || (domVal + ovsVal);
      const cashVal = Math.max(0, nav - domVal - ovsVal);
      const domFutVal = sum(activeList("domesticFutures").map((r) => liveFut(r).valueAbs));
      const ovsFutVal = sum(activeList("overseasFutures").map((r) => liveFut(r).valueAbs));
      const optVal = sum(activeList("options").map((r) => liveOpt(r).valueAbs));
      const parts = [
        { key: "domestic", label: "국내주식", val: domVal },
        { key: "overseas", label: "해외주식", val: ovsVal },
        { key: "domesticFut", label: "국내선물 (노셔널)", val: domFutVal },
        { key: "overseasFut", label: "해외선물 (노셔널)", val: ovsFutVal },
        { key: "opt", label: "옵션 (프리미엄)", val: optVal },
        { key: "cash", label: "현금·기타", val: cashVal },
      ];
      const total = sum(parts.map((p) => p.val)) || 1;   // 파이 비율 분모 = 그로스+현금
      drawPieSVG(document.getElementById("allocPie"), parts.filter((p) => p.val > 0), total, nav);
      const legend = document.getElementById("allocLegend"); legend.innerHTML = "";
      parts.filter((p) => p.val).forEach((p) => {
        const item = document.createElement("div"); item.className = "legend-item";
        item.style.cursor = "pointer";
        if (state.allocSel === p.key) item.style.fontWeight = "700";
        item.addEventListener("click", () => { state.allocSel = (state.allocSel === p.key) ? null : p.key; renderAlloc(); });
        const sw = document.createElement("span"); sw.className = "legend-swatch"; sw.style.background = CHART_COLORS[p.key];
        const lb = document.createElement("span"); lb.className = "legend-label"; lb.textContent = p.label;
        const pc = document.createElement("span"); pc.className = "legend-pct"; pc.textContent = formatPct(p.val / nav);
        const ek = document.createElement("span"); ek.className = "legend-eok"; ek.textContent = formatEok(p.val);
        item.appendChild(sw); item.appendChild(lb); item.appendChild(pc); item.appendChild(ek); legend.appendChild(item);
      });
      renderAllocDetail(nav);
    }

    // ---- 파이 항목 클릭 → 자산군 요약 ----
    const ALLOC_MAP = {
      domestic: { tab: "domesticStock", label: "국내주식" },
      overseas: { tab: "overseasStock", label: "해외주식" },
      domesticFut: { tab: "domesticFutures", label: "국내선물" },
      overseasFut: { tab: "overseasFutures", label: "해외선물" },
      opt: { tab: "options", label: "옵션" },
      cash: { tab: null, label: "현금·기타" },
    };
    function renderAllocDetail(nav) {
      const box = document.getElementById("allocDetail");
      const key = state.allocSel;
      if (!key) { box.innerHTML = '<div class="alloc-detail-empty">다이어그램의 항목을 클릭하면 상세가 표시됩니다</div>'; return; }
      const cfg = ALLOC_MAP[key];
      box.innerHTML = "";
      const head = document.createElement("div"); head.className = "alloc-detail-head";
      const body = document.createElement("div"); body.className = "alloc-detail-body";
      const closeBtn = '<button type="button" class="alloc-close" title="닫기">✕</button>';
      const kv = (k, v) => { const d = document.createElement("div"); d.className = "alloc-detail-row"; d.innerHTML = `<span class="k">${k}</span><span class="v">${v}</span>`; body.appendChild(d); };

      if (key === "cash") {
        const dep = DATA.cashOther.filter((r) => r.isDeposit);
        const mgn = DATA.cashOther.filter((r) => (r.kind || "").includes("증거금"));
        head.innerHTML = `<span>현금·기타</span><span style="display:inline-flex;align-items:center;gap:8px;">${formatEok((DATA.cashDepositValue || 0) + (DATA.marginValue || 0))}${closeBtn}</span>`;
        kv("현금 (예치금+예금)", `${formatEok(DATA.cashDepositValue)} · ${formatPct(DATA.cashDepositWeight)}`);
        kv("위탁증거금", formatEok(DATA.marginValue));
        const list = document.createElement("div"); list.className = "alloc-detail-list";
        dep.concat(mgn).forEach((r) => {
          const d = document.createElement("div"); d.className = "alloc-detail-row";
          d.innerHTML = `<span class="k">${r.stockName} <span style="color:var(--muted);font-size:11px;">${r.kind}</span></span><span class="v">${formatEok(r.valueKrw)}</span>`;
          list.appendChild(d);
        });
        body.appendChild(list);
      } else {
        const tab = cfg.tab;
        const rows = activeList(tab);
        const calc = isOpt(tab) ? liveOpt : isFut(tab) ? liveFut : liveStock;
        const lives = rows.map((r) => ({ r, c: calc(r) }));
        const isDeriv = isFut(tab) || isOpt(tab);
        const totVal = sum(lives.map((x) => isDeriv ? x.c.valueAbs : x.c.value));
        const totPnl = sum(lives.map((x) => x.c.pnl || 0));
        head.innerHTML = `<span>${cfg.label} (${rows.length})</span><span style="display:inline-flex;align-items:center;gap:8px;">${formatEok(totVal)}${closeBtn}</span>`;
        kv("%NAV", formatPct(totVal / nav));
        if (isDeriv) {
          const net = isOpt(tab) ? null : sum(lives.map((x) => dirSign(x.c.direction) * x.c.valueAbs));
          if (net != null) kv("순노출", `${formatEok(net)} · ${formatPct(net / nav)}`);
          kv("매수 / 매도", `${lives.filter((x) => x.c.direction === "매수").length} / ${lives.filter((x) => x.c.direction === "매도").length}`);
        }
        if (isOpt(tab)) {
          const callQty = sum(lives.filter((x) => x.r.optCP === "C").map((x) => Math.abs(num(x.r.qty) || 0)));
          const putQty = sum(lives.filter((x) => x.r.optCP === "P").map((x) => Math.abs(num(x.r.qty) || 0)));
          kv("콜 / 풋 계약수", `${fmt(callQty, "raw")} / ${fmt(putQty, "raw")}`);
        }
        kv("평가손익 합", fmt(totPnl, "won0") + "원");
        // 전체 목록 (표, 순종목비 내림차순)
        const wrap = document.createElement("div"); wrap.className = "alloc-detail-scroll";
        const tbl = document.createElement("table"); tbl.className = "mini-table";
        const optCols = isOpt(tab) ? "<th>C/P</th><th>계약수</th>" : "";
        tbl.innerHTML = `<thead><tr><th style="text-align:left;">종목명</th>${isDeriv ? "<th>방향</th>" : ""}${optCols}<th>평가액</th><th>비중</th><th>손익</th></tr></thead>`;
        const tb = document.createElement("tbody");
        const sorted = lives.slice().sort((a, b) => (b.c.weight || 0) - (a.c.weight || 0));
        sorted.forEach((x) => {
          const tr = document.createElement("tr");
          const pnlV = x.c.pnl;
          const pnlCls = pnlV == null ? "" : (pnlV >= 0 ? "ret-pos" : "ret-neg");
          const dirTd = isDeriv ? `<td class="center"><span class="badge ${x.c.direction === "매수" ? "b" : "s"}" style="min-width:auto;height:18px;font-size:11px;">${x.c.direction}</span></td>` : "";
          const optTd = isOpt(tab) ? `<td class="center">${x.r.optCP || ""}</td><td class="number">${fmt(Math.abs(num(x.r.qty) || 0), "raw")}</td>` : "";
          tr.innerHTML = `<td class="nm" title="${x.r.stockName}">${x.r.stockName}</td>${dirTd}${optTd}`
            + `<td class="number">${formatEok(isDeriv ? x.c.valueAbs : x.c.value)}</td>`
            + `<td class="number">${formatPct(x.c.weight)}</td>`
            + `<td class="number ${pnlCls}">${pnlV == null ? "" : fmt(pnlV, "won0")}</td>`;
          tb.appendChild(tr);
        });
        tbl.appendChild(tb);
        wrap.appendChild(tbl);
        body.appendChild(wrap);
      }
      box.appendChild(head); box.appendChild(body);
      box.querySelector(".alloc-close").addEventListener("click", () => { state.allocSel = null; renderAlloc(); });
    }

    // ---- 익스포저 분해 ----
    // ---- 기초자산별 포지션 구조 ----
    const _OVS_ROOT = /^([A-Z][A-Z0-9 ]*?)[FGHJKMNQUVXZ]\d(?:[CP]\d+(?:\.\d+)?)?$/;
    function ovsRoot(code) {
      const m = _OVS_ROOT.exec((code || "").toUpperCase().trim());
      return m ? m[1].trim() : null;
    }
    const MONTH_CODE = { F:1,G:2,H:3,J:4,K:5,M:6,N:7,Q:8,U:9,V:10,X:11,Z:12 };
    function ovsExpiry(code) {
      // 'CLQ6C90' / 'NQU6' → 월물문자+연도숫자에서 만기월 파생 (Q6 → 2026-08)
      const m = /([FGHJKMNQUVXZ])(\d)(?:[CP]\d+(?:\.\d+)?)?$/.exec((code || "").toUpperCase().trim());
      if (!m) return "";
      const mon = MONTH_CODE[m[1]];
      const nowY = new Date().getFullYear();
      // 연도 한 자리: 현재 십년대 기준 근접 연도 선택
      let y = Math.floor(nowY / 10) * 10 + Number(m[2]);
      if (y < nowY - 2) y += 10;
      return `${y}-${String(mon).padStart(2, "0")}`;
    }
    function cleanOvsName(name) {
      return (name || "").replace(/\s*\(\d{6}\)\s*$/, "").replace(/\s+[CP]\d+(?:\.\d+)?\s*$/i, "").trim();
    }

    function renderNetting() {
      const body = document.getElementById("nettingBody");
      body.innerHTML = "";
      const nav = DATA.nav || 1;
      // U[key] = { name, spot, futLong, futShort, optBullPrem, optBearPrem, bullLegs[], bearLegs[] }
      const U = {};
      const get = (key, name) => (U[key] = U[key] || { name, spot: 0, futLong: 0, futShort: 0, optBullPrem: 0, optBearPrem: 0, futBuyQty: 0, futSellQty: 0, optLegs: {} });

      // ① 현물: 종목별 (개별주식선물 넷팅용)
      activeList("domesticStock").forEach((r) => {
        const c = liveStock(r);
        get(r.stockName, r.stockName).spot += c.value;
      });

      // ② 국내선물
      activeList("domesticFutures").forEach((r) => {
        const c = liveFut(r);
        const kind = r.kind || "";
        let u;
        if (kind === "개별주식선물") u = get(r.futUnder || r.stockName, r.futUnder || r.stockName);
        else if (kind === "주가지수선물") { u = get("__K200__", "KOSPI200"); }
        else if (kind === "코스닥150선물") { u = get("__KQ150__", "KOSDAQ150"); }
        else if (kind === "통화선물") { u = get("__FX__", "USD/KRW (달러선물)"); }
        else u = get(kind, kind);
        const fq = Math.abs(num(r.qty) || 0);
        if (c.direction === "매도") { u.futShort += c.valueAbs; u.futSellQty += fq; }
        else { u.futLong += c.valueAbs; u.futBuyQty += fq; }
      });

      // ③ 해외선물: 코드 루트로 그룹
      activeList("overseasFutures").forEach((r) => {
        const c = liveFut(r);
        const root = ovsRoot(r.code) || cleanOvsName(r.stockName);
        const u = get("OVS:" + root, cleanOvsName(r.stockName));
        const fq = Math.abs(num(r.qty) || 0);
        if (c.direction === "매도") { u.futShort += c.valueAbs; u.futSellQty += fq; }
        else { u.futLong += c.valueAbs; u.futBuyQty += fq; }
      });

      // ④ 옵션: 콜매수/풋매도=상방, 풋매수/콜매도=하방
      activeList("options").forEach((r) => {
        const c = liveOpt(r);
        const isDomIdx = (r.kind || "").endsWith("옵션");     // 국내 주가지수옵션
        let u;
        if (isDomIdx) { u = get("__K200__", "KOSPI200"); }
        else {
          const root = ovsRoot(r.code) || (r.optUnder || cleanOvsName(r.stockName));
          u = get("OVS:" + root, u && u.name ? u.name : (r.optUnder || cleanOvsName(r.stockName)));
        }
        const buy = c.direction === "매수";
        const cp = r.optCP || "?";
        const strike = r.optStrike;
        const legKey = `${cp}|${strike}`;   // 같은 CP·행사가는 매수/매도 상계
        const leg = u.optLegs[legKey] = u.optLegs[legKey] || { cp, strike, buyQty: 0, sellQty: 0, buyPrem: 0, sellPrem: 0, expiry: "" };
        if (!leg.expiry) leg.expiry = r.optExpiry || r.maturity || ovsExpiry(r.code);
        const q = Math.abs(num(r.qty) || 0);
        if (buy) { leg.buyQty += q; leg.buyPrem += c.valueAbs; }
        else { leg.sellQty += q; leg.sellPrem += c.valueAbs; }
      });

      const domStockTotalV = sum(activeList("domesticStock").map((r) => liveStock(r).value));

      // ⑤ 섹션 분리 렌더: 선물 섹션 → 옵션 섹션
      const sectionRow = (label) => {
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 7; td.textContent = label;
        td.style.cssText = "background:#000;color:#ffa028;font-weight:800;font-size:15px;letter-spacing:0.06em;padding:11px 14px;border-top:2px solid #ffa028;border-bottom:1px solid #3d3d3d;";
        tr.appendChild(td); return tr;
      };
      const verdictOf = (spot, bullVal, bearVal) => {
        const hasBull = spot > 0 || bullVal > 0;
        const hasBear = bearVal > 0;
        const net = spot + bullVal - bearVal;
        const gross = Math.abs(spot) + bullVal + bearVal;
        if (hasBull && !hasBear) return { v: "롱 베팅", c: "ret-pos", net };
        if (hasBear && !hasBull) return { v: "숏 베팅", c: "ret-neg", net };
        const tilt = gross ? net / gross : 0;
        if (Math.abs(tilt) < 0.05) return { v: "헤지 · 중립", c: "gap-zero", net };
        if (tilt > 0) return { v: `헤지 · 롱 우위 (${formatPct(tilt)})`, c: "ret-pos", net };
        return { v: `헤지 · 숏 우위 (${formatPct(-tilt)})`, c: "ret-neg", net };
      };
      const pushRow = (name, spot, futNetVal, pillsHtml, net, verdict) => {
        const tr = document.createElement("tr");
        tr.appendChild(cls(textTd(name), "col-name"));
        tr.appendChild(numTd(spot || null, "eok"));
        tr.appendChild(retTd(futNetVal, "eok"));
        const td = document.createElement("td");
        td.style.whiteSpace = "normal"; td.style.maxWidth = "300px";
        td.innerHTML = pillsHtml;
        tr.appendChild(td);
        tr.appendChild(retTd(net, "eok"));
        tr.appendChild(numTd(net / nav, "pct"));
        const vt = textTd(verdict.v); vt.classList.add(verdict.c); vt.style.fontWeight = "700";
        tr.appendChild(vt);
        body.appendChild(tr);
      };

      // --- 선물 섹션 ---
      const futRows = Object.entries(U).filter(([, u]) => u.futLong || u.futShort)
        .map(([key, u]) => {
          const futNet = u.futLong - u.futShort;
          const netQ = u.futBuyQty - u.futSellQty;
          let pill;
          if (u.futBuyQty && u.futSellQty) {
            const detail = `(매수 ${fmt(u.futBuyQty, "raw")} · 매도 ${fmt(u.futSellQty, "raw")})`;
            pill = netQ > 0 ? `<span class="opt-pill up">▲ 선물 순매수 ${fmt(netQ, "raw")}계약 ${detail}</span>`
              : netQ < 0 ? `<span class="opt-pill down">▼ 선물 순매도 ${fmt(-netQ, "raw")}계약 ${detail}</span>`
              : `<span class="opt-pill flat">선물 상쇄 ${detail}</span>`;
          } else if (u.futBuyQty) pill = `<span class="opt-pill up">▲ 선물매수 ${fmt(u.futBuyQty, "raw")}계약</span>`;
          else pill = `<span class="opt-pill down">▼ 선물매도 ${fmt(u.futSellQty, "raw")}계약</span>`;
          const verdict = verdictOf(u.spot, Math.max(futNet, 0), Math.max(-futNet, 0));
          const domestic = !key.startsWith("OVS:");
          return { name: u.name, domestic, spot: u.spot, futNet, pill, net: u.spot + futNet, verdict };
        })
        .sort((a, b) => (b.domestic - a.domestic) || (Math.abs(b.net) - Math.abs(a.net)));
      if (futRows.length) {
        body.appendChild(sectionRow(`선물 (${futRows.length})`));
        const mkFutGroup = (label) => {
          const tr = document.createElement("tr");
          const td = document.createElement("td");
          td.colSpan = 7;
          td.innerHTML = `<span style="display:inline-block;width:9px;height:9px;border-radius:2px;background:#ffa028;margin-right:9px;vertical-align:middle;"></span>${label}`;
          td.style.cssText = "background:linear-gradient(90deg,#3a2a10 0%,#1b1b1b 60%);color:#ffc46b;font-weight:800;font-size:13px;letter-spacing:0.04em;padding:9px 14px;border-left:4px solid #ffa028;border-top:1px solid #2a2a2a;text-transform:uppercase;";
          tr.appendChild(td); return tr;
        };
        let prevDom = null;
        futRows.forEach((x) => {
          if (x.domestic !== prevDom) { body.appendChild(mkFutGroup(x.domestic ? "국내 선물" : "해외 선물")); prevDom = x.domestic; }
          pushRow(x.name, x.spot, x.futNet, x.pill, x.net, x.verdict);
        });
      }

      // --- 옵션 섹션: 전용 열 구성 (기초자산/종류/행사가/포지션/만기월/순노출/판정) ---
      const optLegRows = [];
      Object.entries(U).forEach(([key, u]) => {
        const domestic = !key.startsWith("OVS:");
        Object.values(u.optLegs).forEach((L) => {
          optLegRows.push({ under: u.name, domestic, ...L, netQty: L.buyQty - L.sellQty, netPrem: L.buyPrem - L.sellPrem });
        });
      });
      // 국내 → 해외, 그 안에서 콜 전부 → 풋 전부 (같은 종류 안에서는 기초자산 → 행사가 순)
      optLegRows.sort((a, b) =>
        (b.domestic - a.domestic)
        || (a.cp === b.cp ? 0 : (a.cp === "C" ? -1 : 1))
        || a.under.localeCompare(b.under, "ko")
        || ((a.strike || 0) - (b.strike || 0)));

      const optVerdict = (cp, netQty) => {
        if (netQty === 0) return "상쇄";
        if (cp === "C") return netQty > 0 ? "상승 베팅" : "프리미엄 수취 (상승 제한)";
        return netQty > 0 ? "하락 방어·베팅" : "프리미엄 수취 (하락 감수)";
      };

      if (optLegRows.length) {
        body.appendChild(sectionRow(`옵션 (${optLegRows.length})`));
        // 옵션 전용 헤더 행
        const mkOptHead = () => {
          const hd = document.createElement("tr");
          ["기초자산", "종류", "행사가", "포지션 (계약)", "만기월", "순노출 (프리미엄)", "포지션 판정"].forEach((h, i) => {
            const td = document.createElement("td");
            td.textContent = h;
            td.style.cssText = "background:#0f0f0f;font-weight:700;color:#8a877f;font-size:12px;"
              + (i >= 2 && i <= 5 ? "text-align:right;" : "");
            if (i === 1) td.style.textAlign = "center";
            hd.appendChild(td);
          });
          return hd;
        };
        const mkGroupRow = (label) => {
          const tr = document.createElement("tr");
          const td = document.createElement("td");
          td.colSpan = 7; td.textContent = label;
          td.style.cssText = "background:#1b1b1b;color:#e8e6e1;font-weight:700;font-size:12.5px;padding:5px 12px;border-left:3px solid #ffa028;";
          tr.appendChild(td); return tr;
        };
        body.appendChild(mkOptHead());
        // 국내 옵션 종합 전략 판정
        const domLegs = optLegRows.filter((L) => L.domestic && L.netQty !== 0);
        const callBuy = domLegs.filter((L) => L.cp === "C" && L.netQty > 0);
        const callSell = domLegs.filter((L) => L.cp === "C" && L.netQty < 0);
        const putBuy = domLegs.filter((L) => L.cp === "P" && L.netQty > 0);
        const putSell = domLegs.filter((L) => L.cp === "P" && L.netQty < 0);
        const cB = callBuy.length > 0, cS = callSell.length > 0, pB = putBuy.length > 0, pS = putSell.length > 0;
        const strikesEq = (as, bs) => as.length === 1 && bs.length === 1 && as[0].strike === bs[0].strike;
        const domPrem = sum(domLegs.map((L) => L.netPrem));   // 순프리미엄(음수=수취 우위)
        function classifyDomOptions() {
          if (!domLegs.length) return null;
          // 보유 주식 롱 + 풋 순매수 → 포트폴리오 보험
          if (pB && !cB && !cS && !pS && domStockTotalV > 0)
            return { label: "포트폴리오 보험 (Protective Put)", desc: "주식 롱 보유분의 하락 방어" };
          // 콜매수 + 풋매수 → 스트래들/스트랭글 (변동성 롱)
          if (cB && pB && !cS && !pS) {
            const eq = strikesEq(callBuy, putBuy);
            return { label: eq ? "롱 스트래들" : "롱 스트랭글", desc: "양방향 변동성 확대 베팅" };
          }
          // 콜매도 + 풋매도 → 숏 스트래들/스트랭글 (변동성 숏)
          if (cS && pS && !cB && !pB)
            return { label: strikesEq(callSell, putSell) ? "숏 스트래들" : "숏 스트랭글", desc: "변동성 축소·프리미엄 수취" };
          // 콜만
          if ((cB || cS) && !pB && !pS)
            return cB && !cS ? { label: "콜 방향성 (상승 베팅)", desc: "" }
              : cS && !cB ? { label: "콜 매도 (상승 제한·프리미엄)", desc: "" }
              : { label: "콜 스프레드", desc: "제한적 상승 베팅" };
          // 풋만
          if ((pB || pS) && !cB && !cS)
            return pB && !pS ? { label: "풋 방향성 (하락 베팅/방어)", desc: "" }
              : pS && !pB ? { label: "풋 매도 (하락 감수·프리미엄)", desc: "" }
              : { label: "풋 스프레드", desc: "제한적 하락 베팅" };
          // 콜·풋 혼합 (매수/매도 섞임) → 순프리미엄 부호로 방향 힌트
          return { label: "복합 옵션 전략", desc: domPrem >= 0 ? "프리미엄 지불형" : "프리미엄 수취형" };
        }
        const domVerdict = classifyDomOptions();

        let prevDom = null;
        optLegRows.forEach((L) => {
          if (L.domestic !== prevDom) {
            body.appendChild(mkGroupRow(L.domestic ? "국내 옵션" : "해외 옵션"));
            prevDom = L.domestic;
          }
          const tr = document.createElement("tr");
          tr.appendChild(cls(textTd(L.under), "col-name"));
          const cpTd = textTd(L.cp); cpTd.className = "center"; cpTd.style.fontWeight = "700";
          tr.appendChild(cpTd);
          tr.appendChild(numTd(L.strike, "raw"));
          // 포지션: 매수 + / 매도 − 부호, 양방향이면 내역 병기
          const posTd = document.createElement("td"); posTd.className = "number";
          const both = L.buyQty && L.sellQty;
          const netS = L.netQty > 0 ? `+${fmt(L.netQty, "raw")}` : fmt(L.netQty, "raw");
          posTd.textContent = both ? `${netS} (+${fmt(L.buyQty, "raw")} / -${fmt(L.sellQty, "raw")})` : netS;
          tr.appendChild(posTd);
          tr.appendChild(cls(textTd(L.expiry || ""), "number"));
          tr.appendChild(numTd(L.netPrem, "eok"));
          const vt = textTd(optVerdict(L.cp, L.netQty)); vt.style.fontWeight = "600";
          tr.appendChild(vt);
          body.appendChild(tr);
          // 국내 옵션의 마지막 레그 직후 종합판정 행 삽입
          const isLastDom = L.domestic && domVerdict &&
            (optLegRows.indexOf(L) === optLegRows.map((x) => x.domestic).lastIndexOf(true));
          if (isLastDom) {
            const sr = document.createElement("tr");
            const c1 = document.createElement("td");
            c1.colSpan = 6;
            c1.style.cssText = "text-align:right;font-weight:700;color:#ffa028;background:#0f0f0f;padding:7px 12px;";
            c1.textContent = "국내 옵션 종합 판정 →";
            sr.appendChild(c1);
            const c2 = document.createElement("td");
            c2.style.cssText = "background:#0f0f0f;font-weight:700;color:#e8e6e1;";
            c2.innerHTML = domVerdict.desc
              ? `${domVerdict.label} <span style="font-weight:500;color:var(--muted);font-size:11px;">· ${domVerdict.desc}</span>`
              : domVerdict.label;
            sr.appendChild(c2);
            body.appendChild(sr);
          }
        });
      }
      const optRows = optLegRows;   // hint 용

      const totalRows = futRows.length + optRows.length;
      if (!totalRows) {
        const tr = document.createElement("tr");
        const td = document.createElement("td"); td.colSpan = 7; td.className = "center"; td.textContent = "파생 포지션 없음";
        tr.appendChild(td); body.appendChild(tr);
      }
      document.getElementById("nettingHint").textContent =
        `선물 ${futRows.length} · 옵션 ${optRows.length} 기초자산 · ▲ 오르면 이익 / ▼ 내리면 이익`;
    }

    // ---- 손익: 누적수익률 상/하위 Top 10 ----
    function renderPnl() {
      const rows = activeList("domesticStock").concat(activeList("overseasStock"))
        .map((r) => ({ r, c: liveStock(r) }))
        .filter((x) => x.c.returnPct != null && !Number.isNaN(x.c.returnPct));
      const mkRows = (el, list) => {
        el.innerHTML = "";
        list.forEach((x) => {
          const tr = document.createElement("tr");
          const rp = x.c.returnPct, rpCls = rp >= 0 ? "ret-pos" : "ret-neg";
          const pnlV = x.c.pnl, pnlCls = pnlV == null ? "" : (pnlV >= 0 ? "ret-pos" : "ret-neg");
          tr.innerHTML = `<td class="nm" title="${x.r.stockName}">${x.r.stockName}</td>`
            + `<td class="number ${rpCls}">${(rp >= 0 ? "+" : "") + rp.toFixed(2)}%</td>`
            + `<td class="number">${formatEok(x.c.value)}</td>`
            + `<td class="number ${pnlCls}">${pnlV == null ? "" : fmt(pnlV, "won0")}</td>`;
          el.appendChild(tr);
        });
        if (!list.length) { const tr = document.createElement("tr"); const td = document.createElement("td"); td.colSpan = 4; td.className = "center"; td.textContent = "데이터 없음"; tr.appendChild(td); el.appendChild(tr); }
      };
      const byRet = rows.slice().sort((a, b) => b.c.returnPct - a.c.returnPct);
      mkRows(document.getElementById("pnlTop"), byRet.slice(0, 10));
      mkRows(document.getElementById("pnlBottom"), byRet.slice(-10).reverse());
    }

    // ---- 필터/뷰 ----
    function applyViewMode() {
      // 툴바 제거로 상세표시 토글이 없어져 detail 컬럼을 항상 표시
      document.querySelectorAll("#mainTable .col-hidden").forEach((el) => el.classList.remove("col-hidden"));
    }
    function applyFilters() {
      const n = document.querySelectorAll("#tbody tr").length;
      document.getElementById("visibleCount").textContent = `${n}건`;
    }

    // ---- 엑셀/CSV ----
    function assetsForExport() {
      const out = [];
      TABS.forEach((t) => {
        state.assets[t].forEach((r) => {
          if (isOpt(t)) {
            const c = liveOpt(r);
            out.push({ sourceRow: r.sourceRow, deleted: r.deleted, isNew: r.isNew, category: t, group: r.group, kind: r.kind, sub: r.sub, code: r.code, stockName: r.stockName, qty: r.qty, value: r.deleted ? r.value0 : c.value });
          } else if (isFut(t)) {
            const c = liveFut(r);
            out.push({ sourceRow: r.sourceRow, deleted: r.deleted, isNew: r.isNew, category: t, group: r.group, kind: r.kind, sub: r.sub, code: r.code, ticker: r.ticker, stockName: r.stockName, qty: r.qty, value: r.deleted ? r.value0 : c.value, memo: r.memo });
          } else {
            const c = liveStock(r);
            out.push({ sourceRow: r.sourceRow, deleted: r.deleted, isNew: r.isNew, category: t, group: r.group, kind: r.kind, code: r.code, ticker: r.ticker, stockName: r.stockName, qty: r.qty, price: r.price, value: r.deleted ? r.value0 : c.value });
          }
        });
      });
      return out;
    }

    function downloadExcel() {
      try {
        const X = window.XLSX;
        const wb = X.read(DATA.workbookBase64, { type: "base64" });
        window.BlackOnXlsx.applyEditsToWorkbook(X, wb, DATA, assetsForExport());
        const out = X.write(wb, { bookType: "xlsx", type: "array" });
        const blob = new Blob([out], { type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" });
        triggerDownload(blob, fileStem() + ".xlsx");
      } catch (err) {
        alert("엑셀 생성 실패: " + err.message);
        console.error(err);
      }
    }

    function exportCsv() {
      const tab = state.assetTab;
      const rows = activeRows();
      const heads = HEADERS[tab].map((h) => h[0]).filter((h) => h !== "");
      const lines = [heads];
      rows.forEach((r) => {
        if (isOpt(tab)) { const c = liveOpt(r); lines.push([c.direction, r.stockName, r.optExpiry, r.optCP, r.optStrike, r.qty, c.valueAbs, c.weight, c.pnl]); }
        else if (isFut(tab)) { const c = liveFut(r); lines.push([r.kind, c.direction, r.stockName, r.qty, c.valueAbs, c.weight, r.maturity, c.pnl]); }
        else {
          const c = liveStock(r);
          if (tab === "overseasStock") lines.push([r.ticker, r.stockName, r.qty, c.value, c.weight, c.pnl, c.returnPct, r.weightPrev]);
          else lines.push([r.stockName, r.qty, r.price, c.value, c.weight, c.pnl, c.returnPct, r.weightPrev]);
        }
      });
      const csv = lines.map((row) => row.map(csvEscape).join(",")).join("\n");
      triggerDownload(new Blob(["\ufeff" + csv], { type: "text/csv;charset=utf-8" }), fileStem() + "_" + tab + ".csv");
    }

    function fileStem() { const base = DATA.workbook.replace(/\.xlsx$/i, ""); return base + "_edited"; }
    function triggerDownload(blob, name) {
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a"); a.href = url; a.download = name; a.click();
      setTimeout(() => URL.revokeObjectURL(url), 1000);
    }

    // ---- 포맷/유틸 ----
    function num(v) { if (v === null || v === undefined || v === "") return null; const n = Number(v); return Number.isFinite(n) ? n : null; }
    function parseNumOrNull(v) { const t = (v || "").toString().replace(/,/g, "").trim(); if (!t) return null; const n = Number(t); return Number.isFinite(n) ? n : null; }
    function parsePctOrNull(v) { const t = (v || "").toString().replace("%", "").trim(); if (!t) return null; const n = Number(t); return Number.isFinite(n) ? n / 100 : null; }
    function round(v, d) { if (v == null || Number.isNaN(v)) return ""; const p = Math.pow(10, d); return Math.round(v * p) / p; }
    function sum(a) { return a.reduce((s, v) => s + (Number(v) || 0), 0); }
    function sumDef(a) { const f = a.filter((v) => v != null); return f.length ? sum(f) : null; }
    function formatEok(v) {
      if (v == null || Number.isNaN(v)) return "-";
      const a = Math.abs(v);
      if (a === 0) return "0";
      if (a >= 1e8) return round(v / 1e8, 2).toLocaleString("ko-KR", { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + "억";
      if (a >= 1e4) return Math.round(v / 1e4).toLocaleString("ko-KR") + "만";
      return Math.round(v).toLocaleString("ko-KR") + "원";
    }
    function formatPct(v) { if (v == null || Number.isNaN(v)) return ""; return round(v * 100, 2).toFixed(2) + "%"; }
    function fmt(v, type) {
      if (v == null || v === "" || Number.isNaN(v)) return "";
      const n = Number(v);
      if (type === "eok") return formatEok(n);
      if (type === "pct") return formatPct(n);
      if (type === "ret") return (n >= 0 ? "+" : "") + n.toFixed(2) + "%";
      if (type === "qty") return Math.round(n).toLocaleString("ko-KR");
      if (type === "won0") return Math.round(n).toLocaleString("ko-KR");
      if (type === "raw") return n.toLocaleString("ko-KR", { maximumFractionDigits: 4 });
      return n.toLocaleString("ko-KR");
    }
    function csvEscape(v) { const t = v == null ? "" : String(v); return /[",\n]/.test(t) ? `"${t.replace(/"/g, '""')}"` : t; }
  </script>
</body>
</html>
"""


def write_html(data, output_path):
    base = Path(__file__).resolve().parent
    sheetjs = (base / "vendor" / "xlsx.full.min.js").read_text(encoding="utf-8")
    patchjs = (base / "vendor" / "blackon_xlsx.js").read_text(encoding="utf-8")
    html = HTML_TEMPLATE
    html = html.replace("__SHEETJS__", sheetjs)
    html = html.replace("__BLACKON_XLSX__", patchjs)
    html = html.replace(
        "__DATA_JSON__",
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
    )
    output_path.write_text(html, encoding="utf-8")


def main():
    # 경로: 이 스크립트는 igis/dashboard/generate_black_on_dashboard.py 에 위치
    base_dir = Path(__file__).resolve().parent          # igis/dashboard

    # 입력 워크북 결정: 인자가 있으면 그 파일, 없으면 최신 워크북 자동 선택
    if len(sys.argv) > 1 and not sys.argv[1].startswith("--"):
        workbook_path = Path(sys.argv[1])
        if not workbook_path.is_absolute():
            workbook_path = base_dir / workbook_path
    else:
        candidates = sorted(
            list(base_dir.glob("black_on_*.xlsx")) + list(base_dir.glob("블랙ON_*.xlsx"))
            + list(base_dir.glob("체결화면*.xlsx")),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        candidates = [p for p in candidates if not p.name.startswith("~$")]
        if not candidates:
            print("black_on_*.xlsx / 블랙ON_*.xlsx / 체결화면*.xlsx 파일을 찾을 수 없습니다. (dashboard 폴더에 워크북을 두거나 파일명을 인자로 주세요)")
            sys.exit(1)
        workbook_path = candidates[0]
        print(f"[자동 선택] {workbook_path.name}")

    if not workbook_path.exists():
        print(f"워크북을 찾을 수 없습니다: {workbook_path}")
        sys.exit(1)

    # 출력: dashboard 폴더 안에 black_on_dashboard.html 생성
    if len(sys.argv) > 2 and not sys.argv[2].startswith("--"):
        output_path = Path(sys.argv[2])
        if not output_path.is_absolute():
            output_path = base_dir / output_path
    else:
        output_path = base_dir / "black_on_dashboard.html"

    data = build_data(workbook_path)
    if "--debug" in sys.argv:
        _debug_dump(data)
        return
    write_html(data, output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()