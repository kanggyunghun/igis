# -*- coding: utf-8 -*-
"""black_on_1.xlsx -> black_on_dashboard.html generator.

mp_dashboard 의 레이아웃/CSS 틀을 그대로 계승하되, black_on 펀드(단일 펀드,
국내개별주식 / 미국개별주식 / 해외선물 3개 자산군)의 보유 스냅샷을 표시한다.

데이터 레이어 설계
- 시트는 코드명(9907/4409/4509)이 아니라 헤더 내용으로 자동 분류한다.
- NAV 는 미국시트의 (Amount(KRW) / 종목비%) 로 역산한다.
- 국내/미국 현재비중 = 평가금액 / NAV, 현금·기타 = 100% - 주식비중합.
- 선물은 노셔널이라 평가금액/비중을 산정하지 않고 포지션만 표기한다.
"""
from __future__ import annotations

import base64
import json
import os
import shutil
import statistics
import subprocess
import sys
import tempfile
from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


DATA_START_ROW = 3  # 1-2행은 2단 헤더


def clean_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def clean_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def excel_serial_to_date(value):
    """Excel 날짜 시리얼(1899-12-30 기준) 또는 datetime -> 'YYYY-MM-DD'."""
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    num = clean_number(value)
    if num is None:
        return clean_text(value)
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(num))).strftime("%Y-%m-%d")
    except (OverflowError, ValueError):
        return clean_text(value)


# 통합 단일시트(Sheet1) 1-indexed 컬럼 맵
COL = {
    "no": 1, "group": 2, "kind": 3, "sub": 4,
    "fundCode": 5, "fundName": 6, "code": 7, "name": 8,
    "prevQty": 10, "dQty": 11, "price": 12, "prevValue": 13,
    "pnl": 15, "maturity": 16, "weightPrev": 17,
    "ticker": 28, "value": 30, "weight": 31,
}
SUBTOTAL_KINDS = {"종류별소계", "총계", "소계", "합계"}
CATEGORIES = ["domesticStock", "domesticFutures", "overseasStock", "overseasFutures", "cashOther"]


def classify_row(group, kind):
    """자산그룹/종류로 자산군 분류. 4탭(국내주식/국내선물/해외주식/해외선물) + 현금기타."""
    k = kind or ""
    if k.endswith("선물"):
        return "domesticFutures" if group == "선물그룹" else "overseasFutures"
    if k == "주식":
        return "domesticStock"
    if k == "해외주식":
        return "overseasStock"
    return "cashOther"


def _cell(ws, r, key):
    return ws.cell(r, COL[key]).value


def read_sheet1(ws):
    """통합 단일시트를 자산군별로 분류해 읽는다. 소계/총계/빈 행은 제외.

    - 순종목비(%NAV[당일]) 를 비중으로 직접 사용 (역산 불필요).
    - 보유수량 = 전일보유수량 + 당일증감수량 (선물은 부호 = 방향).
    - 선물 평가액은 부호가 있을 수 있어 절대값(valueAbs)도 함께 보관.
    - 취득원가 = 전일평가액 - 전일평가손익 (두 값 모두 '전일' 컬럼이라 날짜 정합).
    - 수익률 = (당일평가액 - 취득원가) / 취득원가.  (전일손익을 당일평가액에서 빼던 버그 수정)
    """
    cats = {c: [] for c in CATEGORIES}
    for r in range(2, ws.max_row + 1):
        kind = clean_text(_cell(ws, r, "kind"))
        group = clean_text(_cell(ws, r, "group"))
        name = clean_text(_cell(ws, r, "name"))
        if not name or kind in SUBTOTAL_KINDS or group in SUBTOTAL_KINDS or group == "총계":
            continue
        cat = classify_row(group, kind)
        is_futures = cat in ("domesticFutures", "overseasFutures")
        qty = (clean_number(_cell(ws, r, "prevQty")) or 0) + (clean_number(_cell(ws, r, "dQty")) or 0)
        value = clean_number(_cell(ws, r, "value"))
        weight = clean_number(_cell(ws, r, "weight"))
        weight = weight / 100.0 if weight is not None else None
        weight_prev = clean_number(_cell(ws, r, "weightPrev"))
        weight_prev = weight_prev / 100.0 if weight_prev is not None else None
        pnl = clean_number(_cell(ws, r, "pnl"))
        sub = clean_text(_cell(ws, r, "sub"))

        direction = ""
        if is_futures:
            if sub in ("매수", "매도"):
                direction = sub
            elif qty is not None:
                direction = "매도" if qty < 0 else "매수"

        prev_value = clean_number(_cell(ws, r, "prevValue"))  # 전일평가액(원)
        return_pct = None
        if not is_futures and prev_value is not None and pnl is not None:
            cost = prev_value - pnl  # 취득원가 = 전일평가액 − 전일평가손익
            if cost and value is not None:
                return_pct = (value - cost) / cost * 100.0  # 당일평가액 기준 수익률

        cats[cat].append({
            "no": clean_number(_cell(ws, r, "no")),
            "category": cat,
            "group": group,
            "kind": kind,
            "sub": sub,
            "fundCode": clean_text(_cell(ws, r, "fundCode")),
            "fundName": clean_text(_cell(ws, r, "fundName")),
            "code": clean_text(_cell(ws, r, "code")),
            "ticker": clean_text(_cell(ws, r, "ticker")),
            "stockName": name,
            "qty": qty,
            "price": clean_number(_cell(ws, r, "price")),
            "value": value,
            "valueAbs": abs(value) if value is not None else None,
            "prevValue": prev_value,
            "weight": weight,
            "weightPrev": weight_prev,
            "pnl": pnl,
            "returnPct": return_pct,
            "direction": direction,
            "maturity": excel_serial_to_date(_cell(ws, r, "maturity")),
            "sourceRow": r,
        })
    return cats


def derive_nav(cats):
    """순종목비 = |평가액| / NAV 관계로 NAV 역산 (중앙값으로 반올림 영향 최소화)."""
    navs = []
    for recs in cats.values():
        for rec in recs:
            v, w = rec.get("value"), rec.get("weight")
            if v and w and abs(w) > 0.0005:
                navs.append(abs(v) / w)
    return statistics.median(navs) if navs else None


def _os_copy(src: Path, dst: Path):
    """Excel 의 공유잠금에서도 복사되도록 OS 복사 명령을 사용한다.

    절대경로 + PowerShell 로 한글 경로/잠금을 모두 견딘다.
    """
    src, dst = str(src.resolve()), str(dst)
    if os.name == "nt":
        ps = f"Copy-Item -LiteralPath '{src}' -Destination '{dst}' -Force"
        subprocess.run(["powershell", "-NoProfile", "-Command", ps], check=True, capture_output=True)
    else:
        subprocess.run(["cp", src, dst], check=True, capture_output=True)


def load_safe(path: Path):
    """Excel 이 열려 있어 잠겨도 임시 복사본으로 읽는다. (wb, raw_bytes) 반환."""
    try:
        raw = path.read_bytes()
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb, raw
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / f"_blackon_{path.name}"
        try:
            shutil.copy(path, tmp)
        except PermissionError:
            _os_copy(path, tmp)  # Excel 배타 잠금 우회
        raw = tmp.read_bytes()
        wb = openpyxl.load_workbook(tmp, data_only=True)
        return wb, raw


def _signed_exposure(recs):
    """선물 순노출 = Σ 방향부호(매수+/매도−) × |평가액|."""
    total = 0.0
    for r in recs:
        sign = -1 if r["direction"] == "매도" else 1
        total += sign * (r["valueAbs"] or 0)
    return total


def build_data(workbook_path: Path):
    wb, raw = load_safe(workbook_path)
    sheet_name = wb.sheetnames[0]  # 통합 단일시트
    ws = wb[sheet_name]
    cats = read_sheet1(ws)
    nav = derive_nav(cats)

    def vsum(cat):
        return sum((r["value"] or 0) for r in cats[cat])

    def absum(cat):
        return sum((r["valueAbs"] or 0) for r in cats[cat])

    domestic_stock_value = vsum("domesticStock")
    overseas_stock_value = vsum("overseasStock")
    stock_value = domestic_stock_value + overseas_stock_value
    cash_value = vsum("cashOther")

    # 현금 = 외화예치금 + 예금잔고 (증거금 제외).
    # 외화예금/US MARGIN 은 보유수량이 USD 이고 당일평가액(col30)=0 이라, KRW 환산액이
    # col15(전일평가손익) 자리에 음수로 들어있다 → 절대값을 KRW 가치로 사용한다.
    def _cash_krw(r):
        return abs(r["value"]) if r["value"] else (abs(r["pnl"]) if r["pnl"] else 0)

    def _is_deposit(r):
        k = r["kind"] or ""
        return ("예치금" in k or "예금" in k) and "증거금" not in k

    for r in cats["cashOther"]:
        r["valueKrw"] = _cash_krw(r)
        r["isDeposit"] = _is_deposit(r)
    cash_deposit_value = sum(r["valueKrw"] for r in cats["cashOther"] if r["isDeposit"])

    futures_net = _signed_exposure(cats["domesticFutures"]) + _signed_exposure(cats["overseasFutures"])
    futures_gross = absum("domesticFutures") + absum("overseasFutures")
    total_pnl = sum((r["pnl"] or 0) for recs in cats.values() for r in recs)

    stat = workbook_path.stat()
    return {
        "workbook": workbook_path.name,
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbookModifiedAt": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "workbookSize": stat.st_size,
        "sheetName": sheet_name,
        "nav": nav,
        "navSource": "순종목비 역산",
        "domesticStockValue": domestic_stock_value,
        "overseasStockValue": overseas_stock_value,
        "stockValue": stock_value,
        "cashValue": cash_value,
        "cashOtherWeight": (1.0 - stock_value / nav) if nav else None,
        "cashDepositValue": cash_deposit_value,
        "cashDepositWeight": (cash_deposit_value / nav) if nav else None,
        "futuresNet": futures_net,
        "futuresGross": futures_gross,
        "totalPnl": total_pnl,
        "domesticStock": cats["domesticStock"],
        "domesticFutures": cats["domesticFutures"],
        "overseasStock": cats["overseasStock"],
        "overseasFutures": cats["overseasFutures"],
        "cashOther": cats["cashOther"],
        "workbookBase64": base64.b64encode(raw).decode("ascii"),
    }


def _debug_dump(data):
    keys = ["workbook", "sheetName", "nav", "navSource", "domesticStockValue",
            "overseasStockValue", "stockValue", "cashValue", "cashOtherWeight",
            "cashDepositValue", "cashDepositWeight", "futuresNet", "futuresGross", "totalPnl"]
    out = {k: data[k] for k in keys}
    out["counts"] = {c: len(data[c]) for c in CATEGORIES}
    out["base64_len"] = len(data["workbookBase64"])
    out["weightSums"] = {
        c: sum((r["weight"] or 0) for r in data[c])
        for c in CATEGORIES
    }
    out["sample_domStock"] = data["domesticStock"][0] if data["domesticStock"] else None
    out["sample_domFut"] = data["domesticFutures"][0] if data["domesticFutures"] else None
    out["sample_ovsFut"] = data["overseasFutures"][0] if data["overseasFutures"] else None
    out["cashRows"] = [(r["stockName"], r["value"], r["weight"]) for r in data["cashOther"]]
    out["domFut_dirs"] = [(r["stockName"], r["direction"], r["valueAbs"]) for r in data["domesticFutures"]]
    out["ovsFut_dirs"] = [(r["stockName"], r["direction"], r["qty"], r["valueAbs"]) for r in data["overseasFutures"]]
    print(json.dumps(out, ensure_ascii=True, indent=2, default=str))


HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>블랙ON 포트폴리오 대시보드</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9; --panel: #ffffff; --line: #d8dee8; --line-strong: #aab6c6;
      --text: #1d2733; --muted: #687589; --head: #143a5a; --head-2: #1f6f8f;
      --ok: #0f7b59; --warn: #a45b00; --bad: #b3261e;
      --soft-ok: #e5f3ed; --soft-warn: #fff2d8; --soft-bad: #fde7e4; --edit: #fff8d7;
      --shadow: 0 1px 2px rgba(20, 34, 52, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0; background: var(--bg); color: var(--text);
      font-family: "Pretendard Variable", "Pretendard", -apple-system, BlinkMacSystemFont,
        system-ui, "Apple SD Gothic Neo", "Noto Sans KR", "Malgun Gothic", "Segoe UI", Arial, sans-serif;
      font-size: 14px; line-height: 1.45; -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility; font-feature-settings: "tnum" 1;
    }
    header {
      position: sticky; top: 0; z-index: 20; background: var(--panel);
      border-bottom: 1px solid var(--line); box-shadow: var(--shadow);
    }
    .topbar { max-width: 1680px; margin: 0 auto; padding: 14px 18px 12px; display: grid; gap: 10px; }
    .title-row { display: flex; align-items: baseline; justify-content: space-between; gap: 16px; flex-wrap: wrap; }
    h1 { margin: 0; font-size: 20px; font-weight: 700; color: var(--head); }
    .meta { color: var(--muted); font-size: 12px; display: flex; gap: 10px; flex-wrap: wrap; }
    .toolbar { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
    .toolbar-divider { width: 1px; height: 22px; background: var(--line); margin: 0 2px; }
    button, select, input {
      font: inherit; border: 1px solid var(--line-strong); background: #fff; color: var(--text);
      height: 32px; border-radius: 6px;
    }
    button { padding: 0 11px; cursor: pointer; font-weight: 600; }
    button.primary { background: var(--head-2); border-color: var(--head-2); color: #fff; }
    button.accent { background: var(--ok); border-color: var(--ok); color: #fff; }
    button.danger { color: var(--bad); border-color: #e0aaa5; background: #fff; }
    input, select { padding: 0 9px; }
    #searchInput { width: min(260px, 100%); }
    main { max-width: 1680px; margin: 0 auto; padding: 16px 18px 28px; display: grid; gap: 14px; }
    .summary { display: grid; grid-template-columns: repeat(6, minmax(130px, 1fr)); gap: 10px; }
    .metric, .panel { background: var(--panel); border: 1px solid var(--line); box-shadow: var(--shadow); }
    .metric { padding: 11px 12px; min-height: 70px; }
    .metric .label { color: var(--muted); font-size: 12px; margin-bottom: 6px; }
    .metric .value { font-size: 19px; font-weight: 700; color: var(--head); }
    .metric .sub { font-size: 11px; color: var(--muted); margin-top: 2px; font-weight: 600; }
    .metric.bad .value { color: var(--bad); } .metric.ok .value { color: var(--ok); }
    .panel { overflow: hidden; }
    .panel-head {
      display: flex; justify-content: space-between; align-items: center; gap: 12px;
      padding: 10px 12px; background: #eef3f8; border-bottom: 1px solid var(--line);
      color: var(--head); font-weight: 700; min-height: 50px;
    }
    .panel-body { padding: 10px 12px; }
    .workspace { display: grid; grid-template-columns: 300px minmax(0, 1fr); gap: 14px; align-items: stretch; }
    .workspace > section.panel { display: flex; flex-direction: column; max-height: calc(100vh - 130px); min-height: 0; }
    .sector-panel { position: sticky; top: 112px; max-height: calc(100vh - 130px); display: flex; flex-direction: column; overflow: hidden; }
    .sector-panel .panel-body { padding: 8px; overflow: hidden; display: flex; flex-direction: column; gap: 8px; min-height: 0; }
    .sector-nav { display: grid; gap: 4px; }
    .sector-button {
      width: 100%; text-align: left; padding: 9px 11px; border: 1px solid var(--line);
      background: #fff; border-radius: 6px; display: grid; gap: 3px; cursor: pointer; height: auto;
    }
    .sector-button:hover { border-color: var(--head-2); background: #f4faff; }
    .sector-button.active { border-color: var(--head-2); background: #e8f4f8; box-shadow: inset 3px 0 0 var(--head-2); }
    .sector-row { display: flex; justify-content: space-between; align-items: baseline; gap: 8px; }
    .sector-name { font-weight: 700; color: var(--head); font-size: 13.5px; }
    .sector-target { font-weight: 700; font-size: 13.5px; font-variant-numeric: tabular-nums; }
    .sector-count { font-size: 11px; color: var(--muted); }
    .nav-subtotal { display: flex; justify-content: space-between; align-items: center; padding: 6px 11px; font-size: 12.5px; font-weight: 700; color: var(--head); background: #eef3f8; border: 1px solid var(--line); border-radius: 6px; font-variant-numeric: tabular-nums; }
    .nav-subtotal.muted-row { color: var(--muted); font-weight: 600; background: #f7f9fb; }
    .nav-subtotal.total-row { color: var(--head-2); background: #e8f1f7; border-color: #c4d8e6; }
    .nav-divider { height: 0; border-top: 1px dashed var(--line-strong); margin: 6px 2px; }
    .sector-stock-box { border: 1px solid var(--line); border-radius: 6px; background: #fafbfd; display: flex; flex-direction: column; min-height: 0; overflow: hidden; flex: 1 1 auto; }
    .sector-stock-head { padding: 7px 10px; font-weight: 700; color: var(--head); font-size: 12px; background: #eef3f8; border-bottom: 1px solid var(--line); display: flex; justify-content: space-between; }
    .sector-stock-list { margin: 0; padding: 6px; list-style: none; display: grid; gap: 4px; overflow: auto; }
    .sector-stock-list li { display: grid; grid-template-columns: 1fr auto; align-items: center; gap: 6px; padding: 5px 8px; background: #fff; border: 1px solid var(--line); border-radius: 5px; font-size: 12px; }
    .sector-stock-list li.subhead { background: #eef3f8; font-weight: 700; color: var(--head); grid-template-columns: 1fr; }
    .sector-stock-list .stock-name { font-weight: 600; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .sector-stock-list .stock-meta { font-variant-numeric: tabular-nums; font-size: 11px; color: var(--muted); }
    .fund-tabs { display: inline-flex; border: 1px solid var(--line-strong); border-radius: 6px; overflow: hidden; background: #fff; }
    .fund-tab { all: unset; cursor: pointer; padding: 4px 14px; font-weight: 600; font-size: 12.5px; color: var(--muted); border-right: 1px solid var(--line); }
    .fund-tab:last-child { border-right: 0; }
    .fund-tab:hover { background: #f4faff; color: var(--head); }
    .fund-tab.active { background: var(--head-2); color: #fff; }
    .table-wrap { overflow: auto; flex: 1 1 auto; min-height: 0; background: #fff; }
    table { width: 100%; border-collapse: separate; border-spacing: 0; }
    th, td { border-right: 1px solid var(--line); border-bottom: 1px solid var(--line); padding: 8px 10px; vertical-align: middle; white-space: nowrap; }
    th { position: sticky; top: 0; z-index: 5; background: var(--head); color: #fff; font-weight: 700; text-align: center; padding: 10px; }
    tbody td { background: #fff; } tbody tr:nth-child(even) td { background: #fafbfd; }
    tbody tr.hidden { display: none; }
    td.number { text-align: right; font-variant-numeric: tabular-nums; }
    td.center { text-align: center; }
    td.gap-pos { color: var(--ok); font-weight: 700; } td.gap-neg { color: var(--bad); font-weight: 700; } td.gap-zero { color: var(--muted); }
    td.ret-pos { color: var(--ok); font-weight: 600; } td.ret-neg { color: var(--bad); font-weight: 600; }
    .col-hidden { display: none; }
    tbody tr.row-bad td.col-name { box-shadow: inset 3px 0 0 var(--bad); }
    .editable { width: 100%; min-width: 70px; border: 1px solid transparent; background: transparent; color: var(--text); padding: 4px 6px; border-radius: 4px; }
    .editable:hover { background: #f1f5f9; border-color: var(--line); }
    .editable:focus { outline: none; background: var(--edit); border-color: #e1c96a; box-shadow: 0 0 0 2px rgba(225, 201, 106, 0.25); }
    .num-input { text-align: right; }
    .target-input { width: 78px; text-align: right; }
    .memo-input { min-width: 120px; }
    .name-input { min-width: 160px; font-weight: 700; }
    .badge { display: inline-flex; align-items: center; justify-content: center; min-width: 40px; height: 22px; padding: 0 8px; border-radius: 999px; font-size: 12px; font-weight: 700; }
    .badge.b { color: var(--ok); background: var(--soft-ok); } .badge.s { color: var(--bad); background: var(--soft-bad); }
    .diff-cell { display: grid; gap: 3px; }
    .diff-text { text-align: right; font-variant-numeric: tabular-nums; font-weight: 700; }
    .diff-text.gap-pos { color: var(--ok); } .diff-text.gap-neg { color: var(--bad); } .diff-text.gap-zero { color: var(--muted); }
    .diff-bar { height: 4px; background: #eef1f5; border-radius: 2px; position: relative; overflow: hidden; }
    .diff-bar::before { content: ''; position: absolute; top: 0; bottom: 0; left: 50%; width: 1px; background: var(--line-strong); }
    .diff-bar-fill { position: absolute; top: 0; bottom: 0; height: 100%; }
    .diff-bar-fill.pos { left: 50%; background: var(--ok); } .diff-bar-fill.neg { right: 50%; background: var(--bad); }
    tfoot td { position: sticky; bottom: 0; background: #f0f4f8; font-weight: 700; color: var(--head); border-top: 2px solid var(--head-2); }
    .delete-row { width: 30px; padding: 0; }
    .grid-2 { display: grid; grid-template-columns: minmax(0, 1fr) minmax(320px, 0.7fr); gap: 14px; }
    .chart-layout { display: grid; grid-template-columns: minmax(200px, 250px) minmax(0, 1fr); gap: 16px; align-items: center; }
    .exp-wrap { margin-top: 20px; padding-top: 14px; border-top: 1px dashed var(--line); }
    .exp-head { display: flex; justify-content: space-between; align-items: baseline; font-size: 12px; font-weight: 700; color: var(--head); margin-bottom: 18px; }
    .exp-head .muted { font-weight: 600; }
    .exp-track { position: relative; height: 28px; background: #eef1f5; border-radius: 6px; display: flex; overflow: visible; }
    .exp-seg { height: 100%; display: flex; align-items: center; justify-content: center; font-size: 10.5px; color: #fff; font-weight: 700; white-space: nowrap; overflow: hidden; }
    .exp-seg:first-child { border-radius: 6px 0 0 6px; }
    .exp-seg:last-child { border-radius: 0 6px 6px 0; }
    .exp-marker { position: absolute; top: -6px; bottom: -6px; width: 2px; background: var(--head); z-index: 3; }
    .exp-marker-label { position: absolute; top: -16px; transform: translateX(-50%); font-size: 10px; font-weight: 700; color: var(--head); white-space: nowrap; }
    .exp-legend { display: flex; flex-wrap: wrap; gap: 6px 14px; margin-top: 12px; font-size: 11.5px; }
    .exp-legend .legend-item { gap: 6px; }
    .pie-wrap { display: grid; justify-items: center; gap: 9px; }
    .pie-chart { width: min(230px, 100%); aspect-ratio: 1; border-radius: 50%; background: conic-gradient(#d7dde6 0 100%); border: 1px solid var(--line); position: relative; }
    .pie-chart::after { content: ''; position: absolute; width: 50%; height: 50%; top: 25%; left: 25%; background: var(--panel); border-radius: 50%; box-shadow: 0 0 0 4px var(--panel); }
    .pie-center { position: absolute; inset: 0; display: grid; place-items: center; z-index: 2; text-align: center; }
    .pie-center-label { font-size: 11px; color: var(--muted); font-weight: 600; }
    .pie-center-value { font-size: 18px; color: var(--head); font-weight: 800; line-height: 1; margin-top: 2px; }
    .legend { display: grid; gap: 6px; }
    .legend-item { display: flex; align-items: center; gap: 7px; min-width: 0; font-size: 13px; }
    .legend-swatch { width: 11px; height: 11px; border-radius: 3px; flex: 0 0 auto; }
    .legend-label { flex: 1 1 auto; } .legend-val { font-variant-numeric: tabular-nums; font-weight: 700; }
    .fut-summary { display: grid; gap: 6px; }
    .fut-summary .row { display: flex; justify-content: space-between; padding: 6px 8px; border: 1px solid var(--line); border-radius: 6px; background: #fafbfd; }
    @media (max-width: 1100px) {
      .summary { grid-template-columns: repeat(3, minmax(130px, 1fr)); }
      .workspace { grid-template-columns: 1fr; } .grid-2 { grid-template-columns: 1fr; }
      .chart-layout { grid-template-columns: 1fr; } .table-wrap { max-height: none; }
      .sector-panel { position: static; max-height: none; }
    }
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <div class="title-row">
        <h1>블랙ON #1 포트폴리오 대시보드</h1>
        <div class="meta">
          <span id="workbookMeta"></span>
          <span id="navMeta"></span>
          <span id="saveMeta"></span>
        </div>
      </div>
      <div class="toolbar">
        <button class="primary" id="addRowBtn" type="button" title="행 추가 (Ctrl+N)">+ 행 추가</button>
        <button id="saveBtn" type="button" title="브라우저 저장 (Ctrl+S)">💾 저장</button>
        <button class="accent" id="excelBtn" type="button" title="편집분을 반영한 엑셀 다운로드">⬇ 엑셀</button>
        <button id="csvBtn" type="button">CSV</button>
        <button class="danger" id="resetBtn" type="button">초기화</button>
        <span class="toolbar-divider"></span>
        <select id="viewModeSelect" aria-label="표시 모드">
          <option value="simple">간단 표시</option>
          <option value="detail">상세 표시</option>
        </select>
        <input id="searchInput" type="search" placeholder="🔍 종목/Ticker 검색" />
        <button id="clearFiltersBtn" type="button">필터 해제</button>
      </div>
    </div>
  </header>

  <main>
    <section class="summary" aria-label="요약">
      <div class="metric"><div class="label">펀드 NAV</div><div class="value" id="mNav">-</div><div class="sub" id="mNavSub"></div></div>
      <div class="metric"><div class="label">주식 평가액(국내+해외)</div><div class="value" id="mStock">-</div><div class="sub" id="mStockSub"></div></div>
      <div class="metric"><div class="label">현금</div><div class="value" id="mCash">-</div><div class="sub" id="mCashSub">외화예치금+예금잔고</div></div>
      <div class="metric" id="mFutBox"><div class="label">선물 순노출</div><div class="value" id="mFutNet">-</div><div class="sub" id="mFutNetSub"></div></div>
      <div class="metric" id="mPnlBox"><div class="label">총 평가손익</div><div class="value" id="mPnl">-</div><div class="sub" id="mPnlSub"></div></div>
      <div class="metric"><div class="label">종목·포지션 수</div><div class="value" id="mCount">-</div><div class="sub" id="mCountSub"></div></div>
    </section>

    <section class="workspace">
      <aside class="panel sector-panel">
        <div class="panel-head"><span>자산군 ▾</span><span class="muted" id="groupHint"></span></div>
        <div class="panel-body">
          <div class="sector-nav" id="groupNav"></div>
          <div class="sector-stock-box">
            <div class="sector-stock-head"><span id="stockListTitle">종목</span><span class="muted" id="stockListHint" style="font-weight:500;">비중▾</span></div>
            <ul class="sector-stock-list" id="stockList"></ul>
          </div>
        </div>
      </aside>

      <section class="panel">
        <div class="panel-head">
          <div style="display:inline-flex;align-items:center;gap:10px;flex-wrap:wrap;">
            <span id="tableTitle">국내주식</span>
            <div class="fund-tabs" id="assetTabs" role="tablist">
              <button class="fund-tab active" type="button" data-asset="domesticStock">국내주식</button>
              <button class="fund-tab" type="button" data-asset="domesticFutures">국내선물</button>
              <button class="fund-tab" type="button" data-asset="overseasStock">해외주식</button>
              <button class="fund-tab" type="button" data-asset="overseasFutures">해외선물</button>
            </div>
          </div>
          <span class="muted" id="visibleCount"></span>
        </div>
        <div class="table-wrap">
          <table id="mainTable">
            <thead><tr id="theadRow"></tr></thead>
            <tbody id="tbody"></tbody>
            <tfoot><tr id="tfootRow"></tr></tfoot>
          </table>
        </div>
      </section>
    </section>

    <section class="grid-2">
      <section class="panel">
        <div class="panel-head"><span>자산 배분</span><span class="muted">자본 100% + 선물 노셔널</span></div>
        <div class="panel-body">
          <div class="chart-layout">
            <div class="pie-wrap">
              <div class="pie-chart" id="allocPie" role="img" aria-label="자산배분 파이">
                <div class="pie-center"><div><div class="pie-center-label">주식 합</div><div class="pie-center-value" id="pieCenter">-</div></div></div>
              </div>
            </div>
            <div class="legend" id="allocLegend"></div>
          </div>
          <div class="exp-wrap" id="exposureWrap"></div>
        </div>
      </section>
      <section class="panel">
        <div class="panel-head"><span>선물 요약 (국내+해외)</span><span class="muted">노셔널·순노출</span></div>
        <div class="panel-body"><div class="fut-summary" id="futSummary"></div></div>
      </section>
    </section>
  </main>

  <script>__SHEETJS__</script>
  <script>__BLACKON_XLSX__</script>
  <script>
    const DATA = __DATA_JSON__;
    const STORAGE_KEY = `blackon-dashboard:${DATA.workbook}:${DATA.workbookModifiedAt}:${DATA.workbookSize}`;
    const CHART_COLORS = { domestic: "#1f6f8f", overseas: "#7a9f35", cash: "#c46a2b", domesticFut: "#6f5aa7", overseasFut: "#b75b7a" };
    const TABS = ["domesticStock", "domesticFutures", "overseasStock", "overseasFutures"];
    const TAB_LABEL = { domesticStock: "국내주식", domesticFutures: "국내선물", overseasStock: "해외주식", overseasFutures: "해외선물" };
    const isFut = (t) => t === "domesticFutures" || t === "overseasFutures";

    const state = { assetTab: "domesticStock", search: "", group: "all", viewMode: "simple",
      assets: { domesticStock: [], domesticFutures: [], overseasStock: [], overseasFutures: [] } };

    document.addEventListener("DOMContentLoaded", init);

    function init() {
      document.getElementById("workbookMeta").textContent = `${DATA.workbook} | 생성 ${DATA.generatedAt}`;
      const fund = (DATA.domesticStock[0] || {}).fundName || "";
      document.getElementById("navMeta").textContent = `NAV ${formatEok(DATA.nav)}${fund ? " · " + fund : ""}`;
      loadAssets();
      bindEvents();
      render();
    }

    // ---- 편집 상태 빌드/저장 ----
    function mapStock(h) {
      return {
        sourceRow: h.sourceRow, deleted: false, isNew: false, category: h.category,
        group: h.group, kind: h.kind, code: h.code, ticker: h.ticker, stockName: h.stockName,
        qty: h.qty, qty0: h.qty, price: h.price, value0: h.value,
        cost0: (h.prevValue != null && h.pnl != null) ? h.prevValue - h.pnl : null,
        weightPrev: h.weightPrev, targetPct: null, memo: "",
      };
    }
    function mapFut(h) {
      return {
        sourceRow: h.sourceRow, deleted: false, isNew: false, category: h.category,
        group: h.group, kind: h.kind, sub: h.sub, code: h.code, ticker: h.ticker, stockName: h.stockName,
        qty: h.qty, qty0: h.qty, value0: h.value, valueAbs0: h.valueAbs,
        direction: h.direction, maturity: h.maturity, pnl0: h.pnl, weightPrev: h.weightPrev, memo: "",
      };
    }
    function buildAssets() {
      return {
        domesticStock: DATA.domesticStock.map(mapStock),
        overseasStock: DATA.overseasStock.map(mapStock),
        domesticFutures: DATA.domesticFutures.map(mapFut),
        overseasFutures: DATA.overseasFutures.map(mapFut),
      };
    }

    function loadAssets() {
      const stored = localStorage.getItem(STORAGE_KEY);
      if (stored) {
        try { state.assets = JSON.parse(stored); document.getElementById("saveMeta").textContent = "저장본 적용"; return; }
        catch (e) { localStorage.removeItem(STORAGE_KEY); }
      }
      state.assets = buildAssets();
    }

    function saveAssets() {
      localStorage.setItem(STORAGE_KEY, JSON.stringify(state.assets));
      const meta = document.getElementById("saveMeta");
      meta.textContent = `✓ 저장 ${new Date().toLocaleTimeString("ko-KR", { hour12: false })}`;
      meta.style.color = "var(--ok)"; meta.style.fontWeight = "700";
      setTimeout(() => { meta.style.color = ""; meta.style.fontWeight = ""; }, 1500);
    }

    function resetAssets() {
      if (!confirm("브라우저 저장본을 지우고 엑셀 기준으로 되돌릴까요?")) return;
      localStorage.removeItem(STORAGE_KEY);
      state.assets = buildAssets();
      document.getElementById("saveMeta").textContent = "";
      render();
    }

    // ---- 계산 ----
    function liveStock(row) {
      const qty = num(row.qty), price = num(row.price);
      let value;
      if (price != null && price > 0 && qty != null) value = qty * price;
      else if (qty != null && row.qty0) value = row.value0 * qty / row.qty0;
      else value = row.value0 || 0;
      const cost = row.cost0;
      const pnl = cost != null ? value - cost : null;
      const returnPct = cost ? pnl / cost * 100 : null;
      const weight = DATA.nav ? value / DATA.nav : 0;
      const target = row.targetPct;
      const diff = target != null ? target - weight : null;
      const adjValue = target != null && DATA.nav ? target * DATA.nav - value : null;
      const pEff = (price != null && price > 0) ? price : (row.qty0 ? row.value0 / row.qty0 : null);
      const adjQty = adjValue != null && pEff ? adjValue / pEff : null;
      return { value, cost, pnl, returnPct, weight, target, diff, adjQty };
    }
    function liveFut(row) {
      const qty = num(row.qty);
      let value;
      if (qty != null && row.qty0) value = row.value0 * qty / row.qty0;
      else value = row.value0 || 0;
      const valueAbs = Math.abs(value || 0);
      let direction;
      if (qty != null && qty !== 0) direction = qty < 0 ? "매도" : "매수";
      else direction = row.sub || row.direction || "매수";
      const weight = DATA.nav ? valueAbs / DATA.nav : (row.valueAbs0 && DATA.nav ? row.valueAbs0 / DATA.nav : 0);
      return { value, valueAbs, direction, weight, pnl: row.pnl0 };
    }
    const dirSign = (d) => (d === "매도" ? -1 : 1);

    // ---- 이벤트 ----
    function bindEvents() {
      document.getElementById("addRowBtn").addEventListener("click", addRow);
      document.getElementById("saveBtn").addEventListener("click", saveAssets);
      document.getElementById("excelBtn").addEventListener("click", downloadExcel);
      document.getElementById("csvBtn").addEventListener("click", exportCsv);
      document.getElementById("resetBtn").addEventListener("click", resetAssets);
      document.getElementById("clearFiltersBtn").addEventListener("click", () => {
        state.search = ""; state.group = "all";
        document.getElementById("searchInput").value = ""; render();
      });
      document.getElementById("searchInput").addEventListener("input", (e) => { state.search = e.target.value.trim().toLowerCase(); applyFilters(); });
      document.getElementById("viewModeSelect").addEventListener("change", (e) => { state.viewMode = e.target.value; applyViewMode(); });
      document.querySelectorAll("#assetTabs .fund-tab").forEach((tab) => {
        tab.addEventListener("click", () => { state.assetTab = tab.dataset.asset; state.group = "all"; render(); });
      });
      document.addEventListener("keydown", (e) => {
        if ((e.ctrlKey || e.metaKey) && e.key.toLowerCase() === "s") { e.preventDefault(); saveAssets(); }
        if ((e.ctrlKey || e.metaKey) && e.key.toLowerCase() === "n") { e.preventDefault(); addRow(); }
      });
    }

    function addRow() {
      const tab = state.assetTab;
      const base = { sourceRow: null, deleted: false, isNew: true, category: tab, memo: "" };
      if (tab === "domesticStock") Object.assign(base, { group: "주식그룹", kind: "주식", code: "", ticker: "", stockName: "", qty: null, qty0: null, price: null, value0: 0, cost0: null, targetPct: null });
      else if (tab === "overseasStock") Object.assign(base, { group: "해외자산그룹", kind: "해외주식", code: "", ticker: "", stockName: "", qty: null, qty0: null, price: null, value0: 0, cost0: null, targetPct: null });
      else if (tab === "domesticFutures") Object.assign(base, { group: "선물그룹", kind: "", sub: "매수", code: "", ticker: "", stockName: "", qty: null, qty0: null, value0: 0, valueAbs0: 0, direction: "매수", maturity: "", pnl0: null });
      else Object.assign(base, { group: "해외기타그룹", kind: "", sub: "", code: "", ticker: "", stockName: "", qty: null, qty0: null, value0: 0, valueAbs0: 0, direction: "매수", maturity: "", pnl0: null });
      state.assets[tab].push(base);
      render();
      const inputs = document.querySelectorAll("#tbody tr:last-child input");
      if (inputs.length) inputs[0].focus();
    }

    // ---- 렌더 ----
    function render() {
      document.querySelectorAll("#assetTabs .fund-tab").forEach((t) => t.classList.toggle("active", t.dataset.asset === state.assetTab));
      document.getElementById("tableTitle").textContent = TAB_LABEL[state.assetTab];
      renderHead();
      renderBody();
      renderFoot();
      renderGroupNav();
      renderStockList();
      renderMetrics();
      renderAlloc();
      renderFutSummary();
      applyViewMode();
      applyFilters();
    }

    const HEADERS = {
      domesticStock: [["종목명",""],["보유수량",""],["현재가",""],["평가액",""],["비중",""],["목표비중",""],["차이",""],["조정수량",""],["손익","detail"],["수익률",""],["전일비중","detail"],["메모","detail"],["",""]],
      overseasStock: [["Ticker",""],["종목명",""],["보유수량",""],["평가액",""],["비중",""],["목표비중",""],["차이",""],["조정수량",""],["손익","detail"],["수익률",""],["전일비중","detail"],["메모","detail"],["",""]],
      domesticFutures: [["종류",""],["방향",""],["종목명",""],["계약수",""],["평가액",""],["비중",""],["만기","detail"],["손익","detail"],["메모",""],["",""]],
      overseasFutures: [["종류",""],["방향",""],["종목명",""],["계약수",""],["평가액",""],["비중",""],["만기","detail"],["손익","detail"],["메모",""],["",""]],
    };

    function renderHead() {
      const tr = document.getElementById("theadRow");
      tr.innerHTML = "";
      HEADERS[state.assetTab].forEach(([label, cls]) => {
        const th = document.createElement("th");
        th.textContent = label;
        if (cls === "detail") th.classList.add("col-detail");
        tr.appendChild(th);
      });
    }

    function renderBody() {
      const body = document.getElementById("tbody");
      body.innerHTML = "";
      const frag = document.createDocumentFragment();
      const tab = state.assetTab;
      // 평가액 내림차순 정렬 (편집/삭제용 원본 인덱스는 보존)
      const valOf = isFut(tab) ? (r) => liveFut(r).valueAbs : (r) => liveStock(r).value;
      const entries = state.assets[tab].map((row, idx) => ({ row, idx })).filter((e) => !e.row.deleted);
      entries.sort((a, b) => (valOf(b.row) || 0) - (valOf(a.row) || 0));
      entries.forEach(({ row, idx }) => {
        frag.appendChild(isFut(tab) ? futRow(row, idx, tab) : stockRow(row, idx, tab));
      });
      body.appendChild(frag);
      body.querySelectorAll("input[data-field]").forEach((inp) => inp.addEventListener("change", onEdit));
    }

    function onEdit(e) {
      const inp = e.target;
      const idx = Number(inp.dataset.idx), field = inp.dataset.field, kind = inp.dataset.kind;
      const row = state.assets[state.assetTab][idx];
      if (kind === "num") row[field] = parseNumOrNull(inp.value);
      else if (kind === "pct") row[field] = parsePctOrNull(inp.value);
      else row[field] = inp.value;
      render();
    }

    function stockRow(row, idx, tab) {
      const c = liveStock(row);
      const tr = document.createElement("tr");
      tr.dataset.search = `${row.ticker || ""} ${row.stockName || ""}`.toLowerCase();
      if (c.diff != null && Math.abs(c.diff) > 0.01) tr.classList.add("row-bad");
      if (tab === "overseasStock") tr.appendChild(editTd(idx, "ticker", row.ticker, "text"));
      tr.appendChild(editTd(idx, "stockName", row.stockName, "text", "name-input col-name"));
      tr.appendChild(editTd(idx, "qty", row.qty, "num"));
      if (tab === "domesticStock") tr.appendChild(editTd(idx, "price", row.price, "num"));
      tr.appendChild(numTd(c.value, "eok"));
      tr.appendChild(numTd(c.weight, "pct"));
      tr.appendChild(targetCell(idx, row.targetPct));
      tr.appendChild(diffTd(c.diff));
      tr.appendChild(gapTd(c.adjQty, "qty"));
      tr.appendChild(cls(retTd(c.pnl, "won0"), "col-detail"));
      tr.appendChild(retTd(c.returnPct, "ret"));
      tr.appendChild(cls(numTd(row.weightPrev, "pct"), "col-detail"));
      tr.appendChild(cls(editTd(idx, "memo", row.memo, "text", "memo-input"), "col-detail"));
      tr.appendChild(deleteTd(idx, row.stockName));
      return tr;
    }

    function futRow(row, idx, tab) {
      const c = liveFut(row);
      const tr = document.createElement("tr");
      tr.dataset.search = `${row.kind || ""} ${row.stockName || ""} ${row.ticker || ""}`.toLowerCase();
      tr.dataset.group = row.kind || "";
      tr.appendChild(textTd(row.kind));
      tr.appendChild(badgeTd(c.direction));
      tr.appendChild(cls(textTd(row.stockName), "col-name"));
      tr.appendChild(editTd(idx, "qty", row.qty, "num"));
      tr.appendChild(numTd(c.valueAbs, "eok"));
      tr.appendChild(numTd(c.weight, "pct"));
      tr.appendChild(cls(textTd(row.maturity), "col-detail"));
      tr.appendChild(cls(retTd(c.pnl, "won0"), "col-detail"));
      tr.appendChild(editTd(idx, "memo", row.memo, "text", "memo-input"));
      tr.appendChild(deleteTd(idx, row.stockName));
      return tr;
    }

    // ---- 셀 헬퍼 ----
    function cls(td, c) { td.classList.add(c); return td; }
    function textTd(v) { const td = document.createElement("td"); td.textContent = v == null ? "" : v; return td; }
    function numTd(v, type) { const td = document.createElement("td"); td.className = "number"; td.textContent = fmt(v, type); return td; }
    function retTd(v, type) { const td = numTd(v, type); if (v != null && !Number.isNaN(v)) td.classList.add(v >= 0 ? "ret-pos" : "ret-neg"); return td; }
    function badgeTd(dir) {
      const td = document.createElement("td"); td.className = "center";
      const span = document.createElement("span"); span.className = "badge " + (dir === "매수" ? "b" : "s"); span.textContent = dir;
      td.appendChild(span); return td;
    }
    function editTd(idx, field, value, kind, extra) {
      const td = document.createElement("td");
      if (kind === "num") td.className = "number";
      const inp = document.createElement("input");
      inp.className = "editable" + (kind === "num" ? " num-input" : "") + (extra ? " " + extra : "");
      inp.value = (kind === "num") ? (value == null ? "" : value) : (value || "");
      inp.dataset.idx = idx; inp.dataset.field = field; inp.dataset.kind = kind;
      td.appendChild(inp);
      return td;
    }
    function targetCell(idx, target) {
      const td = document.createElement("td"); td.className = "number";
      const inp = document.createElement("input");
      inp.className = "editable target-input";
      inp.value = target == null ? "" : round(target * 100, 2);
      inp.placeholder = "%";
      inp.dataset.idx = idx; inp.dataset.field = "targetPct"; inp.dataset.kind = "pct";
      td.appendChild(inp);
      return td;
    }
    function diffTd(value) {
      const td = document.createElement("td"); td.className = "number";
      if (value == null || Number.isNaN(value)) { td.textContent = ""; return td; }
      const c = Math.abs(value) < 0.0005 ? "gap-zero" : (value > 0 ? "gap-pos" : "gap-neg");
      const wrap = document.createElement("div"); wrap.className = "diff-cell";
      const t = document.createElement("div"); t.className = `diff-text ${c}`; t.textContent = (value >= 0 ? "+" : "") + formatPct(value);
      wrap.appendChild(t);
      const bar = document.createElement("div"); bar.className = "diff-bar";
      if (Math.abs(value) >= 0.0005) { const f = document.createElement("div"); f.className = `diff-bar-fill ${value > 0 ? "pos" : "neg"}`; f.style.width = `${Math.min(Math.abs(value) / 0.03, 1) * 50}%`; bar.appendChild(f); }
      wrap.appendChild(bar); td.appendChild(wrap);
      return td;
    }
    function gapTd(value, type) {
      const td = numTd(value, type);
      if (value == null || Number.isNaN(value)) return td;
      if (Math.abs(value) < 1) td.classList.add("gap-zero"); else td.classList.add(value > 0 ? "gap-pos" : "gap-neg");
      return td;
    }
    function deleteTd(idx, name) {
      const td = document.createElement("td"); td.className = "center";
      const b = document.createElement("button"); b.type = "button"; b.className = "delete-row danger"; b.textContent = "✕"; b.title = "행 삭제";
      b.addEventListener("click", () => {
        if (name && !confirm(`'${name}' 행을 삭제할까요?`)) return;
        const row = state.assets[state.assetTab][idx];
        if (row.isNew || !row.sourceRow) state.assets[state.assetTab].splice(idx, 1);
        else row.deleted = true;
        render();
      });
      td.appendChild(b);
      return td;
    }

    function renderFoot() {
      const tr = document.getElementById("tfootRow");
      tr.innerHTML = "";
      const tab = state.assetTab;
      const heads = HEADERS[tab];
      const rows = activeRows();
      const cells = new Array(heads.length).fill("");
      cells[0] = `합계 (${rows.length})`;
      if (isFut(tab)) {
        const lives = rows.map(liveFut);
        cells[3] = fmt(sum(rows.map((r) => Math.abs(num(r.qty) || 0))), "raw");
        cells[4] = fmt(sum(lives.map((c) => c.valueAbs)), "eok");
        cells[5] = formatPct(sum(lives.map((c) => c.weight)));
        cells[7] = fmt(sum(lives.map((c) => c.pnl || 0)), "won0");
      } else {
        const lives = rows.map(liveStock);
        cells[3] = fmt(sum(lives.map((c) => c.value)), "eok");
        cells[4] = formatPct(sum(lives.map((c) => c.weight)));
        const t = sumDef(rows.map((r) => r.targetPct));
        cells[5] = t == null ? "" : formatPct(t);
        cells[8] = fmt(sum(lives.map((c) => c.pnl || 0)), "won0");
      }
      cells.forEach((txt, i) => { const td = document.createElement("td"); td.textContent = txt; if (i > 0) td.className = "number"; if (heads[i] && heads[i][1] === "detail") td.classList.add("col-detail"); tr.appendChild(td); });
    }

    function activeRows() { return state.assets[state.assetTab].filter((r) => !r.deleted); }
    function activeList(tab) { return state.assets[tab].filter((r) => !r.deleted); }
    function activeCount(tab) { return activeList(tab).length; }
    function tabWeight(tab) {
      const calc = isFut(tab) ? liveFut : liveStock;
      return sum(activeList(tab).map((r) => calc(r).weight));
    }

    // ---- 좌측 패널 ----
    function renderGroupNav() {
      const nav = document.getElementById("groupNav");
      nav.innerHTML = "";
      const mkBtn = (t) => {
        const b = document.createElement("button");
        b.type = "button"; b.className = "sector-button" + (t === state.assetTab ? " active" : "");
        const r1 = document.createElement("div"); r1.className = "sector-row";
        const n = document.createElement("span"); n.className = "sector-name"; n.textContent = TAB_LABEL[t];
        const w = document.createElement("span"); w.className = "sector-target"; w.textContent = formatPct(tabWeight(t));
        r1.appendChild(n); r1.appendChild(w);
        const r2 = document.createElement("div"); r2.className = "sector-row";
        const cnt = document.createElement("span"); cnt.className = "sector-count"; cnt.textContent = `${activeCount(t)}종목${isFut(t) ? " (노셔널)" : ""}`;
        r2.appendChild(cnt);
        b.appendChild(r1); b.appendChild(r2);
        b.addEventListener("click", () => { state.assetTab = t; state.group = "all"; render(); });
        return b;
      };
      const subtotal = (label, value, extra) => {
        const d = document.createElement("div"); d.className = "nav-subtotal" + (extra ? " " + extra : "");
        const a = document.createElement("span"); a.textContent = label;
        const b = document.createElement("span"); b.textContent = value;
        d.appendChild(a); d.appendChild(b);
        return d;
      };
      const divider = () => { const d = document.createElement("div"); d.className = "nav-divider"; return d; };

      const stockW = tabWeight("domesticStock") + tabWeight("overseasStock");
      const cashW = DATA.nav ? Math.max(0, 1 - stockW) : 0;
      const futW = tabWeight("domesticFutures") + tabWeight("overseasFutures");

      nav.appendChild(mkBtn("domesticStock"));
      nav.appendChild(mkBtn("overseasStock"));
      nav.appendChild(subtotal("주식 소계", formatPct(stockW)));
      nav.appendChild(subtotal("현금·기타", formatPct(cashW), "muted-row"));
      nav.appendChild(subtotal("자본 합계", formatPct(stockW + cashW), "total-row"));
      nav.appendChild(divider());
      nav.appendChild(mkBtn("domesticFutures"));
      nav.appendChild(mkBtn("overseasFutures"));
      nav.appendChild(subtotal("선물 노셔널 합", formatPct(futW), "muted-row"));

      document.getElementById("groupHint").textContent = TAB_LABEL[state.assetTab];
    }

    function renderStockList() {
      const list = document.getElementById("stockList");
      const title = document.getElementById("stockListTitle");
      list.innerHTML = "";
      const tab = state.assetTab;
      const rows = activeRows();
      if (isFut(tab)) {
        title.textContent = `종류별 (${rows.length})`;
        document.getElementById("stockListHint").textContent = "계약▾";
        const byK = {};
        rows.forEach((r) => { (byK[r.kind || "기타"] = byK[r.kind || "기타"] || []).push(r); });
        Object.keys(byK).sort().forEach((k) => {
          const head = document.createElement("li"); head.className = "subhead";
          head.textContent = `${k} (${byK[k].length})`; list.appendChild(head);
          byK[k].forEach((r) => {
            const c = liveFut(r);
            const li = document.createElement("li");
            const n = document.createElement("span"); n.className = "stock-name"; n.textContent = r.stockName; n.title = r.stockName;
            const m = document.createElement("span"); m.className = "stock-meta";
            m.innerHTML = `<span class="badge ${c.direction === "매수" ? "b" : "s"}">${c.direction}</span> ${fmt(Math.abs(num(r.qty) || 0), "raw")}`;
            li.appendChild(n); li.appendChild(m); list.appendChild(li);
          });
        });
        return;
      }
      title.textContent = `${TAB_LABEL[tab]} (${rows.length})`;
      document.getElementById("stockListHint").textContent = "비중▾";
      rows.map((r) => ({ r, c: liveStock(r) })).sort((a, b) => b.c.weight - a.c.weight).forEach(({ r, c }) => {
        const li = document.createElement("li");
        const n = document.createElement("span"); n.className = "stock-name"; n.textContent = r.stockName; n.title = r.stockName;
        const m = document.createElement("span"); m.className = "stock-meta";
        const retCls = c.returnPct == null ? "" : (c.returnPct >= 0 ? "gap-pos" : "gap-neg");
        m.innerHTML = `<strong>${formatPct(c.weight)}</strong> <span class="${retCls}">${c.returnPct == null ? "" : (c.returnPct >= 0 ? "+" : "") + round(c.returnPct, 1) + "%"}</span>`;
        li.appendChild(n); li.appendChild(m); list.appendChild(li);
      });
    }

    // ---- 메트릭/차트 ----
    function renderMetrics() {
      const ds = activeList("domesticStock").map(liveStock);
      const os = activeList("overseasStock").map(liveStock);
      const domVal = sum(ds.map((c) => c.value));
      const ovsVal = sum(os.map((c) => c.value));
      const stockVal = domVal + ovsVal;
      const cash = DATA.nav ? 1 - stockVal / DATA.nav : null;
      const df = activeList("domesticFutures").map(liveFut);
      const of = activeList("overseasFutures").map(liveFut);
      const futNet = sum(df.map((c) => dirSign(c.direction) * c.valueAbs)) + sum(of.map((c) => dirSign(c.direction) * c.valueAbs));
      const futGross = sum(df.map((c) => c.valueAbs)) + sum(of.map((c) => c.valueAbs));
      const pnl = sum(ds.map((c) => c.pnl || 0)) + sum(os.map((c) => c.pnl || 0)) + sum(df.map((c) => c.pnl || 0)) + sum(of.map((c) => c.pnl || 0));
      const cnt = TABS.reduce((a, t) => a + activeCount(t), 0);

      document.getElementById("mNav").textContent = formatEok(DATA.nav);
      document.getElementById("mNavSub").textContent = DATA.navSource || "";
      document.getElementById("mStock").textContent = formatEok(stockVal);
      document.getElementById("mStockSub").textContent = DATA.nav ? formatPct(stockVal / DATA.nav) + " of NAV" : "";
      document.getElementById("mCash").textContent = DATA.cashDepositValue == null ? "-" : formatEok(DATA.cashDepositValue);
      document.getElementById("mCashSub").textContent = (DATA.cashDepositWeight == null ? "" : formatPct(DATA.cashDepositWeight) + " · ") + "외화예치금+예금잔고";
      document.getElementById("mFutNet").textContent = formatEok(futNet);
      document.getElementById("mFutNetSub").textContent = `그로스 ${formatEok(futGross)}${DATA.nav ? " · " + formatPct(futNet / DATA.nav) : ""}`;
      document.getElementById("mFutBox").className = "metric " + (futNet >= 0 ? "ok" : "bad");
      document.getElementById("mPnl").textContent = formatEok(pnl);
      document.getElementById("mPnlBox").className = "metric " + (pnl >= 0 ? "ok" : "bad");
      document.getElementById("mPnlSub").textContent = "전일 평가손익 기준";
      document.getElementById("mCount").textContent = cnt + "개";
      document.getElementById("mCountSub").textContent = `주식 ${activeCount("domesticStock") + activeCount("overseasStock")} · 선물 ${activeCount("domesticFutures") + activeCount("overseasFutures")}`;
    }

    function renderAlloc() {
      const domVal = sum(activeList("domesticStock").map((r) => liveStock(r).value));
      const ovsVal = sum(activeList("overseasStock").map((r) => liveStock(r).value));
      const nav = DATA.nav || (domVal + ovsVal);
      const cashVal = Math.max(0, nav - domVal - ovsVal);
      const parts = [
        { key: "domestic", label: "국내주식", val: domVal },
        { key: "overseas", label: "해외주식", val: ovsVal },
        { key: "cash", label: "현금·기타", val: cashVal },
      ];
      let acc = 0; const stops = [];
      parts.forEach((p) => { const a = acc / nav * 100, b = (acc + p.val) / nav * 100; stops.push(`${CHART_COLORS[p.key]} ${a}% ${b}%`); acc += p.val; });
      document.getElementById("allocPie").style.background = `conic-gradient(${stops.join(",")})`;
      document.getElementById("pieCenter").textContent = formatPct((domVal + ovsVal) / nav);
      const legend = document.getElementById("allocLegend"); legend.innerHTML = "";
      parts.forEach((p) => {
        const item = document.createElement("div"); item.className = "legend-item";
        const sw = document.createElement("span"); sw.className = "legend-swatch"; sw.style.background = CHART_COLORS[p.key];
        const lb = document.createElement("span"); lb.className = "legend-label"; lb.textContent = p.label;
        const vl = document.createElement("span"); vl.className = "legend-val"; vl.textContent = `${formatPct(p.val / nav)} · ${formatEok(p.val)}`;
        item.appendChild(sw); item.appendChild(lb); item.appendChild(vl); legend.appendChild(item);
      });
      renderExposureBar(nav, parts);
    }

    function renderExposureBar(nav, capitalParts) {
      const wrap = document.getElementById("exposureWrap");
      if (!wrap) return;
      const dfn = sum(activeList("domesticFutures").map((r) => liveFut(r).valueAbs));
      const ofn = sum(activeList("overseasFutures").map((r) => liveFut(r).valueAbs));
      const dfList = activeList("domesticFutures").map(liveFut);
      const ofList = activeList("overseasFutures").map(liveFut);
      const futNet = sum(dfList.map((c) => dirSign(c.direction) * c.valueAbs)) + sum(ofList.map((c) => dirSign(c.direction) * c.valueAbs));
      const segs = capitalParts.map((p) => ({ key: p.key, label: p.label, val: p.val }))
        .concat([{ key: "domesticFut", label: "국내선물", val: dfn }, { key: "overseasFut", label: "해외선물", val: ofn }]);
      const total = nav + dfn + ofn;
      wrap.innerHTML = "";
      const head = document.createElement("div"); head.className = "exp-head";
      head.innerHTML = `<span>총 익스포저 (NAV 대비)</span><span class="muted">자본 100% + 선물 노셔널 ${formatPct((dfn + ofn) / nav)} · 순노출 ${formatPct(futNet / nav)}</span>`;
      wrap.appendChild(head);
      const track = document.createElement("div"); track.className = "exp-track";
      segs.forEach((s) => {
        if (!s.val) return;
        const seg = document.createElement("div"); seg.className = "exp-seg";
        seg.style.width = (s.val / total * 100) + "%";
        seg.style.background = CHART_COLORS[s.key];
        const pctOfNav = s.val / nav * 100;
        if (pctOfNav >= 7) seg.textContent = formatPct(s.val / nav);
        seg.title = `${s.label} ${formatPct(s.val / nav)} · ${formatEok(s.val)}`;
        track.appendChild(seg);
      });
      const marker = document.createElement("div"); marker.className = "exp-marker"; marker.style.left = (nav / total * 100) + "%";
      const ml = document.createElement("div"); ml.className = "exp-marker-label"; ml.textContent = "자본 100%"; marker.appendChild(ml);
      track.appendChild(marker);
      wrap.appendChild(track);
      const lg = document.createElement("div"); lg.className = "exp-legend";
      segs.forEach((s) => {
        if (!s.val) return;
        const item = document.createElement("div"); item.className = "legend-item";
        const sw = document.createElement("span"); sw.className = "legend-swatch"; sw.style.background = CHART_COLORS[s.key];
        const lb = document.createElement("span"); lb.textContent = `${s.label} ${formatPct(s.val / nav)}`;
        item.appendChild(sw); item.appendChild(lb); lg.appendChild(item);
      });
      wrap.appendChild(lg);
    }

    function renderFutSummary() {
      const wrap = document.getElementById("futSummary"); wrap.innerHTML = "";
      const tabs = [["domesticFutures", "국내선물"], ["overseasFutures", "해외선물"]];
      let allNet = 0, allGross = 0;
      tabs.forEach(([t]) => activeList(t).map(liveFut).forEach((c) => { allNet += dirSign(c.direction) * c.valueAbs; allGross += c.valueAbs; }));
      const head = document.createElement("div"); head.className = "row"; head.style.fontWeight = "700"; head.style.background = "#eef3f8";
      head.innerHTML = `<span>순노출 ${formatEok(allNet)}</span><span class="number">그로스 ${formatEok(allGross)}${DATA.nav ? " · " + formatPct(allGross / DATA.nav) : ""}</span>`;
      wrap.appendChild(head);
      tabs.forEach(([t, label]) => {
        const rows = activeList(t); if (!rows.length) return;
        const lives = rows.map((r) => ({ r, c: liveFut(r) }));
        const net = sum(lives.map((x) => dirSign(x.c.direction) * x.c.valueAbs));
        const sub = document.createElement("div"); sub.className = "row"; sub.style.fontWeight = "700";
        sub.innerHTML = `<span>${label} (${rows.length})</span><span class="number">순 ${formatEok(net)}</span>`;
        wrap.appendChild(sub);
        const byK = {};
        lives.forEach(({ r, c }) => { const k = r.kind || "기타"; const b = byK[k] = byK[k] || { notional: 0, long: 0, short: 0 }; b.notional += c.valueAbs; if (c.direction === "매도") b.short++; else b.long++; });
        Object.keys(byK).sort().forEach((k) => {
          const b = byK[k]; const row = document.createElement("div"); row.className = "row"; row.style.fontSize = "12px";
          row.innerHTML = `<span>· ${k}</span><span class="number">${formatEok(b.notional)} <span class="badge b">매수${b.long}</span> <span class="badge s">매도${b.short}</span></span>`;
          wrap.appendChild(row);
        });
      });
    }

    // ---- 필터/뷰 ----
    function applyViewMode() {
      const detail = state.viewMode === "detail";
      document.querySelectorAll("#mainTable .col-detail").forEach((el) => el.classList.toggle("col-hidden", !detail));
    }
    function applyFilters() {
      const q = state.search;
      let visible = 0;
      document.querySelectorAll("#tbody tr").forEach((tr) => {
        const matchSearch = !q || (tr.dataset.search || "").includes(q);
        const matchGroup = state.group === "all" || (tr.dataset.group || "") === state.group;
        const show = matchSearch && matchGroup;
        tr.classList.toggle("hidden", !show);
        if (show) visible++;
      });
      document.getElementById("visibleCount").textContent = `${visible}건 표시`;
    }

    // ---- 엑셀/CSV ----
    function assetsForExport() {
      const out = [];
      TABS.forEach((t) => {
        state.assets[t].forEach((r) => {
          if (isFut(t)) {
            const c = liveFut(r);
            out.push({ sourceRow: r.sourceRow, deleted: r.deleted, isNew: r.isNew, category: t, group: r.group, kind: r.kind, sub: r.sub, code: r.code, ticker: r.ticker, stockName: r.stockName, qty: r.qty, value: r.deleted ? r.value0 : c.value, memo: r.memo });
          } else {
            const c = liveStock(r);
            out.push({ sourceRow: r.sourceRow, deleted: r.deleted, isNew: r.isNew, category: t, group: r.group, kind: r.kind, code: r.code, ticker: r.ticker, stockName: r.stockName, qty: r.qty, price: r.price, value: r.deleted ? r.value0 : c.value, targetPct: r.targetPct, memo: r.memo });
          }
        });
      });
      return out;
    }

    function downloadExcel() {
      try {
        const X = window.XLSX;
        const wb = X.read(DATA.workbookBase64, { type: "base64" });
        window.BlackOnXlsx.applyEditsToWorkbook(X, wb, DATA, assetsForExport());
        const out = X.write(wb, { bookType: "xlsx", type: "array" });
        const blob = new Blob([out], { type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" });
        triggerDownload(blob, fileStem() + ".xlsx");
      } catch (err) {
        alert("엑셀 생성 실패: " + err.message);
        console.error(err);
      }
    }

    function exportCsv() {
      const tab = state.assetTab;
      const rows = activeRows();
      const heads = HEADERS[tab].map((h) => h[0]).filter((h) => h !== "");
      const lines = [heads];
      rows.forEach((r) => {
        if (isFut(tab)) { const c = liveFut(r); lines.push([r.kind, c.direction, r.stockName, r.qty, c.valueAbs, c.weight, r.maturity, c.pnl, r.memo]); }
        else {
          const c = liveStock(r);
          if (tab === "overseasStock") lines.push([r.ticker, r.stockName, r.qty, c.value, c.weight, r.targetPct, c.diff, c.adjQty, c.pnl, c.returnPct, r.weightPrev, r.memo]);
          else lines.push([r.stockName, r.qty, r.price, c.value, c.weight, r.targetPct, c.diff, c.adjQty, c.pnl, c.returnPct, r.weightPrev, r.memo]);
        }
      });
      const csv = lines.map((row) => row.map(csvEscape).join(",")).join("\n");
      triggerDownload(new Blob(["﻿" + csv], { type: "text/csv;charset=utf-8" }), fileStem() + "_" + tab + ".csv");
    }

    function fileStem() { const base = DATA.workbook.replace(/\.xlsx$/i, ""); return base + "_edited"; }
    function triggerDownload(blob, name) {
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a"); a.href = url; a.download = name; a.click();
      setTimeout(() => URL.revokeObjectURL(url), 1000);
    }

    // ---- 포맷/유틸 ----
    function num(v) { if (v === null || v === undefined || v === "") return null; const n = Number(v); return Number.isFinite(n) ? n : null; }
    function parseNumOrNull(v) { const t = (v || "").toString().replace(/,/g, "").trim(); if (!t) return null; const n = Number(t); return Number.isFinite(n) ? n : null; }
    function parsePctOrNull(v) { const t = (v || "").toString().replace("%", "").trim(); if (!t) return null; const n = Number(t); return Number.isFinite(n) ? n / 100 : null; }
    function round(v, d) { if (v == null || Number.isNaN(v)) return ""; const p = Math.pow(10, d); return Math.round(v * p) / p; }
    function sum(a) { return a.reduce((s, v) => s + (Number(v) || 0), 0); }
    function sumDef(a) { const f = a.filter((v) => v != null); return f.length ? sum(f) : null; }
    function formatEok(v) { if (v == null || Number.isNaN(v)) return "-"; return round(v / 1e8, 2).toLocaleString("ko-KR", { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + "억"; }
    function formatPct(v) { if (v == null || Number.isNaN(v)) return ""; return round(v * 100, 2).toLocaleString("ko-KR", { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + "%"; }
    function fmt(v, type) {
      if (v == null || v === "" || Number.isNaN(v)) return "";
      const n = Number(v);
      if (type === "eok") return formatEok(n);
      if (type === "pct") return formatPct(n);
      if (type === "ret") return (n >= 0 ? "+" : "") + round(n, 2) + "%";
      if (type === "qty") return Math.round(n).toLocaleString("ko-KR");
      if (type === "won0") return Math.round(n).toLocaleString("ko-KR");
      if (type === "raw") return n.toLocaleString("ko-KR", { maximumFractionDigits: 4 });
      return n.toLocaleString("ko-KR");
    }
    function csvEscape(v) { const t = v == null ? "" : String(v); return /[",\n]/.test(t) ? `"${t.replace(/"/g, '""')}"` : t; }
  </script>
</body>
</html>
"""


def write_html(data, output_path):
    base = Path(__file__).resolve().parent
    sheetjs = (base / "vendor" / "xlsx.full.min.js").read_text(encoding="utf-8")
    patchjs = (base / "vendor" / "blackon_xlsx.js").read_text(encoding="utf-8")
    html = HTML_TEMPLATE
    html = html.replace("__SHEETJS__", sheetjs)
    html = html.replace("__BLACKON_XLSX__", patchjs)
    html = html.replace(
        "__DATA_JSON__",
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
    )
    output_path.write_text(html, encoding="utf-8")


def main():
    # 경로: 이 스크립트는 igis/dashboard/generate_black_on_dashboard.py 에 위치
    base_dir = Path(__file__).resolve().parent          # igis/dashboard

    # 입력 워크북 결정: 인자가 있으면 그 파일, 없으면 최신 black_on_*.xlsx 자동 선택
    if len(sys.argv) > 1 and not sys.argv[1].startswith("--"):
        workbook_path = Path(sys.argv[1])
        if not workbook_path.is_absolute():
            workbook_path = base_dir / workbook_path
    else:
        candidates = sorted(
            base_dir.glob("black_on_*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        candidates = [p for p in candidates if not p.name.startswith("~$")]
        if not candidates:
            print("black_on_*.xlsx 파일을 찾을 수 없습니다. (dashboard 폴더에 워크북을 두거나 파일명을 인자로 주세요)")
            sys.exit(1)
        workbook_path = candidates[0]
        print(f"[자동 선택] {workbook_path.name}")

    if not workbook_path.exists():
        print(f"워크북을 찾을 수 없습니다: {workbook_path}")
        sys.exit(1)

    # 출력: dashboard 폴더 안에 black_on_dashboard.html 생성
    if len(sys.argv) > 2 and not sys.argv[2].startswith("--"):
        output_path = Path(sys.argv[2])
        if not output_path.is_absolute():
            output_path = base_dir / output_path
    else:
        output_path = base_dir / "black_on_dashboard.html"

    data = build_data(workbook_path)
    if "--debug" in sys.argv:
        _debug_dump(data)
        return
    write_html(data, output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()