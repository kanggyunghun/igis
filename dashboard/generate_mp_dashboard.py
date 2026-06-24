# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

import openpyxl


FUND_CONFIG = [
    {"key": "income2", "label": "인컴2", "code": "300019", "sheet": "인컴2", "mp_col": 6},
    {"key": "blue1", "label": "블루1", "code": "300037", "sheet": "블루1", "mp_col": 10},
    {"key": "blue3", "label": "블루3", "code": "300085", "sheet": "블루3", "mp_col": 14},
]


def clean_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def clean_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def read_holdings(wb):
    holdings = []
    for fund in FUND_CONFIG:
        ws = wb[fund["sheet"]]
        for row in range(3, ws.max_row + 1):
            name = clean_text(ws.cell(row, 5).value)
            if not name:
                continue
            holdings.append(
                {
                    "fundKey": fund["key"],
                    "fundLabel": fund["label"],
                    "fundCode": clean_text(ws.cell(row, 2).value) or fund["code"],
                    "fundName": clean_text(ws.cell(row, 3).value),
                    "stockCode": clean_text(ws.cell(row, 4).value),
                    "stockName": name,
                    "prevQty": clean_number(ws.cell(row, 6).value),
                    "settleQty": clean_number(ws.cell(row, 10).value),
                    "orderQty": clean_number(ws.cell(row, 11).value),
                    "price": clean_number(ws.cell(row, 12).value),
                    "cost": clean_number(ws.cell(row, 13).value),
                    "value": clean_number(ws.cell(row, 14).value),
                    "returnPct": clean_number(ws.cell(row, 17).value),
                    "dayChangePct": clean_number(ws.cell(row, 18).value),
                    "sourceSheet": fund["sheet"],
                    "sourceRow": row,
                }
            )
    return holdings


def read_funds(wb):
    mp = wb["MP"]
    funds = []
    for fund in FUND_CONFIG:
        col = fund["mp_col"]
        funds.append(
            {
                "key": fund["key"],
                "label": fund["label"],
                "code": fund["code"],
                "navEok": clean_number(mp.cell(1, col).value),
                "bondEok": clean_number(mp.cell(2, col).value),
                "stockAvailableEok": clean_number(mp.cell(3, col).value),
                "stockTargetEok": clean_number(mp.cell(4, col).value),
            }
        )
    return funds


MP_DATA_START_ROW = 7
MP_TOTAL_KEYWORDS = ("총 합계", "총합계", "합계")
MP_STOCK_SUBTOTAL_KEYWORDS = ("주식 소계", "주식소계")
MP_OTHER_SUBTOTAL_KEYWORDS = ("기타 소계", "기타소계")


def _detail_contains(detail: str, keywords) -> bool:
    text = detail.replace(" ", "")
    return any(kw.replace(" ", "") in text for kw in keywords)


def read_mp_rows(wb):
    mp = wb["MP"]
    rows = []
    current_sector = ""
    in_other_block = False

    for row in range(MP_DATA_START_ROW, mp.max_row + 1):
        sector_cell = mp.cell(row, 2).value
        detail = clean_text(mp.cell(row, 3).value)
        name = clean_text(mp.cell(row, 18).value)

        if _detail_contains(detail, MP_TOTAL_KEYWORDS):
            break
        if _detail_contains(detail, MP_OTHER_SUBTOTAL_KEYWORDS):
            break
        if _detail_contains(detail, MP_STOCK_SUBTOTAL_KEYWORDS):
            in_other_block = True
            continue

        if in_other_block:
            current_sector = "기타"
        elif isinstance(sector_cell, str) and sector_cell.strip() and "소계" not in detail:
            current_sector = sector_cell.strip()

        if not name:
            continue

        rows.append(
            {
                "sourceRow": row,
                "no": clean_number(mp.cell(row, 1).value),
                "sector": current_sector,
                "detail": detail,
                "stockName": name,
                "targetPct": clean_number(mp.cell(row, 5).value),
                "memo": "",
            }
        )
    return rows


def unique_stock_names(holdings):
    names = OrderedDict()
    for item in holdings:
        if item["stockName"]:
            names[item["stockName"]] = True
    return list(names.keys())


def read_source_date(wb):
    for sheet_name in wb.sheetnames:
        if "아이타스" in sheet_name:
            ws = wb[sheet_name]
            value = ws.cell(3, 2).value
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            return clean_text(value)
    return ""


def build_data(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    stat = workbook_path.stat()
    holdings = read_holdings(wb)
    return {
        "workbook": workbook_path.name,
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbookModifiedAt": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "workbookSize": stat.st_size,
        "sourceDate": read_source_date(wb),
        "funds": read_funds(wb),
        "holdings": holdings,
        "mpRows": read_mp_rows(wb),
        "stockNames": unique_stock_names(holdings),
        "mpStockCount": clean_number(wb["MP"].cell(3, 1).value),
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>2팀 MP 관리 대시보드</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --line: #d8dee8;
      --line-strong: #aab6c6;
      --text: #1d2733;
      --muted: #687589;
      --head: #143a5a;
      --head-2: #1f6f8f;
      --ok: #0f7b59;
      --warn: #a45b00;
      --bad: #b3261e;
      --soft-ok: #e5f3ed;
      --soft-warn: #fff2d8;
      --soft-bad: #fde7e4;
      --edit: #fff8d7;
      --shadow: 0 1px 2px rgba(20, 34, 52, 0.08);
    }

    * {
      box-sizing: border-box;
    }

    body {
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: "Pretendard Variable", "Pretendard", -apple-system, BlinkMacSystemFont,
        system-ui, "Apple SD Gothic Neo", "Noto Sans KR", "Malgun Gothic",
        "Segoe UI", Arial, sans-serif;
      font-size: 14px;
      line-height: 1.45;
      -webkit-font-smoothing: antialiased;
      -moz-osx-font-smoothing: grayscale;
      text-rendering: optimizeLegibility;
      font-feature-settings: "tnum" 1;
    }

    header {
      position: sticky;
      top: 0;
      z-index: 20;
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      box-shadow: var(--shadow);
    }

    .topbar {
      max-width: 1680px;
      margin: 0 auto;
      padding: 14px 18px 12px;
      display: grid;
      gap: 10px;
    }

    .title-row {
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 16px;
      flex-wrap: wrap;
    }

    h1 {
      margin: 0;
      font-size: 20px;
      letter-spacing: 0;
      font-weight: 700;
      color: var(--head);
    }

    .meta {
      color: var(--muted);
      font-size: 12px;
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
    }

    .toolbar {
      display: flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
    }

    .toolbar-divider {
      width: 1px;
      height: 22px;
      background: var(--line);
      margin: 0 2px;
    }

    .filter-chips {
      display: inline-flex;
      flex-wrap: wrap;
      gap: 6px;
      flex: 1 1 auto;
    }

    .chip {
      display: inline-flex;
      align-items: center;
      gap: 5px;
      padding: 2px 8px;
      height: 22px;
      border-radius: 999px;
      background: #e8f1f7;
      color: var(--head);
      font-size: 12px;
      font-weight: 600;
      border: 1px solid #c4d8e6;
    }

    .chip button {
      all: unset;
      cursor: pointer;
      font-weight: 800;
      color: var(--head-2);
      padding: 0 2px;
      line-height: 1;
    }

    .chip button:hover {
      color: var(--bad);
    }

    button,
    select,
    input {
      font: inherit;
      border: 1px solid var(--line-strong);
      background: #fff;
      color: var(--text);
      height: 32px;
      border-radius: 6px;
    }

    button {
      padding: 0 11px;
      cursor: pointer;
      font-weight: 600;
    }

    button.primary {
      background: var(--head-2);
      border-color: var(--head-2);
      color: #fff;
    }

    button.danger {
      color: var(--bad);
      border-color: #e0aaa5;
      background: #fff;
    }

    input,
    select {
      padding: 0 9px;
    }

    #searchInput {
      width: min(280px, 100%);
    }

    #fundSelect,
    #sectorFundSelect {
      min-width: 116px;
    }

    #sectorSelect,
    #detailSelect {
      min-width: 140px;
      max-width: 200px;
    }

    main {
      max-width: 1680px;
      margin: 0 auto;
      padding: 16px 18px 28px;
      display: grid;
      gap: 14px;
    }

    .summary {
      display: grid;
      grid-template-columns: repeat(6, minmax(130px, 1fr));
      gap: 10px;
    }

    .metric,
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      box-shadow: var(--shadow);
    }

    .metric {
      padding: 11px 12px;
      min-height: 70px;
    }

    .metric .label {
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 6px;
    }

    .metric .value {
      font-size: 20px;
      font-weight: 700;
      color: var(--head);
    }

    .metric.bad .value {
      color: var(--bad);
    }

    .metric.warn .value {
      color: var(--warn);
    }

    .panel {
      overflow: hidden;
    }

    .panel-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      padding: 10px 12px;
      background: #eef3f8;
      border-bottom: 1px solid var(--line);
      color: var(--head);
      font-weight: 700;
      min-height: 50px;
      box-sizing: border-box;
    }

    .panel-body {
      padding: 10px 12px;
    }

    .grid-2 {
      display: grid;
      grid-template-columns: minmax(0, 1.2fr) minmax(360px, 0.8fr);
      gap: 14px;
    }

    .workspace {
      display: grid;
      grid-template-columns: 320px minmax(0, 1fr);
      gap: 14px;
      align-items: stretch;
    }

    .workspace > section.panel {
      display: flex;
      flex-direction: column;
      max-height: calc(100vh - 130px);
      min-height: 0;
    }

    .sector-panel {
      position: sticky;
      top: 112px;
      max-height: calc(100vh - 130px);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .sector-panel .panel-body {
      padding: 8px 8px;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      gap: 8px;
      min-height: 0;
    }

    .sector-nav {
      display: grid;
      gap: 4px;
      overflow: auto;
      max-height: 42vh;
      padding-right: 4px;
    }

    .sector-button {
      width: 100%;
      text-align: left;
      padding: 9px 11px;
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 6px;
      display: grid;
      gap: 3px;
      cursor: pointer;
      height: auto;
    }

    .sector-button:hover {
      border-color: var(--head-2);
      background: #f4faff;
    }

    .sector-button.active {
      border-color: var(--head-2);
      background: #e8f4f8;
      box-shadow: inset 3px 0 0 var(--head-2);
    }

    .sector-row {
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      gap: 8px;
    }

    .sector-name {
      font-weight: 700;
      color: var(--head);
      font-size: 13.5px;
    }

    .sector-target {
      font-weight: 700;
      color: var(--text);
      font-size: 13.5px;
      font-variant-numeric: tabular-nums;
    }

    .sector-count {
      font-size: 11px;
      color: var(--muted);
    }

    .sector-gap {
      font-size: 11.5px;
      font-weight: 600;
      font-variant-numeric: tabular-nums;
    }

    .sector-stock-box {
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fafbfd;
      display: flex;
      flex-direction: column;
      min-height: 0;
      overflow: hidden;
    }

    .sector-stock-head {
      padding: 7px 10px;
      font-weight: 700;
      color: var(--head);
      font-size: 12px;
      background: #eef3f8;
      border-bottom: 1px solid var(--line);
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    .sector-stock-list {
      margin: 0;
      padding: 6px;
      list-style: none;
      display: grid;
      gap: 4px;
      overflow: auto;
      max-height: 32vh;
    }

    .sector-stock-list li {
      display: grid;
      grid-template-columns: 1fr auto;
      align-items: center;
      gap: 6px;
      padding: 5px 8px;
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 5px;
      font-size: 12px;
    }

    .sector-stock-list .stock-name {
      font-weight: 600;
      color: var(--text);
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .sector-stock-list .stock-meta {
      font-variant-numeric: tabular-nums;
      font-size: 11px;
      color: var(--muted);
    }

    .fund-tabs {
      display: inline-flex;
      gap: 0;
      border: 1px solid var(--line-strong);
      border-radius: 6px;
      overflow: hidden;
      background: #fff;
    }

    .fund-tab {
      all: unset;
      cursor: pointer;
      padding: 4px 12px;
      font-weight: 600;
      font-size: 12px;
      color: var(--muted);
      border-right: 1px solid var(--line);
      transition: all 0.1s;
    }

    .fund-tab:last-child { border-right: 0; }
    .fund-tab:hover { background: #f4faff; color: var(--head); }
    .fund-tab.active {
      background: var(--head-2);
      color: #fff;
    }

    .pie-chart {
      width: min(280px, 100%);
      aspect-ratio: 1;
      border-radius: 50%;
      background: conic-gradient(#d7dde6 0 100%);
      border: 1px solid var(--line);
      position: relative;
    }

    .pie-chart::after {
      content: '';
      position: absolute;
      width: 50%;
      height: 50%;
      top: 25%;
      left: 25%;
      background: var(--panel);
      border-radius: 50%;
      border: 1px solid var(--line);
      box-shadow: 0 0 0 4px var(--panel);
    }

    .pie-labels {
      position: absolute;
      inset: 0;
      pointer-events: none;
      z-index: 1;
    }

    .pie-label {
      position: absolute;
      transform: translate(-50%, -50%);
      color: #fff;
      font-size: 12px;
      font-weight: 700;
      white-space: nowrap;
      text-shadow: 0 1px 2px rgba(0,0,0,0.55), 0 0 1px rgba(0,0,0,0.4);
      letter-spacing: -0.1px;
    }

    .pie-center {
      position: absolute;
      inset: 0;
      display: grid;
      place-items: center;
      pointer-events: none;
      z-index: 2;
      text-align: center;
    }

    .pie-center-label {
      font-size: 11px;
      color: var(--muted);
      font-weight: 600;
    }

    .pie-center-value {
      font-size: 22px;
      color: var(--head);
      font-weight: 800;
      line-height: 1;
      margin-top: 2px;
    }

    .col-hidden { display: none; }

    th.fund-grp {
      background: var(--head-2);
      box-shadow: inset 0 -2px 0 #f0c44a;
    }

    .view-selects {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
    }

    .chart-layout {
      display: grid;
      grid-template-columns: minmax(210px, 280px) minmax(0, 1fr);
      gap: 16px;
      align-items: center;
    }

    .pie-wrap {
      display: grid;
      justify-items: center;
      gap: 9px;
    }

    .pie-chart {
      width: min(260px, 100%);
      aspect-ratio: 1;
      border-radius: 50%;
      background: conic-gradient(#d7dde6 0 100%);
      border: 1px solid var(--line);
    }

    .pie-caption {
      color: var(--muted);
      font-size: 12px;
      text-align: center;
    }

    .legend {
      display: grid;
      grid-template-columns: repeat(2, minmax(150px, 1fr));
      gap: 6px 12px;
      margin-top: 10px;
    }

    .legend-item {
      display: flex;
      align-items: center;
      gap: 7px;
      min-width: 0;
    }

    .legend-swatch {
      width: 10px;
      height: 10px;
      border-radius: 3px;
      flex: 0 0 auto;
    }

    .legend-label {
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .focus-note {
      color: var(--muted);
      font-weight: 600;
    }

    .table-wrap {
      overflow: auto;
      flex: 1 1 auto;
      min-height: 0;
      background: #fff;
    }

    table {
      width: 100%;
      border-collapse: separate;
      border-spacing: 0;
    }

    #mpTable { table-layout: fixed; width: 100%; min-width: 1127px; }
    #mpTable th:nth-child(1),  #mpTable td:nth-child(1)  { width: 175px; }   /* 섹터 (방산/우주/로봇 수용) */
    #mpTable th:nth-child(2),  #mpTable td:nth-child(2)  { width: 112px; }   /* 세부 */
    #mpTable th:nth-child(3),  #mpTable td:nth-child(3)  { width: 250px; }   /* 종목명 */
    #mpTable th:nth-child(4),  #mpTable td:nth-child(4)  { width: 86px; }    /* 목표 */
    #mpTable th:nth-child(5),  #mpTable td:nth-child(5)  { width: 96px; }    /* 현재가 */
    #mpTable th:nth-child(6),  #mpTable td:nth-child(6)  { width: 78px; }    /* 상태 */
    #mpTable th:nth-child(7),  #mpTable td:nth-child(7)  { width: 88px; }    /* 평가 (detail) */
    #mpTable th:nth-child(8),  #mpTable td:nth-child(8)  { width: 82px; }    /* 비중 */
    #mpTable th:nth-child(9),  #mpTable td:nth-child(9)  { width: 110px; }   /* 차이(+막대) */
    #mpTable th:nth-child(10), #mpTable td:nth-child(10) { width: 98px; }    /* 조정수량 */
    #mpTable th:nth-child(11), #mpTable td:nth-child(11) { width: 84px; }    /* 수익률 (detail) */
    #mpTable th:nth-child(12), #mpTable td:nth-child(12) { width: 40px; }    /* X */
    #mpTable td input { box-sizing: border-box; }
    #mpTable td input.editable { text-overflow: ellipsis; }
    #mpTable td.col-stockname input { text-overflow: ellipsis; }

    th,
    td {
      border-right: 1px solid var(--line);
      border-bottom: 1px solid var(--line);
      padding: 9px 11px;
      vertical-align: middle;
      white-space: nowrap;
    }

    th {
      position: sticky;
      top: 0;
      z-index: 5;
      background: var(--head);
      color: #fff;
      font-weight: 700;
      text-align: center;
      padding: 11px 11px;
      letter-spacing: 0.2px;
    }

    tbody td { background: #fff; }
    tbody tr:nth-child(even) td { background: #fafbfd; }

    tbody tr.hidden { display: none; }

    td.number {
      text-align: right;
      font-variant-numeric: tabular-nums;
    }

    td.center { text-align: center; }

    td.gap-pos { color: var(--ok); font-weight: 700; }
    td.gap-neg { color: var(--bad); font-weight: 700; }
    td.gap-zero { color: var(--muted); }

    tbody tr.row-bad td.col-stockname {
      box-shadow: inset 3px 0 0 var(--bad);
    }
    tbody tr.row-bad td.col-stockname::before {
      content: '● ';
      color: var(--bad);
      font-size: 10px;
      margin-right: 2px;
    }

    /* Diff inline bar */
    td.col-diff { padding: 6px 11px; }
    .diff-cell {
      display: grid;
      grid-template-columns: 1fr;
      gap: 3px;
      align-items: center;
    }
    .diff-text {
      text-align: right;
      font-variant-numeric: tabular-nums;
      font-weight: 700;
    }
    .diff-text.gap-pos { color: var(--ok); }
    .diff-text.gap-neg { color: var(--bad); }
    .diff-text.gap-zero { color: var(--muted); }
    .diff-bar {
      height: 4px;
      background: #eef1f5;
      border-radius: 2px;
      position: relative;
      overflow: hidden;
    }
    .diff-bar::before {
      content: '';
      position: absolute;
      top: 0;
      bottom: 0;
      left: 50%;
      width: 1px;
      background: var(--line-strong);
      z-index: 1;
    }
    .diff-bar-fill {
      position: absolute;
      top: 0;
      bottom: 0;
      height: 100%;
    }
    .diff-bar-fill.pos {
      left: 50%;
      background: var(--ok);
    }
    .diff-bar-fill.neg {
      right: 50%;
      background: var(--bad);
    }

    /* Consecutive same-sector visual merge */
    tr.same-sector td.col-sector input {
      color: transparent;
    }
    tr.same-sector td.col-sector input:hover,
    tr.same-sector td.col-sector input:focus {
      color: var(--text);
    }

    /* Column width constraints */
    #mpTable th.col-sector,
    #mpTable td.col-sector { min-width: 100px; max-width: 130px; }
    #mpTable th.col-detail-input,
    #mpTable td.col-detail-input { min-width: 100px; max-width: 130px; }
    #mpTable th.col-stockname,
    #mpTable td.col-stockname {
      min-width: 220px;
      font-weight: 700;
    }
    #mpTable td.col-stockname input.name-input { font-weight: 700; }

    tfoot td {
      position: sticky;
      bottom: 0;
      background: #f0f4f8;
      font-weight: 700;
      color: var(--head);
      border-top: 2px solid var(--head-2);
    }

    .legend-item {
      cursor: pointer;
      padding: 2px 4px;
      border-radius: 4px;
      transition: background 0.1s;
    }
    .legend-item:hover { background: #f0f4f8; }
    .legend-item.active { background: #e0eaf2; }

    .editable {
      width: 100%;
      min-width: 92px;
      border: 1px solid transparent;
      background: transparent;
      color: var(--text);
      padding: 4px 6px;
      border-radius: 4px;
      transition: background 0.08s, border-color 0.08s;
    }

    .editable:hover {
      background: #f1f5f9;
      border-color: var(--line);
    }

    .editable:focus {
      outline: none;
      background: var(--edit);
      border-color: #e1c96a;
      box-shadow: 0 0 0 2px rgba(225, 201, 106, 0.25);
    }

    .name-input {
      min-width: 210px;
    }

    .target-input {
      width: 82px;
      text-align: right;
    }
    .target-input:focus { text-align: right; }

    .badge {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 64px;
      height: 24px;
      padding: 0 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
    }

    .badge.ok {
      color: var(--ok);
      background: var(--soft-ok);
    }

    .badge.warn {
      color: var(--warn);
      background: var(--soft-warn);
    }

    .badge.bad {
      color: var(--bad);
      background: var(--soft-bad);
    }

    .mini-table th {
      background: #315b7b;
    }

    .mini-table td:first-child {
      font-weight: 700;
      color: var(--head);
    }

    .validation-list {
      display: grid;
      gap: 10px;
    }

    .validation-box {
      border: 1px solid var(--line);
      background: #fff;
    }

    .validation-box h3 {
      margin: 0;
      padding: 8px 10px;
      font-size: 13px;
      color: var(--head);
      background: #f4f7fa;
      border-bottom: 1px solid var(--line);
    }

    .validation-box ul {
      margin: 0;
      padding: 8px 10px;
      list-style: none;
      display: grid;
      gap: 5px;
      max-height: 145px;
      overflow: auto;
    }

    .validation-box li {
      display: flex;
      justify-content: space-between;
      gap: 10px;
      border-bottom: 1px dashed #e5eaf0;
      padding-bottom: 4px;
    }

    .validation-box li:last-child {
      border-bottom: 0;
      padding-bottom: 0;
    }

    .calc-box {
      padding: 10px;
      display: grid;
      gap: 8px;
    }

    .calc-row {
      display: flex;
      align-items: center;
      gap: 8px;
    }

    .calc-row label {
      flex: 0 0 56px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }

    .calc-row input {
      flex: 1 1 auto;
      min-width: 0;
      padding: 5px 8px;
      border: 1px solid var(--line-strong);
      border-radius: 5px;
      font-size: 13px;
      background: #fff;
    }

    .calc-row input:focus {
      outline: none;
      border-color: var(--head-2);
      box-shadow: 0 0 0 2px rgba(31, 111, 143, 0.18);
    }

    .calc-unit {
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      flex: 0 0 auto;
    }

    .calc-output {
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 6px 14px;
      padding: 8px 4px 0;
      border-top: 1px dashed var(--line);
    }

    .calc-output > div {
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      gap: 6px;
      font-size: 12px;
      color: var(--muted);
    }

    .calc-output strong {
      font-variant-numeric: tabular-nums;
      color: var(--head);
      font-size: 13px;
      font-weight: 700;
    }

    .muted {
      color: var(--muted);
    }

    .delete-row {
      width: 30px;
      padding: 0;
    }

    @media (max-width: 1100px) {
      .summary {
        grid-template-columns: repeat(3, minmax(130px, 1fr));
      }

      .grid-2 {
        grid-template-columns: 1fr;
      }

      .chart-layout {
        grid-template-columns: 1fr;
      }

      .table-wrap {
        max-height: none;
      }
    }

    @media (max-width: 640px) {
      .summary {
        grid-template-columns: repeat(2, minmax(120px, 1fr));
      }

      .legend {
        grid-template-columns: 1fr;
      }

      .topbar,
      main {
        padding-left: 10px;
        padding-right: 10px;
      }

      button,
      select,
      input {
        max-width: 100%;
      }
    }
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <div class="title-row">
        <h1>2팀 MP 관리 대시보드</h1>
        <div class="meta">
          <span id="workbookMeta"></span>
          <span id="sourceMeta"></span>
          <span id="saveMeta"></span>
        </div>
      </div>
      <div class="toolbar">
        <button class="primary" id="addRowBtn" type="button" title="행 추가 (Ctrl+N)">+ 행 추가</button>
        <button id="saveBtn" type="button" title="브라우저 저장 (Ctrl+S)">💾 저장</button>
        <button id="exportBtn" type="button">CSV</button>
        <button class="danger" id="resetBtn" type="button">초기화</button>
        <span class="toolbar-divider" aria-hidden="true"></span>
        <select id="fundSelect" aria-label="테이블 펀드 선택" style="display:none;">
          <option value="income2">인컴2</option>
          <option value="blue1">블루1</option>
          <option value="blue3">블루3</option>
        </select>
        <select id="viewModeSelect" aria-label="표시 모드">
          <option value="simple">간단 표시</option>
          <option value="detail">상세 표시</option>
        </select>
        <select id="sectorSelect" aria-label="섹터 필터">
          <option value="all">전체 섹터</option>
        </select>
        <select id="detailSelect" aria-label="세부 필터">
          <option value="all">전체 세부</option>
        </select>
        <select id="statusFilter" aria-label="상태 필터">
          <option value="all">전체 상태</option>
          <option value="ok">OK</option>
          <option value="missing">원천누락</option>
          <option value="price">현재가누락</option>
        </select>
        <input id="searchInput" type="search" placeholder="🔍 종목/섹터/세부 검색" />
        <button id="clearFiltersBtn" type="button" title="필터 초기화">필터 해제</button>
      </div>
    </div>
  </header>

  <main>
    <section class="summary" aria-label="요약">
      <div class="metric"><div class="label">종목 수</div><div class="value" id="metricRows">0</div></div>
      <div class="metric"><div class="label">목표비중 합계</div><div class="value" id="metricTarget">0.00%</div></div>
      <div class="metric" id="metricMissingBox"><div class="label">원천누락</div><div class="value" id="metricMissing">0</div></div>
      <div class="metric" id="metricPriceBox"><div class="label">현재가누락</div><div class="value" id="metricPriceMissing">0</div></div>
      <div class="metric"><div class="label">원천 보유종목</div><div class="value" id="metricSourceNames">0</div></div>
      <div class="metric" id="metricUnlistedBox"><div class="label">MP 미등록</div><div class="value" id="metricUnlisted">0</div></div>
    </section>

    <section class="workspace">
      <aside class="panel sector-panel">
        <div class="panel-head">
          <span>섹터 ▾</span>
          <span class="muted" id="sectorCount"></span>
        </div>
        <div class="panel-body">
          <div class="sector-nav" id="sectorNav"></div>
          <div class="sector-stock-box">
            <div class="sector-stock-head">
              <span id="sectorStockTitle">전체 종목</span>
              <span class="muted" id="sectorStockHint" style="font-size:11px;font-weight:500;">목표▾</span>
            </div>
            <ul class="sector-stock-list" id="sectorStockList"></ul>
          </div>
        </div>
      </aside>

      <section class="panel">
        <div class="panel-head">
          <div class="view-selects">
            <span id="mpTitle">MP 입력</span>
            <div class="fund-tabs" id="mpFundTabs" role="tablist">
              <button class="fund-tab active" type="button" data-fund="income2" role="tab">인컴2</button>
              <button class="fund-tab" type="button" data-fund="blue1" role="tab">블루1</button>
              <button class="fund-tab" type="button" data-fund="blue3" role="tab">블루3</button>
            </div>
          </div>
          <span class="filter-chips" id="filterChips"></span>
          <span class="muted" id="visibleCount"></span>
        </div>
        <div class="table-wrap">
          <table id="mpTable">
            <thead>
              <tr>
                <th>섹터</th>
                <th>세부</th>
                <th>종목명</th>
                <th>목표</th>
                <th>현재가</th>
                <th>상태</th>
                <th class="fund-grp col-detail" id="fundValueHead">평가</th>
                <th class="fund-grp" id="fundWeightHead">비중</th>
                <th class="fund-grp" id="fundDiffHead">차이</th>
                <th class="fund-grp" id="fundQtyHead">조정수량</th>
                <th class="col-detail" id="fundReturnHead">수익률</th>
                <th></th>
              </tr>
            </thead>
            <tbody id="mpBody"></tbody>
          </table>
        </div>
      </section>
    </section>

    <section class="grid-2">
      <section class="panel">
        <div class="panel-head">
          <div class="view-selects">
            <span>섹터 요약</span>
            <div class="fund-tabs" id="sectorFundTabs" role="tablist">
              <button class="fund-tab active" type="button" data-fund="income2" role="tab">인컴2</button>
              <button class="fund-tab" type="button" data-fund="blue1" role="tab">블루1</button>
              <button class="fund-tab" type="button" data-fund="blue3" role="tab">블루3</button>
            </div>
            <select id="sectorFundSelect" aria-label="섹터 요약 펀드 선택" style="display:none;">
              <option value="income2">인컴2</option>
              <option value="blue1">블루1</option>
              <option value="blue3">블루3</option>
            </select>
          </div>
          <span class="focus-note" id="sectorSummaryNote"></span>
        </div>
        <div class="panel-body">
          <div class="chart-layout">
            <div class="pie-wrap">
              <div class="pie-chart" id="sectorPie" role="img" aria-label="섹터 비중 파이차트">
                <div class="pie-labels" id="pieLabels" aria-hidden="true"></div>
                <div class="pie-center">
                  <div>
                    <div class="pie-center-label" id="pieCenterLabel">인컴2</div>
                    <div class="pie-center-value" id="pieCenterValue">0.00%</div>
                  </div>
                </div>
              </div>
              <div class="pie-caption" id="sectorPieCaption"></div>
            </div>
            <div>
              <table class="mini-table" id="sectorTable">
                <thead>
                  <tr>
                    <th>섹터</th>
                    <th>목표</th>
                    <th id="sectorFundHead">인컴2</th>
                    <th>차이</th>
                  </tr>
                </thead>
                <tbody></tbody>
              </table>
              <div class="legend" id="sectorLegend"></div>
            </div>
          </div>
        </div>
      </section>

      <section class="panel">
        <div class="panel-head"><span>검증</span></div>
        <div class="panel-body validation-list">
          <div class="validation-box">
            <h3>원천누락</h3>
            <ul id="missingList"></ul>
          </div>
          <div class="validation-box">
            <h3>MP 미등록</h3>
            <ul id="unlistedList"></ul>
          </div>
          <div class="validation-box">
            <h3>현재가누락</h3>
            <ul id="priceList"></ul>
          </div>
          <div class="validation-box">
            <h3>주식수 계산기</h3>
            <div class="calc-box">
              <div class="calc-row">
                <label for="calcStock">종목명</label>
                <input id="calcStock" list="stockList" placeholder="종목 검색 / 선택" autocomplete="off" />
              </div>
              <div class="calc-row">
                <label for="calcWeight">비중</label>
                <input id="calcWeight" type="text" inputmode="decimal" placeholder="예: 0.5" autocomplete="off" />
                <span class="calc-unit">%</span>
              </div>
              <div class="calc-output">
                <div><span>현재가</span><strong id="calcPrice">-</strong></div>
                <div><span>인컴2</span><strong id="calcFund_income2">-</strong></div>
                <div><span>블루1</span><strong id="calcFund_blue1">-</strong></div>
                <div><span>블루3</span><strong id="calcFund_blue3">-</strong></div>
              </div>
            </div>
          </div>
        </div>
      </section>
    </section>
  </main>

  <datalist id="stockList"></datalist>
  <datalist id="sectorList"></datalist>
  <datalist id="detailList"></datalist>

  <script>
    const DATA = __DATA_JSON__;
    const STORAGE_KEY = `mp-dashboard:${DATA.workbook}:${DATA.workbookModifiedAt}:${DATA.workbookSize}`;
    const FUND_KEYS = DATA.funds.map((fund) => fund.key);
    const CHART_COLORS = ["#1f6f8f", "#7a9f35", "#c46a2b", "#6f5aa7", "#d29a22", "#4c8b77", "#b75b7a", "#5a78a7", "#9a7b3f", "#668c99", "#8b6b58", "#537a52"];
    const state = {
      rows: [],
      computed: [],
      search: "",
      status: "all",
      selectedSector: "all",
      selectedDetail: "all",
      selectedFund: "income2",
      sectorFund: "income2",
      viewMode: "simple",
    };

    const SECTOR_EXCLUDE = new Set(["기타", "미분류", "미분류(기타)"]);

    const byName = new Map();
    const fundByKey = new Map(DATA.funds.map((fund) => [fund.key, fund]));

    DATA.holdings.forEach((holding) => {
      const name = normalizeName(holding.stockName);
      if (!name) return;
      if (!byName.has(name)) byName.set(name, []);
      byName.get(name).push(holding);
    });

    document.addEventListener("DOMContentLoaded", init);

    function init() {
      document.getElementById("workbookMeta").textContent = `${DATA.workbook} | 생성 ${DATA.generatedAt}`;
      document.getElementById("sourceMeta").textContent = DATA.sourceDate ? `원천 ${DATA.sourceDate}` : "";
      buildDatalist();
      loadRows();
      bindEvents();
      bindCalculator();
      render();
    }

    function bindCalculator() {
      const stockInput = document.getElementById("calcStock");
      const weightInput = document.getElementById("calcWeight");
      const recompute = () => updateCalculator(stockInput.value, weightInput.value);
      stockInput.addEventListener("input", recompute);
      weightInput.addEventListener("input", recompute);
    }

    function parseCalcPct(value) {
      const text = (value || "").toString().trim().replace("%", "");
      if (!text) return 0;
      const n = Number(text);
      if (!Number.isFinite(n)) return 0;
      return n / 100;
    }

    function updateCalculator(stockName, weightStr) {
      const name = normalizeName(stockName);
      const weight = parseCalcPct(weightStr);
      const holdings = byName.get(name) || [];
      const price = firstNumber(holdings.map((item) => item.price));

      document.getElementById("calcPrice").textContent =
        price ? formatNumber(price, 0) + "원" : "-";

      DATA.funds.forEach((fund) => {
        const el = document.getElementById("calcFund_" + fund.key);
        if (!el) return;
        if (!name || !price || !weight || !fund.stockTargetEok) {
          el.textContent = "-";
          return;
        }
        const shares = Math.round(fund.stockTargetEok * 1e8 * weight / price);
        el.textContent = formatNumber(shares, 0) + "주";
      });
    }

    function bindEvents() {
      document.getElementById("addRowBtn").addEventListener("click", addEmptyRow);

      document.getElementById("saveBtn").addEventListener("click", saveRows);
      document.getElementById("resetBtn").addEventListener("click", resetRows);
      document.getElementById("exportBtn").addEventListener("click", exportCsv);
      document.getElementById("clearFiltersBtn").addEventListener("click", clearFilters);
      document.getElementById("searchInput").addEventListener("input", (event) => {
        state.search = event.target.value.trim().toLowerCase();
        applyFilters();
        renderFilterChips();
      });
      document.getElementById("statusFilter").addEventListener("change", (event) => {
        state.status = event.target.value;
        applyFilters();
        renderFilterChips();
      });
      document.getElementById("fundSelect").addEventListener("change", (event) => {
        state.selectedFund = event.target.value;
        render();
      });
      document.getElementById("viewModeSelect").addEventListener("change", (event) => {
        state.viewMode = event.target.value;
        applyViewMode();
      });
      document.querySelectorAll("#sectorFundTabs .fund-tab").forEach((tab) => {
        tab.addEventListener("click", () => {
          state.sectorFund = tab.dataset.fund;
          renderSectorTable();
        });
      });
      document.querySelectorAll("#mpFundTabs .fund-tab").forEach((tab) => {
        tab.addEventListener("click", () => {
          state.selectedFund = tab.dataset.fund;
          render();
        });
      });
      document.getElementById("sectorSelect").addEventListener("change", (event) => {
        state.selectedSector = event.target.value;
        state.selectedDetail = "all";
        render();
      });
      document.getElementById("detailSelect").addEventListener("change", (event) => {
        state.selectedDetail = event.target.value;
        applyFilters();
        renderFilterChips();
      });
      document.getElementById("sectorFundSelect").addEventListener("change", (event) => {
        state.sectorFund = event.target.value;
        renderSectorTable();
      });

      document.addEventListener("keydown", (event) => {
        if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === "s") {
          event.preventDefault();
          saveRows();
        }
        if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === "n") {
          event.preventDefault();
          addEmptyRow();
        }
        if (event.key === "Escape" && document.activeElement === document.getElementById("searchInput")) {
          state.search = "";
          document.getElementById("searchInput").value = "";
          applyFilters();
          renderFilterChips();
        }
      });
    }

    function addEmptyRow() {
      state.rows.push({
        sourceRow: null,
        no: null,
        sector: state.selectedSector !== "all" ? state.selectedSector : "",
        detail: state.selectedDetail !== "all" ? state.selectedDetail : "",
        stockName: "",
        targetPct: 0,
        memo: "",
      });
      render();
      const inputs = document.querySelectorAll(".name-input");
      if (inputs.length) inputs[inputs.length - 1].focus();
    }

    function clearFilters() {
      state.search = "";
      state.status = "all";
      state.selectedSector = "all";
      state.selectedDetail = "all";
      document.getElementById("searchInput").value = "";
      document.getElementById("statusFilter").value = "all";
      render();
    }

    function loadRows() {
      const stored = localStorage.getItem(STORAGE_KEY);
      if (stored) {
        try {
          state.rows = JSON.parse(stored);
          document.getElementById("saveMeta").textContent = "저장본 적용";
          return;
        } catch (error) {
          localStorage.removeItem(STORAGE_KEY);
        }
      }
      state.rows = DATA.mpRows.map((row) => ({ ...row }));
    }

    function saveRows() {
      localStorage.setItem(STORAGE_KEY, JSON.stringify(state.rows));
      const meta = document.getElementById("saveMeta");
      meta.textContent = `✓ 저장 ${new Date().toLocaleTimeString("ko-KR", { hour12: false })}`;
      meta.style.color = "var(--ok)";
      meta.style.fontWeight = "700";
      setTimeout(() => { meta.style.color = ""; meta.style.fontWeight = ""; }, 1500);
    }

    function resetRows() {
      if (!confirm("브라우저 저장본을 지우고 엑셀 기준으로 되돌릴까요?")) return;
      localStorage.removeItem(STORAGE_KEY);
      state.rows = DATA.mpRows.map((row) => ({ ...row }));
      document.getElementById("saveMeta").textContent = "";
      render();
    }

    function buildDatalist() {
      const list = document.getElementById("stockList");
      list.innerHTML = "";
      DATA.stockNames.forEach((name) => {
        const option = document.createElement("option");
        option.value = name;
        list.appendChild(option);
      });
    }

    function render() {
      computeRows();
      populateInputDatalists();
      populateFilterOptions();
      renderSectorNav();
      renderMainTable();
      renderSectorTable();
      renderValidation();
      renderMetrics();
      renderFilterChips();
      applyViewMode();
      applyFilters();
    }

    function applyViewMode() {
      const detail = state.viewMode === "detail";
      document.querySelectorAll("#mpTable .col-detail").forEach((el) => {
        el.classList.toggle("col-hidden", !detail);
      });
    }

    function renderSectorNav() {
      const sectors = getSectorStats();
      const nav = document.getElementById("sectorNav");
      nav.innerHTML = "";
      const eligible = state.computed.filter((r) => r.normalizedName && !SECTOR_EXCLUDE.has(r.sector || "미분류"));
      const totalTarget = sum(eligible.map((r) => r.targetPct));
      const totalCurrent = sum(eligible.map((r) => r.fundValues[state.selectedFund].weight || 0));
      nav.appendChild(sectorButton("all", "전체", eligible.length, totalTarget, totalCurrent));
      sectors.forEach((s) => {
        const cur = s.fundWeights[state.selectedFund] || 0;
        nav.appendChild(sectorButton(s.sector, s.sector, s.count, s.targetPct, cur));
      });
      document.getElementById("sectorCount").textContent = `${sectors.length}개`;
      renderSectorStocks();
    }

    function sectorButton(value, label, count, targetPct, currentPct) {
      const button = document.createElement("button");
      button.type = "button";
      button.className = "sector-button" + (state.selectedSector === value ? " active" : "");

      const row1 = document.createElement("div");
      row1.className = "sector-row";
      const name = document.createElement("span");
      name.className = "sector-name";
      name.textContent = label;
      const target = document.createElement("span");
      target.className = "sector-target";
      target.textContent = formatPct(targetPct);
      row1.appendChild(name);
      row1.appendChild(target);

      const row2 = document.createElement("div");
      row2.className = "sector-row";
      const countEl = document.createElement("span");
      countEl.className = "sector-count";
      countEl.textContent = `${formatNumber(count, 0)}종목`;
      const gap = (targetPct || 0) - (currentPct || 0);
      const gapCls = Math.abs(gap) < 0.0005 ? "gap-zero" : (gap > 0 ? "gap-pos" : "gap-neg");
      const gapEl = document.createElement("span");
      gapEl.className = `sector-gap ${gapCls}`;
      gapEl.textContent = (gap >= 0 ? "+" : "") + formatPct(gap);
      row2.appendChild(countEl);
      row2.appendChild(gapEl);

      button.appendChild(row1);
      button.appendChild(row2);
      button.addEventListener("click", () => {
        state.selectedSector = state.selectedSector === value ? "all" : value;
        state.selectedDetail = "all";
        render();
      });
      return button;
    }

    function renderSectorStocks() {
      const list = document.getElementById("sectorStockList");
      const title = document.getElementById("sectorStockTitle");
      const rows = state.computed
        .filter((row) => state.selectedSector === "all" || (row.sector || "미분류") === state.selectedSector)
        .filter((row) => row.normalizedName)
        .sort((a, b) => b.targetPct - a.targetPct);
      title.textContent = state.selectedSector === "all"
        ? `전체 종목 (${rows.length})`
        : `${state.selectedSector} (${rows.length})`;
      list.innerHTML = "";
      if (!rows.length) {
        const empty = document.createElement("li");
        empty.innerHTML = '<span class="muted">없음</span><span></span>';
        list.appendChild(empty);
        return;
      }
      rows.forEach((row) => {
        const fund = row.fundValues[state.selectedFund];
        const gap = (row.targetPct || 0) - (fund.weight || 0);
        const li = document.createElement("li");
        const name = document.createElement("span");
        name.className = "stock-name";
        name.title = row.stockName;
        name.textContent = row.stockName;
        const meta = document.createElement("span");
        meta.className = "stock-meta";
        const gapCls = Math.abs(gap) < 0.0005 ? "" : (gap > 0 ? " gap-pos" : " gap-neg");
        meta.innerHTML = `<strong>${formatPct(row.targetPct)}</strong> <span class="muted${gapCls}">${gap >= 0 ? "+" : ""}${formatPct(gap)}</span>`;
        li.appendChild(name);
        li.appendChild(meta);
        list.appendChild(li);
      });
    }

    function populateInputDatalists() {
      const sectorSet = new Set();
      const detailSet = new Set();
      state.computed.forEach((r) => {
        if (r.sector && !SECTOR_EXCLUDE.has(r.sector)) sectorSet.add(r.sector);
        if (r.detail) detailSet.add(r.detail);
      });
      fillDatalist("sectorList", [...sectorSet].sort((a, b) => a.localeCompare(b, "ko")));
      fillDatalist("detailList", [...detailSet].sort((a, b) => a.localeCompare(b, "ko")));
    }

    function fillDatalist(id, values) {
      const list = document.getElementById(id);
      list.innerHTML = "";
      values.forEach((v) => {
        const opt = document.createElement("option");
        opt.value = v;
        list.appendChild(opt);
      });
    }

    function populateFilterOptions() {
      const sectorSelect = document.getElementById("sectorSelect");
      const detailSelect = document.getElementById("detailSelect");

      const sectorSet = new Set();
      state.computed.forEach((row) => {
        const sec = row.sector || "";
        if (sec && !SECTOR_EXCLUDE.has(sec)) sectorSet.add(sec);
      });
      const sectors = [...sectorSet].sort((a, b) => a.localeCompare(b, "ko"));

      const detailSet = new Set();
      state.computed.forEach((row) => {
        if (state.selectedSector === "all" || row.sector === state.selectedSector) {
          if (row.detail) detailSet.add(row.detail);
        }
      });
      const details = [...detailSet].sort((a, b) => a.localeCompare(b, "ko"));

      renderSelectOptions(sectorSelect, sectors, state.selectedSector, "전체 섹터");
      renderSelectOptions(detailSelect, details, state.selectedDetail, "전체 세부");
    }

    function renderSelectOptions(select, values, current, allLabel) {
      const previous = current;
      select.innerHTML = "";
      const allOpt = document.createElement("option");
      allOpt.value = "all";
      allOpt.textContent = allLabel;
      select.appendChild(allOpt);
      values.forEach((value) => {
        const opt = document.createElement("option");
        opt.value = value;
        opt.textContent = value;
        select.appendChild(opt);
      });
      select.value = values.includes(previous) || previous === "all" ? previous : "all";
    }

    function renderFilterChips() {
      const wrap = document.getElementById("filterChips");
      wrap.innerHTML = "";
      const chips = [];
      if (state.selectedSector !== "all") chips.push({ label: `섹터: ${state.selectedSector}`, clear: () => { state.selectedSector = "all"; state.selectedDetail = "all"; render(); } });
      if (state.selectedDetail !== "all") chips.push({ label: `세부: ${state.selectedDetail}`, clear: () => { state.selectedDetail = "all"; render(); } });
      if (state.status !== "all") chips.push({ label: `상태: ${state.status}`, clear: () => { state.status = "all"; document.getElementById("statusFilter").value = "all"; render(); } });
      if (state.search) chips.push({ label: `검색: ${state.search}`, clear: () => { state.search = ""; document.getElementById("searchInput").value = ""; render(); } });
      chips.forEach((chip) => {
        const el = document.createElement("span");
        el.className = "chip";
        const label = document.createElement("span");
        label.textContent = chip.label;
        const btn = document.createElement("button");
        btn.type = "button";
        btn.textContent = "×";
        btn.addEventListener("click", chip.clear);
        el.appendChild(label);
        el.appendChild(btn);
        wrap.appendChild(el);
      });
    }

    function computeRows() {
      state.computed = state.rows.map((row, index) => {
        const name = normalizeName(row.stockName);
        const holdings = byName.get(name) || [];
        const price = firstNumber(holdings.map((item) => item.price));
        const targetPct = normalizePct(row.targetPct);
        const fundValues = {};

        FUND_KEYS.forEach((key) => {
          const fund = fundByKey.get(key);
          const fundHoldings = holdings.filter((item) => item.fundKey === key);
          const valueEok = sum(fundHoldings.map((item) => item.value)) / 100000000;
          const weight = fund && fund.stockTargetEok ? valueEok / fund.stockTargetEok : 0;
          const gapQty = price && fund && fund.stockTargetEok
            ? ((targetPct - weight) * fund.stockTargetEok * 100000000) / price
            : null;
          fundValues[key] = {
            valueEok,
            weight,
            gapQty,
            returnPct: firstNumber(fundHoldings.map((item) => item.returnPct)),
          };
        });

        let status = "ok";
        let statusLabel = "OK";
        if (name && holdings.length === 0) {
          status = "missing";
          statusLabel = "원천누락";
        } else if (name && !price) {
          status = "price";
          statusLabel = "현재가누락";
        }

        return {
          ...row,
          index,
          normalizedName: name,
          targetPct,
          holdings,
          price,
          fundValues,
          status,
          statusLabel,
        };
      });
    }

    function getSectorStats(options = {}) {
      const { excludeOther = true } = options;
      const grouped = new Map();
      state.computed.forEach((row) => {
        const sector = row.sector || "미분류";
        if (excludeOther && SECTOR_EXCLUDE.has(sector)) return;
        if (!grouped.has(sector)) {
          grouped.set(sector, {
            sector,
            count: 0,
            targetPct: 0,
            rows: [],
            fundWeights: Object.fromEntries(FUND_KEYS.map((key) => [key, 0])),
          });
        }
        const bucket = grouped.get(sector);
        bucket.count += row.normalizedName ? 1 : 0;
        bucket.targetPct += row.targetPct || 0;
        bucket.rows.push(row);
        FUND_KEYS.forEach((key) => {
          bucket.fundWeights[key] += row.fundValues[key].weight || 0;
        });
      });
      return [...grouped.values()].sort((a, b) => b.targetPct - a.targetPct || a.sector.localeCompare(b.sector, "ko"));
    }

    function renderMainTable() {
      const body = document.getElementById("mpBody");
      body.innerHTML = "";
      const fragment = document.createDocumentFragment();
      const selectedFund = fundByKey.get(state.selectedFund) || DATA.funds[0];

      const titleParts = ["MP 입력"];
      if (state.selectedSector !== "all") titleParts.push(state.selectedSector);
      if (state.selectedDetail !== "all") titleParts.push(state.selectedDetail);
      document.getElementById("mpTitle").textContent = titleParts.join(" › ");
      document.querySelectorAll("#mpFundTabs .fund-tab").forEach((tab) => {
        tab.classList.toggle("active", tab.dataset.fund === selectedFund.key);
      });
      document.getElementById("fundValueHead").textContent = "평가";
      document.getElementById("fundWeightHead").textContent = "비중";
      document.getElementById("fundDiffHead").textContent = "차이";
      document.getElementById("fundQtyHead").textContent = "조정수량";
      document.getElementById("fundReturnHead").textContent = "수익률";

      state.computed.forEach((row) => {
        const tr = document.createElement("tr");
        tr.dataset.index = row.index;
        tr.dataset.status = row.status;
        tr.dataset.sector = row.sector || "미분류";
        tr.dataset.detail = row.detail || "";
        tr.dataset.search = `${row.sector} ${row.detail} ${row.stockName}`.toLowerCase();

        const fund = row.fundValues[selectedFund.key];
        const diff = (row.targetPct || 0) - (fund.weight || 0);
        const absDiff = Math.abs(diff);
        if (row.normalizedName && absDiff > 0.01) tr.classList.add("row-bad");

        const sectorTd = inputCell(row.index, "sector", row.sector, "editable", "sectorList");
        sectorTd.classList.add("col-sector");
        tr.appendChild(sectorTd);
        const detailTd = inputCell(row.index, "detail", row.detail, "editable", "detailList");
        detailTd.classList.add("col-detail-input");
        tr.appendChild(detailTd);
        const nameTd = inputCell(row.index, "stockName", row.stockName, "editable name-input", "stockList");
        nameTd.classList.add("col-stockname");
        const nameInput = nameTd.querySelector("input");
        if (nameInput && row.stockName) nameInput.title = row.stockName;
        tr.appendChild(nameTd);
        tr.appendChild(targetCell(row.index, row.targetPct));
        tr.appendChild(numberCell(row.price, "price"));
        tr.appendChild(statusCell(row.status, row.statusLabel));

        const valueCell = numberCell(fund.valueEok, "eok");
        valueCell.classList.add("col-detail");
        tr.appendChild(valueCell);
        tr.appendChild(numberCell(fund.weight, "pct"));
        const diffTd = diffCell(diff);
        tr.appendChild(diffTd);
        tr.appendChild(gapCell(fund.gapQty, "qty"));
        const retCell = numberCell(fund.returnPct, "ret");
        retCell.classList.add("col-detail");
        tr.appendChild(retCell);

        const deleteTd = document.createElement("td");
        deleteTd.className = "center";
        const button = document.createElement("button");
        button.type = "button";
        button.className = "delete-row danger";
        button.textContent = "X";
        button.title = "행 삭제";
        button.addEventListener("click", () => {
          if (row.stockName && !confirm(`'${row.stockName}' 행을 삭제할까요?`)) return;
          state.rows.splice(row.index, 1);
          render();
        });
        deleteTd.appendChild(button);
        tr.appendChild(deleteTd);
        fragment.appendChild(tr);
      });

      body.appendChild(fragment);
      body.querySelectorAll("input[data-field]").forEach((input) => {
        input.addEventListener("input", storeRowFromInput);
        input.addEventListener("change", commitRowInput);
      });
    }

    function gapCell(value, type) {
      const td = numberCell(value, type);
      if (value === null || value === undefined || Number.isNaN(value)) return td;
      const num = Number(value);
      if (Math.abs(num) < (type === "pct" ? 0.0005 : 1)) td.classList.add("gap-zero");
      else if (num > 0) td.classList.add("gap-pos");
      else td.classList.add("gap-neg");
      return td;
    }

    function diffCell(value) {
      const td = document.createElement("td");
      td.className = "number col-diff";
      if (value === null || value === undefined || Number.isNaN(value)) return td;
      const num = Number(value);
      const cls = Math.abs(num) < 0.0005 ? "gap-zero" : (num > 0 ? "gap-pos" : "gap-neg");
      const wrap = document.createElement("div");
      wrap.className = "diff-cell";
      const text = document.createElement("div");
      text.className = `diff-text ${cls}`;
      text.textContent = formatPct(num);
      wrap.appendChild(text);
      const MAX = 0.03;
      const pct = Math.min(Math.abs(num) / MAX, 1) * 50;
      const bar = document.createElement("div");
      bar.className = "diff-bar";
      if (Math.abs(num) >= 0.0005) {
        const fill = document.createElement("div");
        fill.className = `diff-bar-fill ${num > 0 ? "pos" : "neg"}`;
        fill.style.width = `${pct}%`;
        bar.appendChild(fill);
      }
      wrap.appendChild(bar);
      td.appendChild(wrap);
      return td;
    }

    function inputCell(index, field, value, className, listId) {
      const td = document.createElement("td");
      const input = document.createElement("input");
      input.value = value || "";
      input.className = className;
      input.dataset.index = index;
      input.dataset.field = field;
      if (listId) input.setAttribute("list", listId);
      td.appendChild(input);
      return td;
    }

    function targetCell(index, value) {
      const td = document.createElement("td");
      const input = document.createElement("input");
      input.value = formatPctInput(value);
      input.className = "editable target-input";
      input.dataset.index = index;
      input.dataset.field = "targetPct";
      td.appendChild(input);
      return td;
    }

    function statusCell(status, label) {
      const td = document.createElement("td");
      td.className = "center";
      if (status === "ok") return td;
      const badge = document.createElement("span");
      badge.className = `badge ${status === "price" ? "warn" : "bad"}`;
      badge.textContent = label;
      td.appendChild(badge);
      return td;
    }

    function numberCell(value, type) {
      const td = document.createElement("td");
      td.className = "number";
      if (value === null || value === undefined || Number.isNaN(value)) {
        td.textContent = "";
      } else if (type === "pct") {
        td.textContent = formatPct(value);
      } else if (type === "ret") {
        td.textContent = formatNumber(value, 2) + "%";
      } else if (type === "qty") {
        td.textContent = formatNumber(value, 0);
      } else if (type === "eok") {
        td.textContent = formatNumber(value, 2);
      } else {
        td.textContent = formatNumber(value, 0);
      }
      return td;
    }

    function storeRowFromInput(event) {
      const input = event.currentTarget;
      const index = Number(input.dataset.index);
      const field = input.dataset.field;
      if (!state.rows[index]) return;
      if (field === "targetPct") {
        state.rows[index][field] = parsePct(input.value);
      } else {
        state.rows[index][field] = input.value;
      }
    }

    function commitRowInput(event) {
      storeRowFromInput(event);
      render();
    }

    function renderSectorTable() {
      const grouped = getSectorStats();
      const fund = fundByKey.get(state.sectorFund) || DATA.funds[0];
      const body = document.querySelector("#sectorTable tbody");
      body.innerHTML = "";
      document.getElementById("sectorFundHead").textContent = fund.label;
      document.getElementById("sectorSummaryNote").textContent =
        state.selectedSector === "all" ? "" : `${state.selectedSector} 강조`;

      document.querySelectorAll("#sectorFundTabs .fund-tab").forEach((tab) => {
        tab.classList.toggle("active", tab.dataset.fund === fund.key);
      });

      const sorted = [...grouped].sort((a, b) => (b.fundWeights[fund.key] || 0) - (a.fundWeights[fund.key] || 0));

      sorted.forEach((data) => {
        const tr = document.createElement("tr");
        if (state.selectedSector === data.sector) tr.style.background = "#e8f4f8";
        const sectorTd = document.createElement("td");
        sectorTd.textContent = data.sector;
        sectorTd.style.cursor = "pointer";
        sectorTd.addEventListener("click", () => {
          state.selectedSector = state.selectedSector === data.sector ? "all" : data.sector;
          state.selectedDetail = "all";
          render();
        });
        tr.appendChild(sectorTd);
        tr.appendChild(numberCell(data.targetPct, "pct"));
        tr.appendChild(numberCell(data.fundWeights[fund.key], "pct"));
        const diff = (data.fundWeights[fund.key] || 0) - data.targetPct;
        tr.appendChild(gapCell(-diff, "pct"));
        body.appendChild(tr);
      });

      renderSectorPie(sorted, fund);
    }

    function renderSectorPie(grouped, fund) {
      const pie = document.getElementById("sectorPie");
      const legend = document.getElementById("sectorLegend");
      const caption = document.getElementById("sectorPieCaption");
      const centerLabel = document.getElementById("pieCenterLabel");
      const centerValue = document.getElementById("pieCenterValue");
      const labelsBox = document.getElementById("pieLabels");
      const slices = grouped
        .map((item, index) => ({
          sector: item.sector,
          value: Math.max(item.fundWeights[fund.key] || 0, 0),
          target: item.targetPct,
          color: CHART_COLORS[index % CHART_COLORS.length],
        }))
        .filter((item) => item.value > 0);
      const total = sum(slices.map((item) => item.value));

      centerLabel.textContent = fund.label;
      centerValue.textContent = formatPct(total);
      labelsBox.innerHTML = "";

      if (!total) {
        pie.style.background = "#d7dde6";
        caption.textContent = `${fund.label} 섹터 비중 없음`;
        legend.innerHTML = "";
        return;
      }

      let cursor = 0;
      const segments = slices.map((slice) => {
        const start = cursor;
        const end = cursor + (slice.value / total) * 100;
        cursor = end;
        return `${slice.color} ${start.toFixed(4)}% ${end.toFixed(4)}%`;
      });
      pie.style.background = `conic-gradient(${segments.join(", ")})`;
      caption.textContent = `${fund.label} 주식가능액 대비 주식비중`;

      // Place sector name labels at the geometric center of each slice's visible ring area.
      // Donut visible ring is between ~25% (hole) and 50% (outer) → midpoint 37.5%.
      const LABEL_RADIUS = 37.5;
      const MIN_SHARE = 0.04;
      const MIN_DEG_PER_CHAR = 6.5;  // estimate of arc deg required per Korean char at this radius/font
      cursor = 0;
      slices.forEach((slice) => {
        const share = slice.value / total;
        const startAngle = cursor * 360;
        const endAngle = (cursor + share) * 360;
        cursor += share;
        const arcDeg = share * 360;
        if (share < MIN_SHARE) return;
        if (slice.sector.length * MIN_DEG_PER_CHAR > arcDeg) return;
        const midAngle = (startAngle + endAngle) / 2;
        const rad = (midAngle - 90) * Math.PI / 180;
        const x = 50 + Math.cos(rad) * LABEL_RADIUS;
        const y = 50 + Math.sin(rad) * LABEL_RADIUS;
        const label = document.createElement("div");
        label.className = "pie-label";
        label.textContent = slice.sector;
        label.style.left = x + "%";
        label.style.top = y + "%";
        labelsBox.appendChild(label);
      });

      legend.innerHTML = "";
      slices.forEach((slice) => {
        const item = document.createElement("div");
        item.className = "legend-item" + (state.selectedSector === slice.sector ? " active" : "");
        item.title = `${slice.sector} 필터`;
        item.innerHTML = '<span class="legend-swatch"></span><span class="legend-label"></span><span class="muted"></span>';
        item.querySelector(".legend-swatch").style.background = slice.color;
        item.querySelector(".legend-label").textContent = slice.sector;
        const gap = slice.value - slice.target;
        const gapStr = gap >= 0 ? `+${formatPct(gap)}` : formatPct(gap);
        item.querySelector(".muted").innerHTML =
          `<strong style="color:var(--text);">${formatPct(slice.value)}</strong> <span class="${gap > 0.0005 ? 'gap-neg' : gap < -0.0005 ? 'gap-pos' : ''}">(${gapStr})</span>`;
        item.addEventListener("click", () => {
          state.selectedSector = state.selectedSector === slice.sector ? "all" : slice.sector;
          state.selectedDetail = "all";
          render();
        });
        legend.appendChild(item);
      });
    }

    function renderValidation() {
      const selectedNames = new Set(state.computed.map((row) => row.normalizedName).filter(Boolean));
      const sourceNames = [...byName.keys()];
      const missing = state.computed.filter((row) => row.normalizedName && row.status === "missing");
      const priceMissing = state.computed.filter((row) => row.normalizedName && row.status === "price");
      const unlisted = sourceNames.filter((name) => !selectedNames.has(name));

      renderList("missingList", missing, (row) => [row.stockName, row.sector || ""]);
      renderList("priceList", priceMissing, (row) => [row.stockName, row.sector || ""]);
      renderList("unlistedList", unlisted.slice(0, 80), (name) => {
        const holdings = byName.get(name) || [];
        const funds = [...new Set(holdings.map((item) => item.fundLabel))].join(", ");
        return [name, funds];
      });
    }

    function renderList(id, items, mapper) {
      const list = document.getElementById(id);
      list.innerHTML = "";
      if (!items.length) {
        const li = document.createElement("li");
        li.innerHTML = '<span class="muted">없음</span><span></span>';
        list.appendChild(li);
        return;
      }
      items.forEach((item) => {
        const [left, right] = mapper(item);
        const li = document.createElement("li");
        const l = document.createElement("span");
        const r = document.createElement("span");
        l.textContent = left;
        r.textContent = right;
        r.className = "muted";
        li.appendChild(l);
        li.appendChild(r);
        list.appendChild(li);
      });
    }

    function renderMetrics() {
      const selectedNames = new Set(state.computed.map((row) => row.normalizedName).filter(Boolean));
      const sourceNames = [...byName.keys()];
      const missing = state.computed.filter((row) => row.normalizedName && row.status === "missing").length;
      const priceMissing = state.computed.filter((row) => row.normalizedName && row.status === "price").length;
      const unlisted = sourceNames.filter((name) => !selectedNames.has(name)).length;
      const target = sum(state.computed.map((row) => row.targetPct));

      const a3Count = DATA.mpStockCount;
      document.getElementById("metricRows").textContent = (a3Count !== null && a3Count !== undefined)
        ? formatNumber(a3Count, 0)
        : formatNumber(state.computed.filter((row) => row.normalizedName).length, 0);
      document.getElementById("metricTarget").textContent = formatPct(target);
      document.getElementById("metricMissing").textContent = formatNumber(missing, 0);
      document.getElementById("metricPriceMissing").textContent = formatNumber(priceMissing, 0);
      document.getElementById("metricSourceNames").textContent = formatNumber(sourceNames.length, 0);
      document.getElementById("metricUnlisted").textContent = formatNumber(unlisted, 0);

      toggleMetricState("metricMissingBox", missing, "bad");
      toggleMetricState("metricPriceBox", priceMissing, "warn");
      toggleMetricState("metricUnlistedBox", unlisted, "warn");
    }

    function toggleMetricState(id, count, className) {
      const node = document.getElementById(id);
      node.classList.remove("bad", "warn");
      if (count > 0) node.classList.add(className);
    }

    function applyFilters() {
      let visible = 0;
      let targetSum = 0;
      let weightSum = 0;
      let prevSector = null;
      const fundKey = state.selectedFund;
      document.querySelectorAll("#mpBody tr").forEach((tr) => {
        const statusOk = state.status === "all" || tr.dataset.status === state.status;
        const sectorOk = state.selectedSector === "all" || tr.dataset.sector === state.selectedSector;
        const detailOk = state.selectedDetail === "all" || tr.dataset.detail === state.selectedDetail;
        const searchOk = !state.search || tr.dataset.search.includes(state.search);
        const show = statusOk && sectorOk && detailOk && searchOk;
        tr.classList.toggle("hidden", !show);
        if (show) {
          visible += 1;
          const sector = tr.dataset.sector;
          tr.classList.toggle("same-sector", !!prevSector && sector === prevSector);
          prevSector = sector;
          const row = state.computed[Number(tr.dataset.index)];
          if (row && row.normalizedName) {
            targetSum += row.targetPct || 0;
            weightSum += row.fundValues[fundKey].weight || 0;
          }
        } else {
          tr.classList.remove("same-sector");
        }
      });
      document.getElementById("visibleCount").textContent =
        `보이는 행 ${visible} / ${state.rows.length} · 목표합 ${formatPct(targetSum)} · 현재합 ${formatPct(weightSum)}`;
    }

    function exportCsv() {
      const headers = [
        "섹터",
        "세부",
        "종목명",
        "목표비중",
        "현재가",
        "상태",
        "인컴2 평가",
        "인컴2 비중",
        "인컴2 차이",
        "인컴2 조정수량",
        "블루1 평가",
        "블루1 비중",
        "블루1 차이",
        "블루1 조정수량",
        "블루3 평가",
        "블루3 비중",
        "블루3 차이",
        "블루3 조정수량",
      ];
      const lines = [headers];
      state.computed.forEach((row) => {
        const diff = (key) => (row.targetPct || 0) - (row.fundValues[key].weight || 0);
        lines.push([
          row.sector,
          row.detail,
          row.stockName,
          row.targetPct,
          row.price || "",
          row.statusLabel,
          row.fundValues.income2.valueEok,
          row.fundValues.income2.weight,
          diff("income2"),
          row.fundValues.income2.gapQty,
          row.fundValues.blue1.valueEok,
          row.fundValues.blue1.weight,
          diff("blue1"),
          row.fundValues.blue1.gapQty,
          row.fundValues.blue3.valueEok,
          row.fundValues.blue3.weight,
          diff("blue3"),
          row.fundValues.blue3.gapQty,
        ]);
      });
      const csv = lines.map((row) => row.map(csvEscape).join(",")).join("\n");
      const blob = new Blob(["\ufeff" + csv], { type: "text/csv;charset=utf-8" });
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      const stamp = new Date().toISOString().slice(0, 10);
      link.download = `mp_dashboard_${stamp}.csv`;
      link.click();
      URL.revokeObjectURL(url);
    }

    function normalizeName(value) {
      return (value || "").toString().trim();
    }

    function normalizePct(value) {
      if (value === null || value === undefined || value === "") return 0;
      const number = Number(value);
      if (!Number.isFinite(number)) return 0;
      return number > 1 ? number / 100 : number;
    }

    function parsePct(value) {
      const text = (value || "").toString().trim().replace("%", "");
      if (!text) return 0;
      const number = Number(text);
      if (!Number.isFinite(number)) return 0;
      return number > 1 ? number / 100 : number;
    }

    function formatPctInput(value) {
      return formatNumber(normalizePct(value) * 100, 1) + "%";
    }

    function formatPct(value) {
      return formatNumber((value || 0) * 100, 2) + "%";
    }

    function formatNumber(value, digits) {
      if (value === null || value === undefined || Number.isNaN(value)) return "";
      return Number(value).toLocaleString("ko-KR", {
        minimumFractionDigits: digits,
        maximumFractionDigits: digits,
      });
    }

    function sum(values) {
      return values.reduce((acc, value) => acc + (Number(value) || 0), 0);
    }

    function firstNumber(values) {
      for (const value of values) {
        const number = Number(value);
        if (Number.isFinite(number) && number !== 0) return number;
      }
      return null;
    }

    function csvEscape(value) {
      const text = value === null || value === undefined ? "" : String(value);
      if (/[",\n]/.test(text)) return `"${text.replaceAll('"', '""')}"`;
      return text;
    }
  </script>
</body>
</html>
"""


def write_html(data, output_path):
    html = HTML_TEMPLATE.replace(
        "__DATA_JSON__",
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
    )
    output_path.write_text(html, encoding="utf-8")


def main():
    # 경로: 이 스크립트는 igis/dashboard/generate_mp_dashboard.py 에 위치
    base_dir = Path(__file__).resolve().parent          # igis/dashboard

    # 입력 워크북 결정: 인자가 있으면 그 파일, 없으면 최신 hf2_mp__*.xlsx 자동 선택
    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1])
        if not workbook_path.is_absolute():
            workbook_path = base_dir / workbook_path
    else:
        candidates = sorted(
            base_dir.glob("hf2_mp__*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        candidates = [p for p in candidates if not p.name.startswith("~$")]
        if not candidates:
            print("hf2_mp__*.xlsx 파일을 찾을 수 없습니다. (dashboard 폴더에 워크북을 두거나 파일명을 인자로 주세요)")
            sys.exit(1)
        workbook_path = candidates[0]
        print(f"[자동 선택] {workbook_path.name}")

    if not workbook_path.exists():
        print(f"워크북을 찾을 수 없습니다: {workbook_path}")
        sys.exit(1)

    # 출력: dashboard 폴더 안에 mp_dashboard.html 생성
    if len(sys.argv) > 2:
        output_path = Path(sys.argv[2])
        if not output_path.is_absolute():
            output_path = base_dir / output_path
    else:
        output_path = base_dir / "mp_dashboard.html"

    data = build_data(workbook_path)
    write_html(data, output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()
