#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Post IPO Monitor (v2 — 수급파일 다중일자 비교)
=============================================
2년 이내 신규상장 종목의 수급/거래량/RSI/변동성 모니터링

[v2 변경점]
    - 비교 대상을 "외부 전일 결과 파일"이 아니라 수급.xlsx 안의 다중 일자에서 직접 잡음.
    - 수급.xlsx에서 값이 채워진 가장 최근 날짜 = 기준일(예: 오늘 23일 실행 → 22일),
      그 직전 채워진 날짜 = 비교일(그것의 어제, 예: 19일).
    - 두 날짜 각각에 대해 가격/RSI/거래량 + 수급으로 전체 스코어·등급을 계산.
    - 같은 종목 집합·같은 코드명에서 양일치가 나오므로 K~O 비교열이 결측 없이 채워짐.

데이터 소스:
    - 최초상장일.xlsx: IPO 종목 리스트
    - 수급.xlsx: 기관/외국인 순매수 데이터 (여러 일자 행으로 보관, 대략 5일치)
    - FinanceDataReader: 가격, 거래량, RSI, 변동성

실행:
    python run.py
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
import re

from post_ipo_daily import Config
from post_ipo_daily.utils import print_progress_bar, get_previous_business_day, get_today_business_day


class IPOMonitor:
    """IPO 종목 모니터링 클래스"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self.config.ensure_directories()

        # IPO / 수급
        self._ipo_df = None
        self._supply_df = None            # 기준일 수급
        self._supply_prev_df = None       # 비교일 수급
        self._supply_date = None          # 기준일 (수급 최신 채워진 날짜)
        self._supply_prev_date = None     # 비교일 (그 직전 채워진 날짜)

        # 가격
        self._hist_cache = {}             # {코드: 지표 계산된 일별 히스토리}
        self._mkt_caps = {}               # {코드: 시가총액(억)}
        self._bloomberg_df = None         # 기준일 가격지표
        self._bloomberg_prev_df = None    # 비교일 가격지표

        # 결과
        self._result_df = None            # 기준일 최종 스코어
        self._result_prev_df = None       # 비교일 최종 스코어 (전체_전일자 시트로 사용)

        self._ref_date = None             # 기준일
        self._rsi_cache = {}              # {코드: RSI(14) 시계열}

    # =========================================================================
    # Step 1: IPO 종목 로드
    # =========================================================================
    def load_ipo_universe(self, source: str = 'A') -> pd.DataFrame:
        print("\n" + "=" * 60)
        print("[Step 1] IPO Universe 로드")
        print("=" * 60)
        if source == 'B':
            return self._load_ipo_universe_b()
        return self._load_ipo_universe_a()

    def _load_ipo_universe_a(self) -> pd.DataFrame:
        """버전 A: 최초상장일.xlsx — 2년 이내 상장종목 자동 필터"""
        df = pd.read_excel(self.config.IPO_FILE, skiprows=5)
        df.columns = ['코드', '코드명', '최초상장일', '상장일']

        df['최초상장일_dt'] = pd.to_datetime(
            df['최초상장일'].fillna(0).astype(int).astype(str),
            format='%Y%m%d', errors='coerce'
        )

        from datetime import timedelta
        cutoff = datetime.now() - timedelta(days=self.config.IPO_DAYS_LIMIT)
        recent = df[df['최초상장일_dt'] >= cutoff].copy()
        recent = recent.dropna(subset=['코드'])

        def is_regular_code(code):
            return bool(re.match(r'^A\d{6}$', str(code)))

        regular = recent[recent['코드'].apply(is_regular_code)].copy()

        def is_regular_stock(name):
            if not isinstance(name, str):
                return True
            name_upper = name.upper()
            for kw in self.config.EXCLUDE_KEYWORDS:
                if kw.upper() in name_upper:
                    return False
            return True

        stocks = regular[regular['코드명'].apply(is_regular_stock)].copy()
        stocks = stocks[['코드', '코드명', '최초상장일', '최초상장일_dt']].copy()
        stocks = stocks.drop_duplicates(subset=['코드'], keep='first')
        stocks['days_since_ipo'] = (datetime.now() - stocks['최초상장일_dt']).dt.days
        stocks['ticker_ks'] = stocks['코드'].str[1:] + ' KS Equity'
        stocks['ticker_kq'] = stocks['코드'].str[1:] + ' KQ Equity'
        stocks = stocks.sort_values('최초상장일_dt', ascending=False).reset_index(drop=True)

        self._ipo_df = stocks
        print(f"[버전 A] 기준일: {cutoff.strftime('%Y-%m-%d')} 이후 상장")
        print(f"대상 종목: {len(stocks)}개")
        return stocks

    def _load_ipo_universe_b(self) -> pd.DataFrame:
        """버전 B: __post ipo univ.xlsx — Symbol 열 직접 사용"""
        df = pd.read_excel(self.config.UNIV_FILE, header=1)
        if 'Symbol' not in df.columns:
            print("오류: __post ipo univ.xlsx에 'Symbol' 컬럼이 없습니다.")
            return pd.DataFrame()

        df = df[['Symbol', 'Name']].copy()
        df.columns = ['코드', '코드명']
        df = df.dropna(subset=['코드'])

        def is_regular_code(code):
            return bool(re.match(r'^A\d{6}$', str(code)))

        stocks = df[df['코드'].apply(is_regular_code)].copy()
        stocks['최초상장일'] = None
        stocks['최초상장일_dt'] = pd.NaT
        stocks['days_since_ipo'] = None
        stocks['ticker_ks'] = stocks['코드'].str[1:] + ' KS Equity'
        stocks['ticker_kq'] = stocks['코드'].str[1:] + ' KQ Equity'
        stocks = stocks.reset_index(drop=True)

        self._ipo_df = stocks
        print(f"[버전 B] __post ipo univ.xlsx 로드")
        print(f"대상 종목: {len(stocks)}개")
        return stocks

    # =========================================================================
    # Step 2: 수급 데이터 로드 (다중 일자)
    # =========================================================================
    def load_supply_data(self) -> pd.DataFrame:
        """수급.xlsx에서 기관/외국인 순매수 데이터 로드.
        값이 채워진 가장 최근 날짜 = 기준일, 그 직전 날짜 = 비교일."""
        print("\n" + "=" * 60)
        print("[Step 2] 수급 데이터 로드 (다중 일자)")
        print("=" * 60)

        raw = pd.read_excel(self.config.SUPPLY_FILE, header=None)

        # 메타데이터 (행 8/9/11 → 코드/코드명/아이템코드)
        codes = raw.iloc[8, :].values
        names = raw.iloc[9, :].values
        item_codes = raw.iloc[11, :].values

        # 데이터 (행 14부터)
        data = raw.iloc[14:, :].copy()
        data.columns = range(len(data.columns))
        data[0] = pd.to_datetime(data[0], errors='coerce')
        data = data.dropna(subset=[0])
        data = data.rename(columns={0: '날짜'})

        # 값이 전부 비어있는 행 제거 (06-23 같은 빈 placeholder 방지)
        value_cols = [c for c in data.columns if c != '날짜']
        data = data.dropna(subset=value_cols, how='all')

        if data.empty:
            print("오류: 수급.xlsx에 값이 채워진 날짜가 없습니다.")
            self._supply_df = pd.DataFrame()
            return self._supply_df

        # 값이 채워진 날짜들 (내림차순)
        filled_dates = sorted(data['날짜'].dropna().unique(), reverse=True)
        self._supply_date = pd.to_datetime(filled_dates[0])
        self._supply_prev_date = pd.to_datetime(filled_dates[1]) if len(filled_dates) >= 2 else None

        ref_str = self._supply_date.strftime('%Y-%m-%d')
        prev_str = self._supply_prev_date.strftime('%Y-%m-%d') if self._supply_prev_date is not None else '없음'
        print(f"기준일(최신 채워진 날짜): {ref_str}")
        print(f"비교일(그 직전 날짜)    : {prev_str}")
        if self._supply_prev_date is None:
            print("  ⚠️  비교일이 없어 K~O 비교열은 생성되지 않습니다(수급.xlsx에 이틀치 이상 필요).")

        # 기준일/비교일 각각 수급 테이블 구성
        self._supply_df = self._build_supply_for_date(codes, names, item_codes, data, self._supply_date)
        print(f"기준일 수급: {len(self._supply_df)}개 종목")

        if self._supply_prev_date is not None:
            self._supply_prev_df = self._build_supply_for_date(codes, names, item_codes, data, self._supply_prev_date)
            print(f"비교일 수급: {len(self._supply_prev_df)}개 종목")

        return self._supply_df

    def _build_supply_for_date(self, codes, names, item_codes, data, target_date) -> pd.DataFrame:
        """특정 날짜의 수급 행을 종목별 레코드로 정리"""
        sub = data[data['날짜'] == target_date]
        if sub.empty:
            return pd.DataFrame()
        latest = sub.iloc[0]

        records = []
        i = 1
        while i < len(codes):
            code = codes[i]
            if pd.isna(code) or not str(code).startswith('A'):
                i += 1
                continue

            record = {'코드': code, '코드명': names[i]}
            j = i
            while j < len(codes) and codes[j] == code:
                item_code = item_codes[j]
                value = latest[j] if j < len(latest) else None
                if item_code == 'CI20003020':
                    record['기관_일간'] = value
                elif item_code == 'CI20003021':
                    record['기관_5일'] = value
                elif item_code == 'CI20003022':
                    record['기관_20일'] = value
                elif item_code == 'CI20113020':
                    record['외국인_일간'] = value
                elif item_code == 'CI20113021':
                    record['외국인_5일'] = value
                elif item_code == 'CI20113022':
                    record['외국인_20일'] = value
                j += 1

            records.append(record)
            i = j

        supply_df = pd.DataFrame(records).drop_duplicates(subset=['코드'], keep='first')

        numeric_cols = ['기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']
        for col in numeric_cols:
            if col in supply_df.columns:
                supply_df[col] = pd.to_numeric(supply_df[col], errors='coerce')

        return supply_df

    # =========================================================================
    # Step 3: 가격/거래량 — 히스토리 1회 수집 후 기준일·비교일 두 시점 추출
    # =========================================================================
    @staticmethod
    def _wilder_rsi(close: pd.Series, length: int = 14) -> pd.Series:
        delta = close.diff()
        gain = delta.clip(lower=0)
        loss = -delta.clip(upper=0)
        avg_gain = gain.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
        avg_loss = loss.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
        rs = avg_gain / avg_loss
        return (100 - 100 / (1 + rs)).round(2)

    @staticmethod
    def _load_market_caps() -> dict:
        import FinanceDataReader as fdr
        try:
            listing = fdr.StockListing('KRX')
        except Exception as e:
            print(f"  시가총액 로드 실패(StockListing): {e}")
            return {}
        code_col = next((c for c in ['Code', 'code', 'Symbol'] if c in listing.columns), None)
        cap_col = next((c for c in ['Marcap', 'MarCap', 'marcap', '시가총액'] if c in listing.columns), None)
        if code_col is None or cap_col is None:
            print(f"  시가총액 컬럼 탐색 실패 (컬럼: {list(listing.columns)[:10]})")
            return {}
        caps = {}
        for _, r in listing[[code_col, cap_col]].dropna().iterrows():
            try:
                caps[str(r[code_col]).zfill(6)] = float(r[cap_col]) / 100_000_000
            except (ValueError, TypeError):
                continue
        return caps

    def fetch_bloomberg_data(self, use_prev_day: bool = True) -> pd.DataFrame:
        """가격 히스토리를 1회 수집하여 지표 계산 후, 기준일·비교일 두 시점의 가격지표 생성.
        (use_prev_day 인자는 호환용. 기준일/비교일은 수급.xlsx의 채워진 날짜로 결정됨)"""
        import FinanceDataReader as fdr

        print("\n" + "=" * 60)
        print("[Step 3] 가격/거래량 수집 (히스토리 1회 → 기준일·비교일 추출)")
        print("=" * 60)

        if self._ipo_df is None:
            print("오류: IPO Universe를 먼저 로드하세요.")
            return pd.DataFrame()
        if self._supply_date is None:
            print("오류: 수급 데이터를 먼저 로드하세요.")
            return pd.DataFrame()

        ref_date = self._supply_date
        prev_date = self._supply_prev_date
        self._ref_date = ref_date
        ref_date_str = ref_date.strftime('%Y-%m-%d')
        prev_date_str = prev_date.strftime('%Y-%m-%d') if prev_date is not None else None

        print(f"기준일: {ref_date_str}" + (f"  |  비교일: {prev_date_str}" if prev_date_str else "  |  비교일: 없음"))

        stock_codes = self._ipo_df['코드'].str[1:].tolist()
        print(f"대상: {len(stock_codes)}개 종목")

        self._mkt_caps = self._load_market_caps()
        print(f"시가총액 로드: {len(self._mkt_caps)}개\n")

        # 60일 이동평균까지 두 시점(기준일/비교일) 모두 확보하려면 넉넉히 120일
        start_date = (ref_date - pd.Timedelta(days=120)).strftime('%Y-%m-%d')
        end_date = ref_date_str  # 기준일까지만 사용(이후 데이터는 비교에 불필요)

        self._hist_cache = {}
        self._rsi_cache = {}
        batch_size = self.config.BATCH_SIZE
        total_batches = (len(stock_codes) + batch_size - 1) // batch_size

        for i in range(0, len(stock_codes), batch_size):
            batch_codes = stock_codes[i:i + batch_size]
            batch_num = i // batch_size + 1
            print_progress_bar(batch_num, total_batches, prefix='데이터 수집')

            for code in batch_codes:
                try:
                    hist = None
                    for symbol in (f'KRX:{code}', code):
                        try:
                            hist = fdr.DataReader(symbol, start_date, end_date)
                            if hist is not None and not hist.empty:
                                break
                        except Exception:
                            hist = None

                    if hist is None or hist.empty or len(hist) < 5:
                        continue

                    hist.columns = hist.columns.str.lower()
                    if 'close' not in hist.columns or 'volume' not in hist.columns:
                        continue

                    close = hist['close']
                    volume = hist['volume']

                    hist['rsi_14d'] = self._wilder_rsi(close, 14)
                    hist['mov_avg_10d'] = close.rolling(10).mean()
                    hist['mov_avg_20d'] = close.rolling(20).mean()
                    hist['mov_avg_60d'] = close.rolling(60).mean()

                    sd20 = close.rolling(20).std(ddof=0)
                    hist['bollinger_band_width'] = (sd20 * 4).round(2)
                    hist['volatility_30d'] = (close.pct_change().rolling(30).std() * 100).round(2)
                    hist['volume_avg_20d'] = volume.rolling(20).mean()
                    hist['rvol_20'] = (volume / hist['volume_avg_20d']).round(2)
                    hist['chg_pct_1d'] = (close.pct_change() * 100).round(2)
                    hist['chg_pct_20d'] = (((close - close.shift(20)) / close.shift(20)) * 100).round(2)

                    self._hist_cache[code] = hist
                    self._rsi_cache[code] = hist['rsi_14d']
                except Exception:
                    continue

        if not self._hist_cache:
            print("\n데이터 수집 실패")
            return pd.DataFrame()

        # 기준일/비교일 두 시점의 가격지표 df 생성
        self._bloomberg_df = self._build_bbg_for_date(ref_date)
        print(f"\n기준일 가격지표: {len(self._bloomberg_df)}개 종목 ({ref_date_str})")

        if prev_date is not None:
            self._bloomberg_prev_df = self._build_bbg_for_date(prev_date)
            print(f"비교일 가격지표: {len(self._bloomberg_prev_df)}개 종목 ({prev_date_str})")

        self._save_bloomberg_raw_data(ref_date_str)
        return self._bloomberg_df

    def _build_bbg_for_date(self, target_date) -> pd.DataFrame:
        """캐시된 히스토리에서 target_date 시점(또는 그 이하 마지막 거래일)의 지표 행 추출"""
        target_ts = pd.to_datetime(target_date)
        results = []

        for code, hist in self._hist_cache.items():
            positions = [k for k, d in enumerate(hist.index) if d <= target_ts]
            if not positions:
                continue
            pos = positions[-1]

            row = hist.iloc[pos]
            cur_rsi = row.get('rsi_14d')
            prev_rsi = hist['rsi_14d'].iloc[pos - 1] if pos >= 1 else None
            if pd.notna(cur_rsi) and prev_rsi is not None and pd.notna(prev_rsi):
                rsi_change = round(float(cur_rsi) - float(prev_rsi), 1)
            else:
                rsi_change = None

            close_v = row.get('close')
            vol_v = row.get('volume')
            turnover = (round((close_v * vol_v) / 100_000_000, 1)
                        if (pd.notna(close_v) and pd.notna(vol_v)) else None)

            results.append({
                'ticker': code,
                'px_last': close_v,
                'px_volume': vol_v,
                'turnover': turnover,
                'chg_pct_1d': row.get('chg_pct_1d'),
                'chg_pct_20d': row.get('chg_pct_20d'),
                'rsi_14d': cur_rsi,
                'rsi_prev': prev_rsi,
                'rsi_change': rsi_change,
                'volume_avg_20d': row.get('volume_avg_20d'),
                'rvol_20': row.get('rvol_20'),
                'volatility_30d': row.get('volatility_30d'),
                'mov_avg_10d': row.get('mov_avg_10d'),
                'mov_avg_20d': row.get('mov_avg_20d'),
                'mov_avg_60d': row.get('mov_avg_60d'),
                'bollinger_band_width': row.get('bollinger_band_width'),
                'cur_mkt_cap': self._mkt_caps.get(code),
            })

        if not results:
            return pd.DataFrame()

        df = pd.DataFrame(results).set_index('ticker')
        if 'px_last' in df.columns and 'mov_avg_10d' in df.columns:
            df['above_ma10'] = (df['px_last'] > df['mov_avg_10d']).astype(int)
        if 'px_last' in df.columns and 'mov_avg_20d' in df.columns:
            df['above_ma20'] = (df['px_last'] > df['mov_avg_20d']).astype(int)
        if 'px_last' in df.columns and 'mov_avg_60d' in df.columns:
            df['above_ma60'] = (df['px_last'] > df['mov_avg_60d']).astype(int)
        return df

    def _save_bloomberg_raw_data(self, ref_date_str: str):
        if self._bloomberg_df is None or self._bloomberg_df.empty:
            return
        raw_data_dir = self.config.OUTPUT_DIR / "raw data"
        raw_data_dir.mkdir(parents=True, exist_ok=True)
        filename = f"bloomberg_raw_{ref_date_str.replace('-', '')}.xlsx"
        filepath = raw_data_dir / filename
        try:
            self._bloomberg_df.reset_index().to_excel(filepath, index=False, sheet_name='RawData')
            print(f"  원본 데이터 저장: {filepath}")
        except Exception as e:
            print(f"  원본 저장 실패: {e}")

    # =========================================================================
    # Step 4: 병합 + 스코어 (기준일 / 비교일 각각)
    # =========================================================================
    def merge_data(self) -> pd.DataFrame:
        print("\n" + "=" * 60)
        print("[Step 4] 데이터 병합 및 스코어 계산")
        print("=" * 60)

        if self._ipo_df is None:
            print("오류: IPO Universe가 없습니다.")
            return pd.DataFrame()

        # 기준일
        self._result_df = self._merge_and_score(self._supply_df, self._bloomberg_df, verbose=True)
        print(f"기준일 최종 데이터: {len(self._result_df)}개 종목")
        if '등급' in self._result_df.columns:
            grade_dist = self._result_df['등급'].value_counts().sort_index(ascending=False)
            print(f"기준일 등급 분포: {dict(grade_dist)}")

        # 비교일 (있을 때만)
        if (self._supply_prev_df is not None and self._bloomberg_prev_df is not None
                and not self._bloomberg_prev_df.empty):
            self._result_prev_df = self._merge_and_score(
                self._supply_prev_df, self._bloomberg_prev_df, verbose=False
            )
            print(f"비교일 최종 데이터: {len(self._result_prev_df)}개 종목")
        else:
            self._result_prev_df = None
            print("비교일 데이터 없음 → 전일 비교 생략")

        return self._result_df

    def _merge_and_score(self, supply_df, bloomberg_df, verbose: bool = False) -> pd.DataFrame:
        """IPO + 수급 + 가격지표 병합 후 스코어 계산 (날짜 비의존 공통 로직)"""
        result = self._ipo_df[['코드', '코드명', '최초상장일_dt', 'days_since_ipo', 'ticker_ks']].copy()
        result = result.rename(columns={'최초상장일_dt': '상장일'})

        # 수급 병합
        if supply_df is not None and not supply_df.empty:
            supply_cols = ['코드', '기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']
            supply_cols = [c for c in supply_cols if c in supply_df.columns]
            result = result.merge(supply_df[supply_cols], on='코드', how='left')

        # 가격지표 병합
        if bloomberg_df is not None and not bloomberg_df.empty:
            bbg = bloomberg_df.copy()
            bbg['코드'] = 'A' + bbg.index.astype(str)
            bbg_cols = ['코드', 'px_last', 'px_volume', 'turnover', 'volume_avg_20d', 'rvol_20',
                        'volatility_30d', 'rsi_14d', 'rsi_prev', 'rsi_change',
                        'chg_pct_1d', 'chg_pct_20d',
                        'mov_avg_10d', 'mov_avg_20d', 'mov_avg_60d',
                        'above_ma10', 'above_ma20', 'above_ma60',
                        'bollinger_band_width', 'cur_mkt_cap',
                        'eqy_free_float_pct', 'short_int_ratio']
            bbg_cols = [c for c in bbg_cols if c in bbg.columns]
            result = result.merge(bbg[bbg_cols], on='코드', how='left')

        result = result.drop_duplicates(subset=['코드'], keep='first')

        rename_map = {
            'px_last': '현재가', 'px_volume': '거래량', 'turnover': '거래대금(억)',
            'volume_avg_20d': '20일평균거래량', 'rvol_20': 'RVOL(20)',
            'volatility_30d': '변동성(30D)', 'bollinger_band_width': 'BB폭',
            'rsi_14d': 'RSI(14)', 'rsi_prev': '전일RSI(14)', 'rsi_change': 'RSI변화량',
            'chg_pct_1d': '1일수익률(%)', 'chg_pct_20d': '20일수익률(%)',
            'mov_avg_10d': '10일이평', 'mov_avg_20d': '20일이평', 'mov_avg_60d': '60일이평',
            'above_ma10': '10일선돌파', 'above_ma20': '20일선돌파', 'above_ma60': '60일선돌파',
            'cur_mkt_cap': '시가총액(억)', 'eqy_free_float_pct': '유통비율(%)',
            'short_int_ratio': '공매도비율',
        }
        result = result.rename(columns=rename_map)

        # 시가총액 1000억 미만 제외 (시총은 두 날짜 동일 기준 → 종목 집합 일치)
        if '시가총액(억)' in result.columns:
            result['시가총액(억)'] = pd.to_numeric(result['시가총액(억)'], errors='coerce').round(0)
            before = len(result)
            result = result[result['시가총액(억)'].fillna(0) >= 1000].copy()
            if verbose:
                print(f"  시가총액 1000억 미만 제외: {before - len(result)}개 → {len(result)}개 유지")

        if '상장일' in result.columns:
            result['상장일'] = pd.to_datetime(result['상장일']).dt.strftime('%Y-%m-%d')

        for col in ['기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']:
            if col in result.columns:
                result[col] = (result[col] / 1000).round(1)

        if 'ticker_ks' in result.columns:
            result = result.drop(columns=['ticker_ks'])

        result = self.calculate_scores(result)
        return result

    # =========================================================================
    # Step 5: 스코어링
    # =========================================================================
    def _normalize_score(self, series: pd.Series, higher_is_better: bool = True) -> pd.Series:
        if series.isna().all():
            return pd.Series([50] * len(series), index=series.index)
        pct_rank = series.rank(pct=True, na_option='keep')
        if higher_is_better:
            return (pct_rank * 100).round(1)
        return ((1 - pct_rank) * 100).round(1)

    def calculate_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        result = df.copy()

        # ── 모멘텀 ──
        momentum_parts, momentum_weights = [], []
        if '1일수익률(%)' in result.columns:
            result['_m1'] = self._normalize_score(result['1일수익률(%)'], True)
            momentum_parts.append('_m1'); momentum_weights.append(0.3)
        if '20일수익률(%)' in result.columns:
            result['_m2'] = self._normalize_score(result['20일수익률(%)'], True)
            momentum_parts.append('_m2'); momentum_weights.append(0.3)
        if '10일선돌파' in result.columns:
            result['_m3'] = result['10일선돌파'].fillna(0) * 100
            momentum_parts.append('_m3'); momentum_weights.append(0.2)
        if 'RSI(14)' in result.columns:
            result['_m4'] = self._normalize_score(result['RSI(14)'], True)
            momentum_parts.append('_m4'); momentum_weights.append(0.2)

        if momentum_parts:
            weighted = sum(result[c] * w for c, w in zip(momentum_parts, momentum_weights))
            result['모멘텀스코어'] = (weighted / sum(momentum_weights)).round(1)
        else:
            result['모멘텀스코어'] = 50

        # ── 수급 ──
        supply_components = []
        if '기관_일간' in result.columns:
            result['_s1'] = self._normalize_score(result['기관_일간'], True)
            supply_components.append(('_s1', self.config.SUPPLY_DAILY_WEIGHT))
        if '기관_5일' in result.columns:
            result['_s2'] = self._normalize_score(result['기관_5일'], True)
            supply_components.append(('_s2', self.config.SUPPLY_5D_WEIGHT))
        if '기관_20일' in result.columns:
            result['_s3'] = self._normalize_score(result['기관_20일'], True)
            supply_components.append(('_s3', self.config.SUPPLY_20D_WEIGHT))
        if '외국인_일간' in result.columns:
            result['_s4'] = self._normalize_score(result['외국인_일간'], True)
            supply_components.append(('_s4', self.config.SUPPLY_DAILY_WEIGHT))
        if '외국인_5일' in result.columns:
            result['_s5'] = self._normalize_score(result['외국인_5일'], True)
            supply_components.append(('_s5', self.config.SUPPLY_5D_WEIGHT))
        if '외국인_20일' in result.columns:
            result['_s6'] = self._normalize_score(result['외국인_20일'], True)
            supply_components.append(('_s6', self.config.SUPPLY_20D_WEIGHT))

        if supply_components:
            inst_cols = [c for c, _ in supply_components if c in ['_s1', '_s2', '_s3']]
            frgn_cols = [c for c, _ in supply_components if c in ['_s4', '_s5', '_s6']]
            inst_score = frgn_score = 50
            if inst_cols:
                w = [self.config.SUPPLY_DAILY_WEIGHT, self.config.SUPPLY_5D_WEIGHT, self.config.SUPPLY_20D_WEIGHT][:len(inst_cols)]
                inst_score = sum(result[c] * ww for c, ww in zip(inst_cols, w)) / sum(w)
            if frgn_cols:
                w = [self.config.SUPPLY_DAILY_WEIGHT, self.config.SUPPLY_5D_WEIGHT, self.config.SUPPLY_20D_WEIGHT][:len(frgn_cols)]
                frgn_score = sum(result[c] * ww for c, ww in zip(frgn_cols, w)) / sum(w)
            result['수급스코어'] = ((inst_score + frgn_score) / 2).round(1)
        else:
            result['수급스코어'] = 50

        # ── 거래량 ──
        if 'RVOL(20)' in result.columns:
            result['거래량스코어'] = self._normalize_score(result['RVOL(20)'], True)
        else:
            result['거래량스코어'] = 50

        # ── 종합 ──
        result['종합스코어'] = (
            result['모멘텀스코어'] * self.config.MOMENTUM_WEIGHT +
            result['수급스코어'] * self.config.SUPPLY_WEIGHT +
            result['거래량스코어'] * self.config.VOLUME_WEIGHT
        ).round(1)

        grade = pd.cut(result['종합스코어'], bins=[0, 30, 50, 65, 80, 100],
                       labels=['F', 'D', 'C', 'B', 'A'], include_lowest=True)
        result['등급'] = grade.astype(str)
        result.loc[result['종합스코어'].isna(), '등급'] = '-'

        result = result.drop(columns=[c for c in result.columns if c.startswith('_')])
        return result

    def _create_recommendation(self, df: pd.DataFrame) -> pd.DataFrame:
        if '종합스코어' not in df.columns:
            return None
        top = df.nlargest(self.config.TOP_N_RESULTS, '종합스코어').copy()
        display_cols = [
            '코드', '코드명', '상장일', '현재가', '거래대금(억)',
            '종합스코어', '모멘텀스코어', '수급스코어', '거래량스코어',
            'RSI변화량', 'RSI(14)', '전일RSI(14)',
            'RVOL(20)', '기관_일간', '외국인_일간',
            '10일선돌파', '20일선돌파', '60일선돌파',
            '변동성(30D)', 'BB폭', '유통비율(%)', '공매도비율',
        ]
        display_cols = [c for c in display_cols if c in top.columns]
        return top[display_cols].reset_index(drop=True)

    # =========================================================================
    # Step 6: 보고서 시트 (변경 없음 — 전체_전일자 시트를 INDEX/MATCH로 참조)
    # =========================================================================
    def _create_report_sheet(self, writer, df: pd.DataFrame, has_prev_data: bool = False):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter

        wb = writer.book
        ws = wb.create_sheet('POST IPO 보고서', 0)

        COLOR_NAVY = '1F3864'; COLOR_BLUE = '2E75B6'; COLOR_LIGHT = 'D6E4F0'
        COLOR_WHITE = 'FFFFFF'; COLOR_GOLD = 'C9A84C'; COLOR_GREEN = '70AD47'
        COLOR_ORANGE = 'ED7D31'; COLOR_RED = 'FF0000'; COLOR_GRAY = 'F2F2F2'

        def fill(color): return PatternFill('solid', fgColor=color)
        def font(bold=False, color='000000', size=11):
            return Font(bold=bold, color=color, size=size, name='맑은 고딕')
        def center(): return Alignment(horizontal='center', vertical='center')
        def thin_border():
            s = Side(style='thin', color='BFBFBF')
            return Border(left=s, right=s, top=s, bottom=s)

        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else datetime.now().strftime('%Y-%m-%d')
        total = len(df)

        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        if '등급' in df.columns:
            for g, cnt in df['등급'].value_counts().items():
                if g in grade_counts:
                    grade_counts[g] = int(cnt)
        strong = grade_counts['A'] + grade_counts['B']
        neutral = grade_counts['C']
        weak = grade_counts['D'] + grade_counts['F']
        score_50_plus = int((pd.to_numeric(df.get('종합스코어', pd.Series(dtype=float)), errors='coerce') >= 50).sum())

        col_widths = {1: 3, 2: 18, 3: 14, 4: 10, 5: 10, 6: 10, 7: 10, 8: 12, 9: 12, 10: 10, 11: 14, 12: 3, 13: 16, 14: 10}
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 38
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 12

        ws.merge_cells('B2:I2')
        title_cell = ws['B2']
        title_cell.value = 'POST IPO 모니터링 보고서'
        title_cell.font = Font(bold=True, color=COLOR_WHITE, size=18, name='맑은 고딕')
        title_cell.fill = fill(COLOR_NAVY)
        title_cell.alignment = center()

        ws.merge_cells('B3:I3')
        sub_cell = ws['B3']
        sub_cell.value = f'기준일: {ref_date_str}  |  분석 대상: {total}개 종목  |  2년 이내 신규 상장주'
        sub_cell.font = Font(color=COLOR_WHITE, size=10, name='맑은 고딕')
        sub_cell.fill = fill(COLOR_BLUE)
        sub_cell.alignment = center()

        ws.row_dimensions[5].height = 10
        ws.row_dimensions[6].height = 28
        ws.row_dimensions[7].height = 22
        ws.row_dimensions[8].height = 12

        cards = [
            ('전체 종목', str(total), COLOR_BLUE),
            ('A등급', str(grade_counts['A']), COLOR_GOLD),
            ('B등급', str(grade_counts['B']), COLOR_GREEN),
            ('50점 이상\n종목수', str(score_50_plus), COLOR_NAVY),
        ]
        card_cols = [2, 4, 6, 8]
        for (label, value, color), col in zip(cards, card_cols):
            ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
            ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
            lc = ws.cell(row=6, column=col)
            lc.value = label
            lc.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            lc.fill = fill(color)
            lc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            vc = ws.cell(row=7, column=col)
            vc.value = int(value)
            vc.font = Font(bold=True, color=color, size=20, name='맑은 고딕')
            vc.fill = fill(COLOR_GRAY)
            vc.alignment = center()

        ws.row_dimensions[9].height = 10
        ws.row_dimensions[10].height = 22
        ws.merge_cells('B10:O10')
        sec1 = ws['B10']
        sec1.value = '▶  추천종목  (종합스코어 50점 이상)'
        sec1.font = Font(bold=True, color=COLOR_WHITE, size=11, name='맑은 고딕')
        sec1.fill = fill(COLOR_BLUE)
        sec1.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        headers = ['순위', '종목명', '종합점수', '모멘텀', '수급', '거래량', 'RSI', '등급', 'RSI신호']
        header_cols = list(range(2, 11))
        ws.row_dimensions[11].height = 18
        for h, c in zip(headers, header_cols):
            cell = ws.cell(row=11, column=c)
            cell.value = h
            cell.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            cell.fill = fill(COLOR_NAVY)
            cell.alignment = center()
            cell.border = thin_border()

        grade_colors = {'A': 'FFD700', 'B': COLOR_GREEN, 'C': 'FFC000', 'D': 'ED7D31', 'F': 'FF6B6B'}
        if '종합스코어' in df.columns:
            scored = pd.to_numeric(df['종합스코어'], errors='coerce')
            filtered_rec = df[scored >= 50].sort_values('종합스코어', ascending=False).reset_index(drop=True)
        else:
            filtered_rec = pd.DataFrame()

        if not filtered_rec.empty:
            for idx, (_, row) in enumerate(filtered_rec.iterrows()):
                r = 12 + idx
                ws.row_dimensions[r].height = 17
                row_fill = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)
                grade_str = row.get('등급', '-') if pd.notna(row.get('등급', None)) else '-'

                rsi_cur = pd.to_numeric(row.get('RSI(14)', None), errors='coerce')
                rsi_prev_val = pd.to_numeric(row.get('전일RSI(14)', None), errors='coerce')
                rsi_signal = ('★' if (pd.notna(rsi_cur) and pd.notna(rsi_prev_val)
                                      and rsi_cur >= 65 and rsi_prev_val < 65) else '')

                data = [
                    idx + 1, row.get('코드명', ''), row.get('종합스코어', ''),
                    row.get('모멘텀스코어', ''), row.get('수급스코어', ''),
                    row.get('거래량스코어', ''), row.get('RSI(14)', ''), grade_str, rsi_signal,
                ]
                for val, c in zip(data, header_cols):
                    cell = ws.cell(row=r, column=c)
                    cell.value = val if not (isinstance(val, float) and pd.isna(val)) else '-'
                    cell.font = font(size=9)
                    cell.fill = row_fill
                    cell.alignment = center()
                    cell.border = thin_border()

                grade_cell = ws.cell(row=r, column=9)
                grade_cell.fill = fill(grade_colors.get(grade_str, 'BFBFBF'))
                grade_cell.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')

                if rsi_signal:
                    sig_cell = ws.cell(row=r, column=10)
                    sig_cell.fill = fill('FFF2CC')
                    sig_cell.font = Font(bold=True, color='C9A84C', size=10, name='맑은 고딕')

        if has_prev_data:
            k11 = ws.cell(row=11, column=11)
            k11.value = '등급변화'
            k11.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            k11.fill = fill(COLOR_NAVY); k11.alignment = center(); k11.border = thin_border()
            l11 = ws.cell(row=11, column=12)
            l11.value = '점수변화량'
            l11.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            l11.fill = fill(COLOR_NAVY); l11.alignment = center(); l11.border = thin_border()

            rec_count = len(filtered_rec) if not filtered_rec.empty else 0
            for idx in range(rec_count):
                r = 12 + idx
                row_fill = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)
                k_cell = ws.cell(row=r, column=11)
                prev_grade = ('INDEX(전체_전일자!1:1048576,MATCH(C{r},전체_전일자!B:B,0),'
                              'MATCH("등급",전체_전일자!1:1,0))').format(r=r)
                k_cell.value = (f'=IFERROR(IF(I{r}={prev_grade},"유지",{prev_grade}&" → "&I{r}),"")')
                k_cell.font = font(size=9); k_cell.fill = row_fill
                k_cell.alignment = center(); k_cell.border = thin_border()

                l_cell = ws.cell(row=r, column=12)
                prev_score = ('INDEX(전체_전일자!1:1048576,MATCH(C{r},전체_전일자!B:B,0),'
                              'MATCH("종합스코어",전체_전일자!1:1,0))').format(r=r)
                l_cell.value = f'=IFERROR(D{r}-{prev_score},"")'
                l_cell.number_format = '+0.0;-0.0;0.0'
                l_cell.font = font(size=9); l_cell.fill = row_fill
                l_cell.alignment = center(); l_cell.border = thin_border()

            for col_idx, label in [(13, '모멘텀Δ'), (14, '수급Δ'), (15, '거래량Δ')]:
                hdr = ws.cell(row=11, column=col_idx)
                hdr.value = label
                hdr.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
                hdr.fill = fill(COLOR_NAVY); hdr.alignment = center(); hdr.border = thin_border()

            score_cols = [(13, 'E', '모멘텀스코어'), (14, 'F', '수급스코어'), (15, 'G', '거래량스코어')]
            for idx in range(rec_count):
                r = 12 + idx
                row_fill_val = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)
                for col_idx, src_col, prev_col_name in score_cols:
                    prev_val = ('INDEX(전체_전일자!1:1048576,MATCH(C{r},전체_전일자!B:B,0),'
                                'MATCH("{col}",전체_전일자!1:1,0))').format(r=r, col=prev_col_name)
                    cell = ws.cell(row=r, column=col_idx)
                    cell.value = f'=IFERROR({src_col}{r}-{prev_val},"")'
                    cell.number_format = '+0.0;-0.0;0.0'
                    cell.font = font(size=9); cell.fill = row_fill_val
                    cell.alignment = center(); cell.border = thin_border()

            ws.column_dimensions['K'].width = 14
            ws.column_dimensions['L'].width = 16
            ws.column_dimensions['M'].width = 11
            ws.column_dimensions['N'].width = 10
            ws.column_dimensions['O'].width = 11

        ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=10)
        j6 = ws.cell(row=6, column=10)
        j6.value = 'RSI>65'
        j6.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
        j6.fill = fill(COLOR_NAVY)
        j6.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        j7 = ws.cell(row=7, column=10)
        j7.value = ('=COUNTIF(INDEX(전체!$A:$XFD,0,MATCH("RSI(14)",전체!$1:$1,0)),">=65")')
        j7.font = Font(bold=True, color=COLOR_NAVY, size=20, name='맑은 고딕')
        j7.fill = fill(COLOR_GRAY)
        j7.alignment = center()

        CHART_ROW = 10; CHART_COL_LABEL = 17; CHART_COL_VAL = 18
        ws.cell(row=CHART_ROW, column=CHART_COL_LABEL).value = '구분'
        ws.cell(row=CHART_ROW, column=CHART_COL_VAL).value = '종목수'
        ws.cell(row=CHART_ROW + 1, column=CHART_COL_LABEL).value = '강세 (A+B등급)'
        ws.cell(row=CHART_ROW + 1, column=CHART_COL_VAL).value = strong
        ws.cell(row=CHART_ROW + 2, column=CHART_COL_LABEL).value = '중립 (C등급)'
        ws.cell(row=CHART_ROW + 2, column=CHART_COL_VAL).value = neutral
        ws.cell(row=CHART_ROW + 3, column=CHART_COL_LABEL).value = '약세 (D+F등급)'
        ws.cell(row=CHART_ROW + 3, column=CHART_COL_VAL).value = weak
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 10

        pie = PieChart()
        pie.title = 'Post IPO 모멘텀 현황'
        pie.style = 10
        data_ref = Reference(ws, min_col=CHART_COL_VAL, min_row=CHART_ROW, max_row=CHART_ROW + 3)
        label_ref = Reference(ws, min_col=CHART_COL_LABEL, min_row=CHART_ROW + 1, max_row=CHART_ROW + 3)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(label_ref)
        for i, color in enumerate([COLOR_BLUE, COLOR_ORANGE, 'C00000']):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        pie.width = 13
        pie.height = 10
        ws.add_chart(pie, 'Q14')

    def _set_column_width(self, writer, sheet_name: str):
        from openpyxl.utils import get_column_letter
        excel_width = self.config.EXCEL_COLUMN_WIDTH / 7
        ws = writer.sheets[sheet_name]
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = excel_width

    def _load_rsi_history_from_excel(self) -> pd.DataFrame:
        import glob
        pattern = str(self.config.OUTPUT_DIR / "ipo_monitoring_*.xlsx")
        files = glob.glob(pattern)
        cols = ['날짜', '분석종목수', 'RSI65이상', '과매수비율(%)', '전일대비']
        if not files:
            return pd.DataFrame(columns=cols)
        latest_file = max(files)
        try:
            df = pd.read_excel(latest_file, sheet_name='RSI추이')
            df['날짜'] = pd.to_datetime(df['날짜'])
            return df
        except (ValueError, KeyError):
            return pd.DataFrame(columns=cols)

    def _calculate_rsi_history(self) -> pd.DataFrame:
        cols = ['날짜', '분석종목수', 'RSI65이상', '과매수비율(%)', '전일대비']
        end_date = self._ref_date if self._ref_date else datetime.now()
        end_str = end_date.strftime('%Y-%m-%d')

        existing = self._load_rsi_history_from_excel()
        is_initial = len(existing) < 5
        if is_initial:
            print(f"\nRSI 히스토리 초기화 ({self.config.RSI_CALC_DAYS}일 계산)...")
        else:
            print(f"\nRSI 히스토리 업데이트 (당일)...")

        if not self._rsi_cache:
            print("  RSI 캐시 없음")
            return existing if not existing.empty else pd.DataFrame(columns=cols)

        try:
            rsi_matrix = pd.DataFrame(self._rsi_cache).sort_index()
            rsi_matrix = rsi_matrix[rsi_matrix.index <= pd.to_datetime(end_date)]
            if rsi_matrix.empty:
                return existing if not existing.empty else pd.DataFrame(columns=cols)

            if is_initial:
                target_dates = list(rsi_matrix.index)[-self.config.RSI_CALC_DAYS:]
            else:
                target_dates = [rsi_matrix.index[-1]]

            calc_results = []
            for date in target_dates:
                row_data = rsi_matrix.loc[date]
                valid_count = int(row_data.notna().sum())
                overbought_count = int((row_data >= self.config.RSI_THRESHOLD).sum())
                ratio = (overbought_count / valid_count * 100) if valid_count > 0 else 0
                calc_results.append({
                    '날짜': pd.to_datetime(date), '분석종목수': valid_count,
                    'RSI65이상': overbought_count, '과매수비율(%)': round(ratio, 1),
                })
            calculated = pd.DataFrame(calc_results)

            if is_initial:
                print(f"  초기 {len(calculated)}일 계산 완료")
                history = calculated
            else:
                print(f"  당일 데이터 추가")
                existing = existing[existing['날짜'].dt.strftime('%Y-%m-%d') != end_str]
                history = pd.concat([existing, calculated], ignore_index=True)

            history = history.sort_values('날짜').reset_index(drop=True)
            cutoff_date = end_date - pd.Timedelta(days=self.config.RSI_HISTORY_DAYS)
            history = history[history['날짜'] >= cutoff_date]
            history['전일대비'] = history['RSI65이상'].diff().fillna(0).astype(int)
            print(f"  RSI 히스토리 총 {len(history)}일")
            return history
        except Exception as e:
            print(f"  RSI 히스토리 계산 오류: {e}")
            return existing if not existing.empty else pd.DataFrame(columns=cols)

    # =========================================================================
    # 결과 저장 — 전체_전일자는 메모리상의 비교일 결과를 직접 사용
    # =========================================================================
    def save_results(self, filename: str = None) -> Path:
        print("\n" + "=" * 60)
        print("[Step 6] 결과 저장")
        print("=" * 60)

        if self._result_df is None or self._result_df.empty:
            print("저장할 데이터가 없습니다.")
            return None

        df = self._result_df.copy()
        if filename is None:
            filename = self.config.get_output_filename()
        filepath = self.config.OUTPUT_DIR / filename

        sheet_names = []
        prev_df = self._result_prev_df
        has_prev_data = prev_df is not None and not prev_df.empty

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            recommend_df = self._create_recommendation(df)
            if recommend_df is not None and not recommend_df.empty:
                recommend_df.to_excel(writer, sheet_name='추천종목', index=False)
                sheet_names.append('추천종목')

            if '등급' in df.columns:
                grade_a = df[df['등급'] == 'A'].sort_values('종합스코어', ascending=False)
                if not grade_a.empty:
                    grade_a.to_excel(writer, sheet_name='A등급', index=False)
                    sheet_names.append('A등급')

            df.to_excel(writer, sheet_name='전체', index=False)
            sheet_names.append('전체')

            # 전체_전일자: 비교일(메모리) 결과를 그대로 사용 → 외부 파일·input 불필요
            if has_prev_data:
                prev_df.to_excel(writer, sheet_name='전체_전일자', index=False)
                sheet_names.append('전체_전일자')
                prev_str = self._supply_prev_date.strftime('%Y-%m-%d') if self._supply_prev_date is not None else '?'
                print(f"  전체_전일자 시트 생성 (비교일: {prev_str})")
            else:
                print("  비교일 데이터 없음 → 전체_전일자 시트 생략, K~O 비교열 미생성")

            if 'RSI(14)' in df.columns:
                oversold = df[df['RSI(14)'] < 30].sort_values('RSI(14)')
                if not oversold.empty:
                    oversold.to_excel(writer, sheet_name='RSI과매도', index=False)
                    sheet_names.append('RSI과매도')
                overbought = df[df['RSI(14)'] > 70].sort_values('RSI(14)', ascending=False)
                if not overbought.empty:
                    overbought.to_excel(writer, sheet_name='RSI과매수', index=False)
                    sheet_names.append('RSI과매수')

            if '외국인_일간' in df.columns:
                foreign_top = df.dropna(subset=['외국인_일간']).nlargest(20, '외국인_일간')
                if not foreign_top.empty:
                    foreign_top.to_excel(writer, sheet_name='외국인순매수TOP', index=False)
                    sheet_names.append('외국인순매수TOP')
            if '기관_일간' in df.columns:
                inst_top = df.dropna(subset=['기관_일간']).nlargest(20, '기관_일간')
                if not inst_top.empty:
                    inst_top.to_excel(writer, sheet_name='기관순매수TOP', index=False)
                    sheet_names.append('기관순매수TOP')

            if 'RVOL(20)' in df.columns:
                vol_top = df.dropna(subset=['RVOL(20)']).nlargest(20, 'RVOL(20)')
                if not vol_top.empty:
                    vol_top.to_excel(writer, sheet_name='RVOL_TOP', index=False)
                    sheet_names.append('RVOL_TOP')

            rsi_history = self._calculate_rsi_history()
            if not rsi_history.empty:
                rsi_display = rsi_history.copy()
                rsi_display['날짜'] = rsi_display['날짜'].dt.strftime('%Y-%m-%d')
                rsi_display = rsi_display.sort_values('날짜', ascending=False)
                rsi_display.to_excel(writer, sheet_name='RSI추이', index=False)
                sheet_names.append('RSI추이')

            self._create_report_sheet(writer, df, has_prev_data=has_prev_data)

            wb = writer.book
            try:
                wb.calculation_properties.fullCalcOnLoad = True
                wb.calculation_properties.calcMode = 'auto'
            except Exception:
                pass

            for sheet in sheet_names:
                self._set_column_width(writer, sheet)

        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else '(미정)'
        print(f"저장 완료: {filepath}")
        print(f"  (데이터 기준일: {ref_date_str})")
        return filepath

    def print_summary(self, n: int = 10):
        if self._result_df is None or self._result_df.empty:
            print("데이터가 없습니다.")
            return
        df = self._result_df.copy()

        print("\n" + "=" * 80)
        print("                    IPO 모니터링 요약")
        print("=" * 80)
        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else '(미정)'
        prev_str = self._supply_prev_date.strftime('%Y-%m-%d') if self._supply_prev_date is not None else '없음'
        print(f"기준일: {ref_date_str} (종가)  |  비교일: {prev_str}")
        print(f"실행일: {datetime.now().strftime('%Y-%m-%d')}")
        print(f"총 종목 수: {len(df)}개")

        if '등급' in df.columns:
            grade_dist = df['등급'].value_counts().sort_index(ascending=False)
            print(f"\n등급 분포: {dict(grade_dist)}")

        recommend = self._create_recommendation(df)
        if recommend is not None and not recommend.empty:
            print(f"\n[★ 추천종목 TOP {min(n, len(recommend))}] (종합스코어 기준)")
            print("-" * 80)
            print(f"  {'순위':>4} {'종목명':<12} {'종합':>6} {'모멘텀':>6} {'수급':>6} {'거래량':>6} {'RSI':>6} {'RSI변화':>7}")
            print("-" * 80)
            for i, (_, row) in enumerate(recommend.head(n).iterrows(), 1):
                print(f"  {i:>4} {row['코드명']:<12} "
                      f"{row['종합스코어']:>6.1f} {row.get('모멘텀스코어', 0):>6.1f} "
                      f"{row.get('수급스코어', 0):>6.1f} {row.get('거래량스코어', 0):>6.1f} "
                      f"{row.get('RSI(14)', 0):>6.1f} {row.get('RSI변화량', 0):>+7.1f}")

        if 'RSI(14)' in df.columns:
            oversold = df[df['RSI(14)'] < 30].sort_values('RSI(14)')
            if not oversold.empty:
                print(f"\n[RSI 과매도 (< 30)] {len(oversold)}개")
                print("-" * 70)
                for _, row in oversold.head(n).iterrows():
                    print(f"  {row['코드명']:<15} RSI: {row['RSI(14)']:>5.1f}  변동성: {row.get('변동성(30D)', 0):>5.1f}%")

        if '외국인_일간' in df.columns:
            foreign_top = df.dropna(subset=['외국인_일간']).nlargest(n, '외국인_일간')
            if not foreign_top.empty:
                print(f"\n[외국인 순매수 TOP {n}] (단위: 백만원)")
                print("-" * 70)
                for _, row in foreign_top.iterrows():
                    print(f"  {row['코드명']:<15} 일간: {row['외국인_일간']:>8.1f}  5일: {row.get('외국인_5일', 0):>8.1f}  20일: {row.get('외국인_20일', 0):>8.1f}")

        if '기관_일간' in df.columns:
            inst_top = df.dropna(subset=['기관_일간']).nlargest(n, '기관_일간')
            if not inst_top.empty:
                print(f"\n[기관 순매수 TOP {n}] (단위: 백만원)")
                print("-" * 70)
                for _, row in inst_top.iterrows():
                    print(f"  {row['코드명']:<15} 일간: {row['기관_일간']:>8.1f}  5일: {row.get('기관_5일', 0):>8.1f}  20일: {row.get('기관_20일', 0):>8.1f}")

        if 'RVOL(20)' in df.columns:
            vol_top = df.dropna(subset=['RVOL(20)']).nlargest(n, 'RVOL(20)')
            if not vol_top.empty:
                print(f"\n[RVOL(20) TOP {n}]")
                print("-" * 70)
                for _, row in vol_top.iterrows():
                    print(f"  {row['코드명']:<15} RVOL: {row['RVOL(20)']:>5.2f}x  RSI: {row.get('RSI(14)', 0):>5.1f}")

        print("=" * 80)

    # =========================================================================
    # 전체 실행
    # =========================================================================
    def run(self, source: str = 'A', use_prev_day: bool = True) -> pd.DataFrame:
        """전체 실행.
        비교 대상은 수급.xlsx의 채워진 마지막 두 날짜(기준일/비교일)로 자동 결정.
        외부 전일 결과 파일·input 입력 불필요."""
        self.load_ipo_universe(source=source)
        self.load_supply_data()
        self.fetch_bloomberg_data(use_prev_day=use_prev_day)
        self.merge_data()
        self.print_summary()
        self.save_results()
        return self._result_df


# =============================================================================
# Main
# =============================================================================
if __name__ == "__main__":
    print("run.py를 통해 실행해주세요: python run.py")