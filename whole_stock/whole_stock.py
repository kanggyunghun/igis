#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
전종목_수급.xlsx (stack + update 2시트 구조) 기반 스코어링

워크북 구조:
    - 'update' 시트 : 매일 데이터가이드(DataGuide) 애드인이 새로고침하는 Current 1일치
                      (종목=행, 아이템=열, 날짜컬럼 없음, 상단에 Last Updated 타임스탬프)
    - 'stack'  시트 : 누적 시계열 (날짜=행, 종목×아이템=열)

매일 흐름:
    1) update 시트(오늘 새로고침분)를 읽고 → '직전 거래일(어제, 주말이면 금요일)'로 날짜 스탬프
    2) 기존 누적(워크북 stack 시트 ∪ 별도 백업파일 STACK_FILE)과 합침
    3) (날짜,코드,항목) 중복 제거(update 우선) → 최근 N거래일만 trim
    4) 누적을 ▶ 워크북 'stack' 시트에 직접 적재 + ▶ 별도 백업파일(STACK_FILE)에도 저장
    5) 합쳐진 데이터로 기존 스코어링 파이프라인 그대로 실행

    => 매일 돌리면 update → stack 으로 쌓이고, 다음날 또 돌리면 그 위에 또 쌓이는 구조.
       (KEEP_TRADING_DAYS 만큼만 보존 = 20일 윈도우 + 여유)

실행:
    py -3.12 whole_stock.py                # 직전 거래일 자동 스탬프
    py -3.12 whole_stock.py 2026-06-23     # 날짜 수동 지정(선택)

주의(중요):
    - 워크북 stack 시트에 기록하려면 실행 시점에 해당 .xlsx 가 Excel에서 '닫혀' 있어야 합니다.
      (Excel에 열려 있으면 PermissionError 로 워크북 기록만 건너뜀 → 별도 누적파일은 정상 저장)
    - 누적 안전망은 별도 파일(STACK_FILE)이 담당합니다. (워크북 자동 백업은 생성하지 않음)
    - 워크북 기록을 끄고 별도파일만 쓰려면 WRITE_BACK_TO_WORKBOOK = False.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd


# =============================================================================
# 경로 / 설정
# =============================================================================
# 경로 기준점
#   whole_stock.py 위치: igis/whole_stock/whole_stock.py
#     BASE_DIR  → igis/whole_stock/   (프로젝트 폴더, live 워크북 위치)
#     REPO_ROOT → igis/               (출력 루트)
BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent
CWD = Path.cwd()

# live 워크북 (stack + update 두 시트가 들어있는 파일) — 프로젝트 폴더 우선, 없으면 실행 폴더
INPUT_FILE = BASE_DIR / "전종목_수급.xlsx"
if not INPUT_FILE.exists():
    alternate = CWD / "전종목_수급.xlsx"
    if alternate.exists():
        INPUT_FILE = alternate

# 출력은 igis/outputs/whole_stock 으로 통일 (다른 프로젝트와 동일 규칙)
OUTPUT_DIR = REPO_ROOT / "outputs" / "whole_stock"

# 시트명 (워크북 안)
UPDATE_SHEET = "update"
STACK_SHEET = "stack"

# update 읽은 누적분을 워크북 stack 시트에 직접 기록할지 여부
WRITE_BACK_TO_WORKBOOK = True

# 파이썬이 관리하는 별도 누적 파일 (누적 seed + 워크북 기록 실패 시 안전망)
STACK_FILE = OUTPUT_DIR / "전종목_수급_stack.xlsx"
STACK_FILE_SHEET = "stack"

# 누적 보존 거래일 수 (20일 윈도우 + 여유). 더 길게 보관하려면 늘리면 됨.
KEEP_TRADING_DAYS = 30

PRICE_WEIGHT = 0.50
SUPPLY_WEIGHT = 0.20
VALUE_WEIGHT = 0.30
FDR_LOOKBACK_DAYS = 80

# 내부 필드 ↔ 데이터가이드 라벨 (누적파일/stack 시트 저장 시 사용)
FIELD_ORDER = [
    "close",
    "turnover_avg_5d", "turnover_avg_20d",
    "inst_net_5d", "inst_net_20d",
    "foreign_net_5d", "foreign_net_20d",
    "avg_ret_5d", "avg_ret_20d",
]
FIELD_LABEL = {
    "close": "종가(원)",
    "turnover_avg_5d": "5일평균거래대금(원)",
    "turnover_avg_20d": "20일평균거래대금(원)",
    "inst_net_5d": "5일누적 기관 순매수대금(일간)(만원)",
    "inst_net_20d": "20일누적 기관 순매수대금(일간)(만원)",
    "foreign_net_5d": "5일누적 외국인총합계 순매수대금(일간)(만원)",
    "foreign_net_20d": "20일누적 외국인총합계 순매수대금(일간)(만원)",
    "avg_ret_5d": "1주전대비수익률(%)",
    "avg_ret_20d": "1개월전대비수익률(%)",
}

_CODE_RE = re.compile(r"^A[0-9A-Z]{6}$")   # A + 영숫자 6자리 (우선주·특수종목 포함)
_PERIOD_RE = re.compile(r"(\d+)\s*일")
_DATE_RE = re.compile(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}")


# =============================================================================
# 공통 헬퍼
# =============================================================================
def _is_code(v) -> bool:
    return bool(_CODE_RE.match(str(v).strip()))


def _num(series: pd.Series) -> pd.Series:
    """콤마 포함 문자열도 안전하게 숫자 변환."""
    s = series.astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(s, errors="coerce")


def _xl_safe(v):
    """openpyxl 기록용 값 정리: numpy 타입/NaN/inf/Timestamp 안전 변환."""
    if v is None:
        return None
    if isinstance(v, float):
        return None if (np.isnan(v) or np.isinf(v)) else v
    if isinstance(v, np.floating):
        f = float(v)
        return None if (np.isnan(f) or np.isinf(f)) else f
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, (pd.Timestamp, datetime)):
        return pd.Timestamp(v).strftime("%Y-%m-%d")
    return v


def resolve_item_field(item: str) -> str | None:
    """엑셀 아이템 라벨 → 내부 필드명. (stack/update 시트 라벨 동일하게 매핑)"""
    item = str(item).strip()
    if not item or item in ("아이템명", "nan", "None"):
        return None
    if "종가" in item:
        return "close"

    match = _PERIOD_RE.search(item)
    period = match.group(1) if match else None

    def by_period(field_5d: str, field_20d: str) -> str | None:
        if period == "5":
            return field_5d
        if period == "20":
            return field_20d
        return None

    if "수익" in item:
        if "주" in item:   # 1주전대비 → 단기(≈5거래일)
            return "avg_ret_5d"
        if "월" in item:   # 1개월전대비 → 중기(≈20거래일)
            return "avg_ret_20d"
        return by_period("avg_ret_5d", "avg_ret_20d")
    if "거래대금" in item:
        return by_period("turnover_avg_5d", "turnover_avg_20d")
    if "순매수" in item and "기관" in item:
        return by_period("inst_net_5d", "inst_net_20d")
    if "순매수" in item and ("외국인" in item or "외인" in item):
        return by_period("foreign_net_5d", "foreign_net_20d")
    return None


def normalize_score(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    if series.notna().sum() == 0:
        return pd.Series(50.0, index=series.index)
    pct = series.rank(pct=True, na_option="keep")
    score = pct * 100 if higher_is_better else (1 - pct) * 100
    return score.fillna(50).round(1)


# --- ETF 판별 (FDR ETF 목록 + 종목명 키워드 결합) ---------------------------
ETF_NAME_KEYWORDS = (
    "KODEX", "TIGER", "PLUS", "ACE", "SOL", "RISE", "KOSEF", "ARIRANG", "HANARO",
    "KBSTAR", "KINDEX", "TIMEFOLIO", "KIWOOM", "히어로즈", "마이티", "BNK", "FOCUS",
    "WOORI", "마이다스", "파워", "TREX", "VITA", "스마트", "1Q",
)


def _load_etf_codes() -> set:
    """FDR(FinanceDataReader) ETF/KR 목록의 종목코드 집합 (A###### 형태). 실패 시 빈 집합."""
    try:
        import FinanceDataReader as fdr
        etf = fdr.StockListing("ETF/KR")
        col = next((c for c in ["Symbol", "Code", "code"] if c in etf.columns), None)
        if col is not None:
            return {"A" + str(s).strip().zfill(6) for s in etf[col].dropna()}
    except Exception as e:
        print(f"  (ETF 목록 로드 실패: {e} → 종목명 키워드만 사용)")
    return set()


def classify_etf(codes, names) -> dict:
    """code → True/False (ETF 여부). FDR 목록 OR 종목명 키워드."""
    etf_codes = _load_etf_codes()
    flags = {}
    for code, name in zip(codes, names):
        nm = str(name).upper()
        flags[code] = (code in etf_codes) or any(k.upper() in nm for k in ETF_NAME_KEYWORDS)
    return flags


# =============================================================================
# 날짜 해석 (update 시트는 '직전 거래일 = 어제' 데이터)
# =============================================================================
def _previous_krx_trading_day(ref_dt: pd.Timestamp) -> pd.Timestamp:
    """ref_dt(보통 '오늘') 직전의 KRX(한국거래소) 거래일.
    = update 시트가 담고 있는 '어제' 데이터의 기준일.
    어제가 주말/휴장이면 직전 거래일로 자동 롤백(예: 월요일 실행 → 금요일).
    FDR(FinanceDataReader, KODEX200)로 실제 거래일을 확인하고, 실패 시 주말만 롤백."""
    ref_dt = pd.Timestamp(ref_dt).normalize()
    try:
        import FinanceDataReader as fdr
        start = (ref_dt - pd.Timedelta(days=20)).strftime("%Y-%m-%d")
        end = (ref_dt - pd.Timedelta(days=1)).strftime("%Y-%m-%d")  # 어제까지 조회
        idx = fdr.DataReader("069500", start, end)  # KODEX 200 (거래일 캘린더 대용)
        if idx is not None and not idx.empty:
            idx = idx.reset_index()
            dcol = "Date" if "Date" in idx.columns else idx.columns[0]
            dts = pd.to_datetime(idx[dcol], errors="coerce").dropna()
            dts = dts[dts < ref_dt]  # 당일 제외 = '직전' 거래일
            if len(dts):
                return dts.max().normalize()
    except Exception:
        pass
    # FDR 실패 → 주말만 롤백
    d = ref_dt - pd.Timedelta(days=1)
    while d.weekday() >= 5:  # 토(5)/일(6) → 금요일로
        d -= pd.Timedelta(days=1)
    return d


def _krx_trading_days(start: pd.Timestamp, end: pd.Timestamp) -> list:
    """start ~ end (양끝 포함) 사이의 KRX 거래일 목록.
    FDR(KODEX200)로 실제 거래일을 구하고, 실패 시 주말만 제외(공휴일은 못 거름)."""
    start = pd.Timestamp(start).normalize()
    end = pd.Timestamp(end).normalize()
    if start > end:
        return []
    try:
        import FinanceDataReader as fdr
        idx = fdr.DataReader("069500", start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
        if idx is not None and not idx.empty:
            idx = idx.reset_index()
            dcol = "Date" if "Date" in idx.columns else idx.columns[0]
            dts = pd.to_datetime(idx[dcol], errors="coerce").dropna().dt.normalize()
            dts = dts[(dts >= start) & (dts <= end)]
            if len(dts):
                return sorted(dts.unique().tolist())
    except Exception:
        pass
    # FDR 실패 → 주말만 제외 (공휴일은 못 거르므로 과대 추정 가능 → 경고는 보수적으로)
    return list(pd.bdate_range(start, end))


def detect_stack_gap(stack_last_date, ref_date) -> list:
    """stack 마지막 날짜와 이번 기준일(ref_date) 사이의 '누락된 거래일' 목록.
    - stack_last_date 직후 거래일 ~ ref_date 직전 거래일 까지가 채워져 있어야 함.
    - 연속(어제→오늘)이면 빈 리스트, 중간이 비면 그 날짜들을 반환.
    stack이 비어있으면(첫 실행) 공백 판정 안 함."""
    if stack_last_date is None:
        return []
    stack_last_date = pd.Timestamp(stack_last_date).normalize()
    ref_date = pd.Timestamp(ref_date).normalize()
    if ref_date <= stack_last_date:
        return []
    # 두 날짜 사이(양끝 포함) 거래일 → 양끝 제외하면 '사이에 있어야 할 거래일'
    between = _krx_trading_days(stack_last_date, ref_date)
    gap = [d for d in between if stack_last_date < d < ref_date]
    return gap


def update_equals_stack_latest(update_long: pd.DataFrame, stack_long: pd.DataFrame) -> bool:
    """update 데이터가 stack의 '가장 최근 날짜' 데이터와 (code, field, value) 기준으로 동일한가?
    True  → 단말이 아직 Refresh 안 됨(같은 데이터) → 새로 쌓으면 중복.
    False → 내용이 다르거나 stack이 비어있음(첫 실행) → 정상 누적 대상.
    (전종목의 모든 수치가 우연히 일치할 확률은 0에 가까우므로, 동일=미갱신으로 판단)"""
    if stack_long is None or stack_long.empty or update_long is None or update_long.empty:
        return False
    sd = pd.to_datetime(stack_long["date"], errors="coerce")
    last_date = sd.max()
    if pd.isna(last_date):
        return False

    def _norm(df):
        d = df[["code", "field", "value"]].copy()
        d["value"] = pd.to_numeric(d["value"], errors="coerce").round(4)
        d = d.dropna(subset=["value"])
        return d.sort_values(["code", "field"]).reset_index(drop=True)

    u = _norm(update_long)
    s = _norm(stack_long[sd == last_date])
    if u.empty or len(u) != len(s):
        return False
    merged = u.merge(s, on=["code", "field"], suffixes=("_u", "_s"), how="outer")
    if merged[["value_u", "value_s"]].isna().any().any():
        return False  # 한쪽에만 있는 (code,field) → 다름
    return bool(np.allclose(merged["value_u"], merged["value_s"], rtol=0, atol=1e-6))


def _find_last_updated(raw: pd.DataFrame) -> pd.Timestamp | None:
    """update 시트 상단(최대 8행×6열)에서 Last Updated 타임스탬프 탐색."""
    for i in range(min(8, len(raw))):
        for c in range(min(6, raw.shape[1])):
            v = raw.iloc[i, c]
            if isinstance(v, (datetime, pd.Timestamp)):
                return pd.Timestamp(v)
            if isinstance(v, str) and _DATE_RE.search(v):
                dt = pd.to_datetime(_DATE_RE.search(v).group(0), errors="coerce")
                if pd.notna(dt):
                    return dt
    return None


def resolve_update_date(raw: pd.DataFrame) -> pd.Timestamp:
    """update 시트 데이터는 '직전 거래일(어제, 주말이면 금요일)' 기준으로 스탬프.
    Last Updated 타임스탬프는 기준일 산출의 reference + 신선도 검증/로깅 용도."""
    last_updated = _find_last_updated(raw)
    ref = last_updated if last_updated is not None else pd.Timestamp.now()
    data_date = _previous_krx_trading_day(ref)

    if last_updated is not None:
        if last_updated.normalize() < data_date:
            print(f"  ⚠️ Last Updated({last_updated:%Y-%m-%d})가 산출 기준일({data_date:%Y-%m-%d})보다 과거입니다. "
                  f"워크북이 최신으로 새로고침되지 않았을 수 있습니다.")
        print(f"  update 기준일: {data_date:%Y-%m-%d} (직전 거래일)  |  "
              f"Last Updated: {last_updated:%Y-%m-%d %H:%M}")
    else:
        print(f"  update 기준일: {data_date:%Y-%m-%d} (직전 거래일)  |  "
              f"Last Updated 셀 미발견 → 현재시각 기준")
    return data_date


# =============================================================================
# 시트 로더 (둘 다 tidy long 으로 반환)
# =============================================================================
def _detect_row(header: pd.DataFrame, predicate) -> int | None:
    """헤더 영역에서 predicate(셀)이 가장 많이 True 인 행 인덱스."""
    best, best_i = -1, None
    for i in range(len(header)):
        cnt = sum(predicate(header.iloc[i, c]) for c in range(header.shape[1]))
        if cnt > best:
            best, best_i = cnt, i
    return best_i if best > 0 else None


def load_stack_sheet(path: Path, sheet: str) -> pd.DataFrame:
    """stack 시트(날짜=행, 종목×아이템=열) → tidy long.
    헤더 행 위치는 자동 탐지(원본 워크북 / 파이썬 저장본 둘 다 대응)."""
    raw = pd.read_excel(path, sheet_name=sheet, header=None)

    # 데이터 시작 = col0이 날짜로 파싱되는 첫 행
    data_start = None
    for i in range(len(raw)):
        if pd.notna(pd.to_datetime(raw.iloc[i, 0], errors="coerce")):
            v = str(raw.iloc[i, 0])
            if _DATE_RE.search(v) or isinstance(raw.iloc[i, 0], (datetime, pd.Timestamp)):
                data_start = i
                break
    if data_start is None or data_start == 0:
        raise RuntimeError(f"[{sheet}] 시트에서 날짜 행을 찾지 못했습니다.")

    header = raw.iloc[:data_start]
    code_row = _detect_row(header, _is_code)
    item_row = _detect_row(header, lambda x: resolve_item_field(x) is not None)
    if code_row is None or item_row is None:
        raise RuntimeError(f"[{sheet}] 코드/아이템 헤더 행 탐지 실패")

    # 코드명 행: col0 == '코드명' 우선, 없으면 코드행 +1
    name_row = code_row + 1
    for i in range(len(header)):
        if str(header.iloc[i, 0]).strip() == "코드명":
            name_row = i
            break

    symbols = raw.iloc[code_row]
    names = raw.iloc[name_row]
    items = raw.iloc[item_row]

    data = raw.iloc[data_start:].copy()
    data = data.rename(columns={0: "date"})
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data = data.dropna(subset=["date"]).reset_index(drop=True)

    frames = []
    for col in range(1, raw.shape[1]):
        if not _is_code(symbols.iloc[col]):
            continue
        field = resolve_item_field(items.iloc[col])
        if field is None:
            continue
        nm = names.iloc[col]
        frames.append(pd.DataFrame({
            "date": data["date"],
            "code": str(symbols.iloc[col]).strip(),
            "name": "" if pd.isna(nm) else str(nm),
            "field": field,
            "value": _num(data[col]),
        }))
    if not frames:
        raise RuntimeError(f"[{sheet}] 파싱 가능한 종목 데이터가 없습니다.")
    return pd.concat(frames, ignore_index=True)


def load_update_sheet(path: Path, sheet: str, date_override: pd.Timestamp | None = None):
    """update 시트(종목=행, 아이템=열, 1일치) → tidy long + 기준일."""
    raw = pd.read_excel(path, sheet_name=sheet, header=None)

    data_start = None
    for i in range(len(raw)):
        if _is_code(raw.iloc[i, 0]):
            data_start = i
            break
    if data_start is None:
        raise RuntimeError(f"[{sheet}] 시트에서 종목코드(A######) 행을 찾지 못했습니다.")

    # 아이템 라벨 행: 데이터 위쪽에서 resolve 매핑이 가장 많은 행
    header = raw.iloc[:data_start]
    item_row = _detect_row(header, lambda x: resolve_item_field(x) is not None)
    if item_row is None:
        raise RuntimeError(f"[{sheet}] 아이템 라벨 행 탐지 실패")
    items = raw.iloc[item_row]

    data = raw.iloc[data_start:].reset_index(drop=True)
    codes = data.iloc[:, 0].astype(str).str.strip()
    names = data.iloc[:, 1].fillna("").astype(str)

    data_date = date_override if date_override is not None else resolve_update_date(raw)

    frames = []
    for c in range(2, raw.shape[1]):
        field = resolve_item_field(items.iloc[c])
        if field is None:
            continue
        frames.append(pd.DataFrame({
            "date": data_date,
            "code": codes,
            "name": names,
            "field": field,
            "value": _num(data.iloc[:, c]),
        }))
    if not frames:
        raise RuntimeError(f"[{sheet}] 매핑된 아이템 컬럼이 없습니다. (라벨 확인 필요)")

    long_df = pd.concat(frames, ignore_index=True)
    long_df = long_df[long_df["code"].map(_is_code)].reset_index(drop=True)
    return long_df, pd.Timestamp(data_date)


# =============================================================================
# 적재 (합집합 → 중복제거 → trim → 워크북 stack 시트 + 별도 백업파일 저장)
# =============================================================================
def _build_stack_wide_rows(merged: pd.DataFrame) -> list[list]:
    """누적 long → stack wide 행 리스트.
    [헤더3행] 코드 / 코드명 / 아이템명  +  [날짜행들] yyyy-mm-dd + 값들.
    (load_stack_sheet 가 그대로 다시 읽을 수 있는 포맷)"""
    name_map = merged.sort_values("date").groupby("code")["name"].last()
    codes = sorted(merged["code"].unique())

    wide = merged.pivot_table(index="date", columns=["code", "field"],
                              values="value", aggfunc="last").sort_index()
    col_tuples = [(c, f) for c in codes for f in FIELD_ORDER if (c, f) in wide.columns]
    wide = wide.reindex(columns=pd.MultiIndex.from_tuples(col_tuples))

    n_cols = len(col_tuples) + 1
    if n_cols > 16384:
        print(f"  ⚠️ stack 컬럼 수 {n_cols:,}개가 Excel 한계(16,384)를 초과합니다. "
              f"FIELD_ORDER 축소 또는 종목 필터가 필요할 수 있습니다.")

    rows = [
        ["코드"]     + [c for c, _ in col_tuples],
        ["코드명"]   + [str(name_map.get(c, "")) for c, _ in col_tuples],
        ["아이템명"] + [FIELD_LABEL[f] for _, f in col_tuples],
    ]
    for dt, r in wide.iterrows():
        rows.append([pd.Timestamp(dt).strftime("%Y-%m-%d")] + list(r.values))
    return rows


def _write_stack_wide(merged: pd.DataFrame, stack_path: Path) -> None:
    """누적 long → stack wide 별도 백업파일(STACK_FILE)에 저장."""
    rows = _build_stack_wide_rows(merged)
    out = pd.DataFrame(rows)
    stack_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(stack_path, engine="openpyxl") as w:
        out.to_excel(w, sheet_name=STACK_FILE_SHEET, index=False, header=False)


def write_stack_into_workbook(merged: pd.DataFrame, workbook_path: Path,
                              stack_sheet: str) -> bool:
    """누적 데이터를 ▶ live 워크북의 stack 시트에 직접 기록 ◀.
    누적 안전망은 별도 파일(STACK_FILE)이 담당하므로 워크북 자동 백업은 생성하지 않음."""
    from openpyxl import load_workbook

    rows = _build_stack_wide_rows(merged)
    try:
        wb = load_workbook(workbook_path)  # data_only=False → 수식 보존
        sheet_idx = (wb.sheetnames.index(stack_sheet)
                     if stack_sheet in wb.sheetnames else len(wb.sheetnames))
        if stack_sheet in wb.sheetnames:
            wb.remove(wb[stack_sheet])             # 기존 stack 시트 제거(자리 보존)
        ws = wb.create_sheet(stack_sheet, sheet_idx)  # 같은 위치에 재생성
        for r in rows:
            ws.append([_xl_safe(v) for v in r])
        wb.save(workbook_path)
        n_days = len(rows) - 3
        print(f"  워크북 stack 시트 갱신: {workbook_path.name} "
              f"(시트 '{stack_sheet}', {n_days}거래일 기록)")
        return True
    except PermissionError:
        print(f"  ⚠️ 워크북 기록 실패: 파일이 Excel에서 열려 있습니다. Excel을 닫고 다시 실행하세요.")
        print(f"     (별도 누적파일은 정상 저장됨: {STACK_FILE.name})")
        return False
    except Exception as e:
        print(f"  ⚠️ 워크북 stack 기록 실패: {e}")
        print(f"     (별도 누적파일은 정상 저장됨: {STACK_FILE.name})")
        return False


def accumulate_and_persist(stack_long: pd.DataFrame, update_long: pd.DataFrame,
                           keep_days: int, stack_path: Path) -> pd.DataFrame:
    """누적 + 오늘치 → (날짜,코드,항목) 중복제거(update 우선) → 최근 keep_days 거래일 trim → 백업파일 저장."""
    parts = [p for p in (stack_long, update_long) if p is not None and not p.empty]
    merged = pd.concat(parts, ignore_index=True)  # update 가 뒤 → keep='last' 시 우선
    merged["date"] = pd.to_datetime(merged["date"])
    merged = merged.dropna(subset=["date", "code", "field"])
    merged = merged.drop_duplicates(subset=["date", "code", "field"], keep="last")

    keep_dates = sorted(merged["date"].unique())[-keep_days:]
    merged = merged[merged["date"].isin(keep_dates)].reset_index(drop=True)

    _write_stack_wide(merged, stack_path)
    n_days = merged["date"].nunique()
    print(f"  누적 백업파일 저장: {stack_path.name}  | {merged['code'].nunique():,}종목 × {n_days}거래일 "
          f"({pd.Timestamp(min(keep_dates)):%Y-%m-%d} ~ {pd.Timestamp(max(keep_dates)):%Y-%m-%d})")
    return merged


def long_to_wide(long_df: pd.DataFrame) -> pd.DataFrame:
    """tidy long → (date,code,name) 행 / field 열 wide. (기존 파이프라인 입력 형태)"""
    name_map = long_df.sort_values("date").groupby("code")["name"].last()
    long_df = long_df.copy()
    long_df["name"] = long_df["code"].map(name_map).fillna("")

    wide = (long_df.pivot_table(index=["date", "code", "name"], columns="field",
                                values="value", aggfunc="first")
            .reset_index().sort_values(["code", "date"]))
    wide.columns.name = None

    for f in FIELD_ORDER:           # 누락 필드는 NaN 컬럼으로 보강(파이프라인 KeyError 방지)
        if f not in wide.columns:
            wide[f] = np.nan

    loaded = [c for c in wide.columns if c not in {"date", "code", "name"}]
    print(f"  파싱 종목 수: {wide['code'].nunique():,}개")
    print(f"  데이터 기간 : {wide['date'].min():%Y-%m-%d} ~ {wide['date'].max():%Y-%m-%d}")
    print(f"  로드된 필드 : {sorted(loaded)}")
    return wide


# =============================================================================
# (이하 시계열 피처 / FDR 오버레이 / 스코어링 / 리포트 — 기존 로직 유지)
# =============================================================================
def add_time_series_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values(["code", "date"]).copy()
    grouped = df.groupby("code", group_keys=False)
    df["ma5"] = grouped["close"].transform(lambda s: s.rolling(5, min_periods=3).mean())
    df["ma10"] = grouped["close"].transform(lambda s: s.rolling(10, min_periods=5).mean())
    df["ma20"] = grouped["close"].transform(lambda s: s.rolling(20, min_periods=10).mean())

    if "avg_ret_5d" in df.columns:
        df["ret_5d"] = pd.to_numeric(df["avg_ret_5d"], errors="coerce")
    else:
        df["ret_5d"] = grouped["close"].pct_change(5) * 100
    if "avg_ret_20d" in df.columns:
        df["ret_20d"] = pd.to_numeric(df["avg_ret_20d"], errors="coerce")
    else:
        df["ret_20d"] = grouped["close"].pct_change(20) * 100
    df["ret_1d"] = grouped["close"].pct_change(1) * 100

    df["vol_20d"] = (
        grouped["close"].transform(lambda s: s.pct_change().rolling(20, min_periods=5).std()) * 100
    )
    df["risk_adj_mom"] = df["ret_20d"] / df["vol_20d"].replace(0, np.nan)

    trend_align = (
        0.5 * np.tanh(8 * (df["close"] / df["ma10"] - 1))
        + 0.5 * np.tanh(8 * (df["ma5"] / df["ma10"] - 1))
    )
    df["trend_align_score"] = ((trend_align + 1) * 50).clip(0, 100)

    # 수급은 만원 단위, 거래대금은 원 단위
    df["inst_ratio_5d"] = (df["inst_net_5d"] * 10_000) / (df["turnover_avg_5d"] * 5) * 100
    df["foreign_ratio_5d"] = (df["foreign_net_5d"] * 10_000) / (df["turnover_avg_5d"] * 5) * 100
    df["inst_ratio_20d"] = (df["inst_net_20d"] * 10_000) / (df["turnover_avg_20d"] * 20) * 100
    df["foreign_ratio_20d"] = (df["foreign_net_20d"] * 10_000) / (df["turnover_avg_20d"] * 20) * 100

    df["inst_positive_rate_20d"] = grouped["inst_ratio_5d"].transform(
        lambda s: (s > 0).rolling(20, min_periods=5).mean() * 100
    )
    df["foreign_positive_rate_20d"] = grouped["foreign_ratio_5d"].transform(
        lambda s: (s > 0).rolling(20, min_periods=5).mean() * 100
    )
    df["turnover_rvol"] = df["turnover_avg_5d"] / df["turnover_avg_20d"]
    df["price_data_date"] = df["date"]
    df["price_data_source"] = "Excel"
    df["today_price_applied"] = False
    df["volume"] = np.nan
    df["turnover"] = np.nan
    df["volume_avg_5d"] = np.nan
    df["volume_avg_20d"] = np.nan
    return df.replace([np.inf, -np.inf], np.nan)


def _fetch_fdr_price_metrics(code: str) -> dict | None:
    try:
        import FinanceDataReader as fdr
        numeric_code = str(code).replace("A", "", 1)
        end = datetime.now()
        start = end - timedelta(days=FDR_LOOKBACK_DAYS)
        df = fdr.DataReader(numeric_code, start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
        if df is None or df.empty:
            return None
        df = df.reset_index()
        date_col = "Date" if "Date" in df.columns else df.columns[0]
        df = df.rename(columns={date_col: "date"})
        required = {"date", "Close", "Volume"}
        if not required.issubset(df.columns):
            return None
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
        df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce")
        df = df.dropna(subset=["date", "Close", "Volume"]).sort_values("date")
        if df.empty:
            return None

        close = df["Close"]
        volume = df["Volume"]
        turnover = close * volume
        latest = df.iloc[-1]
        latest_turnover = float(turnover.iloc[-1])

        def tail_mean(series: pd.Series, days: int, min_periods: int):
            if len(series.dropna()) < min_periods:
                return np.nan
            return float(series.tail(days).mean())

        turnover_avg_5d = tail_mean(turnover, 5, 3)
        turnover_avg_20d = tail_mean(turnover, 20, 10)
        volume_avg_5d = tail_mean(volume, 5, 3)
        volume_avg_20d = tail_mean(volume, 20, 10)

        def pct_change(days: int):
            if len(close) <= days or pd.isna(close.iloc[-days - 1]) or close.iloc[-days - 1] == 0:
                return np.nan
            return (close.iloc[-1] / close.iloc[-days - 1] - 1) * 100

        ret_1d = pct_change(1)
        ret_5d = pct_change(5)
        ret_20d = pct_change(20)
        daily_ret = close.pct_change()
        vol_20d = float(daily_ret.tail(20).std() * 100) if daily_ret.notna().sum() >= 5 else np.nan
        risk_adj_mom = ret_20d / vol_20d if pd.notna(vol_20d) and vol_20d else np.nan

        ma5 = tail_mean(close, 5, 3)
        ma10 = tail_mean(close, 10, 5)
        ma20 = tail_mean(close, 20, 10)
        latest_close = float(latest["Close"])
        if pd.notna(ma10) and ma10:
            trend_align = (
                0.5 * np.tanh(8 * (latest_close / ma10 - 1))
                + 0.5 * (np.tanh(8 * (ma5 / ma10 - 1)) if pd.notna(ma5) else 0.0)
            )
            trend_align_score = float(np.clip((trend_align + 1) * 50, 0, 100))
        else:
            trend_align_score = np.nan

        return {
            "code": code,
            "price_data_date": latest["date"],
            "close": latest_close,
            "volume": float(latest["Volume"]),
            "turnover": latest_turnover,
            "turnover_avg_5d": turnover_avg_5d,
            "turnover_avg_20d": turnover_avg_20d,
            "volume_avg_5d": volume_avg_5d,
            "volume_avg_20d": volume_avg_20d,
            "turnover_rvol": latest_turnover / turnover_avg_20d if pd.notna(turnover_avg_20d) and turnover_avg_20d else np.nan,
            "ret_1d": ret_1d, "ret_5d": ret_5d, "ret_20d": ret_20d,
            "vol_20d": vol_20d, "risk_adj_mom": risk_adj_mom,
            "ma5": ma5, "ma10": ma10, "ma20": ma20,
            "trend_align_score": trend_align_score,
        }
    except Exception:
        return None


def apply_fdr_price_overlay(df: pd.DataFrame) -> pd.DataFrame:
    excel_latest_by_code = (
        df.sort_values(["code", "date"]).groupby("code", as_index=False).tail(1).set_index("code")
    )
    codes = excel_latest_by_code.index.tolist()
    if not codes:
        return df

    print("\n[당일 가격/거래량 확인]")
    print(f"FinanceDataReader 조회 대상: {len(codes):,}개 종목")

    metrics = []
    completed = 0
    for code in codes:
        completed += 1
        try:
            result = _fetch_fdr_price_metrics(code)
        except Exception as exc:
            print(f"[FDR] {completed}/{len(codes)} {code}: fetch error: {exc}")
            continue
        if result is not None:
            metrics.append(result)
        if completed % 10 == 0 or completed == len(codes):
            print(f"[FDR] {completed}/{len(codes)}", flush=True)
    print()

    if not metrics:
        print("  FDR 데이터를 가져오지 못해 엑셀 기준 데이터를 사용합니다.")
        return df

    overlay_rows = []
    today = datetime.now().date()
    for item in metrics:
        code = item["code"]
        if code not in excel_latest_by_code.index:
            continue
        base = excel_latest_by_code.loc[code].copy()
        base_date = pd.to_datetime(base["date"])
        price_date = pd.to_datetime(item["price_data_date"])
        if price_date <= base_date:
            continue
        base["code"] = code
        for key, value in item.items():
            if key != "code":
                base[key] = value
        base["date"] = price_date
        base["price_data_source"] = "FinanceDataReader"
        base["today_price_applied"] = price_date.date() == today
        for period, turnover_col in [(5, "turnover_avg_5d"), (20, "turnover_avg_20d")]:
            if pd.notna(base.get(turnover_col)) and base.get(turnover_col) != 0:
                base[f"inst_ratio_{period}d"] = (
                    pd.to_numeric(base.get(f"inst_net_{period}d"), errors="coerce") * 10_000
                    / (base.get(turnover_col) * period) * 100
                )
                base[f"foreign_ratio_{period}d"] = (
                    pd.to_numeric(base.get(f"foreign_net_{period}d"), errors="coerce") * 10_000
                    / (base.get(turnover_col) * period) * 100
                )
        overlay_rows.append(base)

    if not overlay_rows:
        print("  엑셀 기준일보다 최신인 FDR 데이터가 없어 엑셀 기준 데이터를 사용합니다.")
        return df

    overlay_df = pd.DataFrame(overlay_rows).reset_index(drop=True)
    print(f"  당일 가격/거래량 반영: {len(overlay_df):,}개 종목 "
          f"(당일 날짜 일치: {int(overlay_df['today_price_applied'].sum()):,}개)")
    return pd.concat([df, overlay_df], ignore_index=True, sort=False).replace([np.inf, -np.inf], np.nan)


def latest_scoring_frame(df: pd.DataFrame) -> pd.DataFrame:
    latest = df.sort_values(["code", "date"]).groupby("code", as_index=False).tail(1).copy()
    latest["above_ma10"] = latest["close"] > latest["ma10"]
    latest["ma5_above_ma10"] = latest["ma5"] > latest["ma10"]
    latest["both_supply_positive"] = (latest["inst_ratio_5d"] > 0) & (latest["foreign_ratio_5d"] > 0)
    negative_intraday_return_mask = pd.to_numeric(latest["ret_1d"], errors="coerce").lt(0)

    latest["price_score"] = (
        normalize_score(latest["ret_20d"]).clip(2, 98) * 0.40
        + normalize_score(latest["risk_adj_mom"]).clip(2, 98) * 0.25
        + normalize_score(latest["ret_5d"]).clip(2, 98) * 0.20
        + latest["trend_align_score"].fillna(50).clip(0, 100) * 0.15
    ).round(1)
    latest.loc[negative_intraday_return_mask, "price_score"] = (
        latest.loc[negative_intraday_return_mask, "price_score"] * 0.75
    ).round(1)

    latest["supply_score"] = (
        normalize_score(latest["foreign_ratio_5d"]) * 0.25
        + normalize_score(latest["foreign_ratio_20d"]) * 0.15
        + normalize_score(latest["foreign_positive_rate_20d"]) * 0.10
        + normalize_score(latest["inst_ratio_5d"]) * 0.25
        + normalize_score(latest["inst_ratio_20d"]) * 0.15
        + normalize_score(latest["inst_positive_rate_20d"]) * 0.05
        + latest["both_supply_positive"].astype(float) * 100 * 0.05
    ).round(1)

    latest["value_score"] = (
        normalize_score(latest["turnover_rvol"]) * 0.70
        + normalize_score(latest["turnover"].fillna(latest["turnover_avg_5d"])) * 0.30
    ).round(1)
    latest.loc[negative_intraday_return_mask, "value_score"] = (
        latest.loc[negative_intraday_return_mask, "value_score"] * 0.5
    ).round(1)

    latest["total_score"] = (
        latest["price_score"] * PRICE_WEIGHT
        + latest["supply_score"] * SUPPLY_WEIGHT
        + latest["value_score"] * VALUE_WEIGHT
    ).round(1)
    latest["grade"] = pd.cut(
        latest["total_score"], bins=[float("-inf"), 30, 50, 65, 80, float("inf")],
        labels=["F", "D", "C", "B", "A"], right=False,
    ).astype(str)
    return latest.sort_values("total_score", ascending=False, na_position="last")


def _format_table_sheet(ws, freeze_row: int = 2) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    navy, white = "1F3864", "FFFFFF"
    border = Border(left=Side(style="thin", color="D9E2F3"), right=Side(style="thin", color="D9E2F3"),
                    top=Side(style="thin", color="D9E2F3"), bottom=Side(style="thin", color="D9E2F3"))
    ws.freeze_panes = f"A{freeze_row}"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color=white, name="맑은 고딕", size=9)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 80)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 24)


def _create_report_sheet(writer, final, scored, price_ref_date, supply_ref_date) -> None:
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = writer.book
    ws = wb.create_sheet("전종목 보고서", 0)
    ws.sheet_view.showGridLines = False
    color_navy, color_blue, color_light = "1F3864", "2E75B6", "D6E4F0"
    color_white, color_gold, color_green = "FFFFFF", "C9A84C", "70AD47"
    color_orange, color_gray, color_red = "ED7D31", "F2F2F2", "C00000"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="맑은 고딕")
    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))

    for col, width in {1: 2, 2: 6, 3: 14, 4: 18, 5: 11, 6: 9, 7: 10, 8: 10, 9: 12, 10: 11,
                       11: 12, 12: 12, 13: 3, 14: 18, 15: 10, 16: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, h in {1: 6, 2: 42, 3: 23, 4: 8, 5: 8, 6: 28, 7: 22, 8: 10, 9: 8}.items():
        ws.row_dimensions[r].height = h

    def style_range(rng, fc, fo, al=None):
        al = al or center
        for row in ws[rng]:
            for cell in row:
                cell.fill = fill(fc); cell.font = fo; cell.alignment = al; cell.border = border

    ws.merge_cells("B2:L2"); style_range("B2:L2", color_navy, font(True, color_white, 18))
    ws["B2"] = "전종목 수급 스코어링 보고서"
    ws.merge_cells("B3:L3"); style_range("B3:L3", color_blue, font(False, color_white, 10))
    ws["B3"] = (f"가격/거래량 기준일: {price_ref_date}  |  수급 기준일: {supply_ref_date}  |  "
                f"분석 대상: {len(final):,}개 종목  |  가격/수급/거래대금: "
                f"{PRICE_WEIGHT:.0%}/{SUPPLY_WEIGHT:.0%}/{VALUE_WEIGHT:.0%}")

    grade_counts = {g: 0 for g in ["A", "B", "C", "D", "F"]}
    for grade, count in final["등급"].value_counts().items():
        if grade in grade_counts:
            grade_counts[grade] = int(count)
    score_65 = int((pd.to_numeric(final["종합스코어"], errors="coerce") >= 65).sum())
    score_80 = int((pd.to_numeric(final["종합스코어"], errors="coerce") >= 80).sum())
    cards = [("전체 종목", len(final), color_blue), ("A등급", grade_counts["A"], color_gold),
             ("65점 이상", score_65, color_green), ("80점 이상", score_80, color_navy),
             ("당일 가격", int(final["당일가격반영"].fillna(False).sum()), color_orange)]
    for idx, (label, value, color) in enumerate(cards):
        col = 2 + idx * 2
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        for cell in ws.iter_rows(min_row=6, max_row=6, min_col=col, max_col=col + 1):
            for c in cell:
                c.fill = fill(color); c.font = font(True, color_white, 9); c.alignment = center; c.border = border
        for cell in ws.iter_rows(min_row=7, max_row=7, min_col=col, max_col=col + 1):
            for c in cell:
                c.fill = fill(color_gray); c.font = font(True, color, 18); c.alignment = center; c.border = border
        ws.cell(row=6, column=col).value = label
        ws.cell(row=7, column=col).value = value

    ws.merge_cells("B10:L10"); style_range("B10:L10", color_blue, font(True, color_white, 11), left_center)
    ws["B10"] = "[Top] 추천종목  (종합스코어 65점 이상)"
    ws["B10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    headers = ["순위", "코드", "종목명", "종합", "등급", "가격", "수급", "거래대금", "1주수익률", "1개월수익률", "기관/외인"]
    for offset, header in enumerate(headers, start=2):
        cell = ws.cell(row=11, column=offset)
        cell.value = header; cell.font = font(True, color_white, 9); cell.fill = fill(color_navy)
        cell.alignment = center; cell.border = border

    top = final[pd.to_numeric(final["종합스코어"], errors="coerce") >= 65].head(40).copy()
    grade_colors = {"A": color_gold, "B": color_green, "C": "FFC000", "D": color_orange, "F": "FF6B6B"}
    for idx, (_, row) in enumerate(top.iterrows(), start=1):
        er = 11 + idx
        row_fill = fill(color_white if idx % 2 else color_gray)
        data = [idx, row.get("코드", ""), row.get("종목명", ""), row.get("종합스코어", ""),
                row.get("등급", ""), row.get("가격스코어", ""), row.get("수급스코어", ""),
                row.get("거래대금스코어", ""), row.get("1주수익률(%)", ""), row.get("1개월수익률(%)", ""),
                "동시순매수" if bool(row.get("기관외인동시순매수", False)) else ""]
        for offset, value in enumerate(data, start=2):
            cell = ws.cell(row=er, column=offset)
            cell.value = value; cell.font = font(False, "000000", 9); cell.fill = row_fill
            cell.alignment = center; cell.border = border
        gc = ws.cell(row=er, column=6)
        grade = str(row.get("등급", "-"))
        gc.fill = fill(grade_colors.get(grade, "BFBFBF")); gc.font = font(True, color_white, 9)

    chart_row, cl, cv = 10, 14, 15
    ws.cell(row=chart_row, column=cl).value = "구분"
    ws.cell(row=chart_row, column=cv).value = "종목수"
    for idx, (label, value) in enumerate([("강세 (A+B)", grade_counts["A"] + grade_counts["B"]),
                                          ("중립 (C)", grade_counts["C"]),
                                          ("약세 (D+F)", grade_counts["D"] + grade_counts["F"])], start=1):
        ws.cell(row=chart_row + idx, column=cl).value = label
        ws.cell(row=chart_row + idx, column=cv).value = value
    for row in range(chart_row, chart_row + 4):
        for col in [cl, cv]:
            cell = ws.cell(row=row, column=col)
            cell.font = font(row == chart_row, color_white if row == chart_row else "000000", 9)
            cell.fill = fill(color_navy if row == chart_row else color_light)
            cell.alignment = center; cell.border = border
    pie = PieChart(); pie.title = "전종목 스코어 분포"; pie.style = 10
    pie.add_data(Reference(ws, min_col=cv, min_row=chart_row, max_row=chart_row + 3), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=cl, min_row=chart_row + 1, max_row=chart_row + 3))
    for idx, color in enumerate([color_blue, color_orange, color_red]):
        pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = color; pie.series[0].dPt.append(pt)
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True; pie.dataLabels.showCatName = True
    pie.width, pie.height = 12, 9
    ws.add_chart(pie, "N15")


def save_outputs(scored, history, sheet, supply_ref_date):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / f"전종목_수급_스코어링_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_cols = [
        "date", "code", "name", "is_etf", "total_score", "grade", "price_score", "supply_score", "value_score",
        "price_data_date", "price_data_source", "today_price_applied",
        "close", "ret_1d", "ret_5d", "ret_20d", "risk_adj_mom", "vol_20d", "trend_align_score",
        "ma5", "ma10", "ma20", "above_ma10", "ma5_above_ma10",
        "volume", "volume_avg_5d", "volume_avg_20d",
        "turnover", "turnover_avg_5d", "turnover_avg_20d", "turnover_rvol",
        "inst_net_5d", "inst_net_20d", "foreign_net_5d", "foreign_net_20d",
        "inst_ratio_5d", "inst_ratio_20d", "foreign_ratio_5d", "foreign_ratio_20d",
        "inst_positive_rate_20d", "foreign_positive_rate_20d", "both_supply_positive",
    ]
    for col in output_cols:
        if col not in scored.columns:
            scored[col] = pd.NA
    rename = {
        "date": "기준일", "code": "코드", "name": "종목명", "is_etf": "구분",
        "total_score": "종합스코어", "grade": "등급",
        "price_score": "가격스코어", "supply_score": "수급스코어", "value_score": "거래대금스코어",
        "price_data_date": "가격/거래량기준일", "price_data_source": "가격데이터소스", "today_price_applied": "당일가격반영",
        "close": "종가", "ret_1d": "당일수익률(%)", "ret_5d": "1주수익률(%)", "ret_20d": "1개월수익률(%)",
        "risk_adj_mom": "위험조정모멘텀", "vol_20d": "일간변동성20일(%)", "trend_align_score": "추세정렬점수",
        "ma5": "5일선", "ma10": "10일선", "ma20": "20일선", "above_ma10": "10일선상회", "ma5_above_ma10": "5일선>10일선",
        "volume": "당일거래량", "volume_avg_5d": "거래량5일평균", "volume_avg_20d": "거래량20일평균",
        "turnover": "당일거래대금(원)", "turnover_avg_5d": "거래대금5일평균(원)", "turnover_avg_20d": "거래대금20일평균(원)",
        "turnover_rvol": "당일거래대금/20일평균",
        "inst_net_5d": "기관5일순매수(만원)", "inst_net_20d": "기관20일순매수(만원)",
        "foreign_net_5d": "외국인5일순매수(만원)", "foreign_net_20d": "외국인20일순매수(만원)",
        "inst_ratio_5d": "기관5일순매수/거래대금(%)", "inst_ratio_20d": "기관20일순매수/거래대금(%)",
        "foreign_ratio_5d": "외국인5일순매수/거래대금(%)", "foreign_ratio_20d": "외국인20일순매수/거래대금(%)",
        "inst_positive_rate_20d": "기관수급양수비율20일(%)", "foreign_positive_rate_20d": "외국인수급양수비율20일(%)",
        "both_supply_positive": "기관외인동시순매수",
    }
    final = scored[output_cols].rename(columns=rename)
    numeric_cols = final.select_dtypes(include=["number"]).columns
    final[numeric_cols] = final[numeric_cols].round(2)
    recommend = final[pd.to_numeric(final["종합스코어"], errors="coerce") >= 65].copy()
    price_ref_date = pd.to_datetime(scored["price_data_date"], errors="coerce").max().strftime("%Y-%m-%d")
    supply_ref_date_str = pd.to_datetime(supply_ref_date).strftime("%Y-%m-%d")
    summary = pd.DataFrame([
        {"항목": "입력파일", "값": str(INPUT_FILE)},
        {"항목": "사용시트", "값": sheet},
        {"항목": "가격/거래량 기준일", "값": price_ref_date},
        {"항목": "수급 기준일", "값": supply_ref_date_str},
        {"항목": "당일 가격 반영 종목수", "값": int(final["당일가격반영"].fillna(False).sum())},
        {"항목": "대상종목수", "값": len(final)},
        {"항목": "65점 이상", "값": len(recommend)},
        {"항목": "가격/수급/거래대금 가중치", "값": f"{PRICE_WEIGHT:.0%}/{SUPPLY_WEIGHT:.0%}/{VALUE_WEIGHT:.0%}"},
    ])
    a_grade = final[final["등급"] == "A"].copy()
    foreign_top = final.sort_values("외국인5일순매수/거래대금(%)", ascending=False, na_position="last").head(50)
    inst_top = final.sort_values("기관5일순매수/거래대금(%)", ascending=False, na_position="last").head(50)
    value_top = final.sort_values("당일거래대금/20일평균", ascending=False, na_position="last").head(50)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _create_report_sheet(writer, final, scored, price_ref_date, supply_ref_date_str)
        recommend.to_excel(writer, sheet_name="추천종목", index=False)
        a_grade.to_excel(writer, sheet_name="A등급", index=False)
        final.to_excel(writer, sheet_name="전체", index=False)
        foreign_top.to_excel(writer, sheet_name="외국인순매수TOP", index=False)
        inst_top.to_excel(writer, sheet_name="기관순매수TOP", index=False)
        value_top.to_excel(writer, sheet_name="거래대금TOP", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        for sheet_name in ["추천종목", "A등급", "전체", "외국인순매수TOP", "기관순매수TOP", "거래대금TOP", "요약"]:
            _format_table_sheet(writer.sheets[sheet_name])
    return output


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(INPUT_FILE)

    date_override = None
    if len(sys.argv) > 1:
        date_override = pd.to_datetime(sys.argv[1], errors="coerce")
        if pd.isna(date_override):
            print(f"날짜 인자 무시(파싱 실패): {sys.argv[1]}")
            date_override = None

    print("=" * 60)
    print("전종목 수급 스코어링 (update → stack 적재 구조)")
    print("=" * 60)

    # --- 1) 오늘 새로고침된 update 읽기 + 직전 거래일로 날짜 스탬프 ---
    print("\n[1] update 시트 로드")
    update_long, data_date = load_update_sheet(INPUT_FILE, UPDATE_SHEET, date_override)
    print(f"  오늘치 종목 수: {update_long['code'].nunique():,}개  (기준일 {data_date:%Y-%m-%d})")

    # --- 2) 기존 누적 읽기 (워크북 stack 시트 ∪ 별도 백업파일) ---
    print("\n[2] 기존 누적 로드")
    stack_parts = []
    try:
        stack_parts.append(load_stack_sheet(INPUT_FILE, STACK_SHEET))
        print(f"  워크북 stack 시트 로드")
    except Exception as e:
        print(f"  워크북 stack 시트 없음/실패({e}) → 무시")
    if STACK_FILE.exists():
        try:
            stack_parts.append(load_stack_sheet(STACK_FILE, STACK_FILE_SHEET))
            print(f"  백업파일 로드: {STACK_FILE.name}")
        except Exception as e:
            print(f"  백업파일 로드 실패({e}) → 무시")
    stack_long = pd.concat(stack_parts, ignore_index=True) if stack_parts else pd.DataFrame(
        columns=["date", "code", "name", "field", "value"])

    # --- 2-b) 거래일 공백 검사 (빠진 날이 있으면 출력하고 중단) ---
    stack_last_date = None
    if not stack_long.empty:
        sd = pd.to_datetime(stack_long["date"], errors="coerce").dropna()
        if not sd.empty:
            stack_last_date = sd.max().normalize()

    gap_days = detect_stack_gap(stack_last_date, data_date)
    if gap_days:
        print("\n" + "=" * 60)
        print("⛔ 거래일 공백 감지 — 분석을 중단합니다.")
        print("=" * 60)
        print(f"  stack 마지막 날짜 : {stack_last_date:%Y-%m-%d}")
        print(f"  이번 update 기준일: {data_date:%Y-%m-%d}")
        print(f"  누락된 거래일 ({len(gap_days)}일):")
        for d in gap_days:
            print(f"    - {pd.Timestamp(d):%Y-%m-%d}")
        print("-" * 60)
        print("  해결: 같은 폴더의 fill_data.xlsx 에 위 날짜들을 시트로 채운 뒤")
        print("        (시트명 = YYYYMMDD 형식, 각 시트는 update 시트와 동일한 구성)")
        print("        python fill_data.py 를 실행해 전종목_수급.xlsx 의 stack 시트를 보강하세요.")
        print("        그 후 이 스크립트를 다시 실행하면 정상 진행됩니다.")
        print("=" * 60)
        sys.exit(1)

    # --- 2-c) 단말 미갱신 검사 (update가 stack 최신과 내용 동일 → 중단) ---
    if update_equals_stack_latest(update_long, stack_long):
        last_date = pd.to_datetime(stack_long["date"], errors="coerce").max()
        print("\n" + "=" * 60)
        print("⛔ update 내용이 stack 최신 데이터와 동일합니다 — 분석을 중단합니다.")
        print("=" * 60)
        print(f"   stack 최신 날짜   : {last_date:%Y-%m-%d}")
        print(f"   update 산출 기준일: {data_date:%Y-%m-%d}")
        print("-" * 60)
        print("   전종목_수급.xlsx 의 update 시트가 아직 새로고침되지 않았습니다.")
        print("   (같은 데이터를 새 날짜로 중복 적재하지 않도록 멈춥니다.)")
        print("   해결: 전종목_수급.xlsx 를 열어 update 시트를 Refresh(데이터가이드 새로고침)")
        print("         → 저장 → Excel 닫기 → 이 스크립트를 다시 실행하세요.")
        print("=" * 60)
        sys.exit(1)

    # --- 3) 적재(중복제거+trim) → 백업파일 저장 ---
    print("\n[3] 적재 / 중복제거 / trim")
    merged_long = accumulate_and_persist(stack_long, update_long, KEEP_TRADING_DAYS, STACK_FILE)

    # --- 3-b) 워크북 stack 시트에 직접 적재 (요청 동작) ---
    if WRITE_BACK_TO_WORKBOOK:
        print("\n[3-b] update → 워크북 stack 시트 적재")
        write_stack_into_workbook(merged_long, INPUT_FILE, STACK_SHEET)

    # --- 4) wide 변환 후 기존 파이프라인 ---
    print("\n[4] 스코어링 파이프라인")
    wide = long_to_wide(merged_long)
    history = add_time_series_features(wide)
    supply_ref_date = history["date"].max()
    history = apply_fdr_price_overlay(history)
    scored = latest_scoring_frame(history)

    # ETF 판별 (스코어링은 전체 한 풀에서 그대로, 식별용 '구분' 컬럼만 부여)
    etf_flags = classify_etf(scored["code"], scored["name"])
    scored["is_etf"] = scored["code"].map(etf_flags).map({True: "ETF", False: "개별"})
    n_etf = int((scored["is_etf"] == "ETF").sum())
    print(f"  ETF/개별 구분: ETF {n_etf}개 / 개별 {len(scored) - n_etf}개 (전체 동일 풀에서 스코어링)")

    output = save_outputs(scored, history, "stack", supply_ref_date)

    print("\n[TOP 20]")
    cols = ["date", "code", "name", "total_score", "grade", "price_score", "supply_score", "value_score"]
    print(scored[cols].head(20).to_string(index=False))
    print(f"\n저장 완료: {output}")
    print(f"백업 파일 : {STACK_FILE}")


if __name__ == "__main__":
    main()